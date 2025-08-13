# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook

path = "C:\\Users\\fucha\\Desktop\\网盘地址表格.xlsx"
wb = load_workbook(path)
ws = wb.active

#  邮件版
def process_string1(s):
    chinese_pattern = r'《([^《》]+)》'
    size_pattern = r'(\d+[a-zA-Z])'
    pattern = r'《([^《》]+)》([^，]+)，'
    if not s.strip():
        return None, None, None
    try:
        if '：' in s:
            parts = s.split('：')
            link_with_extra = parts[1]
            link = link_with_extra.split(' 提取码')[0]
            size_result = re.search(size_pattern, s)
            title = re.findall(chinese_pattern, s)[0]
            size_value = size_result.group(1)
            first_digit = int(size_value[0]) if size_value[0].isdigit() else None
            if first_digit is not None and first_digit not in [4, 8]:
                result = re.search(pattern, s)
                size_value = result.group(2).strip()
            return link, size_value, title
        else:
            parts = s.split('，')
            link_with_extra = parts[1]
            link = link_with_extra.split(' 提取码')[0]
            size = re.search(size_pattern, s)
            if size:
                size = size.group(1)
            else:
                return None, None, None
            title = re.findall(chinese_pattern, s)[0]
            return link, size, title
    except IndexError:
        return None, None, None
#  微信群版
def process_string2(s):
    chinese_pattern = r'《([^《》]+)》'
    size_pattern = r'【([^【】]+)】'
    if not s.strip():
        return None, None, None
    title = re.search(chinese_pattern, s).group(1)
    size = re.search(size_pattern, s).group(1)
    parts = s.split('】')
    link = parts[1].strip()
    return link, size, title
# 判断插入位置
def determine_column(size):
    if not size:
        return None
    size = size.lower()
    if size == '4k':
        return 5
    elif size == '4m':
        return 3
    elif size == '8m':
        return 2
    elif size == '物料链接':
        return 4
    elif size == '正片介质':
        return 5
    else:
        return None
# 存入表格
def save_to_excel(link, title, column):
    titles = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
    if title in titles:
        if column is not None:
            row_index = titles.index(title) + 1
            ws.cell(row=row_index, column=column).value = link
    else:
        new_row = ws.max_row + 1
        if column is not None:
            ws.cell(row=new_row, column=column).value = link
        ws.cell(row=new_row, column=1).value = title

with open("C:\\Users\\fucha\\Desktop\\网盘地址模板（微信群版）.txt", "r", encoding="utf-8") as f:
    for line in f:
        # link, size, title = process_string1(line)
        link, size, title = process_string2(line)
        column = determine_column(size)
        save_to_excel(link, title, column)
        if link is not None and size is not None and title is not None:
            print(f"标题：{title}")
            print(f"格式：{size}")
            print(f"链接：{link}")
            print()
    wb.save(path)