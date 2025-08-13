import openpyxl
from openpyxl import load_workbook

#   形式：《介质名称》，url

def type1_safe(file_path, input_str):
    wb = load_workbook(file_path)
    sheet = wb.active
    quality_type = input("请输入质量类型（4M/8M/4K）：")
    while True:
        try:
            parts = input_str.split('，https://')
            title = parts[0].strip('《》')
            url = 'https://' + parts[1]
            found = False
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == title:
                    if quality_type == "4M":
                        sheet.cell(row=row_idx, column=3).value = url
                    elif quality_type == "8M":
                        sheet.cell(row=row_idx, column=4).value = url
                    elif quality_type == "4K":
                        sheet.cell(row=row_idx, column=2).value = url
                    found = True
                    break
            if not found:
                new_row = [title, None, None, None, None]
                if quality_type == "4M":
                    new_row[2] = url
                elif quality_type == "8M":
                    new_row[3] = url
                elif quality_type == "4K":
                    new_row[1] = url
                sheet.append(new_row)
        except IndexError:
            print("输入格式错误，请重新输入。")
    wb.save(file_path)
#   三种码率
def type2_safe(file_path, input_str):
    lines = input_str.split('\n')
    title = lines[0].replace('《', '').replace('》', '')
    urls = [line.split('】')[1] for line in lines[1:]]

    wb = load_workbook(file_path)
    ws = wb.active

    if len(urls) >= 3:
        ws.cell(row=ws.max_row + 1, column=1).value = title
        ws.cell(row=ws.max_row, column=2).value = urls[2]
        ws.cell(row=ws.max_row, column=3).value = urls[0]
        ws.cell(row=ws.max_row, column=4).value = urls[1]
    wb.save(file_path)

path = "C:\\Users\\fucha\\Desktop\\网盘地址表格.xlsx"
with open("C:\\Users\\fucha\\Desktop\\网盘地址模板.txt", "r", encoding="utf-8") as f:
    input_str = f.read()
type2_safe(path, input_str)
