# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import shutil

tag_tab = "C:\\Users\\fucha\\Desktop\\腾讯付费命中率.xlsx"
sub_tab = "C:\\Users\\fucha\\Desktop\\腾讯付费总表.xlsx"

# 加载表格
wb_total_filtered = load_workbook(tag_tab)
bt_sheet = wb_total_filtered.active
wb_sub = load_workbook(sub_tab)
st_sheet = wb_sub.active

# base table -> bt
bt_index = 2
# sub table -> st
st_index = 2


# 字典比对，取行号+id
def same_id(bt_sheet, st_sheet):
    bt_ids = {}
    for row_index in range(2, bt_sheet.max_row + 1):
        value = bt_sheet.cell(row=row_index, column=2).value
        if value is None:
            value = bt_sheet.cell(row=row_index, column=3).value
        bt_ids[row_index] = value
    st_ids = {row_index : st_sheet.cell(row=row_index, column=st_index).value for row_index in range(2, st_sheet.max_row + 1)}

    result_dict = {}
    for id1, value1 in bt_ids.items():
        for id2, value2 in st_ids.items():
            if value1 == value2:
                result_dict[id2] = id1

    print(result_dict)
    return  result_dict

# compare(bt_sheet, st_sheet)
result_dict = same_id(bt_sheet, st_sheet)
for key, value in result_dict.items():
    tag = st_sheet.cell(row=value, column=4).value
    if tag:
        if "11" in tag:
            bt_sheet.cell(row=key, column=59, value="R")
            bt_sheet.cell(row=key, column=57, value="T")
        elif "12" in tag:
            bt_sheet.cell(row=key, column=58, value="D")
            bt_sheet.cell(row=key, column=57, value="T")

# # 保存结果
wb_total_filtered.save(tag_tab)