import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font
import os
import platform
import subprocess # For cross-platform file opening
import re
import difflib

# 更改颜色定义常量 - 使用更深的颜色
FUZZY_MATCH_COLORS = {
    'priority_2': "228B22",  # 军绿色 (Forest Green)
    'priority_3_normal': "1E90FF",  # 深海蓝 (Dodger Blue) 
    'priority_3_short': "FF0000"   # 红色 (Red) - 改回原先的红色
}

# 常量定义
ALL_SHEETS_TEXT = "全部sheet" # 添加全部sheet的选项文本

# 新的模糊匹配常量
REPLACE_SYMBOLS_MAP = {
    '：': ':', 
    '（': '(', 
    '）': ')', 
    '！': '!', 
    '－': '-', 
    '—': '-',
    '【': '[',
    '】': ']',
    '？': '?',
    '"': '"',
    '"': '"',
    ''': "'",
    ''': "'",
    '　': ' ',  # 全角空格转半角空格
    '_': '',   # 下划线可以删除
    '-': ''    # 横线可以删除
}

# 更灵活的符号处理正则表达式
SYMBOL_REMOVE_REGEX = re.compile(r'[^\w\s一-龥]')

# 额外的符号替换模式
SYMBOL_REPLACEMENT_PATTERNS = [
    # 括号类符号互换
    (r'\(', '['), (r'\)', ']'),
    (r'\[', '('), (r'\]', ')'),
    # 连字符和下划线互换
    (r'[-_]', ':'), (r'[-_]', ''),
    # 冒号变其他符号
    (r':', '-'), (r':', '_'), (r':', ''),
]

# 中文数字到阿拉伯数字的映射
chinese_to_arabic_map = {
    '一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
    '六': '6', '七': '7', '八': '8', '九': '9', '十': '10',
    '零': '0'
}

# 罗马数字到阿拉伯数字的映射
roman_to_arabic_map = {
    'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5',
    'VI': '6', 'VII': '7', 'VIII': '8', 'IX': '9', 'X': '10',
    'Ⅰ': '1', 'Ⅱ': '2', 'Ⅲ': '3', 'Ⅳ': '4', 'Ⅴ': '5',
    'Ⅵ': '6', 'Ⅶ': '7', 'Ⅷ': '8', 'Ⅸ': '9', 'Ⅹ': '10'
}

# 添加全局缓存
_variation_cache = {}
_match_cache = {}

# --- 新的模糊匹配算法 ---
def safe_str_convert(value):
    """安全地将任何值转换为字符串"""
    if value is None:
        return ''
    try:
        # 处理openpyxl的特殊对象
        if hasattr(value, 'value'):
            return str(value.value) if value.value is not None else ''
        return str(value)
    except Exception:
        return ''

def normalize_for_comparison(text, remove_spaces=False, lower=False, remove_symbols_regex=None):
    """标准化文本用于比较"""
    # 确保输入为字符串
    text = safe_str_convert(text)
    if not text:
        return ''
    
    # 如果 lower 为 True，将文本转为小写
    if lower:
        text = text.lower()
    
    # 使用预定义的 REPLACE_SYMBOLS_MAP 替换特定全角符号为半角
    for full_width, half_width in REPLACE_SYMBOLS_MAP.items():
        text = text.replace(full_width, half_width)
    
    # 如果 remove_spaces 为 True，移除所有空格
    if remove_spaces:
        text = re.sub(r'\s+', '', text)
    
    # 如果提供了 remove_symbols_regex，使用它移除符号
    if remove_symbols_regex:
        text = remove_symbols_regex.sub('', text)
    
    # 去除字符串两端的空白
    return text.strip()

def convert_numbers(text):
    """将文本中的中文数字和罗马数字转换为阿拉伯数字"""
    # 确保输入为字符串
    text = safe_str_convert(text)
    if not text:
        return ''
    
    # 将中文数字转换为阿拉伯数字
    for chinese, arabic in chinese_to_arabic_map.items():
        text = text.replace(chinese, arabic)
    
    # 将罗马数字转换为阿拉伯数字
    for roman, arabic in roman_to_arabic_map.items():
        text = text.replace(roman, arabic)
    
    return text

def normalize_season_format(text):
    """标准化多季专辑格式"""
    if not text:
        return str(text) if text is not None else ''
    
    # 确保输入是字符串
    text = str(text)
    
    # 处理各种季度/部格式，支持更多数字格式
    patterns = [
        # 第x季 格式 (支持数字、汉字、罗马数字)
        (r'第([一二三四五六七八九十0-9IⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩivxlcdm]+)季', r'\1'),
        # 第x部 格式  
        (r'第([一二三四五六七八九十0-9IⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩivxlcdm]+)部', r'\1'),
        # 第x集 格式
        (r'第([一二三四五六七八九十0-9IⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩivxlcdm]+)集', r'\1'),
        # Sx 格式（大小写不敏感）
        (r'S([0-9]+)', r'\1'),
        (r's([0-9]+)', r'\1'),
        # Season x 格式
        (r'Season\s*([0-9]+)', r'\1'),
        (r'season\s*([0-9]+)', r'\1'),
        # 第x代 格式
        (r'第([一二三四五六七八九十0-9IⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩivxlcdm]+)代', r'\1')
    ]
    
    result = text
    for pattern, replacement in patterns:
        result = re.sub(pattern, replacement, result, flags=re.IGNORECASE)
    
    return result

def process_conjunctions(text):
    """处理连词"之"、"的"、"和"的各种变换（简化版）"""
    if not text:
        return [str(text) if text is not None else '']
    
    text = str(text)
    variations = [text]  # 使用列表，限制数量
    
    # 只生成最重要的几个变体，避免组合爆炸
    # 1. 删除连词的版本
    no_conjunctions = re.sub(r'[之的和]', '', text)
    if no_conjunctions != text:
        variations.append(no_conjunctions)
    
    # 2. 只做最常见的替换（限制为3个变体）
    variations.append(text.replace('之', '的'))
    variations.append(text.replace('的', '之'))
    variations.append(text.replace('和', '的'))
    
    # 去重并限制数量
    unique_variations = []
    seen = set()
    for var in variations:
        if var not in seen and len(unique_variations) < 5:
            unique_variations.append(var)
            seen.add(var)
    
    return unique_variations

def generate_priority2_variations(text):
    """生成优先级2的所有可能变体（高度简化版本）"""
    try:
        text = safe_str_convert(text)
        if not text or not text.strip():
            return ['']
        
        variations = [text]  # 使用列表而不是集合，更快
        
        # 1. 基础变换（最重要的几个）
        number_converted = convert_numbers(text)
        if number_converted != text:
            variations.append(number_converted)
        
        season_normalized = normalize_season_format(text)
        if season_normalized != text:
            variations.append(season_normalized)
        
        # 2. 连词处理（最多5个变体）
        conjunction_variations = process_conjunctions(text)
        for conj_var in conjunction_variations[1:4]:  # 只取前3个（除了原文本）
            if conj_var != text and len(variations) < 12:
                variations.append(conj_var)
        
        # 3. 简单的符号和大小写处理
        if len(variations) < 15:
            # 基础标准化
            lower_text = text.lower()
            if lower_text != text:
                variations.append(lower_text)
            
            # 删除符号
            no_symbols = normalize_for_comparison(
                text, 
                remove_spaces=True, 
                lower=True, 
                remove_symbols_regex=SYMBOL_REMOVE_REGEX
            )
            if no_symbols and no_symbols != text.lower():
                variations.append(no_symbols)
            
            # 符号替换（只做最重要的几个）
            symbol_replaced = text.replace('：', ':').replace('（', '(').replace('）', ')')
            if symbol_replaced != text:
                variations.append(symbol_replaced)
        
        # 去重并限制到15个变体
        unique_variations = []
        seen = set()
        for var in variations:
            if var and var.strip() and var not in seen and len(unique_variations) < 15:
                unique_variations.append(var)
                seen.add(var)
        
        return unique_variations if unique_variations else ['']
        
    except Exception as e:
        print(f"警告: 处理文本变体时出错: {e}")
        return [safe_str_convert(text) if text else '']

def priority_1_exact_match(base_name, candidate_name):
    """优先级1：精确匹配"""
    base_name = safe_str_convert(base_name)
    candidate_name = safe_str_convert(candidate_name)
    return base_name == candidate_name

def priority_2_symbol_case_number_match(base_name, candidate_name):
    """优先级2：符号、大小写、数字变体匹配（简化版）"""
    base_name = safe_str_convert(base_name)
    candidate_name = safe_str_convert(candidate_name)
    
    # 生成变体（每个最多15个）
    base_variations = generate_priority2_variations(base_name)
    candidate_variations = generate_priority2_variations(candidate_name)
    
    # 使用集合加速查找
    candidate_set = set(candidate_variations)
    
    # 检查是否有任何变体匹配
    for base_var in base_variations:
        if base_var in candidate_set:
            return True
    
    return False

def calculate_match_score(base_text, candidate_text):
    """计算两个文本的匹配分数，分数越高匹配度越好"""
    if not base_text or not candidate_text:
        return 0
    
    base_lower = base_text.lower()
    candidate_lower = candidate_text.lower()
    
    # 1. 长度相似性分数 (0-40分)
    base_len = len(base_text)
    candidate_len = len(candidate_text)
    max_len = max(base_len, candidate_len)
    if max_len == 0:
        length_score = 40
    else:
        length_diff = abs(base_len - candidate_len)
        length_score = max(0, 40 - (length_diff / max_len) * 40)
    
    # 2. 包含关系分数 (0-30分)
    if base_lower in candidate_lower:
        # 原文本被候选文本包含
        inclusion_score = 30
    elif candidate_lower in base_lower:
        # 候选文本被原文本包含
        inclusion_score = 25
    else:
        # 无直接包含关系，计算最长公共子序列
        inclusion_score = 0
    
    # 3. 字符重叠度分数 (0-20分)
    base_chars = set(base_lower)
    candidate_chars = set(candidate_lower)
    intersection = base_chars & candidate_chars
    union = base_chars | candidate_chars
    if union:
        overlap_score = (len(intersection) / len(union)) * 20
    else:
        overlap_score = 0
    
    # 4. 连续匹配分数 (0-10分) - 寻找最长连续匹配子串
    max_continuous = 0
    base_clean = re.sub(r'[^\w一-龥]', '', base_lower)
    candidate_clean = re.sub(r'[^\w一-龥]', '', candidate_lower)
    
    # 寻找最长公共子串
    for i in range(len(base_clean)):
        for j in range(len(candidate_clean)):
            length = 0
            while (i + length < len(base_clean) and 
                   j + length < len(candidate_clean) and
                   base_clean[i + length] == candidate_clean[j + length]):
                length += 1
            max_continuous = max(max_continuous, length)
    
    if max_continuous > 0:
        min_len = min(len(base_clean), len(candidate_clean))
        if min_len > 0:
            continuous_score = min(10, (max_continuous / min_len) * 10)
        else:
            continuous_score = 0
    else:
        continuous_score = 0
    
    total_score = length_score + inclusion_score + overlap_score + continuous_score
    return total_score

def priority_3_enhanced_selection(base_name_str, potential_matches):
    """优先级3的增强选择算法"""
    if not potential_matches:
        return []
    
    if len(potential_matches) == 1:
        return potential_matches
    
    # 计算每个候选项的匹配分数
    scored_matches = []
    for candidate in potential_matches:
        score = calculate_match_score(base_name_str, candidate)
        scored_matches.append((candidate, score))
    
    # 按分数排序，选择最高分的
    scored_matches.sort(key=lambda x: x[1], reverse=True)
    best_match = scored_matches[0][0]
    
    return [best_match]

def check_short_text_condition(text1, text2):
    """检查是否存在短文本包含情况（用于判断是否需要红色标记）"""
    if not text1 or not text2:
        return False
    
    # 移除符号后计算字符数
    text1_chars = re.sub(r'[^\u4e00-\u9fff\w]', '', text1)
    text2_chars = re.sub(r'[^\u4e00-\u9fff\w]', '', text2)
    
    # 检查是否有一个文本字数<=2且被包含在另一个文本中
    if (len(text1_chars) <= 2 and text1.lower() in text2.lower()) or \
       (len(text2_chars) <= 2 and text2.lower() in text1.lower()):
        return True
    
    return False

def find_fuzzy_matches(base_name_str, candidate_names_list):
    """主协调函数：按优先级查找模糊匹配（优化版）"""
    # 确保 base_name_str 是字符串
    base_name_str = safe_str_convert(base_name_str)
    if not base_name_str:
        return {}
    
    # 将候选列表中的所有元素转换为字符串，过滤掉 None 值
    candidate_names = []
    for name in candidate_names_list:
        name_str = safe_str_convert(name)
        if name_str:  # 过滤掉空字符串
            candidate_names.append(name_str)
    
    if not candidate_names:
        return {}
    
    # 创建缓存键
    candidates_key = tuple(sorted(candidate_names))
    cache_key = (base_name_str, candidates_key)
    
    # 检查缓存
    if cache_key in _match_cache:
        return _match_cache[cache_key]
    
    # 优先级1 (P1 - 精确匹配) - 最快的检查
    candidate_set = set(candidate_names)
    if base_name_str in candidate_set:
        result = {1: [base_name_str]}
        _match_cache[cache_key] = result
        return result
    
    # 优先级2 (P2 - 符号、大小写、数字变体匹配)
    # 使用缓存的变体生成
    if base_name_str not in _variation_cache:
        _variation_cache[base_name_str] = generate_priority2_variations(base_name_str)
    base_variations = _variation_cache[base_name_str]
    
    p2_matches = []
    for candidate in candidate_names:
        if candidate not in _variation_cache:
            _variation_cache[candidate] = generate_priority2_variations(candidate)
        candidate_variations = _variation_cache[candidate]
        
        # 使用集合加速查找
        candidate_set = set(candidate_variations)
        if any(base_var in candidate_set for base_var in base_variations):
            p2_matches.append(candidate)
    
    if p2_matches:
        result = {2: p2_matches}
        _match_cache[cache_key] = result
        return result
    
    # 优先级3 (P3 - 子字符串匹配，增强选择)
    potential_priority_3_matches = []
    base_name_lower = base_name_str.lower()
    
    for candidate in candidate_names:
        candidate_lower = candidate.lower()
        if base_name_lower in candidate_lower or candidate_lower in base_name_lower:
            potential_priority_3_matches.append(candidate)
    
    if not potential_priority_3_matches:
        _match_cache[cache_key] = {}
        return {}
    
    # 使用增强的P3选择算法
    best_matches = priority_3_enhanced_selection(base_name_str, potential_priority_3_matches)
    
    result = {3: best_matches}
    _match_cache[cache_key] = result
    return result

def find_advanced_fuzzy_matches(base_name_str, candidate_names_list, max_matches=3):
    """高级模糊匹配：返回前N个最佳匹配项（带优先级限制）"""
    # 确保 base_name_str 是字符串
    base_name_str = safe_str_convert(base_name_str)
    if not base_name_str:
        return {}
    
    # 将候选列表中的所有元素转换为字符串，过滤掉 None 值
    candidate_names = []
    for name in candidate_names_list:
        name_str = safe_str_convert(name)
        if name_str:  # 过滤掉空字符串
            candidate_names.append(name_str)
    
    if not candidate_names:
        return {}
    
    # 优先级1 (P1 - 精确匹配) - 最快的检查，如果有精确匹配直接返回
    candidate_set = set(candidate_names)
    if base_name_str in candidate_set:
        return {1: [base_name_str]}
    
    # 优先级2 (P2 - 符号、大小写、数字变体匹配)
    if base_name_str not in _variation_cache:
        _variation_cache[base_name_str] = generate_priority2_variations(base_name_str)
    base_variations = _variation_cache[base_name_str]
    
    p2_matches = []
    for candidate in candidate_names:
        if candidate not in _variation_cache:
            _variation_cache[candidate] = generate_priority2_variations(candidate)
        candidate_variations = _variation_cache[candidate]
        
        # 使用集合加速查找
        candidate_set = set(candidate_variations)
        if any(base_var in candidate_set for base_var in base_variations):
            p2_matches.append(candidate)
    
    # 如果有优先级2的匹配，只返回优先级2的结果，不处理优先级3
    if p2_matches:
        # 为优先级2匹配项计算分数并排序
        scored_p2_matches = []
        for candidate in p2_matches:
            score = calculate_match_score(base_name_str, candidate)
            scored_p2_matches.append((candidate, score))
        
        # 按分数排序，取前N个
        scored_p2_matches.sort(key=lambda x: x[1], reverse=True)
        top_p2_matches = [match[0] for match in scored_p2_matches[:max_matches]]
        
        return {2: top_p2_matches}
    
    # 优先级3 (P3 - 子字符串匹配) - 只有在没有优先级2匹配时才处理
    base_name_lower = base_name_str.lower()
    p3_normal_matches = []  # 蓝色匹配（正常包含关系）
    p3_short_matches = []   # 红色匹配（短文本包含）
    
    for candidate in candidate_names:
        candidate_lower = candidate.lower()
        if base_name_lower in candidate_lower or candidate_lower in base_name_lower:
            # 检查是否是短文本包含情况
            bt_norm = normalize_for_comparison(base_name_str)
            matched_norm = normalize_for_comparison(candidate)
            
            if check_short_text_condition(bt_norm, matched_norm):
                p3_short_matches.append(candidate)
            else:
                p3_normal_matches.append(candidate)
    
    # 优先级3的限制：如果有蓝色匹配，就不要红色匹配
    if p3_normal_matches:
        # 只处理蓝色匹配（正常包含关系）
        scored_p3_matches = []
        for candidate in p3_normal_matches:
            score = calculate_match_score(base_name_str, candidate)
            scored_p3_matches.append((candidate, score))
        
        # 按分数排序，取前N个
        scored_p3_matches.sort(key=lambda x: x[1], reverse=True)
        top_p3_matches = [match[0] for match in scored_p3_matches[:max_matches]]
        
        return {3: top_p3_matches}
    
    elif p3_short_matches:
        # 只有红色匹配的情况下才处理
        scored_p3_matches = []
        for candidate in p3_short_matches:
            score = calculate_match_score(base_name_str, candidate)
            scored_p3_matches.append((candidate, score))
        
        # 按分数排序，取前N个
        scored_p3_matches.sort(key=lambda x: x[1], reverse=True)
        top_p3_matches = [match[0] for match in scored_p3_matches[:max_matches]]
        
        return {3: top_p3_matches}
    
    # 没有任何匹配
    return {}

# --- Core Logic ---
def safe_get_cell_value(sheet, row, col):
    """安全地获取单元格值"""
    try:
        cell = sheet.cell(row=row, column=col)
        value = cell.value
        
        # 处理各种类型的值
        if value is None:
            return ''
        
        # 如果是数字，转换为字符串
        if isinstance(value, (int, float)):
            # 对于整数，不显示小数点
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)
        
        # 如果是其他类型，尝试转换为字符串
        return safe_str_convert(value)
    except Exception as e:
        print(f"警告: 读取单元格({row}, {col})时出错: {e}")
        return ''

def safe_set_cell_value(sheet, row, col, value, font=None):
    """安全地设置单元格值"""
    try:
        cell = sheet.cell(row=row, column=col)
        cell.value = value
        if font:
            cell.font = font
        return True
    except Exception as e:
        print(f"警告: 写入单元格({row}, {col})时出错: {e}")
        return False

def process_files(base_path, sub_path, bt_sheet_name, st_sheet_name, bt_key_col_name, st_key_col_name, bt_data_col_names, st_data_col_names, use_fuzzy_match=False, enable_tagging=False, tag_value="T", use_advanced_fuzzy=False, max_matches=3, progress_callback=None):
    """
    Compares two Excel sheets, finds matching keys, and transfers data for multiple columns.
    Handles duplicate keys in the base table.
    Supports processing all sheets in the base workbook when bt_sheet_name is set to ALL_SHEETS_TEXT.
    Supports fuzzy matching with priority-based matching and special formatting.

    Args:
        base_path (str): Path to the base Excel file.
        sub_path (str): Path to the content Excel file.
        bt_sheet_name (str): Name of the sheet in the base table, or ALL_SHEETS_TEXT for all sheets.
        st_sheet_name (str): Name of the sheet in the content table.
        bt_key_col_name (str): Name of the key column in the base table.
        st_key_col_name (str): Name of the key column in the content table.
        bt_data_col_names (list[str]): List of names of data columns in the base table (to be updated).
        st_data_col_names (list[str]): List of names of data columns in the content table (source).
        use_fuzzy_match (bool): Whether to use fuzzy matching.
        enable_tagging (bool): Whether to enable tagging.
        tag_value (str): The value to use for tagging.
        use_advanced_fuzzy (bool): Whether to use advanced fuzzy matching.
        max_matches (int): Maximum number of matches to return for advanced fuzzy matching.
        progress_callback (callable): A callback function to update the progress.

    Returns:
        tuple: (total_base_rows, match_count, dict of per-sheet stats or None, error_message or None)
    """
    try:
        # 清理缓存以释放内存并确保新的处理会话
        global _variation_cache, _match_cache
        _variation_cache.clear()
        _match_cache.clear()
        
        wb_base = openpyxl.load_workbook(base_path)
        wb_sub = openpyxl.load_workbook(sub_path)
        
        # Check if content sheet exists
        if st_sheet_name not in wb_sub.sheetnames:
             return 0, 0, None, f"错误: 在内容表 '{os.path.basename(sub_path)}' 中找不到工作表 '{st_sheet_name}'。可用工作表: {wb_sub.sheetnames}"
        st_sheet = wb_sub[st_sheet_name]
        
        # Find the sub table key column index
        def find_col_index(sheet, col_name, sheet_desc):
            for col_idx in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col_idx).value == col_name:
                    return col_idx
            raise ValueError(f"列 '{col_name}' 在 {sheet_desc} 中未找到。")
            
        st_key_col_idx = find_col_index(st_sheet, st_key_col_name, "内容表")
        
        # Load content table key data
        st_key_map = {}
        for row_index in range(2, st_sheet.max_row + 1):
            try:
                key = safe_get_cell_value(st_sheet, row_index, st_key_col_idx)
                if key and key.strip():  # 确保key不为空且不是纯空格
                    st_key_map[key] = row_index
            except Exception as e:
                print(f"警告: 处理内容表第{row_index}行时出错: {e}，跳过此行")
                continue
        
        # Calculate content table data column indices
        st_data_col_indices = [find_col_index(st_sheet, name, "内容表") for name in st_data_col_names]

        # Process base workbook
        sheets_to_process = []
        if bt_sheet_name == ALL_SHEETS_TEXT:
            # Process all sheets
            sheets_to_process = wb_base.sheetnames
            if not sheets_to_process:
                return 0, 0, None, f"错误: 原表 '{os.path.basename(base_path)}' 不包含任何工作表。"
        else:
            # Process just the selected sheet
            if bt_sheet_name not in wb_base.sheetnames:
                return 0, 0, None, f"错误: 在原表 '{os.path.basename(base_path)}' 中找不到工作表 '{bt_sheet_name}'。可用工作表: {wb_base.sheetnames}"
            sheets_to_process = [bt_sheet_name]
            
        # Initialize counters
        total_base_rows = 0
        total_match_count = 0
        sheet_stats = {}  # 存储每个sheet的统计信息
        
        # 首先计算总行数用于进度计算
        total_rows_to_process = 0
        for sheet_name in sheets_to_process:
            try:
                sheet = wb_base[sheet_name]
                total_rows_to_process += max(0, sheet.max_row - 1)  # 减1是因为不计算表头行
            except:
                continue
        
        processed_rows = 0
        
        # 准备内容表关键字列表以供模糊匹配使用
        st_keys_list = list(st_key_map.keys()) if use_fuzzy_match else []
        
        # Process each sheet in the base workbook
        for sheet_idx, current_sheet_name in enumerate(sheets_to_process):
            try:
                if progress_callback:
                    progress_callback(f"处理工作表 {sheet_idx + 1}/{len(sheets_to_process)}: {current_sheet_name}", processed_rows, total_rows_to_process)
                
                sheet_rows = 0
                sheet_matches = 0
                
                bt_sheet = wb_base[current_sheet_name]
                
                try:
                    # Find key column in this sheet
                    bt_key_col_idx = find_col_index(bt_sheet, bt_key_col_name, f"原表 '{current_sheet_name}'")
                    
                    # Find data columns in this sheet (may fail if sheet structure differs)
                    bt_data_col_indices = []
                    if bt_data_col_names:  # 只有在有数据列名时才查找
                        for name in bt_data_col_names:
                            try:
                                bt_data_col_indices.append(find_col_index(bt_sheet, name, f"原表 '{current_sheet_name}'"))
                            except ValueError:
                                # Skip columns not found in this sheet
                                continue
                    
                    # 如果没有数据列但启用了标记，仍然可以继续处理
                    if not bt_data_col_indices and not enable_tagging:
                        # 只有在既没有数据列又没有启用标记时才跳过
                        sheet_stats[current_sheet_name] = {
                            "rows": 0, 
                            "matches": 0, 
                            "status": f"跳过: 在工作表中找不到任何指定的数据列且未启用标记功能"
                        }
                        continue
                    
                    # 为模糊匹配准备：在关键字列右侧插入一列用于存储匹配的关键字
                    fuzzy_match_col_idx = None
                    if use_fuzzy_match:
                        # 插入新列
                        bt_sheet.insert_cols(bt_key_col_idx + 1)
                        fuzzy_match_col_idx = bt_key_col_idx + 1
                        bt_sheet.cell(row=1, column=fuzzy_match_col_idx, value="模糊匹配关键字")
                        
                        # 更新数据列索引（因为插入了新列）
                        bt_data_col_indices = [idx + 1 if idx > bt_key_col_idx else idx for idx in bt_data_col_indices]
                    
                    # Build key map for this sheet
                    bt_key_map = defaultdict(list)
                    current_sheet_rows = max(0, bt_sheet.max_row - 1)  # 减1是因为不计算表头行
                    for row_index in range(2, bt_sheet.max_row + 1):
                        try:
                            sheet_rows += 1
                            processed_rows += 1
                            
                            # 更新进度
                            if progress_callback and processed_rows % 10 == 0:  # 每10行更新一次进度，避免过于频繁
                                progress_callback(f"处理工作表: {current_sheet_name} (第{processed_rows}/{total_rows_to_process}行)", processed_rows, total_rows_to_process)
                            
                            key = safe_get_cell_value(bt_sheet, row_index, bt_key_col_idx)
                            if key and key.strip():  # 确保key不为空且不是纯空格
                                bt_key_map[key].append(row_index)
                        except Exception as e:
                            print(f"警告: 处理原表第{row_index}行时出错: {e}，跳过此行")
                            continue
                    
                    updated_rows = set()
                    
                    if use_fuzzy_match:
                        # 使用新的模糊匹配逻辑
                        processed_bt_keys = set()
                        
                        # 记录需要插入的新行
                        rows_to_insert = []  # [(insert_after_row, bt_key, matched_st_key, priority), ...]
                        
                        for bt_key in bt_key_map:
                            try:
                                if bt_key in processed_bt_keys:
                                    continue
                                    
                                # 根据是否启用高级模糊匹配选择不同的匹配函数
                                if use_advanced_fuzzy:
                                    match_result = find_advanced_fuzzy_matches(bt_key, st_keys_list, max_matches)
                                else:
                                    match_result = find_fuzzy_matches(bt_key, st_keys_list)
                                
                                if match_result:
                                    bt_rows = bt_key_map[bt_key]
                                    first_bt_row = bt_rows[0]
                                    
                                    # 收集所有匹配项（按优先级排序）
                                    all_matched_keys = []
                                    for priority in sorted(match_result.keys()):
                                        for matched_key in match_result[priority]:
                                            all_matched_keys.append((matched_key, priority))
                                    
                                    if all_matched_keys:
                                        # 处理第一个匹配（使用原行）
                                        first_matched_key, first_priority = all_matched_keys[0]
                                        st_row = st_key_map[first_matched_key]
                                        
                                        # 更新所有具有相同key的行
                                        for bt_row in bt_rows:
                                            # 在模糊匹配列中填入匹配的关键字
                                            safe_set_cell_value(bt_sheet, bt_row, fuzzy_match_col_idx, first_matched_key)
                                            
                                            # 根据优先级设置字体样式
                                            cell = bt_sheet.cell(row=bt_row, column=fuzzy_match_col_idx)
                                            font_style = None
                                            
                                            if first_priority == 2:
                                                font_style = Font(color=FUZZY_MATCH_COLORS['priority_2'], bold=True)  # 军绿色加粗
                                            elif first_priority == 3:
                                                # 检查是否是短文本包含情况
                                                bt_norm = normalize_for_comparison(bt_key)
                                                matched_norm = normalize_for_comparison(first_matched_key)
                                                
                                                if check_short_text_condition(bt_norm, matched_norm):
                                                    font_style = Font(color=FUZZY_MATCH_COLORS['priority_3_short'], bold=True)  # 红色加粗
                                                else:
                                                    font_style = Font(color=FUZZY_MATCH_COLORS['priority_3_normal'], bold=True)  # 深海蓝加粗
                                            
                                            if font_style:
                                                cell.font = font_style
                                            
                                            # 更新数据列（如果有的话）
                                            if bt_data_col_indices:  # 只有在有数据列时才更新
                                                for i in range(len(bt_data_col_indices)):
                                                    bt_data_col_idx = bt_data_col_indices[i]
                                                    if i < len(st_data_col_indices):
                                                        st_data_col_idx = st_data_col_indices[i]
                                                        st_value = safe_get_cell_value(st_sheet, st_row, st_data_col_idx)
                                                        success = safe_set_cell_value(bt_sheet, bt_row, bt_data_col_idx, st_value)
                                                        if success and font_style:
                                                            data_cell = bt_sheet.cell(row=bt_row, column=bt_data_col_idx)
                                                            data_cell.font = font_style
                                                        if not success:
                                                            print(f"警告: 无法更新数据列{i+1}，跳过")
                                            
                                            updated_rows.add(bt_row)
                                        
                                        # 处理其他匹配项（需要插入新行）
                                        for i, (matched_key, priority) in enumerate(all_matched_keys[1:], 1):
                                            rows_to_insert.append((first_bt_row + i - 1, bt_key, matched_key, priority))
                                    
                                    processed_bt_keys.add(bt_key)
                            except Exception as e:
                                print(f"警告: 处理关键字'{bt_key}'时出错: {e}，跳过此关键字")
                                processed_bt_keys.add(bt_key)
                                continue
                        
                        # 插入需要的新行（从后往前插入以保持行号正确）
                        for insert_after_row, bt_key, matched_key, priority in reversed(rows_to_insert):
                            try:
                                bt_sheet.insert_rows(insert_after_row + 1)
                                new_row = insert_after_row + 1
                                
                                # 设置模糊匹配关键字
                                safe_set_cell_value(bt_sheet, new_row, fuzzy_match_col_idx, matched_key)
                                
                                # 设置字体样式
                                cell = bt_sheet.cell(row=new_row, column=fuzzy_match_col_idx)
                                font_style = None
                                
                                if priority == 2:
                                    font_style = Font(color=FUZZY_MATCH_COLORS['priority_2'], bold=True)
                                elif priority == 3:
                                    bt_norm = normalize_for_comparison(bt_key)
                                    matched_norm = normalize_for_comparison(matched_key)
                                    
                                    if check_short_text_condition(bt_norm, matched_norm):
                                        font_style = Font(color=FUZZY_MATCH_COLORS['priority_3_short'], bold=True)
                                    else:
                                        font_style = Font(color=FUZZY_MATCH_COLORS['priority_3_normal'], bold=True)
                                
                                if font_style:
                                    cell.font = font_style
                                
                                # 更新数据列（如果有的话）
                                if bt_data_col_indices:  # 只有在有数据列时才更新
                                    if matched_key in st_key_map:  # 确保匹配的关键字存在于内容表中
                                        st_row = st_key_map[matched_key]
                                        for j in range(len(bt_data_col_indices)):
                                            bt_data_col_idx = bt_data_col_indices[j]
                                            if j < len(st_data_col_indices):
                                                st_data_col_idx = st_data_col_indices[j]
                                                st_value = safe_get_cell_value(st_sheet, st_row, st_data_col_idx)
                                                success = safe_set_cell_value(bt_sheet, new_row, bt_data_col_idx, st_value)
                                                if success and font_style:
                                                    data_cell = bt_sheet.cell(row=new_row, column=bt_data_col_idx)
                                                    data_cell.font = font_style
                                                if not success:
                                                    print(f"警告: 无法更新数据列{j+1}，跳过")
                                    else:
                                        print(f"警告: 匹配的关键字'{matched_key}'在内容表中不存在，跳过数据更新")
                                
                                updated_rows.add(new_row)
                            except Exception as e:
                                print(f"警告: 插入新行时出错: {e}，跳过此行插入")
                                continue
                    else:
                        # 原有的精确匹配逻辑保持不变
                        for bt_key, bt_rows in bt_key_map.items():
                            if bt_key in st_key_map:
                                st_row = st_key_map[bt_key]
                                for bt_row in bt_rows:
                                    # 更新数据列（如果有的话）
                                    if bt_data_col_indices:  # 只有在有数据列时才更新
                                        for i in range(len(bt_data_col_indices)):
                                            bt_data_col_idx = bt_data_col_indices[i]
                                            if i < len(st_data_col_indices):
                                                st_data_col_idx = st_data_col_indices[i]
                                                st_value = safe_get_cell_value(st_sheet, st_row, st_data_col_idx)
                                                safe_set_cell_value(bt_sheet, bt_row, bt_data_col_idx, st_value)
                                    updated_rows.add(bt_row)
                    
                    # 标记功能：无论是否有数据列都可以工作
                    if enable_tagging:
                        # 添加标记列
                        max_col = bt_sheet.max_column + 1
                        if bt_sheet.cell(row=1, column=max_col).value != "匹配标记":
                            bt_sheet.cell(row=1, column=max_col, value="匹配标记")
                        
                        # 标记匹配的行
                        if use_fuzzy_match:
                            # 模糊匹配模式：标记所有匹配的行
                            for row in updated_rows:
                                bt_sheet.cell(row=row, column=max_col, value=tag_value)
                        else:
                            # 精确匹配模式：检查所有行是否有匹配
                            for bt_key, bt_rows in bt_key_map.items():
                                if bt_key in st_key_map:
                                    for bt_row in bt_rows:
                                        bt_sheet.cell(row=bt_row, column=max_col, value=tag_value)
                                        updated_rows.add(bt_row)  # 添加到updated_rows以保持一致性

                    sheet_matches = len(updated_rows)
                    
                    # Update counters
                    total_base_rows += sheet_rows
                    total_match_count += sheet_matches
                    
                    # Save sheet statistics
                    sheet_stats[current_sheet_name] = {
                        "rows": sheet_rows,
                        "matches": sheet_matches,
                        "status": "处理成功"
                    }
                    
                    # Update progress
                    if progress_callback:
                        progress_callback(current_sheet_name, processed_rows, total_rows_to_process)

                except ValueError as ve:
                    # Record error for this sheet but continue with others
                    sheet_stats[current_sheet_name] = {
                        "rows": sheet_rows,
                        "matches": 0,
                        "status": f"错误: {ve}"
                    }
                    continue
                
            except Exception as e:
                print(f"警告: 处理工作表'{current_sheet_name}'时出错: {e}，跳过此工作表")
                sheet_stats[current_sheet_name] = {
                    "rows": sheet_rows,
                    "matches": 0,
                    "status": f"错误: {e}"
                }
                continue
        
        # 最终进度更新
        if progress_callback:
            progress_callback("所有工作表处理完成", total_rows_to_process, total_rows_to_process)
        
        # Save the workbook with all changes
        wb_base.save(base_path)
        return total_base_rows, total_match_count, sheet_stats, None

    except FileNotFoundError as fnf:
        # More specific error message
        missing_file = base_path if not os.path.exists(base_path) else sub_path
        return 0, 0, None, f"错误: 文件未找到: {missing_file}"
    except KeyError as ke: # Handles invalid sheet name from process_files
        return 0, 0, None, f"错误: 无效的工作表名称 - {ke}"
    except ValueError as ve: # Specific error for column not found or internal errors
        return 0, 0, None, f"错误: {ve}"
    except Exception as e:
        # Catch potential permission errors during save
        if isinstance(e, PermissionError):
             return total_base_rows, total_match_count, sheet_stats, f"错误: 无法保存文件 '{os.path.basename(base_path)}'。请确保文件未被其他程序打开，并且您有写入权限。"
        return 0, 0, None, f"发生意外错误: {e}\n类型: {type(e)}"

# --- GUI Application Class ---
class ResultDialog:
    def __init__(self, parent, title, message, file_path=None, matched_count=0):
        self.result = None
        self.file_path = file_path
        self.matched_count = matched_count
        
        # 创建对话框窗口
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("600x400")
        self.dialog.resizable(True, True)
        
        # 设置模态
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 居中显示
        self.dialog.geometry("+%d+%d" % (parent.winfo_rootx() + 50, parent.winfo_rooty() + 50))
        
        # 创建主框架
        main_frame = ttk.Frame(self.dialog, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # 标题标签
        title_label = ttk.Label(main_frame, text=title, font=("Segoe UI", 12, "bold"))
        title_label.pack(anchor="w", pady=(0, 10))
        
        # 创建滚动文本框显示结果
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        # 文本框和滚动条
        self.text_widget = tk.Text(text_frame, wrap=tk.WORD, height=15, width=70)
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.text_widget.yview)
        self.text_widget.configure(yscrollcommand=scrollbar.set)
        
        self.text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 插入文本内容
        self.text_widget.insert("1.0", message)
        self.text_widget.config(state="disabled")  # 设为只读
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))
        
        # 如果有匹配结果且有文件路径，显示打开文件按钮
        if self.matched_count > 0 and self.file_path:
            open_button = ttk.Button(button_frame, text="打开表格", command=self._open_file)
            open_button.pack(side="left", padx=(0, 10))
        
        # 关闭按钮
        close_button = ttk.Button(button_frame, text="关闭", command=self._close)
        close_button.pack(side="right")
        
        # 绑定ESC键关闭对话框
        self.dialog.bind("<Escape>", lambda e: self._close())
        
        # 设置焦点
        close_button.focus()
    
    def _open_file(self):
        """打开文件并关闭对话框"""
        self.result = "open"
        self.dialog.destroy()
    
    def _close(self):
        """关闭对话框"""
        self.result = "close"
        self.dialog.destroy()
    
    def show(self):
        """显示对话框并等待结果"""
        self.dialog.wait_window()
        return self.result

class ExcelComparatorApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel 比对工具 - 增强版")
        master.geometry("920x850") # Adjusted size for additional UI elements

        self.base_headers = []
        self.sub_headers = []
        self.base_sheets = []
        self.sub_sheets = []

        # Store dynamic data column widgets
        self.data_column_widgets = [] # List of tuples: (label, bt_combo, st_combo)

        # --- Create main scrollable frame ---
        # Create a main frame that will contain the canvas and scrollbar
        main_container = ttk.Frame(master)
        main_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create canvas and scrollbar
        self.canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        # Configure scrolling
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        # Pack canvas and scrollbar
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Bind mouse wheel events for scrolling
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_to_mousewheel(event):
            self.canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _unbind_from_mousewheel(event):
            self.canvas.unbind_all("<MouseWheel>")
        
        self.canvas.bind('<Enter>', _bind_to_mousewheel)
        self.canvas.bind('<Leave>', _unbind_from_mousewheel)

        # Now use self.scrollable_frame as the parent for all widgets instead of master

        # --- Style ---
        style = ttk.Style()
        style.configure("TLabel", padding=5)
        style.configure("TButton", padding=5)
        style.configure("TEntry", padding=5)
        style.configure("TCombobox", padding=5)
        style.configure("Header.TLabel", font=("Segoe UI", 9, "bold"))
        style.configure("TLabelframe.Label", padding=(0, 5))

        # --- File Selection Frame ---
        file_frame = ttk.LabelFrame(self.scrollable_frame, text="1. 选择 Excel 文件和工作表", padding=10)
        file_frame.pack(padx=10, pady=(10, 5), fill="x")
        file_frame.columnconfigure(1, weight=1) # Make entry expand

        # Base File
        ttk.Label(file_frame, text="原表（输出表）:").grid(row=0, column=0, sticky="w", padx=(0,5))
        self.base_file_entry = ttk.Entry(file_frame, width=50)
        self.base_file_entry.grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(file_frame, text="浏览...", command=lambda: self._browse_file(self.base_file_entry, self.base_sheet_combo, 'base')).grid(row=0, column=2, padx=5)
        ttk.Label(file_frame, text="工作表:").grid(row=0, column=3, sticky="w", padx=(10,5))
        self.base_sheet_combo = ttk.Combobox(file_frame, state="readonly", width=20)
        self.base_sheet_combo.grid(row=0, column=4, sticky="ew", padx=5)
        self.base_sheet_combo.bind("<<ComboboxSelected>>", lambda event: self._on_sheet_selected(self.base_file_entry.get(), self.base_sheet_combo.get(), 'base'))

        # Sub File
        ttk.Label(file_frame, text="内容表（参考表）:").grid(row=1, column=0, sticky="w", padx=(0,5), pady=(5,0))
        self.sub_file_entry = ttk.Entry(file_frame, width=50)
        self.sub_file_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=(5,0))
        ttk.Button(file_frame, text="浏览...", command=lambda: self._browse_file(self.sub_file_entry, self.sub_sheet_combo, 'sub')).grid(row=1, column=2, padx=5, pady=(5,0))
        ttk.Label(file_frame, text="工作表:").grid(row=1, column=3, sticky="w", padx=(10,5), pady=(5,0))
        self.sub_sheet_combo = ttk.Combobox(file_frame, state="readonly", width=20)
        self.sub_sheet_combo.grid(row=1, column=4, sticky="ew", padx=5, pady=(5,0))
        self.sub_sheet_combo.bind("<<ComboboxSelected>>", lambda event: self._on_sheet_selected(self.sub_file_entry.get(), self.sub_sheet_combo.get(), 'sub'))


        # --- Key Column Selection Frame ---
        key_col_frame = ttk.LabelFrame(self.scrollable_frame, text="2. 关键字列选择", padding=10)
        key_col_frame.pack(padx=10, pady=5, fill="x")
        key_col_frame.columnconfigure(1, weight=1)
        key_col_frame.columnconfigure(2, weight=1)

        # Header Row
        ttk.Label(key_col_frame, text="").grid(row=0, column=0, padx=5, pady=5) # Spacer
        ttk.Label(key_col_frame, text="原表（输出表）", style="Header.TLabel").grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Label(key_col_frame, text="内容表（参考表）", style="Header.TLabel").grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Key Column Row
        ttk.Label(key_col_frame, text="关键字列:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.bt_key_combo = ttk.Combobox(key_col_frame, state="readonly", width=30)
        self.bt_key_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.st_key_combo = ttk.Combobox(key_col_frame, state="readonly", width=30)
        self.st_key_combo.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        # --- Data Column Selection Frame ---
        self.col_frame = ttk.LabelFrame(self.scrollable_frame, text="3. 数据列选择", padding=10)
        self.col_frame.pack(padx=10, pady=5, fill="x")
        self.col_frame.columnconfigure(1, weight=1)
        self.col_frame.columnconfigure(2, weight=1)

        # Header Row for data columns
        ttk.Label(self.col_frame, text="").grid(row=0, column=0, padx=5, pady=5) # Spacer
        ttk.Label(self.col_frame, text="原表（输出表）", style="Header.TLabel").grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Label(self.col_frame, text="内容表（参考表）", style="Header.TLabel").grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Placeholder for dynamic data columns - row index starts at 1 (since header is at row 0)
        self.data_col_start_row = 1

        # Add Data Column Button
        button_frame = ttk.Frame(self.col_frame)
        button_frame.grid(row=self.data_col_start_row, column=0, columnspan=3, pady=10)
        
        self.add_col_button = ttk.Button(button_frame, text="添加数据列", command=self._add_data_column_pair)
        self.add_col_button.pack(side="left", padx=(0, 10))
        
        self.delete_col_button = ttk.Button(button_frame, text="删除数据列", command=self._delete_data_column_pair)
        self.delete_col_button.pack(side="left")
        
        # 初始状态下禁用删除按钮（因为还没有数据列）
        self.delete_col_button.config(state=tk.DISABLED)

        # Add the first data column pair initially
        self._add_data_column_pair()

        # --- 模糊匹配选项框架 ---
        fuzzy_frame = ttk.LabelFrame(self.scrollable_frame, text="4. 匹配选项", padding=10)
        fuzzy_frame.pack(padx=10, pady=5, fill="x")
        
        # 模糊匹配选项行
        fuzzy_options_frame = ttk.Frame(fuzzy_frame)
        fuzzy_options_frame.pack(anchor="w", pady=5, fill="x")
        
        self.use_fuzzy_match = tk.BooleanVar()
        self.fuzzy_checkbox = ttk.Checkbutton(
            fuzzy_options_frame, 
            text="启用模糊匹配（支持符号转换、数字格式转换、包含关系等）", 
            variable=self.use_fuzzy_match,
            command=self._on_fuzzy_match_toggle
        )
        self.fuzzy_checkbox.pack(side="left", anchor="w")
        
        # 高级模糊匹配选项
        self.use_advanced_fuzzy = tk.BooleanVar()
        self.advanced_fuzzy_checkbox = ttk.Checkbutton(
            fuzzy_options_frame,
            text="高级模糊匹配（同优先级内返回前N个最佳匹配）",
            variable=self.use_advanced_fuzzy,
            state=tk.DISABLED  # 初始状态为禁用
        )
        self.advanced_fuzzy_checkbox.pack(side="left", anchor="w", padx=(20, 0))
        
        # 最大匹配条数输入框
        max_matches_frame = ttk.Frame(fuzzy_frame)
        max_matches_frame.pack(anchor="w", pady=2)
        
        # 添加左边距以与高级模糊匹配选项的文本部分对齐
        ttk.Label(max_matches_frame, text="").pack(side="left", padx=(365, 0))  # 与高级模糊匹配文本对齐的空白
        ttk.Label(max_matches_frame, text="最大匹配条数:").pack(side="left", padx=(0, 5))
        self.max_matches_entry = ttk.Entry(max_matches_frame, width=5)
        self.max_matches_entry.insert(0, "3")  # 默认值为3
        self.max_matches_entry.pack(side="left", padx=(0, 5))
        
        # 添加验证函数，只允许输入正整数
        vcmd = (self.master.register(self._validate_max_matches), '%P')
        self.max_matches_entry.config(validate='key', validatecommand=vcmd)
        
        ttk.Label(max_matches_frame, text="(仅在高级模糊匹配时生效)", font=("Segoe UI", 8), foreground="gray").pack(side="left")
        
        # 添加帮助说明
        help_text = ttk.Label(
            fuzzy_frame, 
            text="注意：模糊匹配会在关键字列右侧添加新列，并根据匹配优先级用不同颜色标识结果\n黑色：一毛一样 / 绿色：符号替换 / 蓝色：包含关系 / 红色：离谱的包含关系",
            font=("Segoe UI", 8),
            foreground="gray"
        )
        help_text.pack(anchor="w", pady=(0, 5))

        # 在模糊匹配选项框架中添加打标签选项
        tag_frame = ttk.Frame(fuzzy_frame)
        tag_frame.pack(anchor="w", pady=2, fill="x")
        
        self.enable_tagging = tk.BooleanVar()
        self.tagging_checkbox = ttk.Checkbutton(
            tag_frame, 
            text="启用匹配标记（在最后一列标记匹配的行）", 
            variable=self.enable_tagging
        )
        self.tagging_checkbox.pack(side="left", anchor="w")
        
        # 标记值输入框（在同一行）
        ttk.Label(tag_frame, text="标记值:").pack(side="left", padx=(20, 5))
        self.tag_value_entry = ttk.Entry(tag_frame, width=10)
        self.tag_value_entry.insert(0, "T")
        self.tag_value_entry.pack(side="left")

        # --- Progress Bar Frame (moved above action frame) ---
        progress_frame = ttk.Frame(self.scrollable_frame, padding=10)
        progress_frame.pack(pady=5, fill="x")
        
        # Progress label
        self.progress_label = ttk.Label(progress_frame, text="准备就绪")
        self.progress_label.pack(anchor="w", pady=(0, 5))
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate")
        self.progress_bar.pack(fill="x", pady=(0, 5))

        # --- Action Frame ---
        action_frame = ttk.Frame(self.scrollable_frame, padding=10)
        action_frame.pack(pady=(0, 20), fill="x")  # 增加底部边距，因为这是最后一个元素
        # Make the frame itself expand horizontally
        action_frame.columnconfigure(0, weight=1)
        # Place the button in the center
        self.start_button = ttk.Button(action_frame, text="开始处理", command=self._start_processing, width=15)
        self.start_button.grid(row=0, column=0, pady=5) # Use grid to center within the frame column

        # --- Result Frame ---
        result_frame = ttk.LabelFrame(self.scrollable_frame, text="结果", padding=10)
        result_frame.pack(padx=10, pady=(5, 15), fill="both", expand=True)  # 增加底部边距

        self.result_label = ttk.Label(result_frame, text="请选择文件和列，然后点击 '开始处理'。", wraplength=700, justify="left")
        self.result_label.pack(anchor="nw")

        # Update canvas scroll region after all widgets are added
        self.scrollable_frame.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _add_data_column_pair(self):
        """Adds a new row for selecting a pair of data columns."""
        current_row = self.data_col_start_row + len(self.data_column_widgets) + 1  # +1 for button frame

        # 计算正确的数据列编号（基于当前数据列数量）
        next_col_number = len(self.data_column_widgets) + 1
        
        # Data Column Row
        label_text = f"数据列 ({next_col_number}):"
        data_label = ttk.Label(self.col_frame, text=label_text)
        data_label.grid(row=current_row, column=0, padx=5, pady=5, sticky="w")

        bt_data_combo = ttk.Combobox(self.col_frame, state="readonly", width=30)
        bt_data_combo.grid(row=current_row, column=1, padx=5, pady=5, sticky="ew")
        bt_data_combo['values'] = self.base_headers # Populate with current headers

        st_data_combo = ttk.Combobox(self.col_frame, state="readonly", width=30)
        st_data_combo.grid(row=current_row, column=2, padx=5, pady=5, sticky="ew")
        st_data_combo['values'] = self.sub_headers # Populate with current headers

        # Store widgets for later access/update
        self.data_column_widgets.append((data_label, bt_data_combo, st_data_combo))

        # Move the button frame down
        button_frame = self.add_col_button.master
        button_frame.grid(row=current_row + 1, column=0, columnspan=3, pady=10)
        
        # 启用删除按钮（现在至少有一个数据列）
        self.delete_col_button.config(state=tk.NORMAL)

    def _delete_data_column_pair(self):
        """Deletes the last data column pair."""
        if not self.data_column_widgets:
            return
        
        # Get the last data column widgets
        last_label, last_bt_combo, last_st_combo = self.data_column_widgets.pop()
        
        # Destroy the widgets
        last_label.destroy()
        last_bt_combo.destroy()
        last_st_combo.destroy()
        
        # 重新编号剩余的数据列标签
        self._renumber_data_columns()
        
        # 如果没有数据列了，禁用删除按钮
        if not self.data_column_widgets:
            self.delete_col_button.config(state=tk.DISABLED)
        
        # Move the button frame up
        if self.data_column_widgets:
            last_row = self.data_col_start_row + len(self.data_column_widgets) + 1
        else:
            last_row = self.data_col_start_row + 1
        
        button_frame = self.add_col_button.master
        button_frame.grid(row=last_row, column=0, columnspan=3, pady=10)

    def _renumber_data_columns(self):
        """重新编号所有数据列标签"""
        for i, (label, bt_combo, st_combo) in enumerate(self.data_column_widgets):
            label.config(text=f"数据列 ({i + 1}):")

    def _load_sheets_and_headers(self, file_path, sheet_combo, table_type):
        """Loads sheet names into the combobox and headers for the active sheet."""
        if not file_path or not os.path.exists(file_path):
             # Clear relevant widgets if file path is invalid or cleared
            sheet_combo['values'] = []
            sheet_combo.set('')
            if table_type == 'base':
                self.base_sheets = []
                self.base_headers = []
                self.bt_key_combo['values'] = []
                self.bt_key_combo.set('')
                for _, bt_combo, _ in self.data_column_widgets:
                    bt_combo['values'] = []
                    bt_combo.set('')
            else:
                self.sub_sheets = []
                self.sub_headers = []
                self.st_key_combo['values'] = []
                self.st_key_combo.set('')
                for _, _, st_combo in self.data_column_widgets:
                    st_combo['values'] = []
                    st_combo.set('')
            return

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            active_sheet_name = workbook.active.title # Get active sheet by title
            workbook.close()

            if not sheet_names:
                 messagebox.showwarning("无工作表", f"文件 '{os.path.basename(file_path)}' 不包含任何工作表。")
                 sheet_combo['values'] = []
                 sheet_combo.set('')
                 return

            # 更新下拉框列表
            if table_type == 'base':
                # 为原表添加"全部sheet"选项
                sheet_combo['values'] = [ALL_SHEETS_TEXT] + sheet_names
                self.base_sheets = sheet_names
            else:
                sheet_combo['values'] = sheet_names
                self.sub_sheets = sheet_names

            # 设置默认选中项
            if active_sheet_name in sheet_names:
                 sheet_combo.set(active_sheet_name)
            elif sheet_names: # If active sheet not found, select first
                 sheet_combo.set(sheet_names[0])
                 active_sheet_name = sheet_names[0]
            else: # Should not happen if sheet_names is not empty, but safeguard
                active_sheet_name = None
                sheet_combo.set('')

            # Load headers for the initially selected (active or first) sheet
            if active_sheet_name:
                 self._load_headers(file_path, active_sheet_name, table_type)
            else:
                # Clear headers if no sheet could be selected
                self._clear_headers(table_type)

        except Exception as e:
            messagebox.showerror("错误", f"无法从 '{os.path.basename(file_path)}' 加载工作表列表或初始表头:\n{e}")
            sheet_combo['values'] = []
            sheet_combo.set('')
            self._clear_headers(table_type)

    def _on_sheet_selected(self, file_path, sheet_name, table_type):
        """当用户从下拉框选择工作表时调用"""
        if not file_path:
            self._clear_headers(table_type)
            return
        
        if table_type == 'base' and sheet_name == ALL_SHEETS_TEXT:
            # 如果选择了"全部sheet"，使用第一个sheet的表头作为模板
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            if workbook.sheetnames:
                self._load_headers(file_path, workbook.sheetnames[0], table_type)
            else:
                self._clear_headers(table_type)
        elif file_path and sheet_name:
            self._load_headers(file_path, sheet_name, table_type)
        else:
            # Clear headers if file path or sheet name is missing
            self._clear_headers(table_type)

    def _clear_headers(self, table_type):
         """Clears header comboboxes for the specified table type."""
         if table_type == 'base':
             self.base_headers = []
             self.bt_key_combo['values'] = []
             self.bt_key_combo.set('')
             for _, bt_combo, _ in self.data_column_widgets:
                 bt_combo['values'] = []
                 bt_combo.set('')
         elif table_type == 'sub':
             self.sub_headers = []
             self.st_key_combo['values'] = []
             self.st_key_combo.set('')
             for _, _, st_combo in self.data_column_widgets:
                 st_combo['values'] = []
                 st_combo.set('')

    def _load_headers(self, file_path, sheet_name, table_type):
        """Loads headers from the first row of a specific Excel sheet."""
        if not file_path or not sheet_name:
             self._clear_headers(table_type)
             return

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name not in workbook.sheetnames:
                 # This might happen if the file was changed externally
                 messagebox.showwarning("工作表丢失", f"工作表 '{sheet_name}' 在文件 '{os.path.basename(file_path)}' 中找不到了。请重新选择文件或工作表。")
                 self._clear_headers(table_type)
                 workbook.close()
                 # Also clear the sheet selection as it's invalid now
                 if table_type == 'base':
                     self.base_sheet_combo['values'] = workbook.sheetnames # Update with current sheets
                     self.base_sheet_combo.set('')
                 else:
                    self.sub_sheet_combo['values'] = workbook.sheetnames
                    self.sub_sheet_combo.set('')
                 return

            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1] if cell.value is not None]
            workbook.close()

            if not headers:
                messagebox.showwarning("空工作表或无表头", f"在 '{os.path.basename(file_path)}' 的工作表 '{sheet_name}' 的第一行找不到任何表头，或者工作表为空。")
                self._clear_headers(table_type)
                return

            # Update the corresponding comboboxes
            if table_type == 'base':
                self.base_headers = headers
                self.bt_key_combo['values'] = headers
                # Update all existing base data comboboxes
                for _, bt_combo, _ in self.data_column_widgets:
                     bt_combo['values'] = headers
            elif table_type == 'sub':
                self.sub_headers = headers
                self.st_key_combo['values'] = headers
                 # Update all existing sub data comboboxes
                for _, _, st_combo in self.data_column_widgets:
                     st_combo['values'] = headers

            # Clear current selections when headers change
            # self.bt_key_combo.set('')
            # self.st_key_combo.set('')
            # for _, bt_combo, st_combo in self.data_column_widgets:
            #      bt_combo.set('')
            #      st_combo.set('')

        except Exception as e:
            messagebox.showerror("错误", f"无法从 '{os.path.basename(file_path)}' 的工作表 '{sheet_name}' 读取表头:\n{e}")
            self._clear_headers(table_type)


    def _browse_file(self, entry_widget, sheet_combo, table_type):
        """Opens file dialog, updates entry, loads sheets and initial headers."""
        type_map = {'base': '原表', 'sub': '内容表'}
        filepath = filedialog.askopenfilename(
            title=f"选择 {type_map.get(table_type, table_type.capitalize())} Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls")]
        )
        if filepath:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filepath)
            # Load sheets and headers for the active sheet
            self._load_sheets_and_headers(filepath, sheet_combo, table_type)
        else:
            # If user cancels browse, clear the path and related widgets
             entry_widget.delete(0, tk.END)
             self._load_sheets_and_headers(None, sheet_combo, table_type)

    def _open_file(self, file_path):
         """Opens the specified file using the default application."""
         try:
             # Ensure path is absolute for OS calls
             abs_path = os.path.abspath(file_path)
             if platform.system() == "Windows":
                 os.startfile(abs_path)
             elif platform.system() == "Darwin": # macOS
                 subprocess.call(["open", abs_path])
             else: # Linux variants
                 subprocess.call(["xdg-open", abs_path])
         except Exception as e:
             messagebox.showerror("打开文件失败", f"无法自动打开文件: {e}")
             print(f"无法打开文件: {e}")


    def _start_processing(self):
        """Gets inputs, validates, calls core logic, and updates results."""
        base_path = self.base_file_entry.get()
        sub_path = self.sub_file_entry.get()
        bt_sheet = self.base_sheet_combo.get()
        st_sheet = self.sub_sheet_combo.get()
        bt_key_col = self.bt_key_combo.get()
        st_key_col = self.st_key_combo.get()

        # Gather data column pairs
        bt_data_cols = []
        st_data_cols = []
        has_incomplete_data_pair = False
        for i, (_, bt_combo, st_combo) in enumerate(self.data_column_widgets):
            bt_data = bt_combo.get()
            st_data = st_combo.get()
            if bt_data and st_data: # Only add pair if both are selected
                bt_data_cols.append(bt_data)
                st_data_cols.append(st_data)
            elif bt_data or st_data: # If one is selected but not the other
                has_incomplete_data_pair = True

        # --- Input Validation ---
        if not base_path or not sub_path:
            messagebox.showerror("错误", "请选择原表和内容表 Excel 文件。")
            return
        if not bt_sheet or not st_sheet:
             messagebox.showerror("错误", "请为两个表格选择工作表。")
             return
        if not bt_key_col or not st_key_col:
             messagebox.showerror("错误", "请为两个表格选择关键字列。")
             return
        
        # 检查是否有数据列或启用了标记模式
        if not bt_data_cols and not self.enable_tagging.get():
            messagebox.showerror("错误", "请至少选择一对完整的数据列或启用匹配标记功能。")
            return
        
        if bt_data_cols and not st_data_cols:
            messagebox.showerror("错误", "请为每个原表数据列选择对应的内容表数据列。")
            return
            
        if has_incomplete_data_pair:
            messagebox.showwarning("警告", "检测到未完整选择的数据列对（即原表或内容表的数据列未选），这些不完整的列将被忽略。")

        # Key/Data column name validation (optional but good practice)
        if bt_key_col in bt_data_cols:
             messagebox.showwarning("警告", f"原表的关键字列 ('{bt_key_col}') 也被选作了数据列。确定要这样操作吗？")
        if st_key_col in st_data_cols:
             messagebox.showwarning("警告", f"内容表的关键字列 ('{st_key_col}') 也被选作了数据列。确定要这样操作吗？")
        # Check for duplicate data columns selections for the same table
        if len(bt_data_cols) != len(set(bt_data_cols)):
             messagebox.showwarning("警告", "原表中选择了重复的数据列。请检查选择。")
             # Allow proceeding, but warn
        if len(st_data_cols) != len(set(st_data_cols)):
             messagebox.showwarning("警告", "内容表中选择了重复的数据列。请检查选择。")
             # Allow proceeding, but warn


        # --- Disable UI during processing ---
        self.start_button.config(state=tk.DISABLED)
        self.add_col_button.config(state=tk.DISABLED)
        self.delete_col_button.config(state=tk.DISABLED)
        self.fuzzy_checkbox.config(state=tk.DISABLED)
        self.advanced_fuzzy_checkbox.config(state=tk.DISABLED)
        self.tagging_checkbox.config(state=tk.DISABLED)
        self.progress_label.config(text="准备开始处理...")
        self.master.update_idletasks() # Force UI update

        # --- Call Core Logic ---
        # 获取最大匹配条数
        try:
            max_matches_input = self.max_matches_entry.get().strip()
            if max_matches_input == "":
                max_matches = 3  # 如果输入框为空，使用默认值
            else:
                max_matches = int(max_matches_input)
                if max_matches <= 0:
                    max_matches = 3  # 如果输入值无效，使用默认值
        except ValueError:
            max_matches = 3  # 如果输入值无效，使用默认值
        
        total_rows, matched_count, sheet_stats, error_msg = process_files(
            base_path, sub_path, bt_sheet, st_sheet, bt_key_col, st_key_col, bt_data_cols, st_data_cols, self.use_fuzzy_match.get(), self.enable_tagging.get(), self.tag_value_entry.get(), self.use_advanced_fuzzy.get(), max_matches, self.update_progress
        )

        # --- Update Results ---
        if error_msg:
            self.progress_label.config(text="处理失败")
            messagebox.showerror("处理错误", error_msg)
        else:
            # 构建结果文本
            if bt_sheet == ALL_SHEETS_TEXT:
                # 显示每个sheet的处理结果
                result_text = f"处理完成！\n\n"
                if self.use_fuzzy_match.get():
                    result_text += f"【模糊匹配模式】\n"
                result_text += f"总行数: {total_rows}\n成功匹配并更新的总行数: {matched_count}\n\n各工作表详细情况:\n"
                
                # 按工作表名称排序，并添加详细结果
                for sheet_name in sorted(sheet_stats.keys()):
                    stats = sheet_stats[sheet_name]
                    sheet_result = f"工作表 '{sheet_name}': "
                    
                    if stats['status'] == "处理成功":
                        sheet_result += f"检查 {stats['rows']} 行，匹配更新 {stats['matches']} 行"
                    else:
                        sheet_result += stats['status']
                        
                    result_text += f"{sheet_result}\n"
                    
                result_text += f"\n更新的数据列对数量: {len(bt_data_cols)}"
                if self.use_fuzzy_match.get():
                    result_text += f"\n\n模糊匹配说明:\n• 军绿色加粗：优先级2匹配（符号转换、数字格式等）\n• 深海蓝加粗：优先级3匹配（包含关系）\n• 红色加粗：短文本包含匹配（2字以内）"
                result_text += f"\n\n结果已保存至: {os.path.basename(base_path)}"
            else:
                result_text = f"处理完成！\n\n"
                if self.use_fuzzy_match.get():
                    result_text += f"【模糊匹配模式】\n"
                result_text += (f"原表 '{bt_sheet}' 中检查的总行数: {total_rows}\n"
                              f"成功匹配并更新的行数: {matched_count}\n"
                              f"更新的数据列对数量: {len(bt_data_cols)}")
                if self.use_fuzzy_match.get():
                    result_text += f"\n\n模糊匹配说明:\n• 军绿色加粗：优先级2匹配（符号转换、数字格式等）\n• 深海蓝加粗：优先级3匹配（包含关系）\n• 红色加粗：短文本包含匹配（2字以内）"
                result_text += f"\n\n结果已保存至: {os.path.basename(base_path)}"
                              
            self.progress_label.config(text="处理完成")
            
            # 使用自定义对话框显示结果并询问是否打开文件
            dialog = ResultDialog(self.master, "处理完成", result_text, base_path, matched_count)
            result = dialog.show()
            
            # 根据用户选择决定是否打开文件
            if result == "open":
                self._open_file(base_path)


        # --- Re-enable UI ---
        self.start_button.config(state=tk.NORMAL)
        self.add_col_button.config(state=tk.NORMAL)
        # 只有在有数据列时才启用删除按钮
        if self.data_column_widgets:
            self.delete_col_button.config(state=tk.NORMAL)
        if self.use_fuzzy_match.get():
            self.advanced_fuzzy_checkbox.config(state=tk.NORMAL)
        else:
            self.advanced_fuzzy_checkbox.config(state=tk.DISABLED)
        self.fuzzy_checkbox.config(state=tk.NORMAL)
        self.tagging_checkbox.config(state=tk.NORMAL)

        # Reset progress bar
        self.progress_bar["value"] = 0

    def update_progress(self, status_text, processed_rows, total_rows):
        """Updates the progress bar and label with row-based progress."""
        if total_rows > 0:
            progress = (processed_rows / total_rows) * 100
            self.progress_bar["value"] = progress
            self.progress_label.config(text=f"{status_text} - 进度: {progress:.1f}% ({processed_rows}/{total_rows})")
            self.master.update_idletasks()

    def _on_fuzzy_match_toggle(self):
        """处理模糊匹配复选框的切换"""
        if self.use_fuzzy_match.get():
            # 启用模糊匹配时，启用高级模糊匹配选项
            self.advanced_fuzzy_checkbox.config(state=tk.NORMAL)
        else:
            # 禁用模糊匹配时，禁用并取消选中高级模糊匹配选项
            self.advanced_fuzzy_checkbox.config(state=tk.DISABLED)
            self.use_advanced_fuzzy.set(False)

        # --- Re-enable UI ---
        self.start_button.config(state=tk.NORMAL)
        self.add_col_button.config(state=tk.NORMAL)
        # 只有在有数据列时才启用删除按钮
        if self.data_column_widgets:
            self.delete_col_button.config(state=tk.NORMAL)
        if self.use_fuzzy_match.get():
            self.advanced_fuzzy_checkbox.config(state=tk.NORMAL)
        else:
            self.advanced_fuzzy_checkbox.config(state=tk.DISABLED)
        self.fuzzy_checkbox.config(state=tk.NORMAL)
        self.tagging_checkbox.config(state=tk.NORMAL)

    def _validate_max_matches(self, value):
        """Validate max matches input to ensure only positive integers or empty string are allowed."""
        if value == "" or value.isdigit():
            return True
        return False


# --- Main Execution ---
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparatorApp(root)
    root.mainloop()