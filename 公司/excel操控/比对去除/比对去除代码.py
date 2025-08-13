from openpyxl import load_workbook
import shutil

tag_tab = "C:\\Users\\fucha\\Desktop\\现有介质.xlsx"
sub_tab = "C:\\Users\\fucha\\Desktop\\2025年3月份引入需求.xlsx"
# 总表比对列数
bt_index = 2
# 比对表比对列数
st_index = 3
# 输入要标记的标签
tag_value = "T"
# 加载表格
wb_total_filtered = load_workbook(tag_tab)
bt_sheet = wb_total_filtered.active
wb_sub = load_workbook(sub_tab)
st_sheet = wb_sub.active
# 在 total_filtered.xlsx 最后增加一列
# bt_sheet.cell(row=1, column=bt_sheet.max_column + 1, value='')

# 数组比对，取id
def compare(bt_sheet, st_sheet):
    bt_ids = [bt_sheet.cell(row=row_index, column=bt_index).value for row_index in range(2, bt_sheet.max_row + 1)]
    st_ids = [st_sheet.cell(row=row_index, column=st_index).value for row_index in range(2, st_sheet.max_row + 1)]
    print(f'{bt_ids} \n total: {len(bt_ids)} ')
    print(f'{st_ids} \n total: {len(st_ids)} ')

    diff_ids = list(set(bt_ids) & set(st_ids))
    print(f'{diff_ids} \n total: {len(diff_ids)} ')
# 字典比对，取行号+id
def diff_id(bt_sheet, st_sheet):

    bt_ids = {bt_sheet.cell(row=row_index, column=bt_index).value: row_index for row_index in range(2, bt_sheet.max_row + 1)}
    st_ids = {st_sheet.cell(row=row_index, column=st_index).value: row_index for row_index in range(2, st_sheet.max_row + 1)}
    # print(f'{bt_ids} \n total: {len(bt_ids)} ')
    # print(f'{st_ids} \n total: {len(st_ids)} ')

    diff_ids = list(set(bt_ids) & set(st_ids))
    # print(f'{diff_ids} \n total: {len(diff_ids)} ')

    diff_rows = [bt_ids[i] for i in diff_ids]
    # print(f'{diff_rows} \n total: {len(diff_rows)} ')
    return diff_rows, diff_ids
# 最后一列打标签
def tab_mark(diff_rows):
    n = 0
    for i in diff_rows:
        bt_sheet.cell(row=i, column=bt_sheet.max_column, value=tag_value)
        n += 1
    print(n)
# compare(bt_sheet, st_sheet)
diff_rows, diff_ids = diff_id(bt_sheet, st_sheet)
tab_mark(diff_rows)
# 保存结果
wb_total_filtered.save(tag_tab)