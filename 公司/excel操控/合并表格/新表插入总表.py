from openpyxl import load_workbook
import pypinyin

def get_first_letters(text):
    result = ""
    for char in text:
        pinyin = pypinyin.pinyin(char, style=pypinyin.NORMAL)[0]
        if pinyin:
            first_letter = pinyin[0][0].upper()
            result += first_letter
    return result

def copy_data(main_sheet, patch_sheet, row, patch_id):
    new_row = main_sheet.max_row + 1
    A = get_first_letters(patch_sheet.cell(row=row, column=2).value)
    main_sheet.cell(row=new_row, column=3).value = patch_sheet.cell(row=row, column=2).value
    main_sheet.cell(row=new_row, column=4).value = patch_sheet.cell(row=row, column=2).value
    main_sheet.cell(row=new_row, column=10).value = patch_sheet.cell(row=row, column=10).value
    main_sheet.cell(row=new_row, column=11).value = patch_sheet.cell(row=row, column=19).value
    main_sheet.cell(row=new_row, column=12).value = patch_sheet.cell(row=row, column=60).value
    main_sheet.cell(row=new_row, column=13).value = patch_sheet.cell(row=row, column=16).value
    main_sheet.cell(row=new_row, column=16).value = patch_sheet.cell(row=row, column=12).value
    main_sheet.cell(row=new_row, column=18).value = patch_sheet.cell(row=row, column=14).value
    view = patch_sheet.cell(row=row, column=15).value
    if view:
        main_sheet.cell(row=new_row, column=21).value = "【本视频来源于华视网聚】" + patch_sheet.cell(row=row, column=15).value
    main_sheet.cell(row=new_row, column=34).value = A
    main_sheet.cell(row=new_row, column=35).value = patch_sheet.cell(row=row, column=55).value
    main_sheet.cell(row=new_row, column=39).value = 2
    main_sheet.cell(row=new_row, column=9).value = "MOV"
    main_sheet.cell(row=new_row, column=1).value = 1
    main_sheet.cell(row=new_row, column=2).value = 1

def merge_excels(main_file, patch_file):
    # 加载主表和补丁表
    main_wb = load_workbook(main_file)
    main_sheet = main_wb.active
    patch_wb = load_workbook(patch_file)
    patch_sheet = patch_wb.active

    # 遍历补丁表的每一行
    for row in range(2, patch_sheet.max_row + 1):
        patch_id = patch_sheet.cell(row=row, column=1).value
        copy_data(main_sheet, patch_sheet, row, patch_id)
    # 保存主表
    main_wb.save(main_file)

# 指定文件名
main_file = "C:\\Users\\fucha\\Desktop\\甘肃电影注入模板.xlsx"
patch_file = "C:\\Users\\fucha\\Desktop\\甘肃优先注入表单.xlsx"
merge_excels(main_file, patch_file)



