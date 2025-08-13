# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, Fill, PatternFill
from datetime import datetime

xlsx = "C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\新片表格模板.xlsx"
out_xlsx = "C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\out.xlsx"
test_xlsx = "C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\上传测试表格.xlsx"
summary_xlsx = "C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\新片汇总.xlsx"

workbook = load_workbook(xlsx)
sheet1 = workbook.active
workbook = load_workbook(out_xlsx)
sheet2 = workbook.active
workbook_test = load_workbook(test_xlsx)
sheet3 = workbook_test.active

max_row = 2
Header = {1: '序号', 2: '节目ID', 3: '节目名称', 4: '节目类型', 5: '内容类型', 6: '出品地区', 7: '出品年代', 8: '导演', 9: '主演', 10: '集数', 11: '单集时长', 12: '上线时间', 13: '发行许可证号', 14: '院线上映时间', 15: '免费播出时间', 16: '授权开始时间', 17: '授权结束时间', 18: '是否独家', 19: '表格更新日期'}

# 时间字段列表
time_fields = [14, 15, 16, 17, 12]  # 对应院线上映时间、免费播出时间、授权开始时间、授权结束时间、上线时间
# 是否独家字段
exclusive_field = 18

# 设置字体为等线
font = Font(name='等线')
# 设置对齐方式为居中
alignment = Alignment(horizontal='center', vertical='center')
# 设置边框样式
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def format_time(time_str):
    if not time_str:
        return None
    try:
        # 尝试解析不同格式的时间
        if isinstance(time_str, datetime):
            return time_str.strftime('%Y-%m-%d %H:%M:%S')
        
        # 处理字符串格式的时间
        time_str = str(time_str).strip()
        
        # 处理日期和时间分开的情况（用换行符分隔）
        if '\n' in time_str:
            date_part, time_part = time_str.split('\n', 1)
            date_part = date_part.strip()
            time_part = time_part.strip()
            
            # 尝试不同的日期格式
            try:
                date = datetime.strptime(date_part, '%Y-%m-%d')
            except:
                try:
                    date = datetime.strptime(date_part, '%Y/%m/%d')
                except:
                    return None
            
            # 尝试不同的时间格式
            try:
                time = datetime.strptime(time_part, '%H:%M:%S')
                return datetime.combine(date.date(), time.time()).strftime('%Y-%m-%d %H:%M:%S')
            except:
                return datetime.combine(date.date(), datetime.min.time()).strftime('%Y-%m-%d 00:00:00')
        
        # 处理只有时间的情况
        if ':' in time_str and '-' not in time_str and '/' not in time_str:
            return datetime.strptime(time_str, '%H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
        
        # 处理包含空格的情况
        if ' ' in time_str:
            date_part, time_part = time_str.split(' ', 1)
            if ':' in time_part:
                # 尝试不同的日期格式
                try:
                    return datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                except:
                    try:
                        return datetime.strptime(time_str, '%Y/%m/%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        return datetime.strptime(date_part, '%Y-%m-%d').strftime('%Y-%m-%d 00:00:00')
            else:
                # 尝试不同的日期格式
                try:
                    return datetime.strptime(date_part, '%Y-%m-%d').strftime('%Y-%m-%d 00:00:00')
                except:
                    return datetime.strptime(date_part, '%Y/%m/%d').strftime('%Y-%m-%d 00:00:00')
        else:
            # 尝试不同的日期格式
            try:
                return datetime.strptime(time_str, '%Y-%m-%d').strftime('%Y-%m-%d 00:00:00')
            except:
                return datetime.strptime(time_str, '%Y/%m/%d').strftime('%Y-%m-%d 00:00:00')
    except:
        return None

def format_exclusive_status(value):
    if not value:
        return None
    value = str(value).strip()
    if value == '独家':
        return '是'
    elif value == '非独':
        return '否'
    return value

def format_exclusive_status_test(value):
    if not value:
        return None
    value = str(value).strip()
    if value == '是':
        return '1'
    elif value == '否':
        return '0'
    return value

#   判断表头行
def row_number():
    row_numbers = []    # [1, 4, 7, 10, 13]
    for row in sheet1.iter_rows(min_row=1, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value == "节目ID":
            row_numbers.append(row[0].row)
    print(row_numbers)
    return row_numbers
#   判断每一表头行对应内容行数
def result_dics(row_numbers):
    result_dic = {}     # {1: 1, 4: 1, 7: 1, 10: 1, 13: 2}
    for i in row_numbers:
        temple = 0
        empty = 0
        for row in range(i, sheet1.max_row + 1):
            cell_value = sheet1.cell(row=row, column=1).value
            if cell_value != "节目ID" and cell_value is not None:
                temple += 1
            if cell_value is None:
                empty += 1
            if empty == 1:
                break
        result_dic[i] = temple
    print(result_dic)
    return result_dic
#   目标表格表头
def test_headers():
    test_header = {}
    for col_idx, cell in enumerate(sheet1[row], start=1):
        test_header[col_idx] = cell.value
    return test_header
#   表头比对
def new_dicts(test_header):
    new_dict = {}
    for test_key, test_value in test_header.items():
        for header_key, header_value in Header.items():
            if test_value == header_value:
                new_dict[header_key] = test_key
                break
    return new_dict
#   目标表格内容
def temple_dics():
    temple_dic = {}
    for col_idx, cell in enumerate(sheet1[row], start=1):
        temple_dic[col_idx] = cell.value
    # print(temple_dic)
    return temple_dic
#   最终表格表格内容
def final_dicts(temple_dic):
    final_dict = {}
    for key2, value2 in temple_dic.items():
        if key2 in new_dict.values():
            for key1, value1 in new_dict.items():
                if value1 == key2:
                    final_dict[key1] = value2
                    break
    # print(final_dict)
    return final_dict

# ------main------ #
# 加载或创建汇总文件
if os.path.exists(summary_xlsx):
    workbook_summary = load_workbook(summary_xlsx)
    sheet_summary = workbook_summary.active
else:
    workbook_summary = Workbook()
    sheet_summary = workbook_summary.active

# 获取汇总文件的最后一行序号
last_row = sheet_summary.max_row
last_sequence = 0
if last_row > 1:  # 如果有数据行
    last_sequence = last_row - 1  # 序号从当前行数-1开始

# 检查是否达到Excel行数限制
MAX_EXCEL_ROWS = 1048576
if last_row >= MAX_EXCEL_ROWS:
    print(f"警告：汇总文件已达到Excel最大行数限制（{MAX_EXCEL_ROWS}行）")
    print("建议创建新的汇总文件或清理旧数据")
    exit(1)

# 清空输出表格
for row in sheet2.iter_rows(min_row=2):
    for cell in row:
        cell.value = None
workbook.save(out_xlsx)

# 清空测试表格
for row in sheet3.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

row_numbers = row_number()
result_dic = result_dics(row_numbers)
current_date = datetime.now().strftime('%Y/%m/%d')
new_row_count = 0  # 记录新增的行数

for row in row_numbers:
    test_header = test_headers()
    new_dict = new_dicts(test_header)
    time = result_dic[row]
    i = 1
    while i <= time:
        row += 1
        i += 1
        temple_dic = temple_dics()
        final_dict = final_dicts(temple_dic)
        # 写入out表格
        for col_num, value in final_dict.items():
            # 对时间字段进行格式化
            if col_num in time_fields:
                value = format_time(value)
            # 对是否独家字段进行格式化
            elif col_num == exclusive_field:
                value = format_exclusive_status(value)
            cell = sheet2.cell(row=max_row, column=col_num, value=value)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            
            # 同时写入测试表格（跳过序号字段）
            if col_num != 1:  # 跳过序号字段
                if col_num == exclusive_field:
                    value = format_exclusive_status_test(value)
                # 调整列号（因为跳过了序号字段，所以其他字段的列号需要减1）
                adjusted_col = col_num - 1
                cell_test = sheet3.cell(row=max_row, column=adjusted_col, value=value)
                cell_test.font = font
                cell_test.alignment = alignment
                cell_test.border = thin_border
        
        # 添加iptv和ott字段
        sheet3.cell(row=max_row, column=18, value='1')  # iptv
        sheet3.cell(row=max_row, column=19, value='1')  # ott
        
        # 写入汇总表格
        summary_row = last_row + max_row - 1
        if summary_row > MAX_EXCEL_ROWS:
            print(f"警告：无法添加更多数据，已达到Excel最大行数限制（{MAX_EXCEL_ROWS}行）")
            print("建议创建新的汇总文件或清理旧数据")
            break
            
        for col_num, value in final_dict.items():
            if col_num == 1:  # 序号列
                value = summary_row - 1  # 序号为当前行数-1
            elif col_num == 19:  # 表格更新日期列
                value = current_date
            elif col_num in time_fields:
                value = format_time(value)
            elif col_num == exclusive_field:
                value = format_exclusive_status(value)
            
            cell = sheet_summary.cell(row=summary_row, column=col_num, value=value)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
        
        max_row += 1
        new_row_count += 1

workbook.save(out_xlsx)
workbook_test.save(test_xlsx)
workbook_summary.save(summary_xlsx)

# 自动打开输出文件
os.startfile(test_xlsx)
os.startfile(summary_xlsx)