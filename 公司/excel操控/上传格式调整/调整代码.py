import shutil
import openpyxl
import xlrd
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
# 上述部分为导入所需的库和模块，包括文件操作、Excel 处理和相关样式设置

def read_txt(txt_file):
    """
    函数：read_txt
    作用：以 UTF-8 编码方式读取指定的文本文件，并按行返回内容
    参数：txt_file - 要读取的文本文件路径
    """
    with open(txt_file, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    return lines
# 该函数用于读取文本文件并返回每一行的内容
def write_xls(lines,xlsx_file):
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    # 定义了一个边框样式，用于后续设置单元格边框

    file_end = False
    title_fields = lines[0].strip().split('\t')
    title_len = 0
    rec_type = "E"
    # 初始化一些变量，包括文件结束标志、标题字段、标题长度和记录类型

    with (open('d:\\out.txt', 'w', encoding='utf-8') as file2):
        # 打开一个新的文本文件用于写入中间处理结果
        row_idx = 0
        row_xls = 1
        while not file_end:
            # 开始一个循环，直到文件结束
            line = lines[row_idx].strip()

            if line == "":
                # 如果当前行为空，跳过并处理下一行
                row_idx = row_idx + 1
                if row_idx >= len(lines):
                    file_end = True
                continue

            if line.startswith("序号") or line.startswith("节目ID") or line.startswith("【标题行】"):
                # 处理标题行相关逻辑
                title_fields = line.strip().split('\t')
                title_len = len(title_fields)
                print(title_len, title_fields)
                rec_type = "E"
                if title_len == 13:
                    rec_type = "C"
                if title_len == 14:
                    rec_type = "D"
                if title_len == 15:
                    rec_type = "F"
                if title_len == 17:
                    if title_fields[0] == "序号":
                        rec_type = "A"
                    else:
                        rec_type = "B"
                row_idx = row_idx + 1
                if row_idx >= len(lines):
                    file_end = True
                continue

            if line.startswith("【结束】") or file_end:
                # 处理结束标记
                break

            fields = line.strip().split('\t')
            while len(fields) < title_len:
                # 处理字段长度不足的情况
                row_idx = row_idx + 1
                if row_idx >= len(lines):
                    file_end = True
                    break
                line = line + " " + lines[row_idx].strip()
                fields = line.strip().split('\t')
            row_xls = row_xls + 1
            sheet.row_dimensions[row_xls].height = sheet.row_dimensions[2].height
            file2.write(line + "\n")
            col_add = 0
            for col_idx, field in enumerate(fields,start=1):
                # 根据不同的记录类型进行特定的处理
                if rec_type == "A" and col_idx == 1:
                    col_add = -1
                    continue
                if rec_type == "C":
                    if col_idx == 13:
                        continue
                    if col_idx == 7 or col_idx == 11:
                        cell = sheet.cell(row=row_xls, column=col_idx + col_add, value="")
                        cell.border = border
                        col_add = col_add + 1
                        cell = sheet.cell(row=row_xls, column=col_idx + col_add, value="")
                        cell.border = border
                        col_add = col_add + 1
                if rec_type == "F":
                    if col_idx == 11:
                        col_add = -1
                        continue
                    if col_idx == 14:
                        cell = sheet.cell(row=row_xls, column=col_idx + col_add, value="")
                        cell.border = border
                        col_add = col_add + 1
                        cell = sheet.cell(row=row_xls, column=col_idx + col_add, value="")
                        cell.border = border
                        col_add = col_add + 1

                if rec_type == "B":
                    if col_idx == 17:
                        continue

                if rec_type == "D" and col_idx == 13:
                    cell = sheet.cell(row=row_xls, column=col_idx + col_add, value="")
                    cell.border = border
                    col_add = col_add + 1
                    cell = sheet.cell(row=row_xls, column=col_idx + col_add, value="")
                    cell.border = border
                    col_add = col_add + 1

                cell = sheet.cell(row=row_xls, column=col_idx + col_add, value=field)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            row_idx = row_idx + 1
            if row_idx >= len(lines):
                file_end = True

        workbook.save(xlsx_file)
    # 该函数实现了将文本内容按规则写入 Excel 文件，并处理了各种条件下的单元格设置

txt_file = "C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\新建 文本文档.txt"
xls_template ="C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\新片导入模板.xlsx"
xls_file = "C:\\Users\\fucha\\Desktop\\公司\\4.代码表格\\上传表格\\out.xlsx"
shutil.copyfile(xls_template, xls_file)
# 复制 Excel 模板文件

lines = read_txt(txt_file)
write_xls(lines,xls_file)
# 读取文本文件并调用写入函数将内容写入新的 Excel 文件