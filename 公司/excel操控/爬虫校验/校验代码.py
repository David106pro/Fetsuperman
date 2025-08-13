# -*- coding: utf-8 -*-
from openpyxl import load_workbook

file_path = "C:\\Users\\fucha\\Desktop\\豆瓣爬取模板.xlsx"
wb = load_workbook(file_path)
ws = wb.active
# 数组拆分比对
def compare_actor(entry1, entry2):
    if not entry1 or not entry2:
        return 0
    array1 = [item.strip() for item in entry1.split(',')]
    array2 = [item.strip() for item in entry2.split('/')]
    count = 0
    for item1 in array1:
        for item2 in array2:
            if item1 == item2:
                count += 1
    return count
# 演员比对
def diff_actor(ws, row_index):
    actor1 = ws.cell(row=row_index, column=6).value
    actor2 = ws.cell(row=row_index, column=19).value
    count = compare_actor(actor1, actor2)
    print(f"第 {row_index + 1} 行的相同元素个数: {count}")
    ws.cell(row=row_index, column=29, value=count)  # 将计数器的值写入第 14 列
#   比对函数
def compare_columns(index1, index2, target_index):
    if index1 == 2:
        value1 = ws.cell(row=row, column=index1).value
        if value1 is None:
            value1 = ws.cell(row=row, column=3).value
    else:
        value1 = ws.cell(row=row, column=index1).value
    value2 = ws.cell(row=row, column=index2).value
    if value1 is None or value2 is None:
        ws.cell(row=row, column=target_index).value = 0
    elif value1 == value2:
        ws.cell(row=row, column=target_index).value = 1
    else:
        ws.cell(row=row, column=target_index).value = 0
#   四组比对
def final_compare(row):
    value5 = ws.cell(row=row, column=29).value
    value1 = ws.cell(row=row, column=30).value
    value2 = ws.cell(row=row, column=31).value
    value3 = ws.cell(row=row, column=32).value
    value4 = ws.cell(row=row, column=33).value
    if value1 == value2 == value3 == value4 == value5 == 0:
        ws.cell(row=row, column=34).value = 0

#——————主函数——————#
for row in range(2, ws.max_row+1):
    movie_state = ws.cell(row=row, column=28).value
    if movie_state != "爬取失败":
        diff_actor(ws, row)
        compare_columns(2, 13, 30)   #电影名称
        compare_columns(7, 15, 31)   #电影年代
        compare_columns(8, 16, 32)   #电影地区
        compare_columns(5, 18, 33)   #电影导演
        final_compare(row)

        temple = ws.cell(row=row, column=1).value
    if temple is None:
        break
wb.save(file_path)