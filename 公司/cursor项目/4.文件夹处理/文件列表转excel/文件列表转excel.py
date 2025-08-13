import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os
import openpyxl
from openpyxl import Workbook
import subprocess
import platform
from tkinter import ttk  # 添加 ttk 导入

class FileListToExcel:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("文件列表转Excel工具")
        self.window.geometry("600x450")
        self.window.resizable(False, False)
        
        # 创建主容器
        main_frame = tk.Frame(self.window, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建标题
        title_label = tk.Label(main_frame, text="文件列表转Excel工具", 
                             font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20))
        
        # 创建路径选择框架
        self.path_frame = tk.LabelFrame(main_frame, text="文件夹选择", padx=10, pady=10)
        self.path_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.path_entry = tk.Entry(self.path_frame, width=50)
        self.path_entry.pack(side=tk.LEFT, padx=(5, 10))
        
        self.select_button = tk.Button(self.path_frame, text="选择文件夹", 
                                     command=self.select_folder,
                                     width=12)
        self.select_button.pack(side=tk.LEFT)
        
        # 创建删除字段设置框架
        delete_settings_frame = tk.LabelFrame(main_frame, text="删除字段设置（每行一个）", 
                                            padx=10, pady=10)
        delete_settings_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # 创建文本框和滚动条
        text_frame = tk.Frame(delete_settings_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.delete_text = tk.Text(text_frame, height=10, width=50, 
                                 yscrollcommand=scrollbar.set)
        self.delete_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=self.delete_text.yview)
        
        # 创建操作按钮框架
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        self.start_button = tk.Button(button_frame, text="开始生成Excel", 
                                    command=self.generate_excel,
                                    width=15, height=2,
                                    bg="#4CAF50", fg="white",
                                    font=("Arial", 10, "bold"))
        self.start_button.pack()
        
        self.window.mainloop()
    
    def select_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, folder_path)
    
    def get_delete_fields(self):
        # 获取文本框中的内容，按行分割，去除空行
        text_content = self.delete_text.get("1.0", tk.END).strip()
        fields = [line.strip() for line in text_content.split('\n') if line.strip()]
        return fields

    def open_excel(self, excel_path):
        if platform.system() == 'Darwin':       # macOS
            subprocess.call(('open', excel_path))
        elif platform.system() == 'Windows':    # Windows
            os.startfile(excel_path)
        else:                                   # linux
            subprocess.call(('xdg-open', excel_path))

    def generate_excel(self):
        folder_path = self.path_entry.get()
        if not folder_path:
            messagebox.showerror("错误", "请先选择文件夹！")
            return
        
        try:
            # 获取删除字段
            delete_fields = self.get_delete_fields()
            
            # 创建新的Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "文件列表"
            
            # 添加表头
            ws['A1'] = "文件名称"
            
            # 获取文件列表并写入Excel
            row = 2
            file_count = 0
            for filename in os.listdir(folder_path):
                if os.path.isfile(os.path.join(folder_path, filename)):
                    processed_name = filename
                    # 删除指定字段
                    for field in delete_fields:
                        processed_name = processed_name.replace(field, '')
                    
                    ws[f'A{row}'] = processed_name
                    row += 1
                    file_count += 1
            
            # 生成Excel文件名
            folder_name = os.path.basename(folder_path)
            excel_filename = f"{folder_name}_{file_count}.xlsx"
            excel_path = os.path.join(folder_path, excel_filename)
            
            # 保存Excel文件
            wb.save(excel_path)
            
            messagebox.showinfo("成功", f"Excel文件已生成！\n保存路径：{excel_path}")
            
            # 自动打开Excel文件
            self.open_excel(excel_path)
            
        except Exception as e:
            messagebox.showerror("错误", f"生成Excel文件时出错：\n{str(e)}")

if __name__ == "__main__":
    FileListToExcel()
