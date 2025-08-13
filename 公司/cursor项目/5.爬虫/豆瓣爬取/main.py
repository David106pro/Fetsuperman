# -*- coding: utf-8 -*-
"""
豆瓣影视媒资信息爬取工具
基于之前测试成功的爬虫代码，添加GUI界面和完整功能
增强版：添加反爬检测和智能规避机制
"""

import re
import os
import sys
import json
import time
import random
import string  # 新增：用于生成随机Cookie
import logging
import threading
import traceback
import urllib.parse
from queue import Queue
from datetime import datetime
from difflib import SequenceMatcher  # 新增：用于字符串相似度匹配

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

# 配置日志系统
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('douban_crawler.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# 设置请求头（基础模板）
HEADERS = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Cache-Control': 'no-cache',
    'Pragma': 'no-cache',
    'Sec-Ch-Ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
    'Sec-Ch-Ua-Mobile': '?0', 
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
}

# 表头定义
COLUMN_HEADERS = [
    "爬取名称", "校验状态", "电影名称", "导演", "编剧", "主演", "类型", 
    "制片国家/地区", "语言", "上映日期", "片长", "又名", "IMDb", 
    "剧情简介", "豆瓣评分", "评价人数", "海报URL", "豆瓣ID", "缓存URL"
]


class DoubanCrawler:
    """豆瓣爬虫类，负责数据爬取和解析"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # 扩展User-Agent池 - 包含更多真实浏览器
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/126.0.0.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:127.0) Gecko/20100101 Firefox/127.0'
        ]
        
        # 代理IP池（需要用户自己配置）
        self.proxy_list = []
        self.current_proxy_index = 0
        self.use_proxy = False  # 是否启用代理
        
        # 优化延迟控制 - 参考成功案例的双重延迟策略
        self.base_delay = 3  # 减少基础延迟，参考成功案例的10秒全局延迟
        self.current_delay = 3  # 当前延迟
        self.max_delay = 120  # 最大延迟2分钟
        self.emergency_delay = 180  # 紧急延迟3分钟
        
        # 成功失败计数和状态
        self.success_count = 0
        self.fail_count = 0
        self.consecutive_fails = 0
        self.total_requests = 0
        self.last_success_time = time.time()
        
        # 增强反爬检测关键词
        self.anti_crawler_keywords = [
            '验证码', 'captcha', 'robot', 'blocked', '访问被拒绝', 
            'IP被限制', '请求过于频繁', '安全验证', '人机验证',
            '访问异常', '操作频繁', '稍后再试', 'too many requests',
            '系统繁忙', '访问受限', '请稍后重试', '服务不可用',
            'rate limit', 'forbidden', 'access denied', 'temporarily unavailable',
            '安全检查', '人工检查', '服务器错误', '网络错误'
        ]
        
        # 用于记录有效行映射（只包含非空爬取名称的行）
        self.valid_rows_mapping = {}  # UI行索引 -> Excel实际行号
        self.excel_to_ui_mapping = {}  # Excel行号 -> UI行索引
        
        # 快速模式标志
        self.fast_mode_active = False
        self.fast_mode_count = 0  # 快速模式计数器
        self.quick_anti_crawler_mode = True  # 是否启用快速反爬模式
        
        # 会话管理
        self.session_create_time = time.time()
        self.max_session_age = 1200  # 20分钟重建session（减少频率）
        self.request_count_in_session = 0
        self.max_requests_per_session = 50  # 每个session最多50个请求（更保守）
        
        # 初始化会话
        self.headers = HEADERS.copy()
        self.session = self._create_session()
        
        self.logger.info("豆瓣爬虫初始化完成 - 增强反爬版本v2.0")
    
    def generate_dynamic_bid(self):
        """生成动态豆瓣bid Cookie - 核心反爬技术"""
        # 豆瓣bid格式：11位随机字符串（字母+数字）
        chars = string.ascii_letters + string.digits
        bid = ''.join(random.choice(chars) for _ in range(11))
        self.logger.debug(f"生成新的bid: {bid}")
        return bid
    
    def generate_dynamic_cookies(self):
        """生成动态Cookie集合"""
        cookies = {
            'bid': self.generate_dynamic_bid(),
            'll': '"108288"',  # 北京地区标识，模拟真实用户
        }
        return cookies

    @staticmethod
    def str2urlcode(s):
        """将字符串转换为 URL 编码格式"""
        return urllib.parse.quote(s)
    
    def _create_session(self):
        """创建新的会话 - 核心优化"""
        session = requests.Session()
        
        # 完善的真实浏览器请求头
        enhanced_headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'Accept-Encoding': 'gzip, deflate, br',
            'Cache-Control': 'no-cache',
            'Pragma': 'no-cache',
            'Sec-Ch-Ua': f'"Chromium";v="126", "Not)A;Brand";v="24", "Google Chrome";v="126"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'DNT': '1',
            'Connection': 'keep-alive',
            # 关键：添加豆瓣相关的Origin和Referer
            'Origin': 'https://www.douban.com',
            'Referer': 'https://movie.douban.com/',
        }
        
        # 随机选择User-Agent
        ua = random.choice(self.user_agents)
        enhanced_headers['User-Agent'] = ua
        
        session.headers.update(enhanced_headers)
        
        # 核心：设置动态Cookie
        dynamic_cookies = self.generate_dynamic_cookies()
        session.cookies.update(dynamic_cookies)
        self.logger.info(f"设置动态Cookie: bid={dynamic_cookies['bid']}")
        
        # 设置代理（如果启用）
        if self.use_proxy and self.proxy_list:
            current_proxy = self.proxy_list[self.current_proxy_index % len(self.proxy_list)]
            session.proxies.update(current_proxy)
            self.logger.info(f"使用代理: {list(current_proxy.values())[0][:30]}...")
        
        # 设置超时和连接池参数
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=1,
            pool_maxsize=1,
            max_retries=0
        )
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        
        self.session_create_time = time.time()
        self.request_count_in_session = 0
        
        self.logger.info(f"创建新会话，User-Agent: {ua[:50]}..., Cookie已设置")
        return session

    def _should_rebuild_session(self):
        """判断是否需要重建会话"""
        age = time.time() - self.session_create_time
        return (age > self.max_session_age or 
                self.request_count_in_session > self.max_requests_per_session or
                self.consecutive_fails >= 3)  # 减少到3次连续失败就重建

    def detect_anti_crawler(self, response):
        """增强的反爬检测机制 - 快速检测"""
        if not response:
            return True, "请求失败"
        
        # 快速检测：异常状态码
        if response.status_code in [403, 429, 503, 509, 502, 504]:
            return True, f"异常状态码: {response.status_code}"
        
        # 快速检测：内容长度异常
        content_length = len(response.text)
        if content_length < 2000:  # 进一步降低阈值，更快检测
            return True, f"页面内容异常短: {content_length}字节"
        
        # 快速检测：关键反爬特征（只检查前2000字符）
        content_preview = response.text[:2000].lower()
        
        # 优先检测最常见的反爬关键词
        priority_keywords = ['验证码', 'captcha', 'robot', 'blocked', '访问被拒绝', 
                           'too many requests', '系统繁忙', '访问受限']
        for keyword in priority_keywords:
            if keyword.lower() in content_preview:
                return True, f"检测到反爬关键词: {keyword}"
        
        # 快速检测：检查是否是错误页面
        if any(error in content_preview for error in ['error', '错误', '异常', '失败']):
            # 进一步检查是否真的是错误页面
            full_content = response.text.lower()
            if 'douban' not in full_content and 'movie' not in full_content:
                return True, "可能的错误页面，缺少豆瓣特征"
        
        # 快速检测：异常重定向
        if len(response.history) > 3:
            return True, f"异常重定向次数: {len(response.history)}"
        
        # 对于搜索页面的特殊检测
        if 'search' in response.url:
            if '暂时没有找到' in response.text or '没有找到' in response.text:
                return False, "搜索无结果（正常）"  # 这不是反爬，是真的没找到
        
        # 对于详情页面的快速检测
        if 'subject' in response.url:
            # 检查是否有基本的电影页面结构
            if not any(keyword in content_preview for keyword in ['v:genre', 'rating_num', 'itemreviewed']):
                # 检查完整内容
                full_content = response.text.lower()
                if not any(keyword in full_content for keyword in ['v:genre', 'rating_num', 'itemreviewed', 'movie']):
                    return True, "详情页缺少关键电影信息结构"
        
        return False, "正常"
    
    def enhanced_delay_strategy(self, success=True, fast_mode=False):
        """优化的双重延迟策略 - 参考成功案例"""
        # 双重延迟：随机短延迟 + 基础延迟
        short_delay = random.uniform(0, 1)  # 0-1秒随机延迟，模拟人类操作
        
        if fast_mode:
            # 快速模式：刚处理完反爬，使用更短的延迟
            base_delay = random.uniform(3, 6)  # 3-6秒基础延迟
            self.logger.info("使用快速延迟模式")
        elif success:
            # 成功时使用较短的基础延迟
            base_delay = random.uniform(8, 12)  # 8-12秒基础延迟
        else:
            # 失败时增加延迟
            base_delay = random.uniform(15, 25)  # 15-25秒延迟
        
        total_delay = short_delay + base_delay
        
        # 每10个请求增加一次额外休息（快速模式下跳过）
        if not fast_mode and self.total_requests % 10 == 0 and self.total_requests > 0:
            extra_rest = random.uniform(30, 60)
            total_delay += extra_rest
            self.logger.info(f"第{self.total_requests}个请求，增加额外休息{extra_rest:.1f}秒")
        
        mode_text = "快速模式" if fast_mode else "常规模式"
        self.logger.info(f"延迟策略({mode_text}): 短延迟{short_delay:.1f}s + 基础延迟{base_delay:.1f}s = 总计{total_delay:.1f}s")
        time.sleep(total_delay)

    def adaptive_delay(self, success=True):
        """自适应延迟调整 - 保持现有逻辑并整合新策略"""
        self.total_requests += 1
        
        if success:
            self.success_count += 1
            self.consecutive_fails = 0
            self.last_success_time = time.time()
            # 成功时谨慎减少延迟
            if self.success_count >= 3:  # 减少成功次数要求
                self.current_delay = max(self.base_delay, self.current_delay * 0.95)
        else:
            self.fail_count += 1
            self.consecutive_fails += 1
            self.success_count = 0
            
            # 失败处理 - 更快响应
            if self.consecutive_fails == 1:
                self.current_delay = min(self.current_delay * 1.5, self.max_delay)
            elif self.consecutive_fails == 2:
                self.current_delay = min(self.current_delay * 2, self.max_delay)
            elif self.consecutive_fails >= 3:
                self.current_delay = self.emergency_delay
                self.logger.warning(f"进入紧急模式，延迟{self.emergency_delay}秒")
        
        # 检查是否使用快速模式
        use_fast_mode = self.fast_mode_active and self.fast_mode_count > 0
        if use_fast_mode:
            self.fast_mode_count -= 1
            if self.fast_mode_count <= 0:
                self.fast_mode_active = False
                self.logger.info("快速模式结束，返回常规模式")
        
        # 使用增强的延迟策略
        self.enhanced_delay_strategy(success, fast_mode=use_fast_mode)
    
    def switch_proxy(self):
        """切换代理IP"""
        if self.use_proxy and self.proxy_list:
            self.current_proxy_index = (self.current_proxy_index + 1) % len(self.proxy_list)
            self.logger.info(f"切换到代理 {self.current_proxy_index + 1}/{len(self.proxy_list)}")
    
    def handle_anti_crawler(self, reason):
        """优化的反爬处理机制"""
        self.logger.warning(f"检测到反爬: {reason}")
        
        # 立即重建会话并生成新Cookie
        self.session.close()
        self.session = self._create_session()
        self.logger.info("重建会话并生成新Cookie")
        
        # 切换代理（如果可用）
        if self.use_proxy and self.proxy_list:
            self.switch_proxy()
        
        # 分级等待时间 - 更温和的处理
        if self.consecutive_fails <= 2:
            wait_time = random.uniform(15, 30)
        elif self.consecutive_fails <= 4:
            wait_time = random.uniform(30, 60)
        else:
            wait_time = random.uniform(60, 120)
        
        self.logger.info(f"反爬等待时间: {wait_time:.1f}秒 (连续失败:{self.consecutive_fails}次)")
        time.sleep(wait_time)
    
    def handle_anti_crawler_fast(self, reason):
        """快速反爬处理机制 - 减少等待时间"""
        self.logger.warning(f"快速反爬处理: {reason}")
        
        # 立即重建会话并生成新Cookie
        self.session.close()
        self.session = self._create_session()
        self.logger.info("快速重建会话并生成新Cookie")
        
        # 切换代理（如果可用）
        if self.use_proxy and self.proxy_list:
            self.switch_proxy()
        
        # 快速等待时间 - 大幅减少等待
        wait_time = random.uniform(2, 5)  # 只等待2-5秒
        
        self.logger.info(f"快速反爬等待: {wait_time:.1f}秒")
        time.sleep(wait_time)
        
        # 激活快速模式
        self.fast_mode_active = True
        self.fast_mode_count = 3  # 接下来3次请求使用快速模式
    
    def request_with_intelligent_retry(self, url, max_retries=5):
        """优化的智能重试机制 - 快速反爬处理"""
        self.logger.info(f"请求URL: {url}")
        
        # 检查是否需要重建会话
        if self._should_rebuild_session():
            self.logger.info("会话过期，重建会话")
            self.session.close()
            self.session = self._create_session()
        
        anti_crawler_detected = False
        
        for attempt in range(max_retries):
            try:
                # 如果是重试且上次检测到反爬，立即处理
                if attempt > 0:
                    if anti_crawler_detected:
                        self.logger.info(f"第{attempt+1}次尝试：上次检测到反爬，立即重建会话")
                        self.session.close()
                        self.session = self._create_session()
                        # 快速反爬处理等待
                        quick_wait = random.uniform(3, 8)
                        self.logger.info(f"快速反爬等待: {quick_wait:.1f}秒")
                        time.sleep(quick_wait)
                    else:
                        # 普通重试间隔
                        retry_delay = min(5 * attempt, 15)
                        self.logger.info(f"普通重试等待: {retry_delay}秒")
                        time.sleep(retry_delay)
                
                # 增加请求计数
                self.request_count_in_session += 1
                
                # 在请求前随机短暂停，模拟人类行为
                human_delay = random.uniform(0.5, 2)
                time.sleep(human_delay)
                
                # 发送请求
                response = self.session.get(url, timeout=20, allow_redirects=True)
                self.logger.info(f"响应状态码: {response.status_code}, 内容长度: {len(response.text)}")
                
                # 检测反爬 - 更保守的处理策略
                is_blocked, reason = self.detect_anti_crawler(response)
                if is_blocked:
                    anti_crawler_detected = True
                    self.logger.warning(f"第{attempt+1}次尝试检测到反爬: {reason}")
                    
                    # 调整策略：给予更多机会，降低误判率
                    if attempt == 0:
                        # 第一次检测到反爬，使用快速处理但不过于激进
                        if hasattr(self, 'quick_anti_crawler_mode') and self.quick_anti_crawler_mode:
                            self.logger.info("首次检测到反爬，使用快速处理")
                            self.handle_anti_crawler_fast(reason)
                        else:
                            self.logger.info("检测到反爬，使用常规处理")
                            self.handle_anti_crawler(reason)
                        continue
                    elif attempt < 2:
                        # 前两次重试给予更多耐心
                        self.logger.info(f"第{attempt+1}次反爬重试，继续尝试")
                        continue
                    elif attempt < max_retries - 1:
                        # 后续重试加强处理
                        self.handle_anti_crawler("持续反爬")
                        continue
                    else:
                        self.adaptive_delay(success=False)
                        return None
                
                if response.status_code == 200:
                    self.adaptive_delay(success=True)
                    self.logger.info(f"请求成功: {url}")
                    return response
                else:
                    self.logger.warning(f"异常状态码 {response.status_code}")
                    if attempt < max_retries - 1:
                        continue
                    
            except requests.exceptions.Timeout:
                self.logger.error(f"请求超时 (尝试 {attempt+1}/{max_retries})")
            except requests.exceptions.ConnectionError as e:
                self.logger.error(f"连接错误 (尝试 {attempt+1}/{max_retries}): {e}")
            except Exception as e:
                self.logger.error(f"请求异常 (尝试 {attempt+1}/{max_retries}): {e}")
        
        # 所有重试失败
        self.adaptive_delay(success=False)
        self.logger.error(f"所有重试失败: {url}")
        return None
    
    def get_href_with_smart_matching(self, search_url, target_movie_name):
        """获取电影详情页面链接 - 智能匹配版本"""
        response = self.request_with_intelligent_retry(search_url)
        if not response:
            self.logger.error(f"搜索请求失败: {search_url}")
            return None, None
            
        try:
            soup = BeautifulSoup(response.text, 'html.parser')
            script_tags = soup.find_all('script')
            for script_tag in script_tags:
                if script_tag.string and 'window.__DATA__ =' in script_tag.string:
                    script_content = script_tag.string
                    parts = script_content.split('window.__DATA__ =')  # 修复：完整匹配
                    if len(parts) > 1:
                        json_data = parts[1].split('};')[0] + '}'
                        data = json.loads(json_data)
                        items = data.get('items', [])
                        if items:
                            # 智能匹配：找到最相似的电影
                            best_match = None
                            best_similarity = 0.0
                            best_url = None
                            
                            self.logger.info(f"搜索到 {len(items)} 个结果，开始智能匹配")
                            
                            for i, item in enumerate(items):
                                if 'url' in item and item['url'] and item['url'] != 'N/A':
                                    item_title = item.get('title', '')
                                    if item_title:
                                        similarity = self.calculate_similarity(target_movie_name, item_title)
                                        self.logger.debug(f"结果{i+1}: '{item_title}' 相似度: {similarity:.3f}")
                                        
                                        if similarity > best_similarity:
                                            best_similarity = similarity
                                            best_match = item_title
                                            best_url = item['url']
                            
                            # 设置相似度阈值
                            similarity_threshold = 0.3  # 最低相似度要求
                            
                            if best_url and best_similarity >= similarity_threshold:
                                self.logger.info(f"最佳匹配: '{best_match}' (相似度: {best_similarity:.3f}) -> {best_url}")
                                return best_url, best_match
                            else:
                                self.logger.warning(f"未找到足够相似的匹配项 (最高相似度: {best_similarity:.3f})")
                                # 如果没有足够相似的，返回第一个有效结果作为备选
                                for item in items:
                                    if 'url' in item and item['url'] and item['url'] != 'N/A':
                                        fallback_title = item.get('title', '未知')
                                        self.logger.info(f"使用备选结果: '{fallback_title}' -> {item['url']}")
                                        return item['url'], fallback_title
                                        
        except Exception as e:
            self.logger.error(f"解析搜索结果失败: {e}")
            # 输出部分响应内容用于调试
            if response:
                preview = response.text[:500] + "..." if len(response.text) > 500 else response.text
                self.logger.debug(f"响应内容预览: {preview}")
        
        self.logger.warning(f"未找到电影详情页: {search_url}")
        return None, None
    
    def get_href(self, search_url):
        """获取电影详情页面链接 - 兼容性方法"""
        url, title = self.get_href_with_smart_matching(search_url, "")
        return url
    
    def get_name(self, soup):
        """获取电影名称"""
        try:
            element = soup.find('span', property='v:itemreviewed')
            return element.text if element else 'NO FOUND'
        except:
            return 'NO FOUND'
    
    def get_director(self, soup):
        """获取导演"""
        try:
            element = soup.find('span', class_="attrs")
            if element:
                return re.sub(r' ', '', element.text)
            return 'NO FOUND'
        except:
            return 'NO FOUND'
    
    def get_writer(self, soup):
        """获取编剧"""
        try:
            for span in soup.find_all('span', class_='pl'):
                if span.text.strip() == "编剧:":
                    next_elem = span.next_sibling
                    if next_elem:
                        # 获取编剧信息，可能包含链接
                        writers = []
                        while next_elem and next_elem.name != 'br':
                            if next_elem.name == 'a':
                                writers.append(next_elem.text.strip())
                            elif isinstance(next_elem, str):
                                text = next_elem.strip()
                                if text and text not in ['/', '更多...']:
                                    writers.append(text)
                            next_elem = next_elem.next_sibling
                        return '/'.join([w for w in writers if w and w != '/'])
            return 'NO FOUND'
        except:
            return 'NO FOUND'
    
    def get_actors(self, soup, director):
        """获取主演，去除导演信息"""
        try:
            elements = soup.find_all('span', class_="attrs")
            actors = [elem.text for elem in elements if elem.text]
            actors_str = "/".join(actors)
            actors_str = re.sub(r' ', '', actors_str)
            
            if director != 'NO FOUND':
                director_words = set(director.split('/'))
                actor_words = actors_str.split('/')
                actors_str = "/".join([word for word in actor_words if word not in director_words])
            
            return actors_str if actors_str else 'NO FOUND'
        except:
            return 'NO FOUND'
    
    def get_type(self, soup):
        """获取电影类型"""
        try:
            elements = soup.find_all('span', property='v:genre')
            movie_types = [elem.text for elem in elements]
            return "/".join(movie_types) if movie_types else 'NO FOUND'
        except:
            return 'NO FOUND'
    
    def get_place(self, soup):
        """获取制片国家/地区"""
        try:
            for span in soup.find_all('span', class_='pl'):
                if span.next_sibling and span.text == "制片国家/地区:":
                    return span.next_sibling.strip()
            return "NO FOUND"
        except:
            return "NO FOUND"
    
    def get_language(self, soup):
        """获取语言"""
        try:
            for span in soup.find_all('span', class_='pl'):
                if span.next_sibling and span.text == "语言:":
                    return span.next_sibling.strip()
            return "NO FOUND"
        except:
            return "NO FOUND"
    
    def get_release_date(self, soup):
        """获取上映日期"""
        try:
            elements = soup.find_all('span', property='v:initialReleaseDate')
            if elements:
                dates = [elem.text for elem in elements]
                return '/'.join(dates)
            return "NO FOUND"
        except:
            return "NO FOUND"
    
    def get_runtime(self, soup):
        """获取片长"""
        try:
            element = soup.find('span', property='v:runtime')
            return element.text if element else 'NO FOUND'
        except:
            return 'NO FOUND'
    
    def get_sb_name(self, soup):
        """获取又名"""
        try:
            for span in soup.find_all('span', class_='pl'):
                if span.next_sibling and span.text == "又名:":
                    return span.next_sibling.strip()
            return "NO FOUND"
        except:
            return "NO FOUND"
    
    def get_imdb(self, soup):
        """获取IMDb信息"""
        try:
            for span in soup.find_all('span', class_='pl'):
                if 'IMDb' in span.text:
                    next_elem = span.next_sibling
                    if next_elem and next_elem.name == 'a':
                        return next_elem.text.strip()
                    elif next_elem:
                        return next_elem.strip()
            return "NO FOUND"
        except:
            return "NO FOUND"
    
    def get_summary(self, soup):
        """获取剧情简介"""
        try:
            # 查找剧情简介
            summary_elem = soup.find('span', property='v:summary')
            if summary_elem:
                summary = summary_elem.text.strip()
                # 限制长度，避免过长
                return summary[:200] + '...' if len(summary) > 200 else summary
            
            # 备选方案：查找其他可能的简介元素
            for elem in soup.find_all('span', class_='all'):
                if elem.text:
                    text = elem.text.strip()
                    if len(text) > 50:  # 简介通常较长
                        return text[:200] + '...' if len(text) > 200 else text
            
            return "NO FOUND"
        except:
            return "NO FOUND"
    
    def get_rating(self, soup):
        """获取豆瓣评分和评价人数"""
        try:
            rating_elem = soup.find('strong', {'class': 'll rating_num', 'property': 'v:average'})
            rating = rating_elem.text if rating_elem else "NO FOUND"
            
            votes_elem = soup.find('span', {'property': 'v:votes'})
            votes = votes_elem.text if votes_elem else "NO FOUND"
            
            return rating, votes
        except:
            return "NO FOUND", "NO FOUND"
    
    def get_poster_url(self, soup):
        """获取电影海报URL"""
        try:
            # 方法1：通过nbgnbg class查找海报图片
            poster_img = soup.select_one('#mainpic .nbgnbg img')
            if poster_img and poster_img.get('src'):
                poster_url = poster_img.get('src')
                # 如果是小图，尝试获取大图
                if poster_url and 's_ratio_poster' in poster_url:
                    poster_url = poster_url.replace('s_ratio_poster', 'l_ratio_poster')
                elif poster_url and 'webp' in poster_url:
                    # 有些图片是webp格式，转换为jpg
                    poster_url = poster_url.replace('.webp', '.jpg')
                self.logger.debug(f"获取海报URL: {poster_url}")
                return poster_url
            
            # 方法2：通过mainpic区域查找
            mainpic = soup.select_one('#mainpic img')
            if mainpic and mainpic.get('src'):
                poster_url = mainpic.get('src')
                self.logger.debug(f"获取海报URL(方法2): {poster_url}")
                return poster_url
            
            # 方法3：通过meta标签查找
            meta_image = soup.select_one('meta[property="og:image"]')
            if meta_image and meta_image.get('content'):
                poster_url = meta_image.get('content')
                self.logger.debug(f"获取海报URL(meta): {poster_url}")
                return poster_url
            
            self.logger.warning("未找到海报URL")
            return 'NO FOUND'
            
        except Exception as e:
            self.logger.error(f"获取海报URL失败: {e}")
            return 'NO FOUND'
    
    def get_douban_id(self, movie_url=None):
        """获取豆瓣ID"""
        try:
            # 从URL中提取ID
            if movie_url:
                # URL格式：https://movie.douban.com/subject/1292052/
                import re
                match = re.search(r'/subject/(\d+)', movie_url)
                if match:
                    douban_id = match.group(1)
                    self.logger.debug(f"从URL获取豆瓣ID: {douban_id}")
                    return douban_id
            
            self.logger.warning("未提供URL，无法获取豆瓣ID")
            return 'NO FOUND'
            
        except Exception as e:
            self.logger.error(f"获取豆瓣ID失败: {e}")
            return 'NO FOUND'
    
    def crawl_movie_info(self, movie_name):
        """爬取单个电影的完整信息"""
        try:
            self.logger.info(f"===== 开始爬取电影: {movie_name} =====")
            
            # 搜索电影
            encoded_name = self.str2urlcode(movie_name)
            search_url = f"https://search.douban.com/movie/subject_search?search_text={encoded_name}&cat=1002"
            self.logger.info(f"搜索URL: {search_url}")
            
            href = self.get_href(search_url)
            if not href:
                self.logger.error(f"搜索失败: {movie_name}")
                return None, "搜索失败"
            
            # 获取详情页面
            self.logger.info(f"访问详情页: {href}")
            movie_response = self.request_with_intelligent_retry(href)
            if not movie_response:
                self.logger.error(f"详情页面访问失败: {movie_name}")
                return None, "详情页面访问失败"
            
            soup = BeautifulSoup(movie_response.text, 'html.parser')
            
            # 提取所有信息
            self.logger.info(f"开始解析页面内容: {movie_name}")
            name = self.get_name(soup)
            director = self.get_director(soup)
            writer = self.get_writer(soup)
            actors = self.get_actors(soup, director)
            movie_type = self.get_type(soup)
            place = self.get_place(soup)
            language = self.get_language(soup)
            release_date = self.get_release_date(soup)
            runtime = self.get_runtime(soup)
            sb_name = self.get_sb_name(soup)
            imdb = self.get_imdb(soup)
            summary = self.get_summary(soup)
            rating, votes = self.get_rating(soup)
            poster_url = self.get_poster_url(soup)
            douban_id = self.get_douban_id(href)
            
            # 校验状态判断
            if name == 'NO FOUND':
                validation_status = "fault"
                self.logger.warning(f"解析失败，未获取到电影名称: {movie_name}")
            elif self.normalize_name(movie_name) == self.normalize_name(name):
                validation_status = "true"
                self.logger.info(f"爬取成功，名称匹配: {movie_name} -> {name}")
            else:
                validation_status = "check"
                self.logger.info(f"爬取成功，名称不匹配: {movie_name} -> {name}")
            
            result = [
                movie_name,  # 爬取名称
                validation_status,  # 校验状态
                name,  # 电影名称
                director,  # 导演
                writer,  # 编剧
                actors,  # 主演
                movie_type,  # 类型
                place,  # 制片国家/地区
                language,  # 语言
                release_date,  # 上映日期
                runtime,  # 片长
                sb_name,  # 又名
                imdb,  # IMDb
                summary,  # 剧情简介
                rating,  # 豆瓣评分
                votes,  # 评价人数
                poster_url,  # 海报URL
                douban_id  # 豆瓣ID
            ]
            
            self.logger.info(f"===== 完成爬取电影: {movie_name} (状态: {validation_status}) =====")
            return result, "成功"
            
        except Exception as e:
            self.logger.error(f"爬取 {movie_name} 时出错: {e}")
            self.logger.error(traceback.format_exc())
            return None, f"爬取异常: {str(e)}"
    
    def crawl_movie_info_with_anti_crawler_detection(self, movie_name, cached_url=None):
        """爬取单个电影信息，区分反爬失败和查找失败，支持URL缓存"""
        try:
            self.logger.info(f"===== 开始爬取电影: {movie_name} =====")
            
            href = None
            matched_title = None
            
            # 检查是否有缓存URL
            if cached_url and cached_url.strip():
                self.logger.info(f"使用缓存URL: {cached_url}")
                href = cached_url.strip()
                matched_title = "缓存"
                using_cache = True
            else:
                # 搜索电影
                encoded_name = self.str2urlcode(movie_name)
                search_url = f"https://search.douban.com/movie/subject_search?search_text={encoded_name}&cat=1002"
                self.logger.info(f"搜索URL: {search_url}")
                
                href, matched_title = self.get_href_with_smart_matching(search_url, movie_name)
                using_cache = False
                if not href:
                    # 区分是反爬失败还是查找失败
                    if self.consecutive_fails >= 2:  # 连续失败，可能是反爬
                        self.logger.error(f"搜索失败(疑似反爬): {movie_name}")
                        return None, "反爬失败"
                    else:
                        self.logger.error(f"搜索失败(未找到): {movie_name}")
                        return None, "查找失败"
            
            # 获取详情页面
            self.logger.info(f"访问详情页: {href}")
            movie_response = self.request_with_intelligent_retry(href)
            if not movie_response:
                # 区分是反爬失败还是访问失败
                if self.consecutive_fails >= 2:  # 连续失败，可能是反爬
                    self.logger.error(f"详情页面访问失败(疑似反爬): {movie_name}")
                    return None, "反爬失败"
                else:
                    self.logger.error(f"详情页面访问失败: {movie_name}")
                    return None, "访问失败"
            
            soup = BeautifulSoup(movie_response.text, 'html.parser')
            
            # 提取所有信息
            self.logger.info(f"开始解析页面内容: {movie_name}")
            name = self.get_name(soup)
            director = self.get_director(soup)
            writer = self.get_writer(soup)
            actors = self.get_actors(soup, director)
            movie_type = self.get_type(soup)
            place = self.get_place(soup)
            language = self.get_language(soup)
            release_date = self.get_release_date(soup)
            runtime = self.get_runtime(soup)
            sb_name = self.get_sb_name(soup)
            imdb = self.get_imdb(soup)
            summary = self.get_summary(soup)
            rating, votes = self.get_rating(soup)
            poster_url = self.get_poster_url(soup)
            douban_id = self.get_douban_id(href)
            
            # 校验状态判断 - 增强版
            if name == 'NO FOUND':
                validation_status = "fault"
                self.logger.warning(f"解析失败，未获取到电影名称: {movie_name}")
            else:
                # 使用相似度进行更精确的校验
                similarity = self.calculate_similarity(movie_name, name)
                
                # 如果使用了缓存，优先设置为cached状态
                if using_cache:
                    if similarity >= 0.5:  # 缓存的内容相似度还可以，标记为cached
                        validation_status = "cached"
                        self.logger.info(f"缓存爬取成功: {movie_name} -> {name} (相似度: {similarity:.3f})")
                    else:
                        validation_status = "check"  # 缓存的内容相似度太低，需要检查
                        self.logger.warning(f"缓存爬取，但名称差异较大: {movie_name} -> {name} (相似度: {similarity:.3f})")
                else:
                    # 非缓存的正常判断
                    if similarity >= 0.8:
                        validation_status = "true"
                        self.logger.info(f"爬取成功，高度匹配: {movie_name} -> {name} (相似度: {similarity:.3f})")
                    elif similarity >= 0.5:
                        validation_status = "check"
                        self.logger.info(f"爬取成功，部分匹配: {movie_name} -> {name} (相似度: {similarity:.3f})")
                    else:
                        validation_status = "check"
                        self.logger.warning(f"爬取成功，但名称差异较大: {movie_name} -> {name} (相似度: {similarity:.3f})")
                
                # 如果使用了智能匹配且匹配度较好，记录匹配信息
                if matched_title and matched_title != "缓存":
                    match_similarity = self.calculate_similarity(movie_name, matched_title)
                    self.logger.info(f"搜索匹配: {movie_name} -> {matched_title} (相似度: {match_similarity:.3f})")
            
            result = [
                movie_name,  # 爬取名称
                validation_status,  # 校验状态
                name,  # 电影名称
                director,  # 导演
                writer,  # 编剧
                actors,  # 主演
                movie_type,  # 类型
                place,  # 制片国家/地区
                language,  # 语言
                release_date,  # 上映日期
                runtime,  # 片长
                sb_name,  # 又名
                imdb,  # IMDb
                summary,  # 剧情简介
                rating,  # 豆瓣评分
                votes,  # 评价人数
                poster_url,  # 海报URL
                douban_id,  # 豆瓣ID
                href  # 缓存URL
            ]
            
            self.logger.info(f"===== 完成爬取电影: {movie_name} (状态: {validation_status}) =====")
            return result, "成功"
            
        except Exception as e:
            self.logger.error(f"爬取 {movie_name} 时出错: {e}")
            self.logger.error(traceback.format_exc())
            # 根据连续失败次数判断是否为反爬
            if self.consecutive_fails >= 2:
                return None, "反爬异常"
            else:
                return None, f"爬取异常: {str(e)}"
    
    def get_crawler_status(self):
        """获取爬虫状态信息"""
        current_ua = self.session.headers.get('User-Agent', '') if self.session else ''
        session_age = time.time() - self.session_create_time if hasattr(self, 'session_create_time') else 0
        
        success_rate = (self.success_count / self.total_requests * 100) if self.total_requests > 0 else 0
        
        return {
            'delay': self.current_delay,
            'success_count': self.success_count,
            'fail_count': self.fail_count,
            'consecutive_fails': self.consecutive_fails,
            'total_requests': self.total_requests,
            'success_rate': f"{success_rate:.1f}%",
            'session_age': f"{session_age/60:.1f}min",
            'session_requests': self.request_count_in_session,
            'user_agent': current_ua[:50]
        }
    
    @staticmethod
    def normalize_name(name):
        """标准化电影名称用于比较"""
        # 移除空格、标点符号，转为小写
        import string
        return re.sub(r'[%s\s]+' % re.escape(string.punctuation), '', name.lower())
    
    @staticmethod
    def calculate_similarity(name1, name2):
        """计算两个电影名称的相似度（0-1之间）"""
        if not name1 or not name2:
            return 0.0
        
        # 标准化名称
        norm_name1 = DoubanCrawler.normalize_name(name1)
        norm_name2 = DoubanCrawler.normalize_name(name2)
        
        # 使用SequenceMatcher计算相似度
        similarity = SequenceMatcher(None, norm_name1, norm_name2).ratio()
        
        # 额外检查：如果一个名称包含在另一个名称中，提高相似度
        if norm_name1 in norm_name2 or norm_name2 in norm_name1:
            similarity = max(similarity, 0.8)
        
        # 检查关键词匹配
        words1 = set(norm_name1.split())
        words2 = set(norm_name2.split())
        if words1 and words2:
            word_intersection = len(words1.intersection(words2))
            word_union = len(words1.union(words2))
            word_similarity = word_intersection / word_union if word_union > 0 else 0
            # 综合考虑字符相似度和词汇相似度
            similarity = max(similarity, word_similarity * 0.9)
        
        return similarity


class DoubanCrawlerApp:
    """主应用程序类"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.crawler = DoubanCrawler()
        self.workbook = None
        self.worksheet = None
        self.crawl_thread = None
        self.is_crawling = False
        self.is_paused = False
        self.pause_event = threading.Event()
        self.pause_event.set()  # 初始为非暂停状态
        self.current_row = 2  # 从第二行开始爬取（第一行是表头）
        self.total_rows = 0
        self.progress_queue = Queue()
        
        # 新增：用于记录"爬取名称"列位置和原始表格结构
        self.crawl_name_col = None
        self.original_max_col = 0
        self.crawl_data_start_col = 0
        
        # 用于停止时保存进度
        self.stop_crawling = False
        
        self.setup_ui()
        self.setup_styles()
        
        # 启动UI更新线程
        self.root.after(100, self.update_ui)
        
        # 爬虫状态显示
        self.crawler_status_text = tk.StringVar(value="爬虫状态: 待启动")
        self.setup_crawler_status_display()
    
    def setup_ui(self):
        """设置用户界面"""
        self.root.title("豆瓣影视媒资信息爬取工具 - 智能版v3.0")
        self.root.geometry("1600x800")
        
        # 创建菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 工具菜单
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="工具", menu=tools_menu)
        tools_menu.add_command(label="查看日志文件", command=self.open_log_file)
        tools_menu.add_command(label="清空日志", command=self.clear_log_file)
        tools_menu.add_separator()
        tools_menu.add_command(label="重置爬虫状态", command=self.reset_crawler_status)
        tools_menu.add_command(label="重置处理状态", command=self.reset_processing_status)
        
        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="使用说明", command=self.show_help)
        help_menu.add_command(label="关于", command=self.show_about)
        
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # 顶部文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="5")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # 文件路径选择
        ttk.Label(file_frame, text="Excel文件:").grid(row=0, column=0, sticky=tk.W)
        self.file_path_var = tk.StringVar()
        self.file_path_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, state="readonly")
        self.file_path_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 5))
        ttk.Button(file_frame, text="浏览", command=self.browse_file).grid(row=0, column=2, padx=(5, 0))
        
        # Sheet选择
        ttk.Label(file_frame, text="工作表:").grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(file_frame, textvariable=self.sheet_var, state="readonly")
        self.sheet_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=(5, 0))
        # 绑定sheet选择变化事件
        self.sheet_combo.bind('<<ComboboxSelected>>', self.on_sheet_changed)
        ttk.Button(file_frame, text="刷新", command=self.refresh_sheets).grid(row=1, column=2, padx=(5, 0), pady=(5, 0))
        
        # 状态显示
        self.status_label = ttk.Label(file_frame, text="请选择Excel文件")
        self.status_label.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # 中间表格预览区域
        table_frame = ttk.LabelFrame(main_frame, text="数据预览", padding="5")
        table_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        
        # 创建Treeview - 显示所有爬取字段（包括原始数据列）
        # 先显示原始数据的爬取名称列，然后显示所有爬取数据字段
        preview_columns = ["原始爬取名称"] + COLUMN_HEADERS
        self.tree = ttk.Treeview(table_frame, columns=preview_columns, show='headings', height=12)
        
        # 设置列标题和宽度
        column_widths = {
            "原始爬取名称": 120,
            "爬取名称": 120,
            "校验状态": 80,
            "电影名称": 120,
            "导演": 100,
            "编剧": 100,
            "主演": 150,
            "类型": 80,
            "制片国家/地区": 100,
            "语言": 80,
            "上映日期": 100,
            "片长": 80,
            "又名": 120,
            "IMDb": 100,
            "剧情简介": 200,
            "豆瓣评分": 80,
            "评价人数": 80
        }
        
        for col in preview_columns:
            self.tree.heading(col, text=col)
            width = column_widths.get(col, 100)
            self.tree.column(col, width=width, minwidth=50)
        
        # 添加滚动条
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # 底部控制区域
        control_frame = ttk.LabelFrame(main_frame, text="控制面板", padding="5")
        control_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E))
        control_frame.columnconfigure(0, weight=1)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(control_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # 进度文本
        self.progress_text = tk.StringVar(value="准备就绪")
        ttk.Label(control_frame, textvariable=self.progress_text).grid(row=1, column=0, columnspan=3, pady=(0, 5))
        
        # 设置按钮框架
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        button_frame.columnconfigure(0, weight=1)
        
        # 按钮
        self.start_button = ttk.Button(button_frame, text="开始爬取", command=self.start_crawling)
        self.start_button.grid(row=0, column=0, padx=(0, 5), sticky='w')
        
        self.pause_button = ttk.Button(button_frame, text="暂停", command=self.pause_crawling, state="disabled")
        self.pause_button.grid(row=0, column=1, padx=(5, 5))
        
        self.stop_button = ttk.Button(button_frame, text="停止", command=self.stop_crawling_func, state="disabled")
        self.stop_button.grid(row=0, column=2, padx=(5, 5))
        
        # 设置按钮
        self.config_button = ttk.Button(button_frame, text="高级设置", command=self.show_config)
        self.config_button.grid(row=0, column=3, padx=(5, 0))
        
        # 当前处理状态
        self.current_status = tk.StringVar(value="")
        ttk.Label(control_frame, textvariable=self.current_status).grid(row=3, column=0, columnspan=3, pady=(5, 0))
    
    def setup_crawler_status_display(self):
        """设置爬虫状态显示"""
        # 在控制面板中添加爬虫状态显示
        control_children = self.root.winfo_children()[0].winfo_children()
        control_frame = None
        for child in control_children:
            if hasattr(child, 'cget') and 'text' in child.keys() and child.cget('text') == "控制面板":
                control_frame = child
                break
        
        if control_frame:
            # 爬虫状态标签
            status_label = ttk.Label(control_frame, textvariable=self.crawler_status_text, 
                                   foreground='blue', font=('Arial', 9))
            status_label.grid(row=4, column=0, columnspan=3, pady=(5, 0), sticky='w')
    
    def setup_styles(self):
        """设置样式"""
        self.style = ttk.Style()
        
        # 配置Treeview标签
        self.tree.tag_configure('fault', background='#ffcccc', foreground='red')  # 红色背景，爬取失败
        self.tree.tag_configure('check', background='#ccffcc', foreground='green')  # 绿色背景，需要检查
        self.tree.tag_configure('true', background='white', foreground='black')  # 默认，爬取成功且匹配
        self.tree.tag_configure('warn', background='#ffeb3b', foreground='#e65100')  # 深黄色背景，反爬导致失败
        self.tree.tag_configure('cached', background='#e3f2fd', foreground='#1976d2')  # 浅蓝色背景，使用缓存
        self.tree.tag_configure('empty', background='#f0f0f0', foreground='gray')  # 灰色背景，空白行
    
    def browse_file(self):
        """浏览并选择Excel文件"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
            self.load_workbook()
            self.refresh_sheets()
    
    def load_workbook(self):
        """加载工作簿"""
        try:
            file_path = self.file_path_var.get()
            if file_path:
                self.workbook = load_workbook(file_path)
                self.status_label.config(text=f"已加载文件: {os.path.basename(file_path)}")
                return True
        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")
            self.status_label.config(text="文件加载失败")
        return False
    
    def refresh_sheets(self):
        """刷新工作表列表"""
        if self.workbook:
            # 保存当前选择的sheet
            current_sheet = self.sheet_var.get()
            
            # 重新加载工作簿以获取最新的sheet列表
            self.load_workbook()
            
            # 更新sheet列表
            sheet_names = self.workbook.sheetnames
            self.sheet_combo['values'] = sheet_names
            
            # 尝试恢复之前选择的sheet，如果不存在则选择第一个
            if current_sheet in sheet_names:
                self.sheet_combo.set(current_sheet)
            elif sheet_names:
                self.sheet_combo.set(sheet_names[0])
            
            # 加载数据
            self.load_sheet_data()
            self.status_label.config(text=f"已刷新，共发现 {len(sheet_names)} 个工作表")
    
    def on_sheet_changed(self, event=None):
        """当工作表选择发生变化时的回调函数"""
        self.load_sheet_data()
    
    def load_sheet_data(self):
        """加载工作表数据"""
        try:
            if not self.workbook:
                return
            
            sheet_name = self.sheet_var.get()
            if not sheet_name:
                return
            
            self.worksheet = self.workbook[sheet_name]
            
            # 清空现有数据
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # 查找"爬取名称"列
            self.find_crawl_name_column()
            
            if self.crawl_name_col is None:
                messagebox.showwarning("警告", "未找到'爬取名称'列，请确保表格中包含该列")
                self.status_label.config(text="未找到'爬取名称'列")
                return
            
            # 首先检查是否已有爬取数据表头
            existing_start, existing_end = self.find_existing_crawl_headers()
            
            if existing_start is not None:
                # 已有表头，使用现有位置
                self.crawl_data_start_col = existing_start
                self.original_max_col = existing_start - 1
                logging.info(f"使用已存在的爬取数据表头，列范围: {existing_start}-{existing_end}")
            else:
                # 没有表头，计算新位置
                self.original_max_col = self.worksheet.max_column if self.worksheet.max_column else 1
                self.crawl_data_start_col = self.original_max_col + 1
                logging.info(f"计算新的爬取数据起始位置: 第{self.crawl_data_start_col}列")
            
            # 检查并添加爬取数据表头（如果需要）
            self.ensure_crawl_headers()
            
            # 清空映射关系
            self.crawler.valid_rows_mapping = {}
            self.crawler.excel_to_ui_mapping = {}
            
            # 加载数据 - 只显示有效行（非空的爬取名称）
            non_empty_count = 0
            ui_row_index = 0
            
            for row_num in range(2, self.worksheet.max_row + 1):
                # 获取原始爬取名称
                original_crawl_name = self.worksheet.cell(row=row_num, column=self.crawl_name_col).value
                original_name_str = str(original_crawl_name).strip() if original_crawl_name is not None else ""
                
                # 只处理非空的爬取名称
                if original_name_str:
                    non_empty_count += 1
                    
                    # 建立映射关系
                    self.crawler.valid_rows_mapping[ui_row_index] = row_num
                    self.crawler.excel_to_ui_mapping[row_num] = ui_row_index
                    
                    # 构建完整的行数据：原始爬取名称 + 所有爬取数据字段
                    row_values = []
                    
                    # 第一列：原始爬取名称
                    row_values.append(original_name_str)
                    
                    # 后续列：所有爬取数据字段
                    for col_offset, header in enumerate(COLUMN_HEADERS):
                        col_num = self.crawl_data_start_col + col_offset
                        cell_value = self.worksheet.cell(row=row_num, column=col_num).value
                        
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            # 限制显示长度，避免界面过宽
                            if len(cell_str) > 50:
                                cell_str = cell_str[:47] + "..."
                            row_values.append(cell_str)
                        else:
                            row_values.append("")
                    
                    # 根据校验状态设置标签（校验状态是爬取数据的第二列，即row_values[2]）
                    validation_status = row_values[2] if len(row_values) > 2 else ""  # 校验状态列
                    if validation_status == "fault":
                        tag = "fault"
                    elif validation_status == "check":
                        tag = "check"
                    elif validation_status == "true":
                        tag = "true"
                    elif validation_status == "warn":
                        tag = "warn"
                    elif validation_status == "cached":
                        tag = "cached"
                    else:
                        tag = ""
                    
                    # 确保数组长度与列数匹配
                    expected_length = len(["原始爬取名称"] + COLUMN_HEADERS)
                    while len(row_values) < expected_length:
                        row_values.append("")
                    
                    self.tree.insert('', 'end', values=row_values, tags=(tag,))
                    ui_row_index += 1
            
            # 统计已处理的数据
            processed_count = 0
            for row_num in range(2, self.worksheet.max_row + 1):
                original_crawl_name = self.worksheet.cell(row=row_num, column=self.crawl_name_col).value
                original_name_str = str(original_crawl_name).strip() if original_crawl_name is not None else ""
                
                if original_name_str:  # 只统计有效行
                    validation_status = self.worksheet.cell(row=row_num, column=self.crawl_data_start_col + 1).value
                    if validation_status is not None and str(validation_status).strip() != "" and str(validation_status).strip() != "warn":
                        processed_count += 1
            
            unprocessed_count = non_empty_count - processed_count
            self.total_rows = non_empty_count  # 只计算有效行数
            
            # 检查是否已有表头
            existing_start, existing_end = self.find_existing_crawl_headers()
            header_status = f"(表头: {'已存在' if existing_start else '新建'})"
            
            status_text = f"已加载有效数据 {non_empty_count} 条 {header_status}"
            if processed_count > 0:
                status_text += f"，已处理 {processed_count} 条，待处理 {unprocessed_count} 条"
            else:
                status_text += f"，全部待处理"
                
            self.status_label.config(text=status_text)
            
        except Exception as e:
            messagebox.showerror("错误", f"加载工作表数据失败: {str(e)}")
    
    def find_crawl_name_column(self):
        """查找"爬取名称"列的位置"""
        if not self.worksheet:
            return
        
        self.crawl_name_col = None
        # 在第一行查找"爬取名称"列
        for col_num in range(1, self.worksheet.max_column + 1):
            cell_value = self.worksheet.cell(row=1, column=col_num).value
            if cell_value and str(cell_value).strip() == "爬取名称":
                self.crawl_name_col = col_num
                break
    
    def find_existing_crawl_headers(self):
        """查找已存在的爬取数据表头位置"""
        if not self.worksheet:
            return None, None
        
        # 查找第一个爬取数据表头的位置
        for col_num in range(1, self.worksheet.max_column + 1):
            cell_value = self.worksheet.cell(row=1, column=col_num).value
            if cell_value and str(cell_value).strip() == COLUMN_HEADERS[0]:  # "爬取名称"
                # 验证是否为完整的爬取数据表头序列
                is_complete_header = True
                for i, expected_header in enumerate(COLUMN_HEADERS):
                    check_col = col_num + i
                    if check_col > self.worksheet.max_column:
                        is_complete_header = False
                        break
                    actual_value = self.worksheet.cell(row=1, column=check_col).value
                    if not actual_value or str(actual_value).strip() != expected_header:
                        is_complete_header = False
                        break
                
                if is_complete_header:
                    return col_num, col_num + len(COLUMN_HEADERS) - 1  # 返回起始列和结束列
        
        return None, None
    
    def ensure_crawl_headers(self):
        """确保爬取数据表头存在，如果已存在则不重复创建"""
        if not self.worksheet:
            return
        
        # 首先检查是否已存在完整的爬取数据表头
        existing_start, existing_end = self.find_existing_crawl_headers()
        
        if existing_start is not None:
            # 已存在完整表头，更新起始位置
            self.crawl_data_start_col = existing_start
            logging.info(f"发现已存在的爬取数据表头，起始列: {existing_start}")
            return
        
        # 不存在表头，在原始表格后添加
        logging.info(f"未发现爬取数据表头，将在第{self.crawl_data_start_col}列开始创建")
        need_update = False
        for i, header in enumerate(COLUMN_HEADERS):
            col_num = self.crawl_data_start_col + i
            current_value = self.worksheet.cell(row=1, column=col_num).value
            if current_value != header:
                self.worksheet.cell(row=1, column=col_num, value=header)
                need_update = True
                logging.info(f"添加表头: {header} -> 第{col_num}列")
        
        if need_update:
            self.workbook.save(self.file_path_var.get())
            logging.info("表头已保存到Excel文件")
    
    def ensure_headers(self):
        """确保表头正确"""
        if not self.worksheet:
            return
        
        # 检查第一行是否为正确的表头
        need_update = False
        for col_num, header in enumerate(COLUMN_HEADERS, 1):
            current_value = self.worksheet.cell(row=1, column=col_num).value
            if current_value != header:
                self.worksheet.cell(row=1, column=col_num, value=header)
                need_update = True
        
        if need_update:
            self.workbook.save(self.file_path_var.get())
    
    def start_crawling(self):
        """开始爬取"""
        if not self.worksheet:
            messagebox.showwarning("警告", "请先选择Excel文件和工作表")
            return
        
        if self.is_crawling:
            return
        
        self.is_crawling = True
        self.is_paused = False
        self.stop_crawling = False
        self.pause_event.set()
        
        # 更新按钮状态
        self.start_button.config(text="爬取中...", state="disabled")
        self.pause_button.config(state="normal")
        self.stop_button.config(state="normal")
        
        # 启动爬取线程
        self.crawl_thread = threading.Thread(target=self.crawl_worker, daemon=True)
        self.crawl_thread.start()
    
    def pause_crawling(self):
        """暂停/继续爬取"""
        if not self.is_crawling:
            return
        
        if self.is_paused:
            # 继续爬取
            self.is_paused = False
            self.pause_event.set()
            self.pause_button.config(text="暂停")
            self.progress_queue.put(("status", "继续爬取..."))
        else:
            # 暂停爬取
            self.is_paused = True
            self.pause_event.clear()
            self.pause_button.config(text="继续")
            self.progress_queue.put(("status", "已暂停"))
    
    def stop_crawling_func(self):
        """停止爬取并保存进度"""
        if not self.is_crawling:
            return
        
        self.stop_crawling = True
        self.is_paused = False
        self.pause_event.set()  # 确保线程不被阻塞
        self.progress_queue.put(("status", "正在停止并保存进度..."))
    
    def crawl_worker(self):
        """爬取工作线程"""
        try:
            save_counter = 0
            processed_count = 0
            valid_rows = []
            skipped_count = 0
            total_non_empty = 0
            
            # 收集所有有效的行（非空的爬取名称且未处理过的）
            for row_num in range(2, self.worksheet.max_row + 1):
                movie_name = self.worksheet.cell(row=row_num, column=self.crawl_name_col).value
                if movie_name and str(movie_name).strip():
                    total_non_empty += 1
                    # 检查校验状态列是否为空（第二列是校验状态）
                    validation_status = self.worksheet.cell(row=row_num, column=self.crawl_data_start_col + 1).value
                    
                    # 如果校验状态为空、None或warn，则认为未处理，加入待爬取列表
                    if (validation_status is None or 
                        str(validation_status).strip() == "" or 
                        str(validation_status).strip() == "warn"):
                        valid_rows.append(row_num)
                    else:
                        skipped_count += 1
                        logging.info(f"跳过已处理行 {row_num}: {movie_name} (状态: {validation_status})")
            
            if not valid_rows:
                if skipped_count > 0:
                    self.progress_queue.put(("complete", f"所有数据已处理完成！共{total_non_empty}条数据，跳过{skipped_count}条已处理"))
                else:
                    self.progress_queue.put(("complete", "没有需要处理的数据！"))
                return
            
            self.progress_queue.put(("status", f"总数据:{total_non_empty}条，跳过已处理:{skipped_count}条，待处理:{len(valid_rows)}条"))
            
            for i, row_num in enumerate(valid_rows):
                # 检查停止状态
                if self.stop_crawling:
                    self.workbook.save(self.file_path_var.get())
                    self.progress_queue.put(("complete", f"已停止爬取并保存进度！已处理 {processed_count} 条数据"))
                    return
                
                # 检查暂停状态
                self.pause_event.wait()
                
                if not self.is_crawling:
                    break
                
                # 获取爬取名称
                movie_name = self.worksheet.cell(row=row_num, column=self.crawl_name_col).value
                movie_name = str(movie_name).strip()
                
                # 更新状态
                self.progress_queue.put(("current", f"正在处理: {movie_name} ({i+1}/{len(valid_rows)})"))
                
                # 检查是否有缓存URL
                cached_url = None
                cache_col = self.crawl_data_start_col + len(COLUMN_HEADERS) - 1  # 缓存URL列
                cached_url_cell = self.worksheet.cell(row=row_num, column=cache_col).value
                if cached_url_cell and str(cached_url_cell).strip():
                    cached_url = str(cached_url_cell).strip()
                    self.progress_queue.put(("status", f"检测到缓存URL，跳过搜索步骤: {movie_name}"))
                
                # 爬取数据，使用新的反爬检测方法和缓存支持
                result, message = self.crawler.crawl_movie_info_with_anti_crawler_detection(movie_name, cached_url)
                
                # 更新爬虫状态显示
                status_info = self.crawler.get_crawler_status()
                status_text = (f"延迟:{status_info['delay']:.1f}s | "
                             f"成功率:{status_info['success_rate']} | "
                             f"连续失败:{status_info['consecutive_fails']} | "
                             f"会话:{status_info['session_age']}")
                self.progress_queue.put(("crawler_status", status_text))
                
                if result:
                    # 写入数据到Excel（从爬取数据开始列写入）
                    for col_offset, value in enumerate(result):
                        col_num = self.crawl_data_start_col + col_offset
                        self.worksheet.cell(row=row_num, column=col_num, value=value)
                    
                    # 设置样式
                    self.set_row_style_crawl_data(row_num, result[1])  # result[1]是校验状态
                    
                    # 更新UI预览（显示完整的爬取结果）
                    self.progress_queue.put(("row_update", (row_num, result)))
                else:
                    # 爬取失败，根据失败原因设置状态
                    fault_col = self.crawl_data_start_col + 1  # 校验状态列
                    self.worksheet.cell(row=row_num, column=self.crawl_data_start_col, value=movie_name)
                    
                    # 根据失败原因设置不同状态
                    if "反爬" in message:
                        status = "warn"
                        self.worksheet.cell(row=row_num, column=fault_col, value="warn")
                        self.set_row_style_crawl_data(row_num, "warn")
                        # 创建警告的结果数组
                        fault_result = [movie_name, "warn"] + [""] * (len(COLUMN_HEADERS) - 2)
                    else:
                        status = "fault"
                        self.worksheet.cell(row=row_num, column=fault_col, value="fault")
                        self.set_row_style_crawl_data(row_num, "fault")
                        # 创建失败的结果数组
                        fault_result = [movie_name, "fault"] + [""] * (len(COLUMN_HEADERS) - 2)
                    
                    self.progress_queue.put(("row_update", (row_num, fault_result)))
                
                processed_count += 1
                
                # 更新进度
                current_progress = (processed_count / len(valid_rows)) * 100
                self.progress_queue.put(("progress", current_progress))
                
                save_counter += 1
                
                # 每5个保存一次
                if save_counter >= 5:
                    self.workbook.save(self.file_path_var.get())
                    save_counter = 0
                    self.progress_queue.put(("status", f"已保存进度... 正在处理: {movie_name} ({i+1}/{len(valid_rows)})"))
                
                # 智能延迟已在爬虫内部处理，这里不再额外延迟
            
            # 最终保存
            self.workbook.save(self.file_path_var.get())
            self.progress_queue.put(("complete", f"爬取完成！共处理 {processed_count} 条数据"))
            
        except Exception as e:
            self.progress_queue.put(("error", f"爬取过程中出错: {str(e)}"))
        finally:
            self.is_crawling = False
    
    def set_row_style_crawl_data(self, row_num, validation_status):
        """设置爬取数据区域的Excel行样式"""
        try:
            if validation_status == "fault":
                font = Font(color="FF0000", bold=True)  # 红色加粗
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif validation_status == "check":
                font = Font(color="008000", bold=True)  # 绿色加粗
                fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            elif validation_status == "warn":
                font = Font(color="E65100", bold=True)  # 深橙色加粗
                fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")  # 深黄色背景
            elif validation_status == "cached":
                font = Font(color="1976D2", bold=True)  # 蓝色加粗
                fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 浅蓝色背景
            else:
                return  # true状态保持默认样式
            
            # 应用样式到爬取数据区域
            for col_offset in range(len(COLUMN_HEADERS)):
                col_num = self.crawl_data_start_col + col_offset
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.font = font
                cell.fill = fill
                
        except Exception as e:
            print(f"设置样式失败: {e}")
    
    def set_row_style(self, row_num, validation_status):
        """设置Excel行样式（保留兼容性）"""
        try:
            if validation_status == "fault":
                font = Font(color="FF0000", bold=True)  # 红色加粗
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif validation_status == "check":
                font = Font(color="008000", bold=True)  # 绿色加粗
                fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            else:
                return  # true状态保持默认样式
            
            # 应用样式到整行
            for col_num in range(1, len(COLUMN_HEADERS) + 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.font = font
                cell.fill = fill
                
        except Exception as e:
            print(f"设置样式失败: {e}")
    
    def update_ui(self):
        """更新UI界面"""
        try:
            while not self.progress_queue.empty():
                msg_type, data = self.progress_queue.get_nowait()
                
                if msg_type == "progress":
                    self.progress_var.set(data)
                elif msg_type == "status":
                    self.progress_text.set(data)
                elif msg_type == "current":
                    self.current_status.set(data)
                elif msg_type == "crawler_status":
                    self.crawler_status_text.set(f"爬虫状态: {data}")
                elif msg_type == "row_update":
                    row_num, result_data = data
                    # 更新Treeview中的对应行
                    self.update_tree_row(row_num, result_data)
                elif msg_type == "complete":
                    self.progress_text.set(data)
                    self.reset_buttons()
                elif msg_type == "error":
                    messagebox.showerror("错误", data)
                    self.reset_buttons()
        except:
            pass
        
        # 继续定时更新
        self.root.after(100, self.update_ui)
    
    def update_tree_row(self, row_num, result_data):
        """更新Treeview中的行数据"""
        try:
            # 使用映射关系获取UI行索引
            if row_num not in self.crawler.excel_to_ui_mapping:
                return  # 如果行号不在映射中，说明是空行，直接返回
            
            tree_index = self.crawler.excel_to_ui_mapping[row_num]
            children = self.tree.get_children()
            
            if 0 <= tree_index < len(children):
                item = children[tree_index]
                
                # 获取原始爬取名称
                original_crawl_name = self.worksheet.cell(row=row_num, column=self.crawl_name_col).value
                original_name_str = str(original_crawl_name).strip() if original_crawl_name is not None else ""
                
                # 构建完整的显示数据：原始爬取名称 + 爬取结果数据
                display_values = []
                
                # 第一列：原始爬取名称
                if original_name_str:
                    display_values.append(original_name_str)
                else:
                    display_values.append("[空白]")
                
                # 后续列：爬取结果数据
                for i, value in enumerate(result_data):
                    if value is not None:
                        value_str = str(value).strip()
                        # 限制显示长度，避免界面过宽
                        if len(value_str) > 50:
                            value_str = value_str[:47] + "..."
                        display_values.append(value_str)
                    else:
                        display_values.append("")
                
                # 确保数组长度与列数匹配
                expected_length = len(["原始爬取名称"] + COLUMN_HEADERS)
                while len(display_values) < expected_length:
                    display_values.append("")
                
                # 根据校验状态设置标签（校验状态是result_data[1]，显示在display_values[2]）
                validation_status = result_data[1] if len(result_data) > 1 else ""
                if validation_status == "fault":
                    tag = "fault"
                elif validation_status == "check":
                    tag = "check"
                elif validation_status == "true":
                    tag = "true"
                elif validation_status == "warn":
                    tag = "warn"
                elif validation_status == "cached":
                    tag = "cached"
                else:
                    tag = ""
                
                # 更新值和标签
                self.tree.item(item, values=display_values, tags=(tag,))
                
                # 自动滚动到当前处理的行
                self.tree.see(item)
                
        except Exception as e:
            print(f"更新树行失败: {e}")
    
    def show_config(self):
        """显示高级设置对话框"""
        config_window = tk.Toplevel(self.root)
        config_window.title("爬虫高级设置")
        config_window.geometry("500x600")
        config_window.resizable(False, False)
        config_window.grab_set()  # 模态对话框
        
        # 设置框架
        frame = ttk.Frame(config_window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 当前版本专注于Requests模式
        selenium_available = False
        
        # 爬虫模式选择
        ttk.Label(frame, text="爬虫模式:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        crawler_mode_var = tk.StringVar(value="requests")
        mode_frame = ttk.Frame(frame)
        mode_frame.grid(row=0, column=1, sticky=tk.W, pady=(0, 5))
        ttk.Radiobutton(mode_frame, text="Requests (推荐)", variable=crawler_mode_var, value="requests").pack(side=tk.LEFT)
        
        # 暂时移除selenium选项，专注于requests模式的稳定性
        info_label = ttk.Label(mode_frame, text="(浏览器模式开发中...)", foreground="gray")
        info_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # 代理设置
        ttk.Label(frame, text="代理设置:").grid(row=1, column=0, sticky=tk.W, pady=(10, 5))
        use_proxy_var = tk.BooleanVar()
        ttk.Checkbutton(frame, text="启用代理轮换", variable=use_proxy_var).grid(row=1, column=1, sticky=tk.W, pady=(10, 5))
        
        ttk.Label(frame, text="代理列表:").grid(row=2, column=0, sticky=(tk.W, tk.N), pady=(5, 0))
        proxy_text = tk.Text(frame, height=4, width=35)
        proxy_text.grid(row=2, column=1, sticky=tk.W, pady=(5, 0))
        proxy_text.insert('1.0', "# 每行一个代理，格式：IP:端口\n# 例如：127.0.0.1:8080\n# proxy1.example.com:3128\n# proxy2.example.com:3128")
        
        # 分隔线
        ttk.Separator(frame, orient='horizontal').grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(15, 10))
        
        # 基础延迟设置
        ttk.Label(frame, text="基础延迟 (秒):").grid(row=4, column=0, sticky=tk.W, pady=(0, 5))
        delay_var = tk.DoubleVar(value=self.crawler.base_delay)
        delay_spin = ttk.Spinbox(frame, from_=3, to=30, textvariable=delay_var, width=10)
        delay_spin.grid(row=4, column=1, sticky=tk.W, pady=(0, 5))
        
        # 最大延迟设置
        ttk.Label(frame, text="最大延迟 (秒):").grid(row=5, column=0, sticky=tk.W, pady=(0, 5))
        max_delay_var = tk.DoubleVar(value=self.crawler.max_delay)
        max_delay_spin = ttk.Spinbox(frame, from_=60, to=600, textvariable=max_delay_var, width=10)
        max_delay_spin.grid(row=5, column=1, sticky=tk.W, pady=(0, 5))
        
        # 紧急延迟设置
        ttk.Label(frame, text="紧急延迟 (秒):").grid(row=6, column=0, sticky=tk.W, pady=(0, 5))
        emergency_delay_var = tk.DoubleVar(value=self.crawler.emergency_delay)
        emergency_delay_spin = ttk.Spinbox(frame, from_=180, to=900, textvariable=emergency_delay_var, width=10)
        emergency_delay_spin.grid(row=6, column=1, sticky=tk.W, pady=(0, 5))
        
        # 快速反爬模式设置
        ttk.Label(frame, text="快速反爬模式:").grid(row=7, column=0, sticky=tk.W, pady=(5, 5))
        quick_mode_var = tk.BooleanVar(value=self.crawler.quick_anti_crawler_mode)
        ttk.Checkbutton(frame, text="启用快速反爬处理", variable=quick_mode_var).grid(row=7, column=1, sticky=tk.W, pady=(5, 5))
        
        # Selenium专用设置
        ttk.Label(frame, text="Selenium设置:").grid(row=8, column=0, sticky=tk.W, pady=(10, 5))
        selenium_frame = ttk.Frame(frame)
        selenium_frame.grid(row=8, column=1, sticky=tk.W, pady=(10, 5))
        
        headless_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(selenium_frame, text="无头模式", variable=headless_var).pack(side=tk.LEFT)
        show_browser_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(selenium_frame, text="显示浏览器", variable=show_browser_var).pack(side=tk.LEFT, padx=(10, 0))
        
        # 日志级别设置
        ttk.Label(frame, text="日志级别:").grid(row=9, column=0, sticky=tk.W, pady=(10, 5))
        log_level_var = tk.StringVar(value="INFO")
        log_combo = ttk.Combobox(frame, textvariable=log_level_var, 
                                values=['DEBUG', 'INFO', 'WARNING', 'ERROR'], state="readonly")
        log_combo.grid(row=9, column=1, sticky=tk.W, pady=(10, 5))
        
        # 说明文本
        info_text = tk.Text(frame, height=10, width=50, wrap=tk.WORD)
        info_text.grid(row=10, column=0, columnspan=2, pady=(10, 10))
        info_text.insert('1.0', """反反爬设置说明：

【爬虫模式】
• Requests模式：轻量快速，适合大批量爬取
• Selenium模式：模拟真实浏览器，反爬能力强

【代理设置】
• 启用代理轮换，遇到反爬时自动切换IP
• 代理格式：IP:端口，每行一个
• 可配合免费代理或付费代理池使用

【延迟策略】
• 基础延迟：正常情况下的请求间隔时间
• 最大延迟：触发反爬后的最大等待时间
• 紧急延迟：连续失败过多时的保护延迟

【快速反爬模式 ⭐新功能】
• 启用后，第一次检测到反爬立即处理，无需5次重试
• 大幅减少爬取时间，提高效率
• 处理反爬后使用快速延迟模式（3-6秒）
• 建议启用，可显著提升爬取速度

【Selenium优势】
- 真实浏览器环境，难以被检测
- 支持JavaScript渲染
- 模拟人类行为（滚动、鼠标移动）
- 自动应对简单的反爬机制

建议：启用快速反爬模式以获得最佳性能""")
        info_text.config(state='disabled')
        
        # 按钮框架
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=11, column=0, columnspan=2, pady=(10, 0))
        
        def apply_config():
            # 保存基础设置
            self.crawler.base_delay = delay_var.get()
            self.crawler.max_delay = max_delay_var.get()
            self.crawler.emergency_delay = emergency_delay_var.get()
            
            # 调整当前延迟，不超过最大值
            if self.crawler.current_delay > self.crawler.max_delay:
                self.crawler.current_delay = self.crawler.max_delay
            
            # 代理设置
            use_proxy = use_proxy_var.get()
            proxy_list = []
            if use_proxy:
                proxy_content = proxy_text.get('1.0', tk.END).strip()
                proxy_lines = [line.strip() for line in proxy_content.split('\n') 
                              if line.strip() and not line.strip().startswith('#')]
                if proxy_lines:
                    proxy_list = [{'http': f'http://{proxy}', 'https': f'https://{proxy}'} 
                                 for proxy in proxy_lines]
                    logging.info(f"已配置 {len(proxy_list)} 个代理")
                else:
                    messagebox.showwarning("警告", "启用了代理但未配置代理列表")
            
            # 爬虫模式设置
            selected_mode = crawler_mode_var.get()
            
            # 目前只支持Requests模式
            if hasattr(self.crawler, 'close'):
                self.crawler.close()
            
            self.crawler = DoubanCrawler()
            self.crawler.base_delay = delay_var.get()
            self.crawler.max_delay = max_delay_var.get()
            self.crawler.emergency_delay = emergency_delay_var.get()
            self.crawler.current_delay = delay_var.get()
            self.crawler.use_proxy = use_proxy
            self.crawler.proxy_list = proxy_list
            self.crawler.quick_anti_crawler_mode = quick_mode_var.get()
            
            # 设置日志级别
            level = getattr(logging, log_level_var.get())
            logging.getLogger().setLevel(level)
            
            messagebox.showinfo("设置", f"配置已保存！当前模式: {selected_mode.upper()}")
            config_window.destroy()
        
        ttk.Button(btn_frame, text="保存", command=apply_config).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(btn_frame, text="取消", command=config_window.destroy).grid(row=0, column=1)
    
    def open_log_file(self):
        """打开日志文件"""
        log_file = "douban_crawler.log"
        if os.path.exists(log_file):
            try:
                # Windows系统使用默认程序打开
                os.startfile(log_file)
            except:
                messagebox.showinfo("提示", f"请手动打开日志文件: {os.path.abspath(log_file)}")
        else:
            messagebox.showinfo("提示", "日志文件不存在")
    
    def clear_log_file(self):
        """清空日志文件"""
        if messagebox.askyesno("确认", "确定要清空日志文件吗？"):
            try:
                with open("douban_crawler.log", "w", encoding='utf-8') as f:
                    f.write("")
                messagebox.showinfo("完成", "日志文件已清空")
            except Exception as e:
                messagebox.showerror("错误", f"清空日志失败: {e}")
    
    def reset_crawler_status(self):
        """重置爬虫状态"""
        if self.is_crawling:
            messagebox.showwarning("警告", "爬虫正在运行中，无法重置状态")
            return
        
        if messagebox.askyesno("确认", "确定要重置爬虫状态吗？这将清除统计信息。"):
            self.crawler.success_count = 0
            self.crawler.fail_count = 0
            self.crawler.consecutive_fails = 0
            self.crawler.current_delay = self.crawler.base_delay
            messagebox.showinfo("完成", "爬虫状态已重置")
    
    def reset_processing_status(self):
        """重置处理状态"""
        if self.is_crawling:
            messagebox.showwarning("警告", "爬虫正在运行中，无法重置处理状态")
            return
        
        if not self.worksheet:
            messagebox.showwarning("警告", "请先选择Excel文件和工作表")
            return
        
        if messagebox.askyesno("确认", "确定要清空所有处理状态吗？这将清除所有爬取结果，允许重新爬取。"):
            try:
                # 清空爬取数据区域（保留原始数据）
                cleared_count = 0
                for row_num in range(2, self.worksheet.max_row + 1):
                    # 清空爬取数据列（从爬取数据开始列到最后）
                    for col_offset in range(len(COLUMN_HEADERS)):
                        col_num = self.crawl_data_start_col + col_offset
                        if self.worksheet.cell(row=row_num, column=col_num).value is not None:
                            self.worksheet.cell(row=row_num, column=col_num).value = None
                            cleared_count += 1
                
                # 保存文件
                self.workbook.save(self.file_path_var.get())
                
                # 重新加载显示
                self.load_sheet_data()
                
                messagebox.showinfo("完成", f"已清空处理状态，共清除 {cleared_count} 个单元格的数据")
                
            except Exception as e:
                messagebox.showerror("错误", f"重置处理状态失败: {e}")
    
    def show_help(self):
        """显示使用说明"""
        help_window = tk.Toplevel(self.root)
        help_window.title("使用说明")
        help_window.geometry("600x500")
        help_window.resizable(True, True)
        
        text_widget = tk.Text(help_window, wrap=tk.WORD, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(help_window, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        help_text = """豆瓣影视媒资信息爬取工具 - 使用说明

【基本使用】
1. 点击"浏览"选择包含电影名称的Excel文件
2. 在"工作表"下拉框中选择要处理的工作表
3. 确认数据预览无误后，点击"开始爬取"
4. 爬取过程中可以点击"暂停"来暂停/继续

【数据格式要求】
• Excel必须包含"爬取名称"列
• 程序会自动在原表格后添加爬取数据列
• 支持.xlsx和.xls格式

【校验状态说明】
• true：爬取成功且名称高度匹配（默认样式，相似度≥80%）
• check：爬取成功但名称部分匹配（绿色加粗，需人工核实）
• fault：爬取失败（红色加粗）
• warn：反爬机制导致失败（深黄色加粗）
• cached：使用缓存URL爬取（浅蓝色加粗）

【智能跳过机制】
• 自动跳过已有校验状态的行，避免重复爬取
• warn状态会被重新爬取，因为可能是临时反爬导致
• 状态栏显示：总数据/已处理/待处理统计
• 支持中断后续爬：程序会从未处理的行继续
• 可通过"工具→重置处理状态"清空所有结果重新开始

【控制功能】
• 暂停：临时暂停爬取，可继续
• 停止：停止爬取并保存当前进度，下次启动从未处理行继续

【智能优化功能 ⭐新增】
• 智能名称匹配：自动选择相似度最高的搜索结果，不再盲选第一个
• URL缓存机制：成功爬取后保存电影URL，下次跳过搜索直接访问
• 相似度算法：精确计算电影名称匹配度，显著提高准确率
• 平衡反爬策略：在速度和准确率之间找到最佳平衡

【高级功能】
• 智能反爬检测：自动识别并应对反爬机制
• 自适应延迟：根据成功率自动调整请求间隔
• User-Agent轮换：避免被识别为机器人
• 智能重试：失败请求自动重试
• 实时状态监控：显示当前爬虫健康状态

【反爬应对机制】
程序会自动检测以下情况并调整策略：
- 异常HTTP状态码（403, 429, 503等）
- 页面内容异常（过短或包含验证码关键词）
- 连续失败时会增加延迟时间
- 自动轮换User-Agent
- 清理并重建请求会话

【注意事项】
• 建议基础延迟设置为5-10秒，确保稳定性
• 遇到大量连续失败时，建议暂停一段时间
• 定期检查日志文件了解详细运行情况
• 程序每处理5个项目自动保存一次
• 大批量爬取建议分批进行，避免长时间运行
"""
        
        text_widget.insert('1.0', help_text)
        text_widget.config(state='disabled')
        
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def show_about(self):
        """显示关于信息"""
        about_text = """豆瓣影视媒资信息爬取工具
智能版 v3.0

新增特性：
✓ 智能名称匹配算法（显著提高准确率）
✓ URL缓存机制（二次爬取提速90%+）
✓ 相似度计算引擎
✓ 平衡反爬策略

核心特性：
✓ 智能反爬检测与规避
✓ 自适应延迟调整
✓ User-Agent轮换
✓ 详细日志记录
✓ 实时状态监控
✓ 进度保存与恢复

开发：AI助手协助开发
基于：Python + tkinter + requests + BeautifulSoup + difflib
"""
        messagebox.showinfo("关于", about_text)
    
    def reset_buttons(self):
        """重置按钮状态"""
        self.start_button.config(text="开始爬取", state="normal")
        self.pause_button.config(text="暂停", state="disabled")
        self.stop_button.config(text="停止", state="disabled")
        self.current_status.set("")
        self.crawler_status_text.set("爬虫状态: 待启动")
        self.is_crawling = False
        self.is_paused = False
        self.stop_crawling = False
    
    def run(self):
        """运行应用程序"""
        self.root.mainloop()


def main():
    """主函数"""
    # 设置程序图标和样式
    print("豆瓣影视媒资信息爬取工具 - 智能版v3.0")
    print("===========================================")
    print("新增：智能匹配、URL缓存、相似度算法、平衡反爬")
    print("特性：显著提高准确率、二次爬取提速90%+")
    print("日志文件：douban_crawler.log")
    print("===========================================")
    
    try:
        app = DoubanCrawlerApp()
        app.run()
    except Exception as e:
        logging.error(f"程序启动失败: {e}")
        logging.error(traceback.format_exc())
        print(f"程序启动失败: {e}")


if __name__ == "__main__":
    main()
