import requests
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook

#   定义字典
data = {
    "中文名": "",
    "首播时间": "",
    "制片地区": "",
    "作品类型": "",
    "主要奖项": "",
    "首播电视台": "",
    "导    演": "",
    "主    演": ""
}
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36 Edg/127.0.0.0",

}
#   获取 items
def get_items(url):
    response = requests.get(url, headers=headers)
    html = response.text
    # print(html)
    soup = BeautifulSoup(html, 'html.parser')
    items = soup.find_all('div', class_="itemWrapper_r0diC")
    return items

#   获取数据
def get_data(url, data):
    items = get_items(url)
    for item in items:
        dts = item.find_all('dt', class_="basicInfoItem_pNFko itemName_TQzq3")
        for dt in dts:
            for key in data:
                if dt.text == key:  # 判断 dt 的文本内容（去除空格）与字典中的 key 是否相同
                    dd = dt.find_next_sibling('dd')
                    text = dd.text
                    text = re.sub(r'\[.*?]', '', text)
                    data[key] = text
                    break
    return data

modle_xlsx = "C:\\Users\\fucha\\Desktop\\百度爬取TEST.xlsx"
# 加载工作簿
workbook = load_workbook(modle_xlsx)
# 获取活动工作表
worksheet = workbook.active
start_index = 50  # 目标写入的起始列索引
row_idx = 2
for row_index in range(2, worksheet.max_row + 1):
    n = worksheet.cell(row=row_index, column=1).value
    name = worksheet.cell(row=row_index, column=4).value  # 片名
    if name:
        url = f"https://baike.baidu.com/item/{name}?fromModule=lemma_search-box"
        result = get_data(url, data)
        array = [result[key] for key in result.keys()]
        print(f"####正在爬取第{n}部电视剧")
        print(array)
        columns = [51, 52, 53, 54, 55, 56, 57, 58]
        for i, column in enumerate(columns):
            worksheet.cell(row=row_index, column=column, value=array[i])
    else:
        break

# 保存工作簿
workbook.save(modle_xlsx)