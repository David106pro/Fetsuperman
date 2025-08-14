# -*- coding: utf-8 -*-
from openpyxl import load_workbook

# 加载模板表格
wb_template = load_workbook("C:\\Users\\fucha\\Desktop\\广东IPTV-移动侧底量引入-20240819.xlsx")
ws_template = wb_template.active

# 创建存放表格
wb_out = load_workbook("C:\\Users\\fucha\\Desktop\\out.xlsx")
ws_out = wb_out.active

# 定义行计数器
row_out = 1
def change_value(value):
    if value == "普通话":
        value_new = "国语版"
    elif value == "英语":
        value_new = "英文版"
    else:
        value_new = f"{value}版"
    return value_new

# 遍历模板表格的每一行
for row in ws_template.iter_rows(min_row=1, max_row=ws_template.max_row):
    # 获取第 16 列的值
    col_16_value = row[15].value
    # 如果第 16 列的值不为 None 且包含逗号
    if col_16_value is not None and ',' in col_16_value:
        # 按逗号分割第 16 列的值，得到数组
        col_16_values = col_16_value.split(",")
        for value in col_16_values:
            # 复制 1-15 列的数据
            for col in range(1, 16):
                if col == 3:  # 如果是第三列
                    original_value = row[col - 1].value
                    value_new = change_value(value)
                    ws_out.cell(row=row_out, column=col).value = f"{original_value}[{value_new}]"
                else:
                    ws_out.cell(row=row_out, column=col).value = row[col - 1].value
            # 写入第 16 列的值
            ws_out.cell(row=row_out, column=16).value = value
            # 复制 17-20 列的数据
            for col in range(17, 21):
                ws_out.cell(row=row_out, column=col).value = row[col - 1].value
            # 行计数器加 1
            row_out += 1
    else:
        # 直接复制 1-20 列的数据
        for col in range(1, 21):
            ws_out.cell(row=row_out, column=col).value = row[col - 1].value
        row_out += 1

# 保存存放表格
wb_out.save("C:\\Users\\fucha\\Desktop\\out.xlsx")