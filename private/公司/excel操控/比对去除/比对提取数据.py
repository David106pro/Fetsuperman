import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from collections import defaultdict
import openpyxl
import os
import platform
import subprocess # For cross-platform file opening

# 常量定义
ALL_SHEETS_TEXT = "全部sheet" # 添加全部sheet的选项文本

# --- Core Logic ---
def process_files(base_path, sub_path, bt_sheet_name, st_sheet_name, bt_key_col_name, st_key_col_name, bt_data_col_names, st_data_col_names):
    """
    Compares two Excel sheets, finds matching keys, and transfers data for multiple columns.
    Handles duplicate keys in the base table.
    Supports processing all sheets in the base workbook when bt_sheet_name is set to ALL_SHEETS_TEXT.

    Args:
        base_path (str): Path to the base Excel file.
        sub_path (str): Path to the content Excel file.
        bt_sheet_name (str): Name of the sheet in the base table, or ALL_SHEETS_TEXT for all sheets.
        st_sheet_name (str): Name of the sheet in the content table.
        bt_key_col_name (str): Name of the key column in the base table.
        st_key_col_name (str): Name of the key column in the content table.
        bt_data_col_names (list[str]): List of names of data columns in the base table (to be updated).
        st_data_col_names (list[str]): List of names of data columns in the content table (source).

    Returns:
        tuple: (total_base_rows, match_count, dict of per-sheet stats or None, error_message or None)
    """
    try:
        wb_base = openpyxl.load_workbook(base_path)
        wb_sub = openpyxl.load_workbook(sub_path)
        
        # Check if content sheet exists
        if st_sheet_name not in wb_sub.sheetnames:
             return 0, 0, None, f"错误: 在内容表 '{os.path.basename(sub_path)}' 中找不到工作表 '{st_sheet_name}'。可用工作表: {wb_sub.sheetnames}"
        st_sheet = wb_sub[st_sheet_name]
        
        # Find the sub table key column index
        def find_col_index(sheet, col_name, sheet_desc):
            for col_idx in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col_idx).value == col_name:
                    return col_idx
            raise ValueError(f"列 '{col_name}' 在 {sheet_desc} 中未找到。")
            
        st_key_col_idx = find_col_index(st_sheet, st_key_col_name, "内容表")
        
        # Load content table key data
        st_key_map = {}
        for row_index in range(2, st_sheet.max_row + 1):
             key = st_sheet.cell(row=row_index, column=st_key_col_idx).value
             if key is not None:
                 st_key_map[key] = row_index
                 
        # Calculate content table data column indices
        st_data_col_indices = [find_col_index(st_sheet, name, "内容表") for name in st_data_col_names]

        # Process base workbook
        sheets_to_process = []
        if bt_sheet_name == ALL_SHEETS_TEXT:
            # Process all sheets
            sheets_to_process = wb_base.sheetnames
            if not sheets_to_process:
                return 0, 0, None, f"错误: 原表 '{os.path.basename(base_path)}' 不包含任何工作表。"
        else:
            # Process just the selected sheet
            if bt_sheet_name not in wb_base.sheetnames:
                return 0, 0, None, f"错误: 在原表 '{os.path.basename(base_path)}' 中找不到工作表 '{bt_sheet_name}'。可用工作表: {wb_base.sheetnames}"
            sheets_to_process = [bt_sheet_name]
            
        # Initialize counters
        total_base_rows = 0
        total_match_count = 0
        sheet_stats = {}  # 存储每个sheet的统计信息
        
        # Process each sheet in the base workbook
        for current_sheet_name in sheets_to_process:
            sheet_rows = 0
            sheet_matches = 0
            
            bt_sheet = wb_base[current_sheet_name]
            
            try:
                # Find key column in this sheet
                bt_key_col_idx = find_col_index(bt_sheet, bt_key_col_name, f"原表 '{current_sheet_name}'")
                
                # Find data columns in this sheet (may fail if sheet structure differs)
                bt_data_col_indices = []
                for name in bt_data_col_names:
                    try:
                        bt_data_col_indices.append(find_col_index(bt_sheet, name, f"原表 '{current_sheet_name}'"))
                    except ValueError:
                        # Skip columns not found in this sheet
                        continue
                
                if not bt_data_col_indices:
                    # Skip this sheet if no data columns were found
                    sheet_stats[current_sheet_name] = {
                        "rows": 0, 
                        "matches": 0, 
                        "status": f"跳过: 在工作表中找不到任何指定的数据列"
                    }
                    continue
                
                # Build key map for this sheet
                bt_key_map = defaultdict(list)
                for row_index in range(2, bt_sheet.max_row + 1):
                    sheet_rows += 1
                    key = bt_sheet.cell(row=row_index, column=bt_key_col_idx).value
                    if key is not None:
                        bt_key_map[key].append(row_index)
                
                # Find matching keys and update
                updated_rows = set()
                common_keys = set(bt_key_map.keys()) & set(st_key_map.keys())
                
                for key in common_keys:
                    st_row = st_key_map[key]
                    
                    # Apply to all matching rows in the base table
                    for bt_row in bt_key_map[key]:
                        # Iterate through all data column pairs that exist in this sheet
                        for i in range(len(bt_data_col_indices)):
                            bt_data_col_idx = bt_data_col_indices[i]
                            if i < len(st_data_col_indices):  # Ensure we have a matching content column
                                st_data_col_idx = st_data_col_indices[i]
                                st_value = st_sheet.cell(row=st_row, column=st_data_col_idx).value
                                bt_sheet.cell(row=bt_row, column=bt_data_col_idx, value=st_value)
                        
                        updated_rows.add(bt_row)
                
                sheet_matches = len(updated_rows)
                
                # Update counters
                total_base_rows += sheet_rows
                total_match_count += sheet_matches
                
                # Save sheet statistics
                sheet_stats[current_sheet_name] = {
                    "rows": sheet_rows,
                    "matches": sheet_matches,
                    "status": "处理成功"
                }
                
            except ValueError as ve:
                # Record error for this sheet but continue with others
                sheet_stats[current_sheet_name] = {
                    "rows": sheet_rows,
                    "matches": 0,
                    "status": f"错误: {ve}"
                }
                continue
            
        # Save the workbook with all changes
        wb_base.save(base_path)
        return total_base_rows, total_match_count, sheet_stats, None

    except FileNotFoundError as fnf:
        # More specific error message
        missing_file = base_path if not os.path.exists(base_path) else sub_path
        return 0, 0, None, f"错误: 文件未找到: {missing_file}"
    except KeyError as ke: # Handles invalid sheet name from process_files
        return 0, 0, None, f"错误: 无效的工作表名称 - {ke}"
    except ValueError as ve: # Specific error for column not found or internal errors
        return 0, 0, None, f"错误: {ve}"
    except Exception as e:
        # Catch potential permission errors during save
        if isinstance(e, PermissionError):
             return total_base_rows, total_match_count, sheet_stats, f"错误: 无法保存文件 '{os.path.basename(base_path)}'。请确保文件未被其他程序打开，并且您有写入权限。"
        return 0, 0, None, f"发生意外错误: {e}\n类型: {type(e)}"

# --- GUI Application Class ---
class ExcelComparatorApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel 比对工具")
        master.geometry("750x550") # Adjusted size

        self.base_headers = []
        self.sub_headers = []
        self.base_sheets = []
        self.sub_sheets = []

        # Store dynamic data column widgets
        self.data_column_widgets = [] # List of tuples: (label, bt_combo, st_combo)
        self.next_data_col_num = 1

        # --- Style ---
        style = ttk.Style()
        style.configure("TLabel", padding=5)
        style.configure("TButton", padding=5)
        style.configure("TEntry", padding=5)
        style.configure("TCombobox", padding=5)
        style.configure("Header.TLabel", font=("Segoe UI", 9, "bold"))
        style.configure("TLabelframe.Label", padding=(0, 5))

        # --- File Selection Frame ---
        file_frame = ttk.LabelFrame(master, text="1. 选择 Excel 文件和工作表", padding=10)
        file_frame.pack(padx=10, pady=(10, 5), fill="x")
        file_frame.columnconfigure(1, weight=1) # Make entry expand

        # Base File
        ttk.Label(file_frame, text="原表:").grid(row=0, column=0, sticky="w", padx=(0,5))
        self.base_file_entry = ttk.Entry(file_frame, width=50)
        self.base_file_entry.grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(file_frame, text="浏览...", command=lambda: self._browse_file(self.base_file_entry, self.base_sheet_combo, 'base')).grid(row=0, column=2, padx=5)
        ttk.Label(file_frame, text="工作表:").grid(row=0, column=3, sticky="w", padx=(10,5))
        self.base_sheet_combo = ttk.Combobox(file_frame, state="readonly", width=20)
        self.base_sheet_combo.grid(row=0, column=4, sticky="ew", padx=5)
        self.base_sheet_combo.bind("<<ComboboxSelected>>", lambda event: self._on_sheet_selected(self.base_file_entry.get(), self.base_sheet_combo.get(), 'base'))

        # Sub File
        ttk.Label(file_frame, text="内容表:").grid(row=1, column=0, sticky="w", padx=(0,5), pady=(5,0))
        self.sub_file_entry = ttk.Entry(file_frame, width=50)
        self.sub_file_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=(5,0))
        ttk.Button(file_frame, text="浏览...", command=lambda: self._browse_file(self.sub_file_entry, self.sub_sheet_combo, 'sub')).grid(row=1, column=2, padx=5, pady=(5,0))
        ttk.Label(file_frame, text="工作表:").grid(row=1, column=3, sticky="w", padx=(10,5), pady=(5,0))
        self.sub_sheet_combo = ttk.Combobox(file_frame, state="readonly", width=20)
        self.sub_sheet_combo.grid(row=1, column=4, sticky="ew", padx=5, pady=(5,0))
        self.sub_sheet_combo.bind("<<ComboboxSelected>>", lambda event: self._on_sheet_selected(self.sub_file_entry.get(), self.sub_sheet_combo.get(), 'sub'))


        # --- Column Selection Frame ---
        self.col_frame = ttk.LabelFrame(master, text="2. 选择列", padding=10)
        self.col_frame.pack(padx=10, pady=5, fill="x")
        self.col_frame.columnconfigure(1, weight=1)
        self.col_frame.columnconfigure(2, weight=1)

        # Header Row
        ttk.Label(self.col_frame, text="").grid(row=0, column=0, padx=5, pady=5) # Spacer
        ttk.Label(self.col_frame, text="原表", style="Header.TLabel").grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Label(self.col_frame, text="内容表", style="Header.TLabel").grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Key Column Row
        ttk.Label(self.col_frame, text="关键字列:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.bt_key_combo = ttk.Combobox(self.col_frame, state="readonly", width=30)
        self.bt_key_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.st_key_combo = ttk.Combobox(self.col_frame, state="readonly", width=30)
        self.st_key_combo.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        # Placeholder for dynamic data columns - row index starts at 2
        self.data_col_start_row = 2

        # Add Data Column Button
        self.add_col_button = ttk.Button(self.col_frame, text="添加数据列", command=self._add_data_column_pair)
        # Place button below the key columns initially, will be moved down
        self.add_col_button.grid(row=self.data_col_start_row, column=0, columnspan=3, pady=10)

        # Add the first data column pair initially
        self._add_data_column_pair()


        # --- Action Frame ---
        action_frame = ttk.Frame(master, padding=10)
        action_frame.pack(pady=5, fill="x")
        # Make the frame itself expand horizontally
        action_frame.columnconfigure(0, weight=1)
        # Place the button in the center
        self.start_button = ttk.Button(action_frame, text="开始处理", command=self._start_processing, width=15)
        self.start_button.grid(row=0, column=0, pady=5) # Use grid to center within the frame column

        # --- Result Frame ---
        result_frame = ttk.LabelFrame(master, text="结果", padding=10)
        result_frame.pack(padx=10, pady=(5, 10), fill="both", expand=True)

        self.result_label = ttk.Label(result_frame, text="请选择文件和列，然后点击 '开始处理'。", wraplength=700, justify="left")
        self.result_label.pack(anchor="nw")

    def _add_data_column_pair(self):
        """Adds a new row for selecting a pair of data columns."""
        current_row = self.data_col_start_row + len(self.data_column_widgets)

        # Data Column Row
        label_text = f"数据列 ({self.next_data_col_num}):"
        data_label = ttk.Label(self.col_frame, text=label_text)
        data_label.grid(row=current_row, column=0, padx=5, pady=5, sticky="w")

        bt_data_combo = ttk.Combobox(self.col_frame, state="readonly", width=30)
        bt_data_combo.grid(row=current_row, column=1, padx=5, pady=5, sticky="ew")
        bt_data_combo['values'] = self.base_headers # Populate with current headers

        st_data_combo = ttk.Combobox(self.col_frame, state="readonly", width=30)
        st_data_combo.grid(row=current_row, column=2, padx=5, pady=5, sticky="ew")
        st_data_combo['values'] = self.sub_headers # Populate with current headers

        # Store widgets for later access/update
        self.data_column_widgets.append((data_label, bt_data_combo, st_data_combo))
        self.next_data_col_num += 1

        # Move the button down
        self.add_col_button.grid(row=current_row + 1, column=0, columnspan=3, pady=10)

    def _load_sheets_and_headers(self, file_path, sheet_combo, table_type):
        """Loads sheet names into the combobox and headers for the active sheet."""
        if not file_path or not os.path.exists(file_path):
             # Clear relevant widgets if file path is invalid or cleared
            sheet_combo['values'] = []
            sheet_combo.set('')
            if table_type == 'base':
                self.base_sheets = []
                self.base_headers = []
                self.bt_key_combo['values'] = []
                self.bt_key_combo.set('')
                for _, bt_combo, _ in self.data_column_widgets:
                    bt_combo['values'] = []
                    bt_combo.set('')
            else:
                self.sub_sheets = []
                self.sub_headers = []
                self.st_key_combo['values'] = []
                self.st_key_combo.set('')
                for _, _, st_combo in self.data_column_widgets:
                    st_combo['values'] = []
                    st_combo.set('')
            return

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            active_sheet_name = workbook.active.title # Get active sheet by title
            workbook.close()

            if not sheet_names:
                 messagebox.showwarning("无工作表", f"文件 '{os.path.basename(file_path)}' 不包含任何工作表。")
                 sheet_combo['values'] = []
                 sheet_combo.set('')
                 return

            # 更新下拉框列表
            if table_type == 'base':
                # 为原表添加"全部sheet"选项
                sheet_combo['values'] = [ALL_SHEETS_TEXT] + sheet_names
                self.base_sheets = sheet_names
            else:
                sheet_combo['values'] = sheet_names
                self.sub_sheets = sheet_names

            # 设置默认选中项
            if active_sheet_name in sheet_names:
                 sheet_combo.set(active_sheet_name)
            elif sheet_names: # If active sheet not found, select first
                 sheet_combo.set(sheet_names[0])
                 active_sheet_name = sheet_names[0]
            else: # Should not happen if sheet_names is not empty, but safeguard
                active_sheet_name = None
                sheet_combo.set('')

            # Load headers for the initially selected (active or first) sheet
            if active_sheet_name:
                 self._load_headers(file_path, active_sheet_name, table_type)
            else:
                # Clear headers if no sheet could be selected
                self._clear_headers(table_type)

        except Exception as e:
            messagebox.showerror("错误", f"无法从 '{os.path.basename(file_path)}' 加载工作表列表或初始表头:\n{e}")
            sheet_combo['values'] = []
            sheet_combo.set('')
            self._clear_headers(table_type)

    def _on_sheet_selected(self, file_path, sheet_name, table_type):
        """当用户从下拉框选择工作表时调用"""
        if not file_path:
            self._clear_headers(table_type)
            return
        
        if table_type == 'base' and sheet_name == ALL_SHEETS_TEXT:
            # 如果选择了"全部sheet"，使用第一个sheet的表头作为模板
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            if workbook.sheetnames:
                self._load_headers(file_path, workbook.sheetnames[0], table_type)
            else:
                self._clear_headers(table_type)
        elif file_path and sheet_name:
            self._load_headers(file_path, sheet_name, table_type)
        else:
            # Clear headers if file path or sheet name is missing
            self._clear_headers(table_type)

    def _clear_headers(self, table_type):
         """Clears header comboboxes for the specified table type."""
         if table_type == 'base':
             self.base_headers = []
             self.bt_key_combo['values'] = []
             self.bt_key_combo.set('')
             for _, bt_combo, _ in self.data_column_widgets:
                 bt_combo['values'] = []
                 bt_combo.set('')
         elif table_type == 'sub':
             self.sub_headers = []
             self.st_key_combo['values'] = []
             self.st_key_combo.set('')
             for _, _, st_combo in self.data_column_widgets:
                 st_combo['values'] = []
                 st_combo.set('')

    def _load_headers(self, file_path, sheet_name, table_type):
        """Loads headers from the first row of a specific Excel sheet."""
        if not file_path or not sheet_name:
             self._clear_headers(table_type)
             return

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name not in workbook.sheetnames:
                 # This might happen if the file was changed externally
                 messagebox.showwarning("工作表丢失", f"工作表 '{sheet_name}' 在文件 '{os.path.basename(file_path)}' 中找不到了。请重新选择文件或工作表。")
                 self._clear_headers(table_type)
                 workbook.close()
                 # Also clear the sheet selection as it's invalid now
                 if table_type == 'base':
                     self.base_sheet_combo['values'] = workbook.sheetnames # Update with current sheets
                     self.base_sheet_combo.set('')
                 else:
                    self.sub_sheet_combo['values'] = workbook.sheetnames
                    self.sub_sheet_combo.set('')
                 return

            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1] if cell.value is not None]
            workbook.close()

            if not headers:
                messagebox.showwarning("空工作表或无表头", f"在 '{os.path.basename(file_path)}' 的工作表 '{sheet_name}' 的第一行找不到任何表头，或者工作表为空。")
                self._clear_headers(table_type)
                return

            # Update the corresponding comboboxes
            if table_type == 'base':
                self.base_headers = headers
                self.bt_key_combo['values'] = headers
                # Update all existing base data comboboxes
                for _, bt_combo, _ in self.data_column_widgets:
                     bt_combo['values'] = headers
            elif table_type == 'sub':
                self.sub_headers = headers
                self.st_key_combo['values'] = headers
                 # Update all existing sub data comboboxes
                for _, _, st_combo in self.data_column_widgets:
                     st_combo['values'] = headers

            # Clear current selections when headers change
            # self.bt_key_combo.set('')
            # self.st_key_combo.set('')
            # for _, bt_combo, st_combo in self.data_column_widgets:
            #      bt_combo.set('')
            #      st_combo.set('')

        except Exception as e:
            messagebox.showerror("错误", f"无法从 '{os.path.basename(file_path)}' 的工作表 '{sheet_name}' 读取表头:\n{e}")
            self._clear_headers(table_type)


    def _browse_file(self, entry_widget, sheet_combo, table_type):
        """Opens file dialog, updates entry, loads sheets and initial headers."""
        type_map = {'base': '原表', 'sub': '内容表'}
        filepath = filedialog.askopenfilename(
            title=f"选择 {type_map.get(table_type, table_type.capitalize())} Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls")]
        )
        if filepath:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filepath)
            # Load sheets and headers for the active sheet
            self._load_sheets_and_headers(filepath, sheet_combo, table_type)
        else:
            # If user cancels browse, clear the path and related widgets
             entry_widget.delete(0, tk.END)
             self._load_sheets_and_headers(None, sheet_combo, table_type)

    def _open_file(self, file_path):
         """Opens the specified file using the default application."""
         try:
             # Ensure path is absolute for OS calls
             abs_path = os.path.abspath(file_path)
             if platform.system() == "Windows":
                 os.startfile(abs_path)
             elif platform.system() == "Darwin": # macOS
                 subprocess.call(["open", abs_path])
             else: # Linux variants
                 subprocess.call(["xdg-open", abs_path])
         except Exception as e:
             self.result_label.config(text=self.result_label.cget("text") + f"\n无法自动打开文件: {e}")


    def _start_processing(self):
        """Gets inputs, validates, calls core logic, and updates results."""
        base_path = self.base_file_entry.get()
        sub_path = self.sub_file_entry.get()
        bt_sheet = self.base_sheet_combo.get()
        st_sheet = self.sub_sheet_combo.get()
        bt_key_col = self.bt_key_combo.get()
        st_key_col = self.st_key_combo.get()

        # Gather data column pairs
        bt_data_cols = []
        st_data_cols = []
        has_incomplete_data_pair = False
        for i, (_, bt_combo, st_combo) in enumerate(self.data_column_widgets):
            bt_data = bt_combo.get()
            st_data = st_combo.get()
            if bt_data and st_data: # Only add pair if both are selected
                bt_data_cols.append(bt_data)
                st_data_cols.append(st_data)
            elif bt_data or st_data: # If one is selected but not the other
                has_incomplete_data_pair = True

        # --- Input Validation ---
        if not base_path or not sub_path:
            messagebox.showerror("错误", "请选择原表和内容表 Excel 文件。")
            return
        if not bt_sheet or not st_sheet:
             messagebox.showerror("错误", "请为两个表格选择工作表。")
             return
        if not bt_key_col or not st_key_col:
             messagebox.showerror("错误", "请为两个表格选择关键字列。")
             return
        if not bt_data_cols or not st_data_cols:
            messagebox.showerror("错误", "请至少选择一对完整的数据列（原表和内容表都需要选择）。")
            return
        if has_incomplete_data_pair:
            messagebox.showwarning("警告", "检测到未完整选择的数据列对（即原表或内容表的数据列未选），这些不完整的列将被忽略。")

        # Key/Data column name validation (optional but good practice)
        if bt_key_col in bt_data_cols:
             messagebox.showwarning("警告", f"原表的关键字列 ('{bt_key_col}') 也被选作了数据列。确定要这样操作吗？")
        if st_key_col in st_data_cols:
             messagebox.showwarning("警告", f"内容表的关键字列 ('{st_key_col}') 也被选作了数据列。确定要这样操作吗？")
        # Check for duplicate data columns selections for the same table
        if len(bt_data_cols) != len(set(bt_data_cols)):
             messagebox.showwarning("警告", "原表中选择了重复的数据列。请检查选择。")
             # Allow proceeding, but warn
        if len(st_data_cols) != len(set(st_data_cols)):
             messagebox.showwarning("警告", "内容表中选择了重复的数据列。请检查选择。")
             # Allow proceeding, but warn


        # --- Disable UI during processing ---
        self.start_button.config(state=tk.DISABLED)
        self.add_col_button.config(state=tk.DISABLED)
        self.result_label.config(text="处理中... 请稍候。")
        self.master.update_idletasks() # Force UI update

        # --- Call Core Logic ---
        total_rows, matched_count, sheet_stats, error_msg = process_files(
            base_path, sub_path, bt_sheet, st_sheet, bt_key_col, st_key_col, bt_data_cols, st_data_cols
        )

        # --- Update Results ---
        if error_msg:
            self.result_label.config(text=error_msg)
            messagebox.showerror("处理错误", error_msg)
        else:
            if bt_sheet == ALL_SHEETS_TEXT:
                # 显示每个sheet的处理结果
                result_text = f"处理完成！\n\n总行数: {total_rows}\n成功匹配并更新的总行数: {matched_count}\n\n各工作表详细情况:\n"
                
                # 按工作表名称排序，并添加详细结果
                for sheet_name in sorted(sheet_stats.keys()):
                    stats = sheet_stats[sheet_name]
                    sheet_result = f"工作表 '{sheet_name}': "
                    
                    if stats['status'] == "处理成功":
                        sheet_result += f"检查 {stats['rows']} 行，匹配更新 {stats['matches']} 行"
                    else:
                        sheet_result += stats['status']
                        
                    result_text += f"{sheet_result}\n"
                    
                result_text += f"\n更新的数据列对数量: {len(bt_data_cols)}\n结果已保存至: {os.path.basename(base_path)}"
            else:
                result_text = (f"处理完成！\n"
                              f"原表 '{bt_sheet}' 中检查的总行数: {total_rows}\n"
                              f"成功匹配并更新的行数: {matched_count}\n"
                              f"更新的数据列对数量: {len(bt_data_cols)}\n"
                              f"结果已保存至: {os.path.basename(base_path)}")
                              
            self.result_label.config(text=result_text)
            # Ask before opening
            if matched_count > 0 and messagebox.askyesno("打开结果?", f"处理完成。是否打开更新后的文件？\n{base_path}"):
                self._open_file(base_path)


        # --- Re-enable UI ---
        self.start_button.config(state=tk.NORMAL)
        self.add_col_button.config(state=tk.NORMAL)


# --- Main Execution ---
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparatorApp(root)
    root.mainloop()