from openpyxl import load_workbook
import shutil

tag_tab = "C:\\Users\\fucha\\Desktop\\【已筛】华视网聚内容片单20241127.xlsx"
sub_tab = "C:\\Users\\fucha\\Desktop\\大片单列表-20241223.xlsx"
# 总表比对列数
bt_index = 1
# 比对表比对列数
st_index = 1
# 输入要标记的标签
tag_value = "T"
# 加载表格
wb_total_filtered = load_workbook(tag_tab)
bt_sheet = wb_total_filtered.active
wb_sub = load_workbook(sub_tab)
st_sheet = wb_sub.active


# 字典比对，取行号+id，支持重复id
def diff_id(bt_sheet, st_sheet):
    # 使用列表存储每个id对应的所有行号
    bt_ids = {}
    st_ids = {}

    # 收集总表中的id和对应的所有行号
    for row_index in range(2, bt_sheet.max_row + 1):
        id_value = bt_sheet.cell(row=row_index, column=bt_index).value
        if id_value not in bt_ids:
            bt_ids[id_value] = []
        bt_ids[id_value].append(row_index)

    # 收集子表中的id和对应的所有行号
    for row_index in range(2, st_sheet.max_row + 1):
        id_value = st_sheet.cell(row=row_index, column=st_index).value
        if id_value not in st_ids:
            st_ids[id_value] = []
        st_ids[id_value].append(row_index)

    # 找出共同的id
    common_ids = set(bt_ids.keys()) & set(st_ids.keys())

    # 收集所有匹配行的行号
    all_matching_rows = []
    for id_value in common_ids:
        all_matching_rows.extend(bt_ids[id_value])

    print(f'找到的匹配ID数量: {len(common_ids)}')
    print(f'找到的匹配行数: {len(all_matching_rows)}')

    return all_matching_rows, list(common_ids)


# 最后一列打标签
def tab_mark(diff_rows):
    n = 0
    for i in diff_rows:
        bt_sheet.cell(row=i, column=bt_sheet.max_column, value=tag_value)
        n += 1
    print(f'已标记行数: {n}')


# 执行比对和标记
diff_rows, diff_ids = diff_id(bt_sheet, st_sheet)
tab_mark(diff_rows)
# 保存结果
wb_total_filtered.save(tag_tab)