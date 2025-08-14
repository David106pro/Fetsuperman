#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
北京运营表格生成脚本
功能：根据输入表格生成送审表和上线表
作者：AI助手
版本：1.0
"""

import os
import sys
import json
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from pathlib import Path
import logging
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import glob
import subprocess
import re

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('北京运营表格生成.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class BeijingTableGenerator:
    """北京运营表格生成器主类"""
    
    def __init__(self):
        """初始化配置"""
        self.config_file = "setting/config.json"
        self.settings_window = None  # 设置窗口引用，用于防止重复打开
        self.load_config()
        # 获取默认路径并更新配置
        default_paths = self.get_default_paths()
        for key, value in default_paths.items():
            if not self.config.get(key):  # 如果配置中没有该路径或为空
                self.config[key] = value
        self.setup_ui()
        
    def load_config(self):
        """加载配置文件"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
            else:
                # 默认配置
                self.config = {
                    "总表路径": "",
                    "专辑注入路径": "",
                    "剧集注入路径": "",
                    "审核表文件夹": r"C:\Users\fucha\Desktop\北京项目\3、审核结果",
                    "送审表输出路径": r"C:\Users\fucha\Desktop\北京项目\1、送审",
                    "上线表输出路径": r"C:\Users\fucha\Desktop\北京项目\2、上线",
                    "是否备份总表": False,
                    "是否统一审批状态格式": False
                }
                self.save_config()
        except Exception as e:
            logger.error(f"加载配置文件失败: {e}")
            self.config = {}
    
    def get_default_paths(self):
        """获取默认路径，如果文件不存在则返回空"""
        default_paths = {}
        
        # 总表默认路径
        total_table_path = r"C:\Users\fucha\Desktop\【北京OTT总表】提审&发上线总表.xlsx"
        if os.path.exists(total_table_path):
            default_paths["总表路径"] = total_table_path
        else:
            default_paths["总表路径"] = ""
        
        # 专辑注入默认路径
        album_path = r"C:/Users/fucha/Desktop/北京输入表格/北京专辑注入.csv"
        if os.path.exists(album_path):
            default_paths["专辑注入路径"] = album_path
        else:
            default_paths["专辑注入路径"] = ""
        
        # 剧集注入默认路径
        episode_path = r"C:/Users/fucha/Desktop/北京输入表格/北京子集注入.csv"
        if os.path.exists(episode_path):
            default_paths["剧集注入路径"] = episode_path
        else:
            default_paths["剧集注入路径"] = ""
        
        # 审核表默认路径 - 获取文件夹中最新的xlsx文件
        review_folder = r"C:\Users\fucha\Desktop\北京项目\3、审核结果"
        latest_review_file = self.get_latest_file_from_folder(review_folder, "*.xlsx")
        default_paths["审核表路径"] = latest_review_file if latest_review_file else ""
        
        # 送审表输出路径 - 获取文件夹中最新的xlsx文件
        submit_folder = r"C:\Users\fucha\Desktop\北京项目\1、送审"
        latest_submit_file = self.get_latest_file_from_folder(submit_folder, "*.xlsx")
        default_paths["送审表输出路径"] = latest_submit_file if latest_submit_file else ""
        
        # 上线表输出路径 - 获取文件夹中最新的xlsx文件
        online_folder = r"C:\Users\fucha\Desktop\北京项目\2、上线"
        latest_online_file = self.get_latest_file_from_folder(online_folder, "*.xlsx")
        default_paths["上线表输出路径"] = latest_online_file if latest_online_file else ""
        
        return default_paths
    
    def get_latest_file_from_folder(self, folder_path, pattern):
        """从文件夹中获取最新的文件"""
        try:
            if not os.path.exists(folder_path):
                return None
            
            # 查找所有匹配的文件
            pattern_path = os.path.join(folder_path, pattern)
            files = glob.glob(pattern_path)
            
            if not files:
                return None
            
            # 按修改时间排序，返回最新的文件
            latest_file = max(files, key=os.path.getmtime)
            return latest_file
            
        except Exception as e:
            logger.error(f"获取最新文件失败: {e}")
            return None
    
    def save_config(self):
        """保存配置文件"""
        try:
            os.makedirs("setting", exist_ok=True)
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"保存配置文件失败: {e}")
    
    def setup_ui(self):
        """设置UI界面"""
        self.root = tk.Tk()
        self.root.title("北京运营表格生成工具")
        self.root.geometry("900x820")  # 增加页面高度
        
        # 设置样式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 自定义勾选项样式，使用对钩符号
        style.configure('Custom.TCheckbutton', 
                       indicatorcolor='white',
                       indicatorrelief='raised',
                       indicatordiameter=15)
        
        # 设置勾选项的选中和未选中状态
        style.map('Custom.TCheckbutton',
                 indicatorcolor=[('selected', 'green'), ('!selected', 'white')],
                 indicatorrelief=[('pressed', 'sunken'), ('!pressed', 'raised')])
        
        # 强制更新样式
        style.theme_use('clam')
        
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="北京运营表格生成工具", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 10))
        
        # 创建内容框架
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # 总表路径选择
        self.create_total_table_section(content_frame)
        
        # CMS表格导入
        self.create_cms_section(content_frame)
        
        # 审核表导入
        self.create_review_section(content_frame)
        
        # 送审表生成
        self.create_submit_section(content_frame)
        
        # 上线表生成
        self.create_online_section(content_frame)
        
        # 底部按钮区域
        self.create_bottom_buttons(main_frame)
    
    def create_total_table_section(self, parent):
        """创建总表路径选择区域"""
        frame = ttk.LabelFrame(parent, text="总表路径选择", padding="5")
        frame.pack(fill="x", padx=5, pady=5)
        
        # 总表路径 - 一行布局
        path_frame = ttk.Frame(frame)
        path_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(path_frame, text="总表路径:", width=12).pack(side="left", padx=(5, 5), pady=5)
        
        self.total_table_path_var = tk.StringVar(value=self.config.get("总表路径", ""))
        self.total_table_entry = ttk.Entry(path_frame, textvariable=self.total_table_path_var)
        self.total_table_entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(path_frame, text="浏览...", command=self.browse_total_table, width=10).pack(side="right", padx=5, pady=5)
    
    def create_cms_section(self, parent):
        """创建CMS表格导入区域"""
        frame = ttk.LabelFrame(parent, text="CMS表格导入", padding="5")
        frame.pack(fill="x", padx=5, pady=5)
        
        # 专辑注入路径 - 一行布局
        album_frame = ttk.Frame(frame)
        album_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(album_frame, text="专辑注入路径:", width=12).pack(side="left", padx=(5, 5), pady=5)
        
        self.album_path_var = tk.StringVar(value=self.config.get("专辑注入路径", ""))
        self.album_entry = ttk.Entry(album_frame, textvariable=self.album_path_var)
        self.album_entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(album_frame, text="浏览...", command=self.browse_album, width=10).pack(side="right", padx=5, pady=5)
        
        # 剧集注入路径 - 一行布局
        episode_frame = ttk.Frame(frame)
        episode_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(episode_frame, text="剧集注入路径:", width=12).pack(side="left", padx=(5, 5), pady=5)
        
        self.episode_path_var = tk.StringVar(value=self.config.get("剧集注入路径", ""))
        self.episode_entry = ttk.Entry(episode_frame, textvariable=self.episode_path_var)
        self.episode_entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(episode_frame, text="浏览...", command=self.browse_episode, width=10).pack(side="right", padx=5, pady=5)
        
        # 勾选项
        checkbox_frame = ttk.Frame(frame)
        checkbox_frame.pack(fill="x", padx=5, pady=5)
        
        self.import_album_var = tk.BooleanVar()
        self.import_episode_var = tk.BooleanVar()
        
        ttk.Checkbutton(checkbox_frame, text="导入专辑数据", variable=self.import_album_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
        ttk.Checkbutton(checkbox_frame, text="导入剧集数据", variable=self.import_episode_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
    
    def create_review_section(self, parent):
        """创建审核表导入区域"""
        frame = ttk.LabelFrame(parent, text="审核表导入", padding="5")
        frame.pack(fill="x", padx=5, pady=5)
        
        # 审核表路径 - 一行布局
        review_frame = ttk.Frame(frame)
        review_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(review_frame, text="审核表路径:", width=12).pack(side="left", padx=(5, 5), pady=5)
        
        # 获取最新的审核表文件路径
        latest_review_file = self.get_latest_review_file()
        review_path = self.config.get("审核表路径", latest_review_file if latest_review_file else "")
        
        self.review_path_var = tk.StringVar(value=review_path)
        self.review_entry = ttk.Entry(review_frame, textvariable=self.review_path_var)
        self.review_entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(review_frame, text="浏览...", command=self.browse_review, width=10).pack(side="right", padx=5, pady=5)
        
        # 勾选项
        checkbox_frame = ttk.Frame(frame)
        checkbox_frame.pack(fill="x", padx=5, pady=5)
        
        self.import_album_review_var = tk.BooleanVar()
        self.import_episode_review_var = tk.BooleanVar()
        self.supplement_status_var = tk.BooleanVar()
        
        ttk.Checkbutton(checkbox_frame, text="导入节目集审核结果", variable=self.import_album_review_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
        ttk.Checkbutton(checkbox_frame, text="导入节目审核结果", variable=self.import_episode_review_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
        ttk.Checkbutton(checkbox_frame, text="总表审核状态补充", variable=self.supplement_status_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
    
    def create_submit_section(self, parent):
        """创建送审表生成区域"""
        frame = ttk.LabelFrame(parent, text="送审表生成", padding="5")
        frame.pack(fill="x", padx=5, pady=5)
        
        # 输出路径 - 一行布局
        output_frame = ttk.Frame(frame)
        output_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(output_frame, text="送审表输出路径:", width=12).pack(side="left", padx=(5, 5), pady=5)
        
        self.submit_path_var = tk.StringVar(value=self.config.get("送审表输出路径", ""))
        self.submit_entry = ttk.Entry(output_frame, textvariable=self.submit_path_var)
        self.submit_entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(output_frame, text="浏览...", command=self.browse_submit, width=10).pack(side="right", padx=5, pady=5)
        
        # 勾选项
        checkbox_frame = ttk.Frame(frame)
        checkbox_frame.pack(fill="x", padx=5, pady=5)
        
        self.generate_episode_submit_var = tk.BooleanVar()
        self.generate_album_submit_var = tk.BooleanVar()
        
        ttk.Checkbutton(checkbox_frame, text="生成剧集送审表", variable=self.generate_episode_submit_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
        ttk.Checkbutton(checkbox_frame, text="生成专辑送审表", variable=self.generate_album_submit_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
    
    def create_online_section(self, parent):
        """创建上线表生成区域"""
        frame = ttk.LabelFrame(parent, text="上线表生成", padding="5")
        frame.pack(fill="x", padx=5, pady=5)
        
        # 输出路径 - 一行布局
        output_frame = ttk.Frame(frame)
        output_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(output_frame, text="上线表输出路径:", width=12).pack(side="left", padx=(5, 5), pady=5)
        self.online_path_var = tk.StringVar(value=self.config.get("上线表输出路径", ""))
        self.online_entry = ttk.Entry(output_frame, textvariable=self.online_path_var)
        self.online_entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(output_frame, text="浏览...", command=self.browse_online, width=10).pack(side="right", padx=5, pady=5)
        
        # 勾选项
        checkbox_frame = ttk.Frame(frame)
        checkbox_frame.pack(fill="x", padx=5, pady=5)
        
        self.generate_online_var = tk.BooleanVar()
        ttk.Checkbutton(checkbox_frame, text="生成上线表", variable=self.generate_online_var, 
                       style='Custom.TCheckbutton').pack(side="left", padx=10)
    
    def create_bottom_buttons(self, parent):
        """创建底部按钮区域"""
        bottom_frame = ttk.Frame(parent)
        bottom_frame.pack(fill="x", padx=10, pady=10)
        
        # 左侧状态显示
        self.status_label = ttk.Label(bottom_frame, text="准备就绪")
        self.status_label.pack(side="left", padx=10, pady=5)
        
        # 中间执行按钮
        button_frame = ttk.Frame(bottom_frame)
        button_frame.pack(side="left", expand=True, pady=5)
        
        ttk.Button(button_frame, text="执行勾选项", command=self.execute_selected, 
                  width=15).pack(side="left", padx=5, pady=5)
        ttk.Button(button_frame, text="清除全部选择", command=self.clear_all_selections, 
                  width=15).pack(side="left", padx=5, pady=5)
        
        # 右侧设置按钮（图标形式）
        settings_button = ttk.Button(bottom_frame, text="⚙", command=self.open_settings, 
                                    width=5)
        settings_button.pack(side="right", padx=10, pady=5)
    

    
    def clear_all_selections(self):
        """清除全部选择"""
        try:
            # 清除所有勾选项
            self.import_album_var.set(False)
            self.import_episode_var.set(False)
            self.import_album_review_var.set(False)
            self.import_episode_review_var.set(False)
            self.supplement_status_var.set(False)
            self.generate_episode_submit_var.set(False)
            self.generate_album_submit_var.set(False)
            self.generate_online_var.set(False)
            
            self.log_message("已清除全部选择")
            self.status_label.configure(text="已清除全部选择")
            
        except Exception as e:
            self.log_message(f"清除选择失败: {e}")
    
    def log_message(self, message):
        """添加日志消息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        print(log_entry)
        logger.info(message)
    
    def browse_total_table(self):
        """浏览总表文件"""
        filename = filedialog.askopenfilename(
            title="选择总表文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.total_table_path_var.set(filename)
    
    def browse_album(self):
        """浏览专辑注入文件"""
        filename = filedialog.askopenfilename(
            title="选择专辑注入文件",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.album_path_var.set(filename)
    
    def browse_episode(self):
        """浏览剧集注入文件"""
        filename = filedialog.askopenfilename(
            title="选择剧集注入文件",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.episode_path_var.set(filename)
    
    def browse_review(self):
        """浏览审核表文件"""
        filename = filedialog.askopenfilename(
            title="选择审核表文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.review_path_var.set(filename)
    
    def browse_submit(self):
        """浏览送审表输出文件"""
        filename = filedialog.askopenfilename(
            title="选择送审表输出文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.submit_path_var.set(filename)
    
    def browse_online(self):
        """浏览上线表输出文件"""
        filename = filedialog.askopenfilename(
            title="选择上线表输出文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.online_path_var.set(filename)
    
    def open_settings(self):
        """打开设置窗口"""
        # 检查是否已有设置窗口打开
        if self.settings_window is not None:
            try:
                # 尝试将窗口置顶
                self.settings_window.lift()
                self.settings_window.focus_force()
                return
            except tk.TclError:
                # 如果窗口已被销毁，重置引用
                self.settings_window = None
        
        # 创建设置窗口
        self.settings_window = tk.Toplevel(self.root)
        self.settings_window.title("设置")
        self.settings_window.geometry("600x500")
        self.settings_window.resizable(True, True)
        
        # 设置窗口关闭时的处理
        def on_settings_close():
            if self.settings_window is not None:
                self.settings_window.destroy()
                self.settings_window = None
        
        self.settings_window.protocol("WM_DELETE_WINDOW", on_settings_close)
        
        # 创建设置界面
        main_frame = ttk.Frame(self.settings_window, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 创建主内容框架，填充整个窗口
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill="both", expand=True)
        
        # 路径设置区域
        path_frame = ttk.LabelFrame(content_frame, text="路径设置", padding="10")
        path_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 总表路径设置 - 一行布局
        self.create_setting_item_compact(path_frame, "总表路径", self.total_table_path_var)
        
        # 专辑注入路径设置 - 一行布局
        self.create_setting_item_compact(path_frame, "专辑注入路径", self.album_path_var)
        
        # 剧集注入路径设置 - 一行布局
        self.create_setting_item_compact(path_frame, "剧集注入路径", self.episode_path_var)
        
        # 审核表路径设置 - 一行布局
        self.create_setting_item_compact(path_frame, "审核表路径", self.review_path_var)
        
        # 送审表输出路径设置 - 一行布局
        self.create_setting_item_compact(path_frame, "送审表输出路径", self.submit_path_var)
        
        # 上线表输出路径设置 - 一行布局
        self.create_setting_item_compact(path_frame, "上线表输出路径", self.online_path_var)
        
        # 功能设置区域
        function_frame = ttk.LabelFrame(content_frame, text="功能设置", padding="10")
        function_frame.pack(fill="x", padx=5, pady=5)
        
        # 功能选项放在同一行
        checkbox_frame = ttk.Frame(function_frame)
        checkbox_frame.pack(fill="x", padx=5, pady=5)
        
        # 是否备份总表
        self.backup_total_table_var = tk.BooleanVar(value=self.config.get("是否备份总表", False))
        backup_check = ttk.Checkbutton(checkbox_frame, text="是否备份总表", 
                                      variable=self.backup_total_table_var, style='Custom.TCheckbutton')
        backup_check.pack(side="left", padx=(5, 20), pady=5)
        
        # 审核状态格式统一
        self.unify_status_format_var = tk.BooleanVar(value=self.config.get("审核状态格式统一", False))
        unify_check = ttk.Checkbutton(checkbox_frame, text="审核状态格式统一", 
                                     variable=self.unify_status_format_var, style='Custom.TCheckbutton')
        unify_check.pack(side="left", padx=5, pady=5)
        

        
        # 保存和返回按钮
        button_frame = ttk.Frame(content_frame)
        button_frame.pack(fill="x", pady=20)
        
        # 按钮居中
        center_frame = ttk.Frame(button_frame)
        center_frame.pack(expand=True)
        
        ttk.Button(center_frame, text="保存", command=lambda: self.save_settings(self.settings_window)).pack(side="left", padx=10)
        ttk.Button(center_frame, text="返回", command=on_settings_close).pack(side="left", padx=10)
    
    def create_setting_item(self, parent, label, variable):
        """创建设置项"""
        frame = ttk.Frame(parent)
        frame.pack(fill="x", pady=5)
        
        ttk.Label(frame, text=f"{label}:").pack(anchor="w", padx=5, pady=2)
        
        entry = ttk.Entry(frame, textvariable=variable, width=50)
        entry.pack(side="left", padx=5, pady=5, fill="x", expand=True)
        
        ttk.Button(frame, text="浏览", 
                  command=lambda: self.browse_setting_path(variable, label)).pack(side="right", padx=5, pady=5)
    
    def create_setting_item_compact(self, parent, label, variable):
        """创建紧凑型设置项（一行布局）"""
        frame = ttk.Frame(parent)
        frame.pack(fill="x", pady=2)
        
        # 标签和输入框在同一行
        ttk.Label(frame, text=f"{label}:", width=15).pack(side="left", padx=(5, 5), pady=2)
        
        entry = ttk.Entry(frame, textvariable=variable)
        entry.pack(side="left", padx=5, pady=2, fill="x", expand=True)
        
        ttk.Button(frame, text="浏览", width=8,
                  command=lambda: self.browse_setting_path(variable, label)).pack(side="right", padx=5, pady=2)
    
    def browse_setting_path(self, variable, label):
        """浏览设置路径"""
        # 根据标签选择不同的文件类型
        if "专辑注入" in label or "剧集注入" in label:
            # 专辑和剧集注入文件支持CSV和Excel
            filename = filedialog.askopenfilename(
                title=f"选择{label}",
                filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                variable.set(filename)
        elif "审核表" in label:
            # 审核表只支持Excel
            filename = filedialog.askopenfilename(
                title=f"选择{label}",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                variable.set(filename)
        elif "总表" in label:
            # 总表只支持Excel
            filename = filedialog.askopenfilename(
                title=f"选择{label}",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                variable.set(filename)
        elif "输出路径" in label:
            # 输出路径选择文件夹
            folder = filedialog.askdirectory(title=f"选择{label}")
            if folder:
                variable.set(folder)
        else:
            # 其他文件类型
            filename = filedialog.askopenfilename(title=f"选择{label}")
            if filename:
                variable.set(filename)
    
    def save_settings(self, window):
        """保存设置"""
        self.config["总表路径"] = self.total_table_path_var.get()
        self.config["专辑注入路径"] = self.album_path_var.get()
        self.config["剧集注入路径"] = self.episode_path_var.get()
        self.config["审核表路径"] = self.review_path_var.get()
        self.config["送审表输出路径"] = self.submit_path_var.get()
        self.config["上线表输出路径"] = self.online_path_var.get()
        self.config["是否备份总表"] = self.backup_total_table_var.get()
        self.config["审核状态格式统一"] = self.unify_status_format_var.get()

        
        self.save_config()
        self.log_message("设置已保存")
        window.destroy()
    
    def check_files_accessibility(self):
        """检查文件是否可访问（未被打开）"""
        try:
            files_to_check = []
            
            # 检查总表
            total_table_path = self.total_table_path_var.get()
            if total_table_path and os.path.exists(total_table_path):
                files_to_check.append(("总表", total_table_path))
            
            # 检查专辑注入表
            album_path = self.album_path_var.get()
            if album_path and os.path.exists(album_path):
                files_to_check.append(("专辑注入表", album_path))
            
            # 检查剧集注入表
            episode_path = self.episode_path_var.get()
            if episode_path and os.path.exists(episode_path):
                files_to_check.append(("剧集注入表", episode_path))
            
            # 检查审核表
            review_path = self.review_path_var.get()
            if review_path and os.path.exists(review_path):
                files_to_check.append(("审核表", review_path))
            
            # 尝试打开每个文件进行写入测试
            opened_files = []
            for name, path in files_to_check:
                try:
                    # 尝试以写入模式打开文件
                    if path.endswith('.xlsx'):
                        import openpyxl
                        wb = openpyxl.load_workbook(path)
                        wb.close()
                    elif path.endswith('.csv'):
                        # CSV文件通常不会被Excel锁定
                        pass
                    else:
                        # 其他格式文件
                        with open(path, 'r+b') as f:
                            pass
                except PermissionError:
                    opened_files.append((name, path))
                except Exception as e:
                    # 其他错误也认为是文件被占用
                    opened_files.append((name, path))
            
            return opened_files
            
        except Exception as e:
            self.log_message(f"检查文件访问性失败: {e}")
            return []
    
    def close_opened_files(self, opened_files):
        """关闭已打开的文件"""
        try:
            import subprocess
            import psutil
            
            for name, path in opened_files:
                try:
                    # 查找占用该文件的进程
                    for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                        try:
                            if proc.info['open_files']:
                                for file_info in proc.info['open_files']:
                                    if file_info.path == path:
                                        # 尝试关闭进程
                                        proc.terminate()
                                        proc.wait(timeout=5)
                                        self.log_message(f"已关闭占用{name}的进程: {proc.info['name']}")
                                        break
                        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.TimeoutExpired):
                            continue
                except Exception as e:
                    self.log_message(f"关闭{name}失败: {e}")
            
            # 等待一下让文件释放
            import time
            time.sleep(1)
            
        except Exception as e:
            self.log_message(f"关闭文件失败: {e}")
    
    def show_file_access_dialog(self, opened_files):
        """显示文件访问对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("文件被占用")
        dialog.geometry("500x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"500x300+{x}+{y}")
        
        # 创建内容
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="检测到以下文件正在被使用：", font=("Arial", 12, "bold"))
        title_label.pack(pady=(0, 10))
        
        # 文件列表
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        for name, path in opened_files:
            file_label = ttk.Label(list_frame, text=f"• {name}: {os.path.basename(path)}", 
                                 wraplength=450, justify=tk.LEFT)
            file_label.pack(anchor=tk.W, pady=2)
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        def close_files_and_continue():
            self.close_opened_files(opened_files)
            dialog.destroy()
            # 重新检查文件访问性
            remaining_files = self.check_files_accessibility()
            if remaining_files:
                self.show_file_access_dialog(remaining_files)
            else:
                # 继续执行操作
                self.continue_execution()
        
        def cancel_operation():
            dialog.destroy()
            self.status_label.configure(text="操作已取消")
        
        # 关闭文件按钮
        close_button = ttk.Button(button_frame, text="关闭表格", 
                                command=close_files_and_continue, style="Accent.TButton")
        close_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 取消按钮
        cancel_button = ttk.Button(button_frame, text="取消", command=cancel_operation)
        cancel_button.pack(side=tk.RIGHT)
        
        # 设置焦点
        close_button.focus_set()
    
    def continue_execution(self):
        """继续执行操作（在文件检查通过后调用）"""
        try:
            # 加载总表
            if not self.load_total_table():
                self.status_label.configure(text="准备就绪")
                return
            
            # 执行勾选的操作
            if self.import_album_var.get():
                self.import_album_data()
            
            if self.import_episode_var.get():
                self.import_episode_data()
            
            if self.import_album_review_var.get():
                self.import_album_review_data()
            
            if self.import_episode_review_var.get():
                self.import_episode_review_data()
            
            if self.supplement_status_var.get():
                self.supplement_album_status()
            
            if self.generate_episode_submit_var.get():
                self.generate_episode_submit_table()
            
            if self.generate_album_submit_var.get():
                self.generate_album_submit_table()
            
            if self.generate_online_var.get():
                self.generate_online_table()
            
            # 保存总表
            self.save_total_table()
            
            self.log_message("勾选的操作执行完成！")
            self.status_label.configure(text="执行完成")
            
        except Exception as e:
            self.log_message(f"执行过程中出现错误: {e}")
            logger.error(f"执行错误: {e}", exc_info=True)
            self.status_label.configure(text="执行失败")
    
    def execute_selected(self):
        """执行勾选的操作"""
        self.status_label.configure(text="正在执行...")
        self.log_message("开始执行勾选的操作...")
        
        try:
            # 初始校验：检查文件是否被占用
            opened_files = self.check_files_accessibility()
            if opened_files:
                self.show_file_access_dialog(opened_files)
                return
            
            # 备份总表（如果启用）
            if self.config.get("是否备份总表", False):
                total_table_path = self.total_table_path_var.get()
                if total_table_path and os.path.exists(total_table_path):
                    self.backup_total_table(total_table_path)
            
            # 继续执行操作
            self.continue_execution()
            
        except Exception as e:
            self.log_message(f"执行过程中出现错误: {e}")
            logger.error(f"执行错误: {e}", exc_info=True)
            self.status_label.configure(text="执行失败")
    
    def execute_all(self):
        """执行所有操作"""
        self.status_label.configure(text="正在执行...")
        self.log_message("开始执行所有操作...")
        
        try:
            # 初始校验：检查文件是否被占用
            opened_files = self.check_files_accessibility()
            if opened_files:
                self.show_file_access_dialog(opened_files)
                return
            
            # 备份总表（如果启用）
            if self.config.get("是否备份总表", False):
                total_table_path = self.total_table_path_var.get()
                if total_table_path and os.path.exists(total_table_path):
                    self.backup_total_table(total_table_path)
            
            # 加载总表
            if not self.load_total_table():
                self.status_label.configure(text="准备就绪")
                return
            
            # 执行所有操作
            self.import_album_data()
            self.import_episode_data()
            self.import_album_review_data()
            self.import_episode_review_data()
            self.supplement_album_status()
            self.generate_episode_submit_table()
            self.generate_album_submit_table()
            self.generate_online_table()
            
            # 保存总表
            self.save_total_table()
            
            self.log_message("所有操作执行完成！")
            self.status_label.configure(text="执行完成")
            
        except Exception as e:
            self.log_message(f"执行过程中出现错误: {e}")
            logger.error(f"执行错误: {e}", exc_info=True)
            self.status_label.configure(text="执行失败")
    
    def load_total_table(self):
        """加载总表"""
        try:
            total_table_path = self.total_table_path_var.get()
            if not total_table_path or not os.path.exists(total_table_path):
                self.log_message("总表文件不存在，请检查路径")
                return False
            
            self.log_message(f"正在加载总表: {total_table_path}")
            
            # 读取总表的两个sheet
            self.total_album_df = pd.read_excel(total_table_path, sheet_name="专辑")
            self.total_episode_df = pd.read_excel(total_table_path, sheet_name="剧集")
            
            self.log_message("总表加载成功")
            return True
            
        except Exception as e:
            self.log_message(f"加载总表失败: {e}")
            return False
    
    def save_total_table(self):
        """保存总表"""
        try:
            total_table_path = self.total_table_path_var.get()
            
            # 检查是否需要备份总表
            if self.config.get("是否备份总表", False):
                self.backup_total_table(total_table_path)
            
            with pd.ExcelWriter(total_table_path, engine='openpyxl') as writer:
                # 保存专辑数据，设置专辑ID列为文本格式
                self.total_album_df.to_excel(writer, sheet_name="专辑", index=False)
                worksheet = writer.sheets["专辑"]
                # 设置所有ID列为文本格式，避免科学计数法
                for col_num, col_name in enumerate(self.total_album_df.columns, 1):
                    if 'ID' in col_name or 'id' in col_name:
                        for row_num in range(2, len(self.total_album_df) + 2):  # Excel行号从1开始，数据从第2行开始
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = '@'  # 文本格式
                            # 强制将数值转换为字符串，确保显示为完整数字
                            if not pd.isna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.value = str(int(cell.value))
                
                # 设置专辑表审核状态颜色
                self.set_review_status_colors(worksheet, self.total_album_df)
                
                # 调整专辑表列宽
                self.adjust_column_width(worksheet, self.total_album_df)
                
                # 保存剧集数据，设置剧集ID列为文本格式
                self.total_episode_df.to_excel(writer, sheet_name="剧集", index=False)
                worksheet = writer.sheets["剧集"]
                # 设置所有ID列为文本格式，避免科学计数法
                for col_num, col_name in enumerate(self.total_episode_df.columns, 1):
                    if 'ID' in col_name or 'id' in col_name:
                        for row_num in range(2, len(self.total_episode_df) + 2):  # Excel行号从1开始，数据从第2行开始
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = '@'  # 文本格式
                            # 强制将数值转换为字符串，确保显示为完整数字
                            if not pd.isna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.value = str(int(cell.value))
                
                # 设置剧集表审核状态颜色
                self.set_review_status_colors(worksheet, self.total_episode_df)
                
                # 调整剧集表列宽
                self.adjust_column_width(worksheet, self.total_episode_df)
            
            self.log_message("总表保存成功")
            
            # 自动打开总表
            self.open_file(total_table_path)
            
        except Exception as e:
            self.log_message(f"保存总表失败: {e}")
    
    def set_review_status_colors(self, worksheet, df):
        """设置审核状态列的字体颜色"""
        try:
            # 查找审核状态列
            review_status_columns = []
            for col_num, col_name in enumerate(df.columns, 1):
                if '审核状态' in col_name:
                    review_status_columns.append((col_num, col_name))
            
            if not review_status_columns:
                return
            
            # 为每个审核状态列设置颜色和字体样式
            for col_num, col_name in review_status_columns:
                for row_num in range(len(df)):
                    cell_value = df.iloc[row_num][col_name]
                    if not pd.isna(cell_value):
                        # 获取对应的字体样式（颜色和加粗）
                        font_style = self.get_review_status_font(cell_value)
                        if font_style:
                            # 设置单元格字体样式
                            cell = worksheet.cell(row=row_num + 2, column=col_num)  # Excel行号从1开始，数据从第2行开始
                            cell.font = font_style
            
            self.log_message(f"审核状态颜色设置完成，处理了 {len(review_status_columns)} 个审核状态列")
            
        except Exception as e:
            self.log_message(f"设置审核状态颜色失败: {e}")
    
    def adjust_column_width(self, worksheet, df):
        """调整Excel表格列宽，使其适应表头和数据内容"""
        try:
            for col_num, col_name in enumerate(df.columns, 1):
                # 计算列名长度（中文字符按2个字符宽度计算）
                header_length = self.calculate_string_width(str(col_name))
                
                # 计算该列所有数据的最大长度
                max_data_length = 0
                for row_num in range(len(df)):
                    cell_value = df.iloc[row_num][col_name]
                    if not pd.isna(cell_value):
                        cell_length = self.calculate_string_width(str(cell_value))
                        max_data_length = max(max_data_length, cell_length)
                
                # 取表头长度和数据最大长度的较大值
                max_length = max(header_length, max_data_length)
                
                # 限制最大列宽为表头字数的两倍
                max_allowed_width = header_length * 2
                max_length = min(max_length, max_allowed_width)
                
                # 设置列宽（增加适当空间，确保内容不被截断）
                # Excel列宽单位约为字符宽度的0.7倍，所以需要适当调整
                adjusted_width = max_length * 1.2 + 3
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = adjusted_width
            
            self.log_message("列宽调整完成")
            
        except Exception as e:
            self.log_message(f"列宽调整失败: {e}")
    
    def calculate_string_width(self, text):
        """计算字符串的显示宽度（中文字符按2个字符宽度计算）"""
        width = 0
        for char in text:
            # 中文字符、全角字符按2个字符宽度计算
            if '\u4e00' <= char <= '\u9fff' or '\uff01' <= char <= '\uff5e':
                width += 2
            else:
                width += 1
        return width
    
    def clean_review_status(self, status_value):
        """清洗审核状态，确保只有五个标准选项"""
        if pd.isna(status_value) or status_value == '':
            return ''
        
        status_str = str(status_value).strip()
        
        # 使用正则表达式匹配标准审核状态
        # 通过、驳回、审核中、待重审、待提审
        if re.search(r'通过', status_str):
            return '通过'
        elif re.search(r'驳回', status_str):
            return '驳回'
        elif re.search(r'审核中', status_str):
            return '审核中'
        elif re.search(r'待重审', status_str):
            return '待重审'
        elif re.search(r'待提审', status_str):
            return '待提审'
        else:
            # 如果都不匹配，根据设置决定是否统一格式
            if self.config.get("审核状态格式统一", False):
                # 尝试根据相似性匹配到最接近的标准状态
                if re.search(r'通过|成功|完成', status_str):
                    return '通过'
                elif re.search(r'驳回|拒绝|不通过|失败', status_str):
                    return '驳回'
                elif re.search(r'审核|审查|处理中', status_str):
                    return '审核中'
                elif re.search(r'重审|重新|再次', status_str):
                    return '待重审'
                elif re.search(r'提审|提交|待审', status_str):
                    return '待提审'
                else:
                    # 如果无法匹配，默认为"待提审"
                    self.log_message(f"未知的审核状态: {status_str}，已统一为'待提审'")
                    return '待提审'
            else:
                # 如果不统一格式，保留原始值
                return status_str
    
    def get_review_status_color(self, status_value):
        """获取审核状态对应的字体颜色"""
        if pd.isna(status_value) or status_value == '':
            return None
        
        status_str = str(status_value).strip()
        
        # 审核状态颜色映射（按照图片中的颜色方案）
        color_mapping = {
            '通过': '008000',    # 绿色
            '驳回': 'FF0000',    # 红色
            '审核中': '000000',  # 默认黑色
            '待提审': 'FFD700',  # 黄色
            '待重审': 'FF8C00'   # 橘色
        }
        
        return color_mapping.get(status_str, None)
    
    def get_review_status_font(self, status_value):
        """获取审核状态对应的字体样式（颜色和加粗）"""
        if pd.isna(status_value) or status_value == '':
            return None
        
        status_str = str(status_value).strip()
        
        # 审核状态颜色映射（按照图片中的颜色方案）
        color_mapping = {
            '通过': '008000',    # 绿色
            '驳回': 'FF0000',    # 红色
            '审核中': '000000',  # 默认黑色
            '待提审': 'FFD700',  # 黄色
            '待重审': 'FF8C00'   # 橘色
        }
        
        color_hex = color_mapping.get(status_str, None)
        if color_hex:
            # 驳回状态加粗，其他状态正常
            bold = (status_str == '驳回')
            return Font(color=color_hex, bold=bold)
        
        return None
    
    def open_file(self, file_path):
        """自动打开文件"""
        try:
            if os.path.exists(file_path):
                if sys.platform == "win32":
                    os.startfile(file_path)
                elif sys.platform == "darwin":  # macOS
                    subprocess.run(["open", file_path])
                else:  # Linux
                    subprocess.run(["xdg-open", file_path])
                self.log_message(f"已自动打开文件: {file_path}")
        except Exception as e:
            self.log_message(f"自动打开文件失败: {e}")
    
    def import_album_data(self):
        """导入专辑数据"""
        try:
            album_path = self.album_path_var.get()
            if not album_path or not os.path.exists(album_path):
                self.log_message("专辑注入文件不存在，跳过导入")
                return
            
            self.log_message("正在导入专辑数据...")
            
            # 读取专辑注入数据
            if album_path.endswith('.csv'):
                album_df = pd.read_csv(album_path)
            else:
                album_df = pd.read_excel(album_path)
            
            # 数据映射 - 根据需求调整字段映射
            mapping = {
                '专辑ID': '专辑ID',
                '频道属性': '频道属性',
                '批次号': '批次号',
                '集数': '集数',
                '注入行为': '注入行为',
                '注入状态': '注入状态',
                '发送时间': '发送时间',
                '回执时间': '回执时间',
                '创建时间': '创建时间',
                '更新时间': '更新时间',
                '专辑名称': '专辑名称',
                '二级类型': '二级分类'
            }
            
            # 处理数据
            for _, row in album_df.iterrows():
                album_id = row.get('专辑ID')
                if pd.isna(album_id):
                    continue
                
                # 查找总表中是否存在该专辑ID
                mask = self.total_album_df['专辑ID'] == album_id
                
                if mask.any():
                    # 专辑ID存在，检查审核状态
                    current_status = self.total_album_df.loc[mask, '审核状态（专辑）'].iloc[0]
                    
                    if current_status == "驳回":
                        # 驳回状态改为待重审
                        self.total_album_df.loc[mask, '审核状态（专辑）'] = "待重审"
                    
                    # 更新数据
                    for source_col, target_col in mapping.items():
                        if source_col in row and not pd.isna(row[source_col]):
                            # 处理时间格式
                            if '时间' in source_col and not pd.isna(row[source_col]):
                                try:
                                    time_value = pd.to_datetime(row[source_col], dayfirst=True)
                                    formatted_time = time_value.strftime('%Y-%m-%d %H:%M:%S')
                                    self.total_album_df.loc[mask, target_col] = formatted_time
                                except:
                                    self.total_album_df.loc[mask, target_col] = row[source_col]
                            else:
                                self.total_album_df.loc[mask, target_col] = row[source_col]
                else:
                    # 专辑ID不存在，添加新行
                    new_row = {}
                    for source_col, target_col in mapping.items():
                        if source_col in row and not pd.isna(row[source_col]):
                            if '时间' in source_col:
                                try:
                                    time_value = pd.to_datetime(row[source_col], dayfirst=True)
                                    formatted_time = time_value.strftime('%Y-%m-%d %H:%M:%S')
                                    new_row[target_col] = formatted_time
                                except:
                                    new_row[target_col] = row[source_col]
                            else:
                                new_row[target_col] = row[source_col]
                    
                    new_row['审核状态（专辑）'] = "待提审"
                    self.total_album_df = pd.concat([self.total_album_df, pd.DataFrame([new_row])], ignore_index=True)
            
            self.log_message("专辑数据导入完成")
            
        except Exception as e:
            self.log_message(f"导入专辑数据失败: {e}")
    
    def import_episode_data(self):
        """导入剧集数据"""
        try:
            episode_path = self.episode_path_var.get()
            if not episode_path or not os.path.exists(episode_path):
                self.log_message("剧集注入文件不存在，跳过导入")
                return
            
            self.log_message("正在导入剧集数据...")
            
            # 读取剧集注入数据
            if episode_path.endswith('.csv'):
                episode_df = pd.read_csv(episode_path)
            else:
                episode_df = pd.read_excel(episode_path)
            
            # 数据映射 - 根据需求调整字段映射
            mapping = {
                '专辑ID': '专辑ID',
                '剧集ID': '剧集ID',
                '频道属性': '频道属性',
                '批次号': '批次号',
                '专辑名称': '专辑名称',
                '剧集名称': '剧集名称',
                '注入行为': '注入行为',
                '注入状态': '注入状态',
                '发送时间': '发送时间',
                '回执时间': '回执时间',
                '创建时间': '创建时间',
                '更新时间': '更新时间',
                '二级类型': '二级分类'
            }
            
            # 处理数据
            for _, row in episode_df.iterrows():
                episode_id = row.get('剧集ID')
                if pd.isna(episode_id):
                    continue
                
                # 查找总表中是否存在该剧集ID
                mask = self.total_episode_df['剧集ID'] == episode_id
                
                if mask.any():
                    # 剧集ID存在，检查审核状态
                    current_status = self.total_episode_df.loc[mask, '审核状态（剧集）'].iloc[0]
                    
                    if current_status == "驳回":
                        # 驳回状态改为待重审
                        self.total_episode_df.loc[mask, '审核状态（剧集）'] = "待重审"
                    
                    # 更新数据
                    for source_col, target_col in mapping.items():
                        if source_col in row and not pd.isna(row[source_col]):
                            # 处理时间格式
                            if '时间' in source_col and not pd.isna(row[source_col]):
                                try:
                                    time_value = pd.to_datetime(row[source_col], dayfirst=True)
                                    formatted_time = time_value.strftime('%Y-%m-%d %H:%M:%S')
                                    self.total_episode_df.loc[mask, target_col] = formatted_time
                                except:
                                    self.total_episode_df.loc[mask, target_col] = row[source_col]
                            else:
                                self.total_episode_df.loc[mask, target_col] = row[source_col]
                else:
                    # 剧集ID不存在，添加新行
                    new_row = {}
                    for source_col, target_col in mapping.items():
                        if source_col in row and not pd.isna(row[source_col]):
                            if '时间' in source_col:
                                try:
                                    time_value = pd.to_datetime(row[source_col], dayfirst=True)
                                    formatted_time = time_value.strftime('%Y-%m-%d %H:%M:%S')
                                    new_row[target_col] = formatted_time
                                except:
                                    new_row[target_col] = row[source_col]
                            else:
                                new_row[target_col] = row[source_col]
                    
                    new_row['审核状态（剧集）'] = "待提审"
                    self.total_episode_df = pd.concat([self.total_episode_df, pd.DataFrame([new_row])], ignore_index=True)
            
            self.log_message("剧集数据导入完成")
            
        except Exception as e:
            self.log_message(f"导入剧集数据失败: {e}")
    
    def get_latest_review_file(self):
        """获取审核表文件"""
        try:
            # 首先尝试从UI路径获取
            review_path = self.review_path_var.get()
            if review_path and os.path.exists(review_path):
                self.log_message(f"使用UI设置的审核表路径: {review_path}")
                return review_path
            
            # 如果UI路径不存在，从默认文件夹获取最新文件
            review_folder = r"C:\Users\fucha\Desktop\北京项目\3、审核结果"
            if os.path.exists(review_folder):
                import glob
                pattern_path = os.path.join(review_folder, "*.xlsx")
                files = glob.glob(pattern_path)
                
                # 排除临时文件（以~$开头的文件）
                files = [f for f in files if not os.path.basename(f).startswith('~$')]
                
                if files:
                    latest_file = max(files, key=os.path.getmtime)
                    self.log_message(f"从文件夹获取最新审核表: {os.path.basename(latest_file)}")
                    return latest_file
            
            self.log_message("未找到审核表文件")
            return None
            
        except Exception as e:
            self.log_message(f"获取审核表文件失败: {e}")
            return None
    
    def import_album_review_data(self):
        """导入专辑审核数据"""
        try:
            review_file = self.get_latest_review_file()
            if not review_file:
                self.log_message("未找到审核表文件，跳过导入")
                return
            
            self.log_message(f"正在导入专辑审核数据: {review_file}")
            
            # 读取审核表
            try:
                review_df = pd.read_excel(review_file, sheet_name="节目集审核结果")
            except:
                self.log_message("未找到'节目集审核结果'sheet，跳过导入")
                return
            
            self.log_message(f"节目集审核结果包含 {len(review_df)} 行数据")
            self.log_message(f"节目集审核结果列名: {list(review_df.columns)}")
            
            # 数据映射 - 根据实际字段名调整
            mapping = {
                '节目集ID': '专辑ID（平台）',
                '节目集名称': '专辑名称',
                '审核状态': '审核状态（专辑）',
                '审核结果': '审核结果（专辑）',
                '不通过详情': '不通过详情（专辑）',
                '审核时间': '审核时间',
                '国家': '国家',
                '推送时间': '推送时间'
            }
            
            # 处理数据
            updated_count = 0
            for _, row in review_df.iterrows():
                album_name = row.get('节目集名称')
                if pd.isna(album_name):
                    continue
                
                # 根据专辑名称查找总表
                mask = self.total_album_df['专辑名称'] == album_name
                
                if mask.any():
                    # 更新数据
                    for source_col, target_col in mapping.items():
                        if source_col in row and not pd.isna(row[source_col]):
                            # 对审核状态字段进行清洗
                            if '审核状态' in target_col:
                                cleaned_value = self.clean_review_status(row[source_col])
                                self.total_album_df.loc[mask, target_col] = cleaned_value
                            # 对ID字段进行格式处理，避免科学计数法
                            elif 'ID' in target_col or 'id' in target_col:
                                id_value = row[source_col]
                                if isinstance(id_value, (int, float)):
                                    # 确保ID显示为完整数字，不转换为科学计数法
                                    self.total_album_df.loc[mask, target_col] = str(int(id_value))
                                else:
                                    self.total_album_df.loc[mask, target_col] = str(id_value)
                            else:
                                self.total_album_df.loc[mask, target_col] = row[source_col]
                    updated_count += 1
                else:
                    self.log_message(f"未找到匹配的专辑名称: {album_name}")
            
            self.log_message(f"专辑审核数据导入完成，更新了 {updated_count} 条记录")
            
        except Exception as e:
            self.log_message(f"导入专辑审核数据失败: {e}")
            import traceback
            self.log_message(f"错误详情: {traceback.format_exc()}")
    
    def import_episode_review_data(self):
        """导入剧集审核数据"""
        try:
            review_file = self.get_latest_review_file()
            if not review_file:
                self.log_message("未找到审核表文件，跳过导入")
                return
            
            self.log_message(f"正在导入剧集审核数据: {review_file}")
            
            # 读取审核表
            try:
                review_df = pd.read_excel(review_file, sheet_name="节目审核结果")
            except:
                self.log_message("未找到'节目审核结果'sheet，跳过导入")
                return
            
            self.log_message(f"节目审核结果包含 {len(review_df)} 行数据")
            self.log_message(f"节目审核结果列名: {list(review_df.columns)}")
            
            # 数据映射 - 根据实际字段名调整
            mapping = {
                '节目名称': '剧集名称',
                '节目集名称': '专辑名称',
                '节目ID': '剧集ID(平台)',  # 修正映射字段
                '审核状态': '审核状态（剧集）',
                '不通过详情': '不通过详情（剧集）',
                '单集时长': '单集时长'  # 新增单集时长字段映射
                # 注意：审核表中没有'单集时长（平台）'字段
            }
            
            # 处理数据
            updated_count = 0
            for _, row in review_df.iterrows():
                episode_name = row.get('节目名称')
                if pd.isna(episode_name):
                    continue
                
                # 根据剧集名称查找总表
                mask = self.total_episode_df['剧集名称'] == episode_name
                
                if mask.any():
                    # 更新数据
                    for source_col, target_col in mapping.items():
                        if source_col in row and not pd.isna(row[source_col]):
                            # 对审核状态字段进行清洗
                            if '审核状态' in target_col:
                                cleaned_value = self.clean_review_status(row[source_col])
                                self.total_episode_df.loc[mask, target_col] = cleaned_value
                            # 对单集时长字段进行单位换算（秒转小时，保留两位小数）
                            elif source_col == '单集时长':
                                try:
                                    duration_seconds = float(row[source_col])
                                    duration_hours = duration_seconds / 3600  # 秒转小时
                                    formatted_duration = round(duration_hours, 2)  # 保留两位小数
                                    self.total_episode_df.loc[mask, target_col] = formatted_duration
                                    self.log_message(f"单集时长转换: {duration_seconds}秒 → {formatted_duration}小时")
                                except (ValueError, TypeError) as e:
                                    self.log_message(f"单集时长转换失败: {row[source_col]}, 错误: {e}")
                                    # 转换失败时保留原值
                                    self.total_episode_df.loc[mask, target_col] = row[source_col]
                            # 对ID字段进行格式处理，避免科学计数法
                            elif 'ID' in target_col or 'id' in target_col:
                                id_value = row[source_col]
                                if isinstance(id_value, (int, float)):
                                    # 确保ID显示为完整数字，不转换为科学计数法
                                    self.total_episode_df.loc[mask, target_col] = str(int(id_value))
                                else:
                                    self.total_episode_df.loc[mask, target_col] = str(id_value)
                            else:
                                self.total_episode_df.loc[mask, target_col] = row[source_col]
                    updated_count += 1
                else:
                    self.log_message(f"未找到匹配的剧集名称: {episode_name}")
            
            self.log_message(f"剧集审核数据导入完成，更新了 {updated_count} 条记录")
            
        except Exception as e:
            self.log_message(f"导入剧集审核数据失败: {e}")
            import traceback
            self.log_message(f"错误详情: {traceback.format_exc()}")
    
    def supplement_album_status(self):
        """补充专辑审核状态到剧集表"""
        try:
            self.log_message("正在补充专辑审核状态...")
            
            updated_count = 0
            # 根据专辑ID将专辑表的数据补充到剧集表
            for _, album_row in self.total_album_df.iterrows():
                album_id = album_row.get('专辑ID')
                album_status = album_row.get('审核状态（专辑）')
                album_id_platform = album_row.get('专辑ID（平台）')
                
                if pd.isna(album_id):
                    continue
                
                # 在剧集表中查找对应的专辑ID
                mask = self.total_episode_df['专辑ID'] == album_id
                if mask.any():
                    # 补充审核状态（专辑）
                    if not pd.isna(album_status):
                        self.total_episode_df.loc[mask, '审核状态（专辑）'] = album_status
                    
                    # 补充专辑ID（平台）
                    if not pd.isna(album_id_platform):
                        # 确保ID格式正确，避免科学计数法
                        if isinstance(album_id_platform, (int, float)):
                            formatted_id = str(int(album_id_platform))
                        else:
                            formatted_id = str(album_id_platform)
                        self.total_episode_df.loc[mask, '专辑ID（平台）'] = formatted_id
                    
                    updated_count += 1
            
            self.log_message(f"专辑审核状态补充完成，更新了 {updated_count} 条记录")
            
        except Exception as e:
            self.log_message(f"补充专辑审核状态失败: {e}")
            import traceback
            self.log_message(f"错误详情: {traceback.format_exc()}")
    
    def generate_episode_submit_table(self):
        """生成剧集送审表"""
        try:
            self.log_message("正在生成剧集送审表...")
            
            # 筛选送审时间为"T"的剧集
            submit_mask = self.total_episode_df['送审时间'] == 'T'
            submit_data = self.total_episode_df[submit_mask].copy()
            
            if submit_data.empty:
                self.log_message("没有需要送审的剧集数据")
                return
            
            # 创建送审表
            submit_table = []
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            for _, row in submit_data.iterrows():
                # 确定节目集状态和节目状态（严格按需求）
                episode_status = row.get('审核状态（剧集）', '')
                album_status = row.get('审核状态（专辑）', '')
                
                # 根据需求逻辑确定状态（包含待提审）
                if episode_status in ['待审核', '待重审', '待提审'] and album_status in ['待审核', '待重审', '待提审']:
                    album_table_status = "送审"
                    episode_table_status = "送审"
                elif album_status == "通过" and episode_status in ['待审核', '待重审', '待提审']:
                    album_table_status = "通过"
                    episode_table_status = "送审"
                else:
                    album_table_status = ""
                    episode_table_status = ""
                submit_row = {
                    '节目集id': row.get('专辑ID', ''),
                    '节目集名称': row.get('专辑名称', ''),
                    '节目集状态': album_table_status,
                    '节目ID': row.get('剧集ID', ''),  # 修正：使用剧集ID而不是专辑ID（平台）
                    '节目名称': row.get('剧集名称', ''),
                    '节目状态': episode_table_status,
                    '一级分类': row.get('频道属性', ''),
                    '二级分类': row.get('二级分类', ''),
                    'CPCODE': 'BJ_JZYS',
                    '备注': ''
                }
                submit_table.append(submit_row)
            
            # 创建DataFrame
            submit_df = pd.DataFrame(submit_table)
            
            # 保存送审表
            output_path = self.submit_path_var.get()
            if not output_path:
                output_path = os.path.expanduser("~/Desktop")
            os.makedirs(output_path, exist_ok=True)
            current_date_str = datetime.now().strftime('%Y%m%d')
            filename = f"【送审】极智云视（北京）(BJ_JZYS）--{current_date_str}.xlsx"
            filepath = os.path.join(output_path, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                submit_df.to_excel(writer, index=False)
                worksheet = writer.sheets["Sheet1"]
                # 设置所有ID列为文本格式，避免科学计数法
                for col_num, col_name in enumerate(submit_df.columns, 1):
                    if 'ID' in col_name or 'id' in col_name:
                        for row_num in range(2, len(submit_df) + 2):  # Excel行号从1开始，数据从第2行开始
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = '@'  # 文本格式
                            if not pd.isna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.value = str(int(cell.value))
                # 调整送审表列宽
                self.adjust_column_width(worksheet, submit_df)
                # 添加全部边框
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.border = border
            
            # 更新总表中的送审时间和审核状态
            self.total_episode_df.loc[submit_mask, '送审时间'] = current_date
            # 将剧集审核状态更新为"审核中"
            self.total_episode_df.loc[submit_mask, '审核状态（剧集）'] = "审核中"
            
            # 同时更新对应的专辑审核状态为"审核中"
            for _, row in submit_data.iterrows():
                album_id = row.get('专辑ID')
                if not pd.isna(album_id):
                    album_mask = self.total_album_df['专辑ID'] == album_id
                    if album_mask.any():
                        self.total_album_df.loc[album_mask, '审核状态（专辑）'] = "审核中"
            
            self.log_message(f"剧集送审表生成完成: {filepath}")
            # 自动打开生成的送审表
            self.open_file(filepath)
        except Exception as e:
            self.log_message(f"生成剧集送审表失败: {e}")
    
    def generate_album_submit_table(self):
        """生成专辑送审表"""
        try:
            self.log_message("正在生成专辑送审表...")
            
            # 筛选送审时间为"T"的专辑
            submit_mask = self.total_album_df['送审时间'] == 'T'
            submit_data = self.total_album_df[submit_mask].copy()
            
            if submit_data.empty:
                self.log_message("没有需要送审的专辑数据")
                return
            
            # 创建送审表
            submit_table = []
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            for _, row in submit_data.iterrows():
                submit_row = {
                    '节目集id': row.get('专辑ID', ''),
                    '节目集名称': row.get('专辑名称', ''),
                    '节目集状态': '送审',
                    '节目ID': '',
                    '节目名称': '',
                    '节目状态': '',  # 专辑送审表中节目状态不填充
                    '一级分类': row.get('频道属性', ''),
                    '二级分类': row.get('二级分类', ''),
                    'CPCODE': 'BJ_JZYS',
                    '备注': ''
                }
                submit_table.append(submit_row)
            
            # 创建DataFrame
            submit_df = pd.DataFrame(submit_table)
            
            # 保存送审表
            output_path = self.submit_path_var.get()
            if not output_path:
                output_path = os.path.expanduser("~/Desktop")
            
            os.makedirs(output_path, exist_ok=True)
            
            current_date_str = datetime.now().strftime('%Y%m%d')
            filename = f"【送审】极智云视（北京）(BJ_JZYS）--{current_date_str}.xlsx"
            filepath = os.path.join(output_path, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                submit_df.to_excel(writer, index=False)
                worksheet = writer.sheets["Sheet1"]
                
                # 设置所有ID列为文本格式，避免科学计数法
                for col_num, col_name in enumerate(submit_df.columns, 1):
                    if 'ID' in col_name or 'id' in col_name:
                        for row_num in range(2, len(submit_df) + 2):  # Excel行号从1开始，数据从第2行开始
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = '@'  # 文本格式
                            # 强制将数值转换为字符串，确保显示为完整数字
                            if not pd.isna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.value = str(int(cell.value))
                
                # 调整送审表列宽
                self.adjust_column_width(worksheet, submit_df)
                
                # 为所有单元格添加边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 获取最大行和最大列
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # 为所有单元格添加边框
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
            
            # 更新总表中的送审时间和审核状态
            self.total_album_df.loc[submit_mask, '送审时间'] = current_date
            # 将专辑审核状态更新为"审核中"
            self.total_album_df.loc[submit_mask, '审核状态（专辑）'] = "审核中"
            
            self.log_message(f"专辑送审表生成完成: {filepath}")
            
            # 自动打开生成的送审表
            self.open_file(filepath)
            
        except Exception as e:
            self.log_message(f"生成专辑送审表失败: {e}")
    
    def generate_online_table(self):
        """生成上线表"""
        try:
            self.log_message("正在生成上线表...")
            
            # 筛选申请上线时间为"T"的剧集
            online_mask = self.total_episode_df['申请上线时间'] == 'T'
            online_data = self.total_episode_df[online_mask].copy()
            
            if online_data.empty:
                self.log_message("没有需要申请上线的数据")
                return
            
            # 创建上线表
            online_table = []
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            for _, row in online_data.iterrows():
                online_row = {
                    'CID': row.get('专辑ID', ''),
                    '应用名称': 'CP-极智云视（北京）',
                    '节目集id': row.get('专辑ID（平台）', ''),
                    'cpCode': 'BJ_JZYS',
                    '节目集名称': row.get('专辑名称', ''),
                    '一级分类': row.get('频道属性', ''),
                    '节目ID': row.get('剧集ID(平台)', ''),
                    '节目名称': row.get('剧集名称', ''),
                    '备注': ''
                }
                online_table.append(online_row)
            
            # 创建DataFrame
            online_df = pd.DataFrame(online_table)
            
            # 保存上线表
            output_path = self.online_path_var.get()
            if not output_path:
                output_path = os.path.expanduser("~/Desktop")
            
            os.makedirs(output_path, exist_ok=True)
            
            current_date_str = datetime.now().strftime('%Y%m%d')
            filename = f"申请上线定价--极智云视（北京）（BJ_JZYS）--{current_date_str}.xlsx"
            filepath = os.path.join(output_path, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                online_df.to_excel(writer, index=False)
                worksheet = writer.sheets["Sheet1"]
                # 设置所有ID列为文本格式，避免科学计数法
                for col_num, col_name in enumerate(online_df.columns, 1):
                    if 'ID' in col_name or 'id' in col_name:
                        for row_num in range(2, len(online_df) + 2):  # Excel行号从1开始，数据从第2行开始
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = '@'  # 文本格式
                            # 强制将数值转换为字符串，确保显示为完整数字
                            if not pd.isna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.value = str(int(cell.value))
                
                # 调整上线表列宽
                self.adjust_column_width(worksheet, online_df)
                
                # 为所有单元格添加边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 获取最大行和最大列
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # 为所有单元格添加边框
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
            
            # 更新总表中的申请上线时间
            self.total_episode_df.loc[online_mask, '申请上线时间'] = current_date
            
            self.log_message(f"上线表生成完成: {filepath}")
            
            # 自动打开生成的上线表
            self.open_file(filepath)
            
        except Exception as e:
            self.log_message(f"生成上线表失败: {e}")
    
    def backup_total_table(self, total_table_path):
        """备份总表"""
        try:
            if not os.path.exists(total_table_path):
                return
            
            # 备份到总表同路径文件夹
            backup_dir = os.path.dirname(total_table_path)
            
            # 生成固定的备份文件名
            filename = os.path.basename(total_table_path)
            name, ext = os.path.splitext(filename)
            backup_filename = f"{name}_备份{ext}"
            backup_path = os.path.join(backup_dir, backup_filename)
            
            # 复制文件（覆盖现有备份）
            import shutil
            shutil.copy2(total_table_path, backup_path)
            
            self.log_message(f"总表已备份到: {backup_path}")
            
        except Exception as e:
            self.log_message(f"备份总表失败: {e}")
    
    def run(self):
        """运行程序"""
        self.root.mainloop()

def main():
    """主函数"""
    try:
        app = BeijingTableGenerator()
        app.run()
    except Exception as e:
        logger.error(f"程序运行错误: {e}", exc_info=True)
        print(f"程序运行错误: {e}")
        input("按回车键退出...")

if __name__ == "__main__":
    main() 