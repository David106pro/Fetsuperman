# -*- coding: utf-8 -*-
from openpyxl import load_workbook
def copy_data(main_sheet, patch_sheet, row):
    new_row = main_sheet.max_row + 1
    main_sheet.cell(row=new_row, column=1).value = new_row-1
    main_sheet.cell(row=new_row, column=2).value = patch_sheet.cell(row=row, column=1).value
    main_sheet.cell(row=new_row, column=3).value = patch_sheet.cell(row=row, column=3).value
    main_sheet.cell(row=new_row, column=5).value = "8M"
    main_sheet.cell(row=new_row, column=6).value = patch_sheet.cell(row=row, column=9).value
    main_sheet.cell(row=new_row, column=7).value = patch_sheet.cell(row=row, column=45).value
    main_sheet.cell(row=new_row, column=8).value = patch_sheet.cell(row=row, column=46).value
    main_sheet.cell(row=new_row, column=9).value = patch_sheet.cell(row=row, column=47).value
    main_sheet.cell(row=new_row, column=11).value = patch_sheet.cell(row=row, column=10).value
    main_sheet.cell(row=new_row, column=12).value = patch_sheet.cell(row=row, column=12).value
    main_sheet.cell(row=new_row, column=13).value = patch_sheet.cell(row=row, column=11).value
    main_sheet.cell(row=new_row, column=14).value = patch_sheet.cell(row=row, column=17).value
    main_sheet.cell(row=new_row, column=15).value = patch_sheet.cell(row=row, column=14).value
    main_sheet.cell(row=new_row, column=16).value = patch_sheet.cell(row=row, column=49).value
    main_sheet.cell(row=new_row, column=17).value = patch_sheet.cell(row=row, column=18).value
    main_sheet.cell(row=new_row, column=18).value = "2024/8/15"
    main_sheet.cell(row=new_row, column=20).value = "移动（基础包）"


def merge_excels(main_file, patch_file):
    # 加载主表和补丁表
    main_wb = load_workbook(main_file)
    main_sheet = main_wb.active
    patch_wb = load_workbook(patch_file)
    patch_sheet = patch_wb.active

    # 遍历补丁表的每一行
    for row in range(2, patch_sheet.max_row + 1):
        copy_data(main_sheet, patch_sheet, row)
    # 保存主表
    main_wb.save(main_file)

# 指定文件名
main_file = "C:\\Users\\fucha\\Desktop\\底量引入模板.xlsx"
patch_file = "C:\\Users\\fucha\\Desktop\\广东原表.xlsx"
merge_excels(main_file, patch_file)