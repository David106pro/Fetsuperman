import time
import traceback
import requests
import random
import re
import os
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, Playwright
from openpyxl import load_workbook

xlsx = "C:\\Users\\fucha\\Desktop\\猫眼模板.xlsx"
path_to_extension = "./my-extension"  # 定义扩展路径
user_data_dir = "/tmp/test-user-data-dir"  # 定义用户数据目录
headers = {  # 定义请求头
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36 Edg/127.0.0.0",
    "cookie": "__mta=143505930.1721729537864.1722391669891.1722398395719.56; uuid_n_v=v1; uuid=09EDF75048DC11EFAD2C9D84A9502795C74339EF0D0246B69AA625DDDBABD5AB; _lxsdk_cuid=190906c187fc8-02326d3fde828b-4c657b58-151800-190906c187fc8; WEBDFPID=1uxz77v921v05uu7yyy3xw2852014z7x809v6x7954z97958v86v0749-2037168061264-1721808060272KMUCGQQfd79fef3d01d5e9aadc18ccd4d0c95071445; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; _lxsdk=09EDF75048DC11EFAD2C9D84A9502795C74339EF0D0246B69AA625DDDBABD5AB; _csrf=e621735b58dac37e972a4bb926f6a547551dab7f73b922a119caf2bd348f5558; Hm_lvt_703e94591e87be68cc8da0da7cbd0be2=1722330825,1722332208,1722391670,1722398388; HMACCOUNT=020341BA99F5C478; __mta=143505930.1721729537864.1722332208176.1722398387968.60; Hm_lpvt_703e94591e87be68cc8da0da7cbd0be2=1722398396; _lxsdk_s=19106ed243d-c3c-34f-1af%7C%7C4",
    "referer": "https://www.maoyan.com/query?kw=%E5%9B%9E%E4%B9%A1%E9%80%97%E5%84%BF"
}
devices = ['iPhone X', 'iPhone 11', 'iPhone 12', 'iPhone 13', 'iPhone 14']  # 定义设备列表
# 模拟人工访问界面
def run(name, playwright: Playwright):  # 定义一个名为 run 的函数，接收 Playwright 对象作为参数
    browser = playwright.chromium.launch(channel= "msedge", headless=False, chromium_sandbox=True)  # 启动 Edge 浏览器实例，有界面，启用沙箱
    device = playwright.devices[random.choice(devices)]  # 从 devices 中随机选择一个设备配置
    context = browser.new_context(  # 根据所选设备配置创建新的浏览器上下文
        **device
    )
    page = context.new_page()  # 在上下文里创建新页面

    url = "https://i.maoyan.com/apollo/search?searchtype=movie&$from=canary"  # 定义要访问的网址
    # 使用 setExtraHTTPHeaders 方法设置请求头
    page.set_extra_http_headers(headers=headers)  # 为当前页面设置额外的 HTTP 请求头
    page.goto(url)  # 让页面跳转到指定网址
    time.sleep(0.5)

    # 模拟输入操作
    page.type('//*[@id="app"]/div/div[4]/div/div[1]/div[1]/input', name)
    # 模拟点击操作
    page.click('//*[@id="app"]/div/div[4]/div/div[2]/div/div/div/div[1]/div/div[1]')
    if "tfz.maoyan.com" in page.url:
        print("检测到人工验证，等待 30 秒...")
        time.sleep(30)
    time.sleep(random.randint(3,5))
    # time.sleep(10)
    html = page.content()  # 获取当前页面的 HTML 内容
    # print(html)  # 打印 HTML 内容
    soup = BeautifulSoup(html, 'html.parser')
    context.close()
    return soup  # 返回解析后的页面内容
# 爬取票房
def parse_box(soup):
    data_box = soup.find('div', class_='data-box')  # 查找具有特定类名的 div 元素
    if data_box is not None:  # 如果找到
        items = data_box.find_all('div', class_='item')  # 在找到的元素中查找所有具有特定类名的 div 元素
        last_value = items[-1].find('div', class_='value').text  # 获取最后一个元素中的特定类名的 div 元素的文本内容
        return last_value  # 返回文本内容
    else:
        return ""  # 如果未找到，返回特定字符串
# 爬取海报 src
def parse_post(soup):
    imgs = soup.find_all("img")  # 查找所有的 <img> 标签
    for img in imgs:  # 遍历找到的所有图片
        src = img["src"]  # 获取图片的 src 属性值
        if "https://p0.pipi.cn/" not in src:  # 如果 src 属性值中不包含特定字符串，跳过
            continue
        src = replace_numbers(src)
        return src  # 返回替换后的 src 值
# 正则表达式：图片大小转换
def replace_numbers(src):
    pattern = r"w/(\d+)/h/(\d+)"
    replaced = re.sub(pattern, "w/464/h/644", src)
    return replaced
# 存储海报到本地 D 盘
def save_post(src, picture_name):
    # 将文件地址改为 D 盘，并新建一个文件夹：“猫眼海报”
    file_path = f"D:/猫眼海报/{picture_name}.jpg"
    with open(file_path, "wb") as f:  # 以二进制写入模式打开或创建一个文件
        resp_img = requests.get(src)  # 发送 GET 请求获取图片数据
        if resp_img:  # 如果获取成功
            f.write(resp_img.content)  # 将获取到的图片数据写入文件
            file_size = os.path.getsize(file_path)
            file_size_mb = round(file_size / (1024 * 1024), 2)
            save_state = "存储完毕"
            return file_size_mb, save_state  # 返回存储成功的消息
        else:
            save_state = "存储失败"
            file_size_mb = None
            return file_size_mb,save_state  # 如果获取失败，返回存储失败的消息
# 爬取评分+评论人数
def get_movie_info(soup):
    left_div = soup.find('div', class_='left-section')  # 查找具有特定类名的 div 元素
    if left_div:
        spans = left_div.find_all('span')  # 在找到的元素中查找所有的 span 元素
        span_infos = [span.get_text() for span in spans]  # 获取所有 span 元素的文本内容
        if '人想看' in ''.join(span_infos):  # 根据文本内容判断
            score = 'no found'  # 设置特定值
            rating_count = 'no found'  # 设置特定值
        elif '人评' in ''.join(span_infos):  # 根据文本内容判断
            values = [info.split()[0] for info in span_infos if info.split()]  # 处理文本内容
            score = values[0] if len(values) > 0 else 'no found'  # 设置特定值
            rating_count = values[1] if len(values) > 1 else 'no found'  # 设置特定值
        else:
            values = [info.split()[0] for info in span_infos if info.split()]  # 处理文本内容
            score = values[0] if len(values) > 0 else 'no found'  # 设置特定值
            rating_count = 'no found'  # 设置特定值
        return score, rating_count  # 返回处理后的结果
    else:
        score = 'no found'  # 设置特定值
        rating_count = 'no found'  # 设置特定值
        return score, rating_count  # 返回处理后的结果
#   爬取电影名称
def get_movie_name(soup):
    div = soup.find('div', class_="movie-cn-name")
    if div:
        h1 = div.find('h1')
        if h1:
            return h1.text
    return "no found"
#   爬取上映时间+地点
def get_show_time_and_place(soup):
    div = soup.find('div', class_="movie-show-time")
    if div:
        if '/' in div.span.text:
            time_place_duration = div.span.text.split('/')
            time_and_place = time_place_duration[0].strip()
            duration = time_place_duration[1].strip()
        else:
            time_and_place = div.span.text
            duration = "no found"
        # 以下赋值语句无论“/”在不在里面都运行
        show_place = ''.join([char for char in time_and_place if '\u4e00' <= char <= '\u9fff'])
        show_time = time_and_place.replace(show_place, '').strip()
        year = show_time.split('-')[0]
        return show_time, show_place, duration, year
    else:
        return "no found", "no found", "no found", "no found"
#   爬取主演
def get_actors(soup):
    div = soup.find('div', class_="actors")
    if div:
        actors = [re.sub(r'\s*[/]\s*', ',', a.text.strip()) for a in div.find_all('a')]
        actors = [re.sub(r'\s*&\nbsp;\s*', ',', actor) for actor in actors]
        actors = [re.sub(r'\s*&nbsp;\s*', '', actor) for actor in actors]
        actors = [re.sub(r'\s*,\s*$', '', actor) for actor in actors]
        result = '/'.join(actors)  # 将数组中的所有值用“/”链接并存入 result 变量
        return result
    else:
        return "no found"
#   爬取猫眼id
def maoyan_id(soup):
    link = soup.find('meta', {'name':'share:wechat:message:link'})['content']
    return link.split('?')[0]
#   函数调用
def crawl_movie_info(soup, row_index, picture_name):
    print(f"####正在爬取电影NO.{row_index}")
    src = parse_post(soup)  # 解析海报 src
    box = parse_box(soup)     # 票房
    save_state, file_size_mb = save_post(src, picture_name)  # 存储海报
    score, rating_count = get_movie_info(soup)  # 获取评分和评论人数
    movie_name = get_movie_name(soup)   # 电影名字
    show_time, show_place, duration, year = get_show_time_and_place(soup)
    actors = get_actors(soup)
    id = maoyan_id(soup)
    return movie_name, score, rating_count, show_time, show_place, duration, year, actors, save_state, file_size_mb, id, src, box
#   表格存储
def save_to_excel(result, ws_xlsx, row_index):
    columns = [6,7,8,9,10,11,12,13,14,15,16,17,18]
    for i, column in enumerate(columns):
        if result[i] != "no found":
            ws_xlsx.cell(row=row_index, column=column, value=result[i])
    print("表格存储完毕")

wb_xlsx = load_workbook(xlsx)
ws_xlsx = wb_xlsx.active
for row_index in range(2, ws_xlsx.max_row + 1):  # 从第二行开始，跳过标题行
    # n = ws_xlsx.cell(row=row_index, column=1).value
    movie_id = ws_xlsx.cell(row=row_index, column=1).value  # 媒资号
    movie_name = ws_xlsx.cell(row=row_index, column=2).value  # 片名
    movie_state = ws_xlsx.cell(row=row_index, column=15).value  # 海报存取状态
    # if movie_id is None:
    #     break
    if movie_state != "存储完毕":
        with sync_playwright() as playwright:
            try:
                soup = run(movie_name, playwright)
                movie_box = parse_box(soup)
                result = crawl_movie_info(soup, row_index, movie_id)
                for item in result:
                    print(item, end="\t")
                save_to_excel(result, ws_xlsx, row_index)
            except Exception as e:
                ws_xlsx.cell(row=row_index, column=15, value="抓取失败")  # 海报存取状态
                print(f"发生异常: 《{movie_name}》{e} {traceback.format_exc()}")
    if row_index % 5 == 0:
        wb_xlsx.save(xlsx)
wb_xlsx.save(xlsx)
