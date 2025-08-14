# -*- coding: utf-8 -*-
import time
import traceback
import requests
import random
import re
import os
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, Playwright
from openpyxl import load_workbook

xlsx = "C:\\Users\\fucha\\Desktop\\爬取通用模板.xlsx"
path_to_extension = "./my-extension"  # 定义扩展路径
user_data_dir = "/tmp/test-user-data-dir"  # 定义用户数据目录
headers = {  # 定义请求头
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36 Edg/127.0.0.0",
    "cookie": "__mta=143505930.1721729537864.1722391669891.1722398395719.56; uuid_n_v=v1; uuid=09EDF75048DC11EFAD2C9D84A9502795C74339EF0D0246B69AA625DDDBABD5AB; _lxsdk_cuid=190906c187fc8-02326d3fde828b-4c657b58-151800-190906c187fc8; WEBDFPID=1uxz77v921v05uu7yyy3xw2852014z7x809v6x7954z97958v86v0749-2037168061264-1721808060272KMUCGQQfd79fef3d01d5e9aadc18ccd4d0c95071445; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; _lxsdk=09EDF75048DC11EFAD2C9D84A9502795C74339EF0D0246B69AA625DDDBABD5AB; _csrf=e621735b58dac37e972a4bb926f6a547551dab7f73b922a119caf2bd348f5558; Hm_lvt_703e94591e87be68cc8da0da7cbd0be2=1722330825,1722332208,1722391670,1722398388; HMACCOUNT=020341BA99F5C478; __mta=143505930.1721729537864.1722332208176.1722398387968.60; Hm_lpvt_703e94591e87be68cc8da0da7cbd0be2=1722398396; _lxsdk_s=19106ed243d-c3c-34f-1af%7C%7C4",
    "referer": "https://www.maoyan.com/query?kw=%E5%9B%9E%E4%B9%A1%E9%80%97%E5%84%BF"
}
devices = ['iPhone X', 'iPhone 11', 'iPhone 12', 'iPhone 13', 'iPhone 14']  # 定义设备列表

# 模拟人工访问界面
def run(name, playwright: Playwright):  # 定义一个名为 run 的函数，接收 Playwright 对象作为参数
    browser = playwright.chromium.launch(channel= "msedge", headless=False,  # 启动 Edge 浏览器实例，有界面，启用沙箱
                                         chromium_sandbox=True)
    device = playwright.devices[random.choice(devices)]  # 从 devices 中随机选择一个设备配置
    context = browser.new_context(  # 根据所选设备配置创建新的浏览器上下文
        **device
    )
    page = context.new_page()  # 在上下文里创建新页面

    url = "https://i.maoyan.com/apollo/search?searchtype=movie&$from=canary"  # 定义要访问的网址
    # 使用 setExtraHTTPHeaders 方法设置请求头
    page.set_extra_http_headers(headers=headers)  # 为当前页面设置额外的 HTTP 请求头
    page.goto(url)  # 让页面跳转到指定网址
    time.sleep(0.5)
    # 模拟输入操作
    page.type('//*[@id="app"]/div/div[4]/div/div[1]/div[1]/input', name)
    # 模拟点击操作
    page.click('//*[@id="app"]/div/div[4]/div/div[2]/div/div/div/div[1]/div/div[1]')
    # time.sleep(random.randint(3, 5))# 随机等待 3 到 8 秒

    html = page.content()  # 获取当前页面的 HTML 内容
    time.sleep(15)
    context.close()
    return html  # 返回解析后的页面内容
# 爬取名称
def get_title(soup):
    # 找到name属性为share:wechat:message:desc的meta标签，获取其content属性里的文本内容
    title = soup.find('meta', attrs={'name': 'share:wechat:message:title'}).get('content')
    if title:
        # 假设brief_introduction是提取到的文本内容
        title = re.sub(r'[\u2f00]', '', title)
        title = title.encode('gbk', 'replace')  # 使用'replace'参数，遇到无法编码的字符用'?'替代
        title = title.decode('gbk')
        title = re.findall(r'《(.*?)》', title)[0]
        return title
    else:
        return "NO FOUND"
# 爬取演员
def get_act(soup):
    # 找到name属性为share:wechat:message:desc的meta标签，获取其content属性里的文本内容
    act = soup.find('meta', attrs={'name': 'keywords'}).get('content')
    if act:
        # 假设brief_introduction是提取到的文本内容
        act = re.sub(r'[\u2f00]', '', act)
        act = act.encode('gbk', 'replace')  # 使用'replace'参数，遇到无法编码的字符用'?'替代
        act = act.decode('gbk')
        act = re.sub(r'^.*?,|,选座购票.*$', '', act)
        return act
    else:
        return "NO FOUND"
# 爬取简介
def get_introduce(soup):
    # 找到name属性为share:wechat:message:desc的meta标签，获取其content属性里的文本内容
    introduce = soup.find('meta', attrs={'name': 'share:wechat:message:desc'}).get('content')
    if introduce:
        # 提取出简介部分（去除前面可能有的类似“简介|”这样的前缀，实际可根据具体格式调整）
        introduce = introduce.split("|")[-1].strip()
        # 假设brief_introduction是提取到的文本内容
        introduce = re.sub(r'[\u2f00]', ' ', introduce)
        introduce = introduce.encode('gbk', 'replace')  # 使用'replace'参数，遇到无法编码的字符用'?'替代
        introduce =introduce.decode('gbk')
        return introduce
    else:
        return "NO FOUND"
#   函数调用
def crawl_movie_info(html,row_index):
    print(f"####正在爬取电影NO.{row_index}")
    soup = BeautifulSoup(html, 'lxml')
    introduce = get_introduce(soup)
    title = get_title(soup)
    act = get_act(soup)
    return title, act, introduce
#   表格存储
def save_to_excel(result, ws_xlsx, row_index):
    columns = [4,5,6]
    for i, column in enumerate(columns):
        if result[i] != "NO FOUND":
            ws_xlsx.cell(row=row_index, column=column, value=result[i])
            ws_xlsx.cell(row=row_index, column=7, value="存储完毕")
    print("表格存储完毕")


wb_xlsx = load_workbook(xlsx)
ws_xlsx = wb_xlsx.active

for row_index in range(2, ws_xlsx.max_row + 1):  # 从第二行开始，跳过标题行
    movie_name = ws_xlsx.cell(row=row_index, column=3).value  # 片名
    movie_state = ws_xlsx.cell(row=row_index, column=7).value  # 存取状态
    if movie_state != "存储完毕":
        with sync_playwright() as playwright:
            try:
                html = run(movie_name, playwright)
                result = crawl_movie_info(html, row_index)
                for item in result:
                    print(item, end="\t")
                save_to_excel(result, ws_xlsx, row_index)
            except Exception as e:
                ws_xlsx.cell(row=row_index, column=7, value="抓取失败")  # 海报存取状态
                print(f"发生异常: 《{movie_name}》{e} {traceback.format_exc()}")
    if row_index % 5 == 0:
        wb_xlsx.save(xlsx)
wb_xlsx.save(xlsx)