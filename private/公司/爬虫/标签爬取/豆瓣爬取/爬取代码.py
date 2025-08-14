from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import re
import os
import urllib.parse
import json
import shutil


# 定义一个函数，将字符串转换为 URL 编码
def str2urlcode(s):
    """
    此函数接受一个字符串参数 s，使用 urllib.parse.quote 方法将其转换为 URL 编码并返回
    """
    return urllib.parse.quote(s)


# 定义 Excel 模板文件和输出文件的路径
# xls_template = "C:\\Users\\fucha\\PycharmProjects\\python\\file\\豆瓣标签爬取模板.xlsx"
# xls_file = "C:\\Users\\fucha\\PycharmProjects\\python\\file\\豆瓣标签输出表格.xlsx"

xls_template = "C:\\Users\\fucha\\Desktop\\豆瓣标签爬取模板.xlsx"
xls_file = "C:\\Users\\fucha\\Desktop\\豆瓣标签输出表格.xlsx"

# 复制模板文件为输出文件
shutil.copyfile(xls_template, xls_file)
# 加载输出文件的工作簿
workbook = load_workbook(xls_file)
# 获取活动工作表
sheet = workbook.active
# 初始化电影名称和类型变量
movie_name = None
movie_type = None

# 设置请求头
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Cookie': 'bid=SQbrZnxAM7M; _pk_id.100001.2939=c909a5c953dcc3fd.1721572772.; ll="108288"; dbcl2="282351081:aRSLWCaig7A"; ck=oryP; _pk_ref.100001.2939=%5B%22%22%2C%22%22%2C1722297151%2C%22https%3A%2F%2Faccounts.douban.com%2F%22%5D; _pk_ses.100001.2939=1; push_noty_num=0; push_doumail_num=0; __utma=30149280.17505400.1721572772.1721572772.1722297151.2; __utmc=30149280; __utmz=30149280.1722297151.2.2.utmcsr=accounts.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/; __utmt=1; __utmv=30149280.28235; __yadk_uid=kr0dBbMfDdFrJkI7HyeP7C5zg1MLI7ld; __utmb=30149280.6.9.1722297166961',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0'
}

# 从 Excel 表格中读取数据
# 从第 2 行开始读取电影名称列的数据
row_idx = 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    # 获取当前行的电影名称
    movie_name = row[3]
    # 初始化电影类型为 None
    movie_title = None
    movie_type = None
    movie_average = None
    movie_votes = None
    movie_href = None
    # 如果电影名称为空，结束循环
    if movie_name == None:
        break
    # 如果当前行的第 9 列（索引为 8）为空
    # if row[9] is None:
    if row_idx is not None:
        # 对电影名称进行 URL 编码
        encoded_name = str2urlcode(movie_name)
        # 构建搜索 URL
        search_url = f"https://search.douban.com/movie/subject_search?search_text={encoded_name}&cat=1002"

        # 发送 GET 请求获取 HTML 内容
        response = requests.get(search_url)
        # 如果请求成功（状态码为 200）
        if response.status_code == 200:
            # 使用 BeautifulSoup 解析 HTML 内容
            soup = BeautifulSoup(response.text, 'html.parser')
            # 初始化电影响应为 None
            movie_response = None
            # 查找所有的 <script> 标签
            script_tags = soup.find_all('script')
            for script_tag in script_tags:
                # 如果 <script> 标签的字符串内容包含 'window.__DATA__ ='
                if script_tag.string and 'window.__DATA__ =' in script_tag.string:
                    # 提取 JavaScript 内容
                    script_content = script_tag.string
                    # 按特定字符串分割内容
                    parts = script_content.split('indow.__DATA__ =')
                    if len(parts) > 1:
                        # 找到 JSON 数据部分
                        json_data = parts[1].split('};')[0] + '}'
                        # 解析 JSON 数据
                        data = json.loads(json_data)

                        # 提取 URL 列表
                        urls = []
                        for item in data['items']:
                            if 'url' in item:
                                urls.append(item['url'])

                        # 如果获取到了 URL 列表且不为空
                        if len(urls) > 0:
                            # 获取第一个 URL
                            href = urls[0]
                            # 带着请求头发送 GET 请求获取电影详情页
                            movie_response = requests.get(href, headers=headers)
                            # 如果请求成功
                            if movie_response.status_code == 200:
                                # 使用 BeautifulSoup 解析电影详情页的 HTML 内容
                                movie_soup = BeautifulSoup(movie_response.text, 'html.parser')
                                # 查找具有特定属性的 <span> 标签来获取电影类型
                                title_span = movie_soup.find('span', property='v:itemreviewed')
                                genre_span = movie_soup.find('span', property='v:genre')
                                average_strong = movie_soup.find('strong', property='v:average')
                                votes_span = movie_soup.find('span', property='v:votes')
                                # 如果找到了，获取其文本作为电影类型，否则设为 'NO FOUND'
                                movie_title = title_span.text if title_span else 'NO FOUND'
                                movie_type = genre_span.text if genre_span else 'NO FOUND'
                                movie_average = average_strong.text if average_strong else 'NO FOUND'
                                movie_votes = votes_span.text if votes_span else 'NO FOUND'
                                movie_href = href if title_span else ''
                            else:
                                # 如果请求电影详情页失败，设置电影类型为相应提示信息
                                movie_type = f"无法访问 {href}"

        # 打印电影的链接、名称和类型
        print(row_idx, movie_name, movie_type, movie_average, movie_votes)
        # 如果电影类型不是 'NO FOUND'，将其写入 Excel 表格的第 10 列
        # if movie_type!= "NO FOUND":
        #    sheet.cell(row=row_idx, column=10, value=movie_type)
        # 将电影类型写入 Excel 表格的第 11 列
        sheet.cell(row=row_idx, column=46, value=movie_type)
        sheet.cell(row=row_idx, column=47, value=movie_average)
        sheet.cell(row=row_idx, column=48, value=movie_votes)
        sheet.cell(row=row_idx, column=49, value=movie_title)
        sheet.cell(row=row_idx, column=50, value=movie_href)
        # 行索引递增
        row_idx = row_idx + 1

# 保存修改后的 Excel 文件
workbook.save(xls_file)

"""
<h1>
    <span property="v:itemreviewed">我要成为超级巨星 วันหนึ่งจะเป็นซุปตาร์</span>
        <span class="year">(2015)</span>
</h1>

<div id="mainpic" class="">
    <a class="nbgnbg" href="https://movie.douban.com/subject/26357829/photos?type=R" title="点击看更多海报">
        <img src="https://img1.doubanio.com/view/photo/s_ratio_poster/public/p2256695789.webp" title="点击看更多海报" alt="วันหนึ่งจะเป็นซุปตาร์" rel="v:image" />
   </a>
                    <p class="gact">
                        <a href="https://movie.douban.com/subject/26357829/edit">
                            更新描述或海报
                        </a>
                    </p>
</div>

<div id="info">


        <span><span class='pl'>主演</span>: <span class='attrs'><a href="https://www.douban.com/personage/27499440/" rel="v:starring">普提查·克瑟辛</a> / <a href="https://www.douban.com/personage/27549903/" rel="v:starring">阿芮根妲·马哈乐沙空</a> / <a href="/subject_search?search_text=Woonsen%20Virithipa%20Pakdeeprasong" rel="v:starring">Woonsen Virithipa Pakdeeprasong</a> / <a href="https://www.douban.com/personage/27545691/" rel="v:starring">协塔朋·平朋</a> / <a href="/subject_search?search_text=Kittiphong%20Tantichinanon" rel="v:starring">Kittiphong Tantichinanon</a> / <a href="/subject_search?search_text=Tyler%20Harrison" rel="v:starring">Tyler Harrison</a> / <a href="/subject_search?search_text=Bom%20Anuruk%20Boonpermpoon" rel="v:starring">Bom Anuruk Boonpermpoon</a></span></span><br/>
        <span class="pl">类型:</span> <span property="v:genre">喜剧</span> / <span property="v:genre">爱情</span><br/>

        <span class="pl">制片国家/地区:</span> 泰国<br/>
        <span class="pl">语言:</span> 泰语<br/>
        <span class="pl">首播:</span> <span property="v:initialReleaseDate" content="2015-08-03(泰国)">2015-08-03(泰国)</span><br/>

        <span class="pl">集数:</span> 26<br/>
        <span class="pl">单集片长:</span> 50分钟<br/>
        <span class="pl">又名:</span> 下一站天后 / I Wanna Be Sup’Tar<br/>

</div>

<div class="rating_self clearfix" typeof="v:Rating">
    <strong class="ll rating_num" property="v:average">7.6</strong>
    <span property="v:best" content="10.0"></span>
    <div class="rating_right ">
        <div class="ll bigstar bigstar40"></div>
        <div class="rating_sum">
                <a href="comments" class="rating_people">
                    <span property="v:votes">10398</span>人评价
                </a>
        </div>
    </div>
</div>
"""