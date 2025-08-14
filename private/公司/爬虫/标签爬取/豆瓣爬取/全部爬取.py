from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import re
import os
import urllib.parse
import json

# 定义一个函数，将字符串转换为 URL 编码
def str2urlcode(s):
    return urllib.parse.quote(s)

# 定义 Excel 模板文件的路径
xlsx = "C:\\Users\\fucha\\PycharmProjects\\python\\file\\豆瓣标签爬取模板.xlsx"
# 加载模板文件的工作簿
workbook = load_workbook(xlsx)
sheet = workbook.active

# 设置请求头
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Cookie': 'bid=_PFUYl8B7qY; ll="108288"; _pk_id.100001.4cf6=0d53d319030ff1a7.1720419526.; _vwo_uuid_v2=D7538082B8EDB72774B12DFB37F81BD1B|294f9e79497f5751092275e93394e8e5; __yadk_uid=L8nVjMFI5LDAlMd02u6LYgUpC0mJ6hCQ; dbcl2="281958561:gfTbwIjn9IQ"; push_noty_num=0; push_doumail_num=0; __utmv=30149280.28195; ck=GEkJ; __utmc=30149280; __utmc=223695111; frodotk_db="6af6b99d54891ac040e42cd7fd9b5040"; __utmz=30149280.1721603612.22.9.utmcsr=search.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/movie/subject_search; ap_v=0,6.0; __utma=30149280.718252523.1720419527.1721606853.1721612671.24; __utma=223695111.311806802.1720419527.1721606856.1721612676.23; __utmz=223695111.1721612676.23.10.utmcsr=search.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/movie/subject_search; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1721615526%2C%22https%3A%2F%2Fsearch.douban.com%2Fmovie%2Fsubject_search%3Fsearch_text%3D1%2F2%E6%AE%B5%E6%83%85%26cat%3D1002%22%5D; _pk_ses.100001.4cf6=1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0'
}

# 从 Excel 表格中读取数据
row_idx = 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    movie_name = row[3]
    if movie_name is None:
        break
    encoded_name = str2urlcode(movie_name)
    search_url = f"https://search.douban.com/movie/subject_search?search_text={encoded_name}&cat=1002"

    response = requests.get(search_url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        movie_response = None
        script_tags = soup.find_all('script')
        for script_tag in script_tags:
            if script_tag.string and 'window.__DATA__ =' in script_tag.string:
                script_content = script_tag.string
                parts = script_content.split('indow.__DATA__ =')
                if len(parts) > 1:
                    json_data = parts[1].split('};')[0] + '}'
                    data = json.loads(json_data)

                    urls = []
                    for item in data['items']:
                        if 'url' in item:
                            urls.append(item['url'])

                    if len(urls) > 0:
                        href = urls[0]
                        movie_response = requests.get(href, headers=headers)
                        if movie_response.status_code == 200:
                            movie_soup = BeautifulSoup(movie_response.text, 'html.parser')
                            genre_span = movie_soup.find('span', property='v:genre')
                            movie_type = genre_span.text if genre_span else 'NO FOUND'
                        else:
                            movie_type = f"无法访问 {href}"

        print(f"{movie_name} {movie_type}")
        sheet.cell(row=row_idx, column=11, value=movie_type)  # 将标签内容写入第 K 列（假设 K 为 11）
        row_idx += 1

# 保存修改后的 Excel 文件
workbook.save(xlsx)