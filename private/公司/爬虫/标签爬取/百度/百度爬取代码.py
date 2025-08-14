import shutil
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import re
import os

xls_template = "C:\\Users\\fucha\\PycharmProjects\\python\\file\\百度标签爬取模板.xlsx"
xls_file = "C:\\Users\\fucha\\PycharmProjects\\python\\file\\百度标签输出表格.xlsx"

shutil.copyfile(xls_template, xls_file)
workbook = load_workbook(xls_file)
sheet = workbook.active
movie_name = None
movie_type = None

row_idx = 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    movie_name = row[3]
    movie_type = None
    if row[9] is None:
        url = f"https://baike.baidu.com/item/{movie_name}?fromModule=lemma_search-box"
        response = requests.get(url)
        html = response.text
        soup = BeautifulSoup(html,'html.parser')
        # print(html)
        found_items = soup.find_all('div', class_="itemWrapper_wncpX")

        for item in found_items:
            dts = item.find_all('dt', class_="basicInfoItem_EAC5k itemName_ObiPN")
            for dt in dts:
                if dt.text == "类    型":
                    dd = dt.find_next_sibling('dd')
                    text = dd.text
                    # text = re.sub(r'\[\d+\]', '', text)
                    # result = text.strip()
                    movie_type = text.strip()
                    break
        if movie_type == None:
            movie_type = "no found"
        #movie_type = "喜剧 / 动作 / 西部 / 冒险[15]"
        movie_type = movie_type.replace(" / ", ",")
        movie_type = movie_type.replace("/", ",")
        movie_type = movie_type.replace("，", ",")
        movie_type = movie_type.replace("、", ",")
        movie_type = movie_type.replace(",,", ",")
        movie_type = movie_type.replace(" ", ",")
        movie_type = re.sub(r'\[\d+\]', '', movie_type)
        print(row_idx - 1,row[3],movie_type)
        if movie_type != "no,found":
            sheet.cell(row=row_idx, column=10, value=movie_type)
        sheet.cell(row=row_idx, column=11, value=movie_type)
    row_idx = row_idx + 1

workbook.save(xls_file)
