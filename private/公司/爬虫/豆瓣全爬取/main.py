# -*- coding: utf-8 -*-
import re
import os
import sys
import json
import requests
import urllib.parse
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 设置请求头
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Cookie': 'bid=_PFUYl8B7qY; ll="108288"; _pk_id.100001.4cf6=0d53d319030ff1a7.1720419526.; _vwo_uuid_v2=D7538082B8EDB72774B12DFB37F81BD1B|294f9e79497f5751092275e93394e8e5; __yadk_uid=L8nVjMFI5LDAlMd02u6LYgUpC0mJ6hCQ; dbcl2="281958561:gfTbwIjn9IQ"; push_noty_num=0; push_doumail_num=0; __utmv=30149280.28195; ck=GEkJ; __utmc=30149280; __utmc=223695111; frodotk_db="6af6b99d54891ac040e42cd7fd9b5040"; __utmz=30149280.1721603612.22.9.utmcsr=search.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/movie/subject_search; ap_v=0,6.0; __utma=30149280.718252523.1720419527.1721606853.1721612671.24; __utma=223695111.311806802.1720419527.1721606856.1721612676.23; __utmz=223695111.1721612676.23.10.utmcsr=search.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/movie/subject_search; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1721615526%2C%22https%3A%2F%2Fsearch.douban.com%2Fmovie%2Fsubject_search%3Fsearch_text%3D1%2F2%E6%AE%B5%E6%83%85%26cat%3D1002%22%5D; _pk_ses.100001.4cf6=1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0'
}
# 名字转换
def str2urlcode(s):
    """将字符串转换为 URL 编码格式。"""
    return urllib.parse.quote(s)
# 获取href
def get_href(search_url):
    response = requests.get(search_url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        script_tags = soup.find_all('script')
        for script_tag in script_tags:
            if script_tag.string and 'window.__DATA__ =' in script_tag.string:
                script_content = script_tag.string
                parts = script_content.split('indow.__DATA__ =')
                if len(parts) > 1:
                    json_data = parts[1].split('};')[0] + '}'
                    data = json.loads(json_data)
                    urls = []
                    for item in data['items']:
                        if 'url' in item:
                            urls.append(item['url'])
                    if len(urls) > 0:
                        href = urls[0]
                        # print(href)
                        return href
    else:
        return None
# 获取电影名
def get_name(soup):
    temple = soup.find('span', property='v:itemreviewed')
    name = temple.text if temple else 'NO FOUND'
    return name
# 获取类型
def get_type(soup):
    """从电影详情页面获取电影类型。"""
    temple = soup.find_all('span', property='v:genre')
    movie_type = [i.text if temple else 'NO FOUND' for i in temple]
    type = "/".join(movie_type)
    return type
# 获取年份
def get_year(soup):
    """从电影详情页面获取电影年份。"""
    temple = soup.find('span', class_="year")
    year = temple.text if temple else 'NO FOUND'
    year = re.sub(r'\(|\)', '', year)
    return year
# 获取导演
def get_director(soup):
    """从电影详情页面获取电影导演。"""
    temple = soup.find('span', class_="attrs")
    director = temple.text if temple else 'NO FOUND'
    director = re.sub(r' ', '', director)
    return director
# 获取主演
def get_actors(soup, director):
    """从电影详情页面获取电影主演，并去除导演信息。"""
    temple = soup.find_all('span', class_="attrs")
    actors = [i.text if temple else 'NO FOUND' for i in temple]
    actors = "/".join(actors)
    actors = re.sub(r' ', '', actors)
    words1 = set(director.split('/'))
    words2 = actors.split('/')

    actors = "/".join([word for word in words2 if word not in words1])
    return actors
# 获取地点
def get_place(soup):
    for span in soup.find_all('span', class_='pl'):
        if span.next_sibling and span.text == "制片国家/地区:":
            return span.next_sibling.strip()
    return "NO FOUND"
# 获取集数
def get_count(soup):
    for span in soup.find_all('span', class_='pl'):
        if span.next_sibling and "集数" in span.text:
            return span.next_sibling.strip()
    return "NO FOUND"
# 获取单集片长
def get_time(soup):
    for span in soup.find_all('span', class_='pl'):
        if span.next_sibling and span.text == "单集片长:":
            return span.next_sibling.strip()
    return "NO FOUND"
# 获取语言
def get_language(soup):
    for span in soup.find_all('span', class_='pl'):
        if span.next_sibling and span.text == "语言:":
            return span.next_sibling.strip()
    return "NO FOUND"
# 获取又名
def get_sb_name(soup):
    for span in soup.find_all('span', class_='pl'):
        if span.next_sibling and span.text == "又名:":
            return span.next_sibling.strip()
    return "NO FOUND"
# 获取猫眼评分+评论人数
def get_rating(soup):
    rating_element = soup.find('strong', {'class': 'll rating_num', 'property': 'v:average'})
    rating = rating_element.text if rating_element else "NO FOUND"
    votes_element = soup.find('span', {'property': 'v:votes'})
    votes = votes_element.text if votes_element else "NO FOUND"
    return rating, votes
# 爬取海报 src
def get_post(soup):
    imgs = soup.find_all("img")  # 查找所有的 <img> 标签
    for img in imgs:
        src = img["src"]
        if "doubanio.com/view/photo/s_ratio_poster" in src:
            return src
        else:
            return "NO FOUND"
# 存储海报到本地 D 盘
def save_post(src, picture_name):
    # 将文件地址改为 D 盘，并新建一个文件夹：“猫眼海报”
    file_path = f"D:/豆瓣海报/{picture_name}.jpg"
    with open(file_path, "wb") as f:  # 以二进制写入模式打开或创建一个文件
        if src != "NO FOUND":
            resp_img = requests.get(src)  # 发送 GET 请求获取图片数据
            if resp_img:  # 如果获取成功
                f.write(resp_img.content)  # 将获取到的图片数据写入文件
                file_size = os.path.getsize(file_path)
                file_size_mb = round(file_size / (1024 * 1024), 2)
                save_state = "存储完毕"
                return file_size_mb, save_state  # 返回存储成功的消息
            else:
                save_state = "存储失败"
                file_size_mb = None
                return file_size_mb,save_state  # 如果获取失败，返回存储失败的消息
        else:
            return 0,"未找到海报"
#   函数调用
def use_Function(movie_response, row_index, href, picture_name):
    print(f"#####正在爬取电影NO.{row_index}{movie_name}")
    soup = BeautifulSoup(movie_response.text, 'html.parser')
    name = get_name(soup)
    typ = get_type(soup)
    year = get_year(soup)
    director = get_director(soup)
    actors = get_actors(soup, director)
    place = get_place(soup)
    count = get_count(soup)
    time = get_time(soup)
    language = get_language(soup)
    sb_name = get_sb_name(soup)
    rating, votes = get_rating(soup)
    src = get_post(soup)
    file_size_mb, save_state = save_post(src, picture_name)
    try:
        print(src)
        print(f"电影名称：{name}    年份：{year}    制片地区：{place} 类型：{typ}")
        print(f"导演：{director}   主演：{actors}")
        print(f"集数：{count}  单集片长：{time} 语言：{language}   又名：{sb_name}    评分：{rating} 评论人数：{votes}")
        print(f"海报大小：{file_size_mb}   海报存取状态：{save_state}   海报src：{src}")
    except UnicodeEncodeError:
        print("无法正常打印")
    return name, sb_name, year, place, typ, director, actors, count, time, language, rating, votes,  file_size_mb, src, href, save_state
#   表格存储
def save_to_excel(result, xlsx, row_index):
    for i,column in enumerate(range(13, 29)):
        if result[i] != "NO FOUND":
            sheet.cell(row=row_index, column=column, value=result[i])
    print("表格存储完毕")

# 加载模板文件的工作簿
xlsx = "C:\\Users\\fucha\\Desktop\\豆瓣爬取模板.xlsx"
workbook = load_workbook(xlsx)
sheet = workbook.active
#——————主函数——————#
for row_index in range(2, sheet.max_row + 1):  # 从第二行开始，跳过标题行
    movie_id = sheet.cell(row=row_index, column=1).value  # 媒资号
    movie_name = sheet.cell(row=row_index, column=2).value  # 片名
    picture_name = movie_id
    if movie_name is None:
        movie_name = sheet.cell(row=row_index, column=3).value
    movie_state = sheet.cell(row=row_index, column=28).value  # 海报存取状态
    if movie_state != "存储完毕":
        encoded_name = str2urlcode(movie_name)
        search_url = f"https://search.douban.com/movie/subject_search?search_text={encoded_name}&cat=1002"
        href = get_href(search_url)
        if href:
            movie_response = requests.get(href, headers=headers)
            if movie_response.status_code == 200:
                result = use_Function(movie_response, row_index, href, picture_name)
                save_to_excel(result, sheet, row_index)
            else:
                movie_type = f"无法访问 {movie_name}"
        else:
            sheet.cell(row=row_index, column=28, value="爬取失败")
    workbook.save(xlsx)