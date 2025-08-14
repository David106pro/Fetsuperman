import os
import re
import json
import requests
import urllib.parse
import difflib
from bs4 import BeautifulSoup
import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk # Import ttk for Tabview if needed (CTK has its own)
from PIL import Image, ImageTk, ImageFilter
from io import BytesIO
import traceback
import threading
import pandas as pd
import time
import random

TARGET_PREVIEW_V_HEIGHT = 80 # Target height for Vertical/Original previews
TARGET_PREVIEW_H_WIDTH = 120 # Target width for Horizontal previews

class MultiPlatformImageDownloader(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.config_file = "config.json"
        settings = self.load_settings() # Load settings early

        # Store filename format
        self.filename_format = settings.get("filename_format", "{标题}_{类型}_{图片尺寸}")
        
        # 设置窗口
        self.title("视频图片下载器")
        self.geometry("1400x800")  # 增加宽度以容纳侧边栏
        
        # 配置主窗口网格
        self.grid_columnconfigure(0, weight=0)  # 左侧工具栏 - 固定宽度
        self.grid_columnconfigure(1, weight=1)  # 右侧内容区 - 可扩展
        self.grid_rowconfigure(0, weight=1)     # 整体行 - 可扩展

        # === 创建左侧工具栏 ===
        self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        self.sidebar_frame.grid_propagate(False)  # 防止自动调整大小
        
        # 工具栏标题
        self.sidebar_title = ctk.CTkLabel(
            self.sidebar_frame, 
            text="功能选择", 
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.sidebar_title.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        # 创建工具栏按钮 - 上对齐的三个
        self.classic_button = ctk.CTkButton(
            self.sidebar_frame,
            text="经典爬取",
            height=40,
            command=lambda: self.switch_page("classic")
        )
        self.classic_button.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        
        self.batch_button = ctk.CTkButton(
            self.sidebar_frame,
            text="批量爬取", 
            height=40,
            command=lambda: self.switch_page("batch")
        )
        self.batch_button.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        
        self.delete_button = ctk.CTkButton(
            self.sidebar_frame,
            text="批量删除错误海报",
            height=40,
            command=lambda: self.switch_page("delete")
        )
        self.delete_button.grid(row=3, column=0, padx=20, pady=10, sticky="ew")
        
        # 占位空间 - 用于将设置按钮推到底部
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        
        # 设置按钮 - 下对齐
        self.sidebar_settings_button = ctk.CTkButton(
            self.sidebar_frame,
            text="默认设置",
            height=40,
            command=lambda: self.switch_page("settings")
        )
        self.sidebar_settings_button.grid(row=5, column=0, padx=20, pady=(10, 20), sticky="ew")

        # === 创建右侧内容区域 ===
        self.content_frame = ctk.CTkFrame(self)
        self.content_frame.grid(row=0, column=1, padx=(0, 10), pady=10, sticky="nsew")
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=1)

        # 当前页面状态
        self.current_page = "classic"
        
        # 初始化页面变量
        self.pages = {}
        
        # 初始化平台选择变量
        self.selected_platform = settings.get("default_platform", "爱奇艺")
        self.platform_var = ctk.StringVar(value=self.selected_platform)
        
        # 初始化精确搜索变量
        self.precise_search_var = ctk.BooleanVar(value=settings.get("default_precise", False))
        
        # 初始化图片类型选择变量
        self.orientation_var = ctk.StringVar(value=settings.get("default_download_type", "全部"))
        
        # 尺寸预设选择
        self.size_presets = {
            "基础尺寸": {"vertical": (412, 600), "horizontal": (528, 296)}, # Renamed
            "河南尺寸": {"vertical": (525, 750), "horizontal": (257, 145)},
            "甘肃尺寸": {"vertical": (412, 600), "horizontal": (562, 375)},
            "陕西尺寸": {"vertical": (245, 350), "horizontal": (384, 216)},
            "云南尺寸": {"vertical": (262, 360), "horizontal": (412, 230)}, # Added Yunnan
            "自定义尺寸": {"vertical": (0, 0), "horizontal": (0, 0)} # 0,0 signifies custom input needed
        }
        
        # 存储结果
        self.results = []
        self.result_checkboxes = []
        self.result_vars = []
        
        # 全选/取消全选状态
        self.select_all_state = True
        
        # 批量模式相关状态初始化
        self.is_batch_processing = False
        self.batch_paused = False
        self.current_batch_row = 0
        self.batch_df = None
        self.excel_file_path = ""
        self.current_table_mode = "batch"
        
        # 批量删除相关变量
        self.delete_excel_file_path = ""
        self.delete_df = None
        self.selected_delete_key_column = None  # 选择的删除标准列
        self.delete_column_checkboxes = {}  # 存储列头复选框
        
        # 优先级选择状态
        self.selected_priority_index = -1
        self.priority_frames = []
        
        # 设置爱奇艺请求头 (Load cookie from settings)
        self.iqiyi_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Referer": "https://www.iqiyi.com/",
            "Accept": "application/json, text/plain, */*",
            "Cookie": settings.get("iqiyi_cookie", "") # Load from settings
        }

        # 腾讯视频请求头 (Load cookie from settings)
        self.tencent_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36",
            "Content-Type": "application/json",
            "Referer": "https://v.qq.com/",
            "Origin": "https://v.qq.com",
            "Cookie": settings.get("tencent_cookie", "") # Load from settings
        }

        # 优酷视频请求头 (Load cookie from settings)
        self.youku_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36 Edg/135.0.0.0",
            "Referer": "https://www.youku.com/",
            "Cookie": settings.get("youku_cookie", "") # Load from settings
        }

        # 创建各个页面内容
        self.create_classic_page()
        self.create_batch_page()
        self.create_delete_page()
        self.create_settings_page()
        
        # 默认显示经典爬取页面
        self.switch_page("classic")

        # 初始化尺寸标签和自定义输入框状态
        self.on_preset_change("基础尺寸")

    def create_classic_page(self):
        """创建经典爬取页面"""
        # 创建经典爬取页面框架
        self.classic_page = ctk.CTkFrame(self.content_frame)
        
        # 配置网格
        self.classic_page.grid_columnconfigure(0, weight=1)
        self.classic_page.grid_rowconfigure(0, weight=0)  # 搜索区域
        self.classic_page.grid_rowconfigure(1, weight=0)  # 状态栏
        self.classic_page.grid_rowconfigure(2, weight=1)  # 结果区域 - 可扩展
        self.classic_page.grid_rowconfigure(3, weight=0)  # 下载选项 - 固定在底部
        
        # 创建搜索框架
        self.search_frame = ctk.CTkFrame(self.classic_page)
        self.search_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.search_frame.grid_columnconfigure(0, weight=0) # Platform Radio Frame
        self.search_frame.grid_columnconfigure(2, weight=1) # Search Entry gets most weight
        self.search_frame.grid_columnconfigure(3, weight=0) # Search Button
        self.search_frame.grid_columnconfigure(4, weight=0) # Precise Switch

        # 平台选择
        self.platform_frame = ctk.CTkFrame(self.search_frame, fg_color="transparent")
        self.platform_frame.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")

        platform_label = ctk.CTkLabel(self.platform_frame, text="平台:")
        platform_label.pack(side="left", padx=(0, 5))

        ctk.CTkRadioButton(
            self.platform_frame, text="爱奇艺", variable=self.platform_var, value="爱奇艺",
            command=self.on_platform_change
        ).pack(side="left", padx=5)
        ctk.CTkRadioButton(
            self.platform_frame, text="腾讯视频", variable=self.platform_var, value="腾讯视频",
            command=self.on_platform_change
        ).pack(side="left", padx=5)
        ctk.CTkRadioButton(
            self.platform_frame, text="优酷视频", variable=self.platform_var, value="优酷视频",
            command=self.on_platform_change
        ).pack(side="left", padx=5)

        # 搜索关键词输入
        search_label = ctk.CTkLabel(self.search_frame, text="关键词:")
        search_label.grid(row=0, column=1, padx=(10, 5), pady=10, sticky="w")

        self.search_entry = ctk.CTkEntry(self.search_frame, placeholder_text="请输入搜索关键词（例如：嘟当曼）")
        self.search_entry.grid(row=0, column=2, padx=(0, 10), pady=10, sticky="ew")
        self.search_entry.bind("<Return>", lambda event: self.search())
        
        # 搜索按钮
        self.search_button = ctk.CTkButton(self.search_frame, text="搜索", command=self.search)
        self.search_button.grid(row=0, column=3, padx=10, pady=10)

        # 精确搜索开关
        self.precise_search_switch = ctk.CTkSwitch(
            self.search_frame,
            text="精确搜索",
            variable=self.precise_search_var
        )
        self.precise_search_switch.grid(row=0, column=4, padx=10, pady=10)
        
        # 状态标签
        self.status_label = ctk.CTkLabel(self.classic_page, text="准备就绪")
        self.status_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        
        # 结果框架（带滚动条）
        self.results_frame = ctk.CTkScrollableFrame(self.classic_page)
        self.results_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        self.results_frame.grid_columnconfigure(0, weight=4)  # Title column gets more weight
        self.results_frame.grid_columnconfigure(1, weight=1)  # Original preview column
        self.results_frame.grid_columnconfigure(2, weight=1)  # Vertical preview column
        self.results_frame.grid_columnconfigure(3, weight=1)  # Horizontal preview column

        # 下载选项框架
        self.download_frame = ctk.CTkFrame(self.classic_page)
        self.download_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")
        
        # 配置下载框架网格
        self.download_frame.grid_columnconfigure(0, weight=0) # Preset Label
        self.download_frame.grid_columnconfigure(1, weight=2) # Preset Combo
        self.download_frame.grid_columnconfigure(2, weight=0) # Labels
        self.download_frame.grid_columnconfigure(3, weight=2) # Entries
        self.download_frame.grid_columnconfigure(4, weight=1) # Radio Button Frame
        
        # 尺寸预设标签
        self.preset_label = ctk.CTkLabel(self.download_frame, text="尺寸预设:")
        self.preset_label.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w")
        
        # 尺寸预设下拉框
        preset_options = list(self.size_presets.keys())
        self.preset_combo = ctk.CTkComboBox(self.download_frame, values=preset_options, command=self.on_preset_change)
        self.preset_combo.grid(row=0, column=1, columnspan=3, padx=(0, 10), pady=5, sticky="ew")
        self.preset_combo.set("基础尺寸")
        
        # 图片类型选择
        self.orientation_frame = ctk.CTkFrame(self.download_frame)
        self.orientation_frame.grid(row=0, column=4, rowspan=2, padx=10, pady=5, sticky="ne")
        
        self.all_radio = ctk.CTkRadioButton(
            self.orientation_frame, text="全部下载", variable=self.orientation_var, value="全部",
            command=self.update_size_entries
        )
        self.all_radio.pack(anchor="w", pady=2)
        
        self.vertical_radio = ctk.CTkRadioButton(
            self.orientation_frame, text="仅下竖图", variable=self.orientation_var, value="竖图",
            command=self.update_size_entries
        )
        self.vertical_radio.pack(anchor="w", pady=2)
        
        self.horizontal_radio = ctk.CTkRadioButton(
            self.orientation_frame, text="仅下横图", variable=self.orientation_var, value="横图",
            command=self.update_size_entries
        )
        self.horizontal_radio.pack(anchor="w", pady=2)

        # 尺寸输入框
        self.v_size_label = ctk.CTkLabel(self.download_frame, text="竖图尺寸:")
        self.v_size_label.grid(row=1, column=0, padx=(10, 5), pady=5, sticky="w")
        self.v_size_entry = ctk.CTkEntry(self.download_frame, placeholder_text="宽度x高度")
        self.v_size_entry.grid(row=1, column=1, padx=(0, 10), pady=5, sticky="ew")

        self.h_size_label = ctk.CTkLabel(self.download_frame, text="横图尺寸:")
        self.h_size_label.grid(row=1, column=2, padx=(10, 5), pady=5, sticky="w")
        self.h_size_entry = ctk.CTkEntry(self.download_frame, placeholder_text="宽度x高度")
        self.h_size_entry.grid(row=1, column=3, padx=(0, 10), pady=5, sticky="ew")

        # 下载路径
        self.path_label = ctk.CTkLabel(self.download_frame, text="下载路径:")
        self.path_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.path_entry = ctk.CTkEntry(self.download_frame)
        # 使用加载的默认路径
        settings = self.load_settings()
        default_path_loaded = settings.get("default_path", os.path.join(os.path.expanduser("~"), "Desktop"))
        self.path_entry.insert(0, default_path_loaded)
        self.path_entry.grid(row=2, column=1, columnspan=3, padx=10, pady=10, sticky="ew")
        self.browse_button = ctk.CTkButton(self.download_frame, text="浏览", command=self.select_directory)
        self.browse_button.grid(row=2, column=4, padx=10, pady=10, sticky="e")

        # 下载按钮
        self.download_button = ctk.CTkButton(self.download_frame, text="下载选中项", command=self.download_selected)
        self.download_button.grid(row=3, column=0, columnspan=5, padx=10, pady=10, sticky="ew")

        # 存储页面
        self.pages["classic"] = self.classic_page

    def create_batch_page(self):
        """创建批量爬取页面"""
        self.batch_page = ctk.CTkFrame(self.content_frame)
        
        # 加载设置
        settings = self.load_settings()
        
        # === B1. 批量任务配置区 ===
        self.batch_config_frame = ctk.CTkFrame(self.batch_page)
        self.batch_config_frame.pack(fill="x", padx=10, pady=(10, 5))
        self.batch_config_frame.grid_columnconfigure(1, weight=2)  # Excel Entry列
        self.batch_config_frame.grid_columnconfigure(4, weight=1)  # Sheet Combo列
        
        # Excel文件和Sheet选择 - 一行显示
        excel_label = ctk.CTkLabel(self.batch_config_frame, text="Excel文件:")
        excel_label.grid(row=0, column=0, padx=(5, 5), pady=10, sticky="w")
        
        self.batch_excel_entry = ctk.CTkEntry(self.batch_config_frame, state="readonly")
        self.batch_excel_entry.grid(row=0, column=1, padx=(0, 5), pady=10, sticky="ew")
        
        self.batch_excel_button = ctk.CTkButton(
            self.batch_config_frame, 
            text="选择文件", 
            width=80,
            command=self.select_excel_file
        )
        self.batch_excel_button.grid(row=0, column=2, padx=(0, 15), pady=10)
        
        # Sheet选择
        sheet_label = ctk.CTkLabel(self.batch_config_frame, text="Sheet:")
        sheet_label.grid(row=0, column=3, padx=(0, 5), pady=10, sticky="w")
        
        self.batch_sheet_combo = ctk.CTkComboBox(
            self.batch_config_frame, 
            values=["请先选择Excel文件"],
            width=150,
            command=self.on_sheet_change
        )
        self.batch_sheet_combo.grid(row=0, column=4, padx=(0, 5), pady=10, sticky="ew")
        
        # === B2. 批量数据显示与反馈区 ===
        # 表格预览
        self.batch_table_frame = ctk.CTkFrame(self.batch_page)
        self.batch_table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # 表格标题栏
        self.table_title_frame = ctk.CTkFrame(self.batch_table_frame, fg_color="transparent")
        self.table_title_frame.pack(fill="x", padx=5, pady=5)
        
        self.table_title_label = ctk.CTkLabel(
            self.table_title_frame, 
            text="📊 批量爬取数据预览", 
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.table_title_label.pack(side="left")
        
        # 使用ttk.Treeview创建表格
        style = ttk.Style()
        style.theme_use("clam")  # 使用较现代的主题
        
        # 配置红色加粗标签用于标题不一致项
        style.configure("RedBold.Treeview", foreground="red", font=("TkDefaultFont", 9, "bold"))
        style.map("RedBold.Treeview", 
                 foreground=[('selected', 'white')])  # 选中时保持白色文字
        
        # 创建Treeview（支持动态列配置）
        self.batch_tree = ttk.Treeview(
            self.batch_table_frame, 
            show="headings",
            height=15
        )
        
        # 配置标题不一致的行样式（红色加粗）
        self.batch_tree.tag_configure("mismatch", foreground="red", font=("TkDefaultFont", 9, "bold"))
        
        # 初始化为批量爬取模式
        self.current_table_mode = "batch"
        self.setup_table_columns("batch")
        
        # 添加滚动条
        tree_scrollbar = ttk.Scrollbar(self.batch_table_frame, orient="vertical", command=self.batch_tree.yview)
        self.batch_tree.configure(yscrollcommand=tree_scrollbar.set)
        
        # 布局表格和滚动条
        self.batch_tree.pack(side="left", fill="both", expand=True)
        tree_scrollbar.pack(side="right", fill="y")
        
        # === 批量模式下载选项框架（采用原版UI样式）===
        self.batch_download_frame = ctk.CTkFrame(self.batch_page)
        self.batch_download_frame.pack(fill="x", padx=10, pady=(5, 5))
        
        # 配置下载框架网格（采用原版布局）
        self.batch_download_frame.grid_columnconfigure(0, weight=0) # Preset Label
        self.batch_download_frame.grid_columnconfigure(1, weight=2) # Preset Combo
        self.batch_download_frame.grid_columnconfigure(2, weight=0) # Labels
        self.batch_download_frame.grid_columnconfigure(3, weight=2) # Entries
        self.batch_download_frame.grid_columnconfigure(4, weight=1) # Radio Button Frame

        # 尺寸预设标签
        self.batch_preset_label = ctk.CTkLabel(self.batch_download_frame, text="尺寸预设:")
        self.batch_preset_label.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w")
        
        # 尺寸预设下拉框
        preset_options = list(self.size_presets.keys())
        self.batch_preset_combo = ctk.CTkComboBox(self.batch_download_frame, values=preset_options, command=self.on_batch_preset_change)
        self.batch_preset_combo.grid(row=0, column=1, columnspan=3, padx=(0, 10), pady=5, sticky="ew")
        self.batch_preset_combo.set(settings.get("batch_default_size", "基础尺寸"))
        
        # 图片类型选择 - 批量模式专用（采用原版样式）
        self.batch_orientation_var = ctk.StringVar(value="全部") # 批量模式默认为全部
        self.batch_orientation_frame = ctk.CTkFrame(self.batch_download_frame)
        self.batch_orientation_frame.grid(row=0, column=4, rowspan=2, padx=10, pady=5, sticky="ne")
        
        self.batch_all_radio = ctk.CTkRadioButton(
            self.batch_orientation_frame, text="全部下载", variable=self.batch_orientation_var, value="全部",
            command=self.update_batch_size_entries
        )
        self.batch_all_radio.pack(anchor="w", pady=2)
        
        self.batch_vertical_radio = ctk.CTkRadioButton(
            self.batch_orientation_frame, text="仅下竖图", variable=self.batch_orientation_var, value="竖图",
            command=self.update_batch_size_entries
        )
        self.batch_vertical_radio.pack(anchor="w", pady=2)
        
        self.batch_horizontal_radio = ctk.CTkRadioButton(
            self.batch_orientation_frame, text="仅下横图", variable=self.batch_orientation_var, value="横图",
            command=self.update_batch_size_entries
        )
        self.batch_horizontal_radio.pack(anchor="w", pady=2)

        # Size Entries（采用原版样式）
        self.batch_v_size_label = ctk.CTkLabel(self.batch_download_frame, text="竖图尺寸:")
        self.batch_v_size_label.grid(row=1, column=0, padx=(10, 5), pady=5, sticky="w")
        self.batch_v_size_entry = ctk.CTkEntry(self.batch_download_frame, placeholder_text="宽度x高度")
        self.batch_v_size_entry.grid(row=1, column=1, padx=(0, 10), pady=5, sticky="ew")

        self.batch_h_size_label = ctk.CTkLabel(self.batch_download_frame, text="横图尺寸:")
        self.batch_h_size_label.grid(row=1, column=2, padx=(10, 5), pady=5, sticky="w")
        self.batch_h_size_entry = ctk.CTkEntry(self.batch_download_frame, placeholder_text="宽度x高度")
        self.batch_h_size_entry.grid(row=1, column=3, padx=(0, 10), pady=5, sticky="ew")

        # Download Path Row - 横图路径
        self.batch_h_path_label = ctk.CTkLabel(self.batch_download_frame, text="横图路径:")
        self.batch_h_path_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.batch_h_path_entry = ctk.CTkEntry(self.batch_download_frame)
        self.batch_h_path_entry.insert(0, settings.get("batch_horizontal_path", r"D:\海报需求\第十一批\横图"))
        self.batch_h_path_entry.grid(row=2, column=1, columnspan=3, padx=10, pady=5, sticky="ew")
        self.batch_h_browse_button = ctk.CTkButton(self.batch_download_frame, text="浏览", command=lambda: self.browse_batch_path(self.batch_h_path_entry))
        self.batch_h_browse_button.grid(row=2, column=4, padx=10, pady=5, sticky="e")

        # Download Path Row - 竖图路径
        self.batch_v_path_label = ctk.CTkLabel(self.batch_download_frame, text="竖图路径:")
        self.batch_v_path_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.batch_v_path_entry = ctk.CTkEntry(self.batch_download_frame)
        self.batch_v_path_entry.insert(0, settings.get("batch_vertical_path", r"D:\海报需求\第十一批\竖图"))
        self.batch_v_path_entry.grid(row=3, column=1, columnspan=3, padx=10, pady=5, sticky="ew")
        self.batch_v_browse_button = ctk.CTkButton(self.batch_download_frame, text="浏览", command=lambda: self.browse_batch_path(self.batch_v_path_entry))
        self.batch_v_browse_button.grid(row=3, column=4, padx=10, pady=5, sticky="e")
        
        # 初始化尺寸显示 - 设置默认值
        default_preset = settings.get("batch_default_size", "基础尺寸")
        self.on_batch_preset_change(default_preset)
        
        # === 控制按钮区域 ===
        self.control_frame = ctk.CTkFrame(self.batch_page, fg_color="transparent")
        self.control_frame.pack(fill="x", padx=10, pady=(10, 5))
        
        # 进度条
        self.batch_progress_bar = ctk.CTkProgressBar(self.control_frame)
        self.batch_progress_bar.pack(fill="x", pady=(0, 5))
        self.batch_progress_bar.set(0)
        
        self.batch_progress_label = ctk.CTkLabel(self.control_frame, text="就绪")
        self.batch_progress_label.pack(pady=(0, 10))
        
        # 操作按钮（居中显示）
        button_frame = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        button_frame.pack(expand=True, fill="x")
        
        # 创建居中容器
        center_frame = ctk.CTkFrame(button_frame, fg_color="transparent")
        center_frame.pack(expand=True)
        
        self.batch_start_button = ctk.CTkButton(
            center_frame, 
            text="开始批量爬取", 
            width=120,
            state="disabled",
            command=self.start_batch_crawling
        )
        self.batch_start_button.pack(side="left", padx=(0, 10))
        
        self.batch_pause_button = ctk.CTkButton(
            center_frame, 
            text="暂停爬取", 
            width=120,
            state="disabled",
            command=self.toggle_batch_pause
        )
        self.batch_pause_button.pack(side="left")
        
        self.pages["batch"] = self.batch_page

    def create_delete_page(self):
        """创建批量删除错误海报页面"""
        self.delete_page = ctk.CTkFrame(self.content_frame)
        
        # 配置网格
        self.delete_page.grid_columnconfigure(0, weight=1)
        self.delete_page.grid_rowconfigure(0, weight=0)  # 文件选择区域
        self.delete_page.grid_rowconfigure(1, weight=1)  # 表格显示区域
        self.delete_page.grid_rowconfigure(2, weight=0)  # 路径选择区域
        self.delete_page.grid_rowconfigure(3, weight=0)  # 删除按钮区域
        
        # === 文件选择区域 ===
        self.delete_file_frame = ctk.CTkFrame(self.delete_page)
        self.delete_file_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.delete_file_frame.grid_columnconfigure(1, weight=2)  # Excel Entry列
        self.delete_file_frame.grid_columnconfigure(4, weight=1)  # Sheet Combo列
        
        # Excel文件和Sheet选择 - 一行显示
        excel_label = ctk.CTkLabel(self.delete_file_frame, text="Excel文件:")
        excel_label.grid(row=0, column=0, padx=(5, 5), pady=10, sticky="w")
        
        self.delete_excel_path_entry = ctk.CTkEntry(self.delete_file_frame, state="readonly")
        self.delete_excel_path_entry.grid(row=0, column=1, padx=(0, 5), pady=10, sticky="ew")
        
        self.delete_excel_browse_button = ctk.CTkButton(
            self.delete_file_frame, 
            text="选择文件", 
            width=80,
            command=self.browse_delete_excel_file
        )
        self.delete_excel_browse_button.grid(row=0, column=2, padx=(0, 15), pady=10)
        
        # Sheet选择
        sheet_label = ctk.CTkLabel(self.delete_file_frame, text="Sheet:")
        sheet_label.grid(row=0, column=3, padx=(0, 5), pady=10, sticky="w")
        
        self.delete_sheet_combo = ctk.CTkComboBox(
            self.delete_file_frame, 
            values=["请先选择Excel文件"],
            width=150,
            command=self.on_delete_sheet_change
        )
        self.delete_sheet_combo.grid(row=0, column=4, padx=(0, 5), pady=10, sticky="ew")
        
        # === 表格显示区域 ===
        self.delete_table_frame = ctk.CTkFrame(self.delete_page)
        self.delete_table_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        self.delete_table_frame.grid_columnconfigure(0, weight=1)
        self.delete_table_frame.grid_rowconfigure(0, weight=0)  # 列选择区域
        self.delete_table_frame.grid_rowconfigure(1, weight=1)  # 表格区域
        
        # === 删除标准列选择区域 ===
        self.delete_column_frame = ctk.CTkFrame(self.delete_table_frame)
        self.delete_column_frame.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        # 提示标签
        self.delete_key_label = ctk.CTkLabel(self.delete_column_frame, text="选择删除标准列:")
        self.delete_key_label.pack(side="left", padx=10, pady=5)
        
        # 创建表格 - 启用多选模式
        self.delete_table = ttk.Treeview(self.delete_table_frame, show="headings", selectmode="extended")
        self.delete_table.grid(row=1, column=0, sticky="nsew")
        
        # 绑定点击事件用于多选功能
        self.delete_table.bind("<Button-1>", self.on_delete_table_click)
        
        # 添加滚动条
        delete_scrollbar_y = ttk.Scrollbar(self.delete_table_frame, orient="vertical", command=self.delete_table.yview)
        delete_scrollbar_y.grid(row=1, column=1, sticky="ns")
        self.delete_table.configure(yscrollcommand=delete_scrollbar_y.set)
        
        delete_scrollbar_x = ttk.Scrollbar(self.delete_table_frame, orient="horizontal", command=self.delete_table.xview)
        delete_scrollbar_x.grid(row=2, column=0, sticky="ew")
        self.delete_table.configure(xscrollcommand=delete_scrollbar_x.set)
        
        # === 路径选择区域 ===
        self.delete_path_frame = ctk.CTkFrame(self.delete_page)
        self.delete_path_frame.grid(row=2, column=0, padx=10, pady=10, sticky="ew")
        
        # 配置路径框架网格
        self.delete_path_frame.grid_columnconfigure(0, weight=0) # Label
        self.delete_path_frame.grid_columnconfigure(1, weight=2) # Entry
        self.delete_path_frame.grid_columnconfigure(2, weight=0) # Browse Button
        self.delete_path_frame.grid_columnconfigure(3, weight=0) # Label
        self.delete_path_frame.grid_columnconfigure(4, weight=2) # Entry
        self.delete_path_frame.grid_columnconfigure(5, weight=0) # Browse Button
        
        # 横图路径选择
        self.delete_h_path_label = ctk.CTkLabel(self.delete_path_frame, text="横图路径:")
        self.delete_h_path_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.delete_h_path_entry = ctk.CTkEntry(self.delete_path_frame)
        settings = self.load_settings()
        self.delete_h_path_entry.insert(0, settings.get("delete_horizontal_path", r"D:\海报需求\第十一批\横图"))
        self.delete_h_path_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        self.delete_h_browse_button = ctk.CTkButton(
            self.delete_path_frame, 
            text="浏览", 
            width=60,
            command=lambda: self.browse_delete_path(self.delete_h_path_entry)
        )
        self.delete_h_browse_button.grid(row=0, column=2, padx=10, pady=5, sticky="e")
        
        # 竖图路径选择
        self.delete_v_path_label = ctk.CTkLabel(self.delete_path_frame, text="竖图路径:")
        self.delete_v_path_label.grid(row=0, column=3, padx=10, pady=5, sticky="w")
        self.delete_v_path_entry = ctk.CTkEntry(self.delete_path_frame)
        self.delete_v_path_entry.insert(0, settings.get("delete_vertical_path", r"D:\海报需求\第十一批\竖图"))
        self.delete_v_path_entry.grid(row=0, column=4, padx=10, pady=5, sticky="ew")
        self.delete_v_browse_button = ctk.CTkButton(
            self.delete_path_frame, 
            text="浏览", 
            width=60,
            command=lambda: self.browse_delete_path(self.delete_v_path_entry)
        )
        self.delete_v_browse_button.grid(row=0, column=5, padx=10, pady=5, sticky="e")
        
        # === 删除按钮区域 ===
        self.delete_button_frame = ctk.CTkFrame(self.delete_page)
        self.delete_button_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")
        
        # 状态标签
#        self.delete_status_label = ctk.CTkLabel(self.delete_button_frame, text="请选择Excel文件查看错误海报")
#        self.delete_status_label.pack(side="left", padx=10, pady=10)
        
        # 操作按钮（居中显示）
        button_frame = ctk.CTkFrame(self.delete_button_frame, fg_color="transparent")
        button_frame.pack(expand=True, fill="x")
        
        # 创建居中容器
        center_frame = ctk.CTkFrame(button_frame, fg_color="transparent")
        center_frame.pack(expand=True)
        
        # 全选/取消全选按钮
        self.select_all_delete_button = ctk.CTkButton(
            center_frame, 
            text="全选", 
            width=120,
            command=self.toggle_select_all_delete
        )
        self.select_all_delete_button.pack(side="left", padx=(0, 10))
        
        # 删除按钮
        self.delete_selected_button = ctk.CTkButton(
            center_frame, 
            text="开始删除", 
            width=120,
            command=self.delete_selected_files
        )
        self.delete_selected_button.pack(side="left")
        
        self.pages["delete"] = self.delete_page

    def create_settings_page(self):
        """创建设置页面"""
        self.settings_page = ctk.CTkFrame(self.content_frame)
        
        # 配置网格
        self.settings_page.grid_columnconfigure(0, weight=1)
        self.settings_page.grid_rowconfigure(0, weight=1)
        
        # 创建滚动框架
        self.settings_scroll_frame = ctk.CTkScrollableFrame(self.settings_page)
        self.settings_scroll_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.settings_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # 页面标题
        title_label = ctk.CTkLabel(self.settings_scroll_frame, text="默认设置", font=ctk.CTkFont(size=20, weight="bold"))
        title_label.grid(row=0, column=0, pady=(10, 20))
        
        # 创建选项卡视图
        self.settings_tabview = ctk.CTkTabview(self.settings_scroll_frame)
        self.settings_tabview.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        
        # 基础设置选项卡
        tab_basic = self.settings_tabview.add("基础设置")
        self.create_basic_settings_tab(tab_basic)
        
        # 批量设置选项卡
        tab_batch = self.settings_tabview.add("批量设置")
        self.create_batch_settings_tab(tab_batch)
        
        # Cookie设置选项卡
        tab_cookies = self.settings_tabview.add("Cookie设置")
        self.create_cookies_settings_tab(tab_cookies)
        
        # 保存按钮
        save_button = ctk.CTkButton(
            self.settings_scroll_frame, 
            text="保存设置", 
            command=self.save_page_settings,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        save_button.grid(row=2, column=0, pady=20)
        
        self.pages["settings"] = self.settings_page

    def switch_page(self, page):
        """切换页面"""
        # 隐藏所有页面
        for page_frame in self.pages.values():
            page_frame.grid_remove()
        
        # 显示选中页面
        if page in self.pages:
            self.pages[page].grid(row=0, column=0, sticky="nsew")
            self.current_page = page
            
            # 更新按钮状态
            self.update_sidebar_buttons(page)

    def update_sidebar_buttons(self, active_page):
        """更新侧边栏按钮状态"""
        # 重置所有按钮状态
        buttons = {
            "classic": self.classic_button,
            "batch": self.batch_button,
            "delete": self.delete_button,
            "settings": self.sidebar_settings_button
        }
        
        for page_name, button in buttons.items():
            if page_name == active_page:
                button.configure(fg_color=("gray75", "gray25"))  # 激活状态
            else:
                button.configure(fg_color=("gray90", "gray20"))  # 默认状态

    def on_preset_change(self, choice):
        """当尺寸预设变化时更新尺寸输入框状态和内容"""
        if choice == "自定义尺寸":
            # Enable entries
            self.v_size_entry.configure(state="normal")
            self.h_size_entry.configure(state="normal")
            # Set placeholders
            self.v_size_entry.delete(0, 'end')
            self.v_size_entry.insert(0, "例如: 500x700")
            self.h_size_entry.delete(0, 'end')
            self.h_size_entry.insert(0, "例如: 700x500")
        else:
            # Get preset dimensions
            vertical_size = self.size_presets[choice]["vertical"]
            horizontal_size = self.size_presets[choice]["horizontal"]
            
            # Format text
            v_text = f"{vertical_size[0]}x{vertical_size[1]}"
            h_text = f"{horizontal_size[0]}x{horizontal_size[1]}"

            # Update and disable vertical entry
            self.v_size_entry.configure(state='normal')
            self.v_size_entry.delete(0, 'end')
            self.v_size_entry.insert(0, v_text)
            self.v_size_entry.configure(state='disabled')

            # Update and disable horizontal entry
            self.h_size_entry.configure(state='normal')
            self.h_size_entry.delete(0, 'end')
            self.h_size_entry.insert(0, h_text)
            self.h_size_entry.configure(state='disabled')
    
    def update_size_entries(self):
        """No longer needed as on_preset_change handles everything"""
        pass
    
    def toggle_select_all(self, button):
        """切换全选/取消全选状态"""
        if self.select_all_state:  # 当前是全选状态
            for var in self.result_vars:
                var.set(True)
            button.configure(text="取消全选")
            self.select_all_state = False
        else:  # 当前是取消全选状态
            for var in self.result_vars:
                var.set(False)
            button.configure(text="全选")
            self.select_all_state = True
    
    def get_sized_url(self, original_url, width, height, img_type=""): # Add img_type
        """根据平台和图片类型获取调整尺寸的 URL (尝试构建，不保证有效)"""
        if not original_url or width <= 0 or height <= 0:
             return original_url

        platform = self.selected_platform

        if platform == "爱奇艺":
            # 爱奇艺: _width_height suffix (横竖图都适用)
            return self.build_iqiyi_sized_url(original_url, width, height) # Renamed original function
        elif platform == "腾讯视频":
            # 腾讯视频: 始终返回原始 URL (智能缩放裁剪处理)
            return original_url
        elif platform == "优酷视频":
            if img_type == "竖图":
                # 优酷竖图: 尝试构建 OSS URL
                return self.build_youku_vertical_sized_url(original_url, width, height)
            else: # 优酷横图
                # 优酷横图: 始终返回原始 URL (智能缩放裁剪处理)
                return original_url
        else: # Unknown platform
            return original_url

    def build_iqiyi_sized_url(self, original_url, width, height):
        """构建爱奇艺带尺寸参数的URL (_width_height suffix)"""
        if not original_url or not width or not height:
            return original_url
        try:
            last_dot_index = original_url.rfind('.')
            query_start_index = original_url.rfind('?')
            # Handle cases with or without query parameters
            end_index = query_start_index if query_start_index != -1 else len(original_url)
            if last_dot_index == -1 or last_dot_index < original_url.rfind('/'):
                base_url = original_url[:end_index]
                extension = ""
            else:
                 # Ensure the dot is before any query params
                 if query_start_index != -1 and last_dot_index > query_start_index:
                      # Dot is part of query params, treat as no extension
                      base_url = original_url[:end_index]
                      extension = "" # No extension part
                 else:
                      base_url = original_url[:last_dot_index]
                      extension = original_url[last_dot_index:end_index] # Include dot
                 # Remove potential old suffix before adding new one
                 base_url = re.sub(r'_\d+_\d+$', '', base_url)
                 # Add query params back if they existed
                 query_string = original_url[query_start_index:] if query_start_index != -1 else ""
            
            sized_url = f"{base_url}_{width}_{height}{extension}{query_string}"
            return sized_url
        except Exception as e:
            print(f"构造爱奇艺尺寸URL出错: {e}")
            return original_url
        
    def search(self):
        """根据选择的平台执行搜索并获取结果"""
        search_term = self.search_entry.get().strip()
        if not search_term:
            messagebox.showerror("错误", "请输入搜索关键词")
            return
            
        platform = self.platform_var.get() # Read from platform_var
        self.status_label.configure(text=f"正在 {platform} 中搜索 '{search_term}'...")
        self.update()
        
        # 清除之前的结果
        self.clear_results()
        
        # 根据选择的平台调用相应的搜索函数
        try:
            if platform == "爱奇艺":
                results_list = self.search_iqiyi(search_term)
            elif platform == "腾讯视频":
                results_list = self.search_tencent(search_term)
            elif platform == "优酷视频":
                results_list = self.search_youku(search_term)
            else:
                self.status_label.configure(text=f"不支持的平台: {platform}")
                return

            # 更新结果
            self.results = results_list

            if not self.results:
                self.status_label.configure(text=f"在{platform}中未找到结果")
                return

            # 更新GUI显示
            self.update_results_list()

        except Exception as e:
            print(f"搜索错误: {e}")
            print(traceback.format_exc())
            self.status_label.configure(text=f"搜索出错: {str(e)}")
            messagebox.showerror("错误", f"搜索过程中发生错误: {str(e)}")

    def search_iqiyi(self, search_term):
        """执行爱奇艺搜索"""
        results_list = []
        try:
            # 构建API URL
            encoded_term = urllib.parse.quote(search_term)
            api_url = f"https://mesh.if.iqiyi.com/portal/lw/search/homePageV3?key={encoded_term}&current_page=1&pageNum=1&pageSize=25"
            
            print(f"爱奇艺 API URL: {api_url}")
            
            # 发送请求
            response = requests.get(api_url, headers=self.iqiyi_headers, timeout=20)
            print(f"爱奇艺API响应状态码: {response.status_code}")
            
            if response.status_code == 200:
                # 解析JSON响应
                data_dict = response.json()
                print("爱奇艺API响应JSON已成功解析")
                    
                # 保存调试文件
                try:
                    with open("debug_iqiyi_response.json", "w", encoding="utf-8") as f:
                        json.dump(data_dict, f, ensure_ascii=False, indent=4)
                    print("完整的爱奇艺API JSON响应已保存到debug_iqiyi_response.json")
                except Exception as e_save:
                    print(f"警告：保存爱奇艺JSON响应失败: {e_save}")

                # 检查搜索模式
                if self.precise_search_var.get():
                    # 处理精确搜索结果
                    print("处理爱奇艺精确搜索结果...")
                    try:
                        # 获取templates列表
                        templates = data_dict.get('data', {}).get('templates', [])
                        if templates:
                            # 获取albumInfo对象
                            album_info = templates[0].get('albumInfo')
                            if album_info:
                                # 提取title和img
                                title = album_info.get('title')
                                img_url = album_info.get('img') or album_info.get('imgH')

                                if title and img_url:
                                    # 清理URL
                                    if img_url.startswith('//'): img_url = 'https:' + img_url
                                    results_list.append((title, img_url))
                                    print(f"  提取到爱奇艺精确结果: Title='{title}', ImgURL='{img_url}'")
                    except Exception as e:
                        print(f"错误：提取爱奇艺精确搜索结果时出错: {e}")
                        print(traceback.format_exc())
                else:
                    # 处理宽泛搜索结果
                    print("处理爱奇艺宽泛搜索结果...")
                    try:
                        templates = data_dict.get('data', {}).get('templates', [])
                        if templates:
                            # 获取intentAlbumInfos列表
                            intent_album_infos = templates[0].get('intentAlbumInfos', [])
                            print(f"找到 {len(intent_album_infos)} 个intentAlbumInfos项目")
                        
                            # 遍历列表提取信息
                            for item in intent_album_infos:
                                title = item.get('title')
                                img_url = item.get('img')

                                if title and img_url:
                                    # 清理URL
                                    if img_url.startswith('//'): img_url = 'https:' + img_url
                                    results_list.append((title, img_url))
                                    # 打印前几个结果用于验证
                                    if len(results_list) <= 5:
                                        print(f"  提取到爱奇艺结果: Title='{title}', ImgURL='{img_url}'")
                    except Exception as e:
                        print(f"错误：提取爱奇艺宽泛搜索结果时出错: {e}")
                        print(traceback.format_exc())
            else:
                print(f"爱奇艺API请求失败，状态码: {response.status_code}")
                print(f"响应内容: {response.text[:500]}...")
                raise requests.RequestException(f"爱奇艺API请求失败，状态码: {response.status_code}，可能是Cookie过期")

        except Exception as e:
            print(f"爱奇艺搜索出错: {e}")
            print(traceback.format_exc())
            raise

        print(f"爱奇艺搜索完成，共找到 {len(results_list)} 个结果")
        return results_list
    
    def search_tencent(self, search_term):
        """执行腾讯视频搜索"""
        results_list = []
        try:
            # 构建API URL
            api_url = "https://pbaccess.video.qq.com/trpc.videosearch.mobile_search.MultiTerminalSearch/MbSearch?vplatform=2"
            
            # 构建POST请求的JSON负载
            payload = {
                "version": "25031901",
                "clientType": 1,
                "filterValue": "",
                "uuid": "75D75495-4CF1-4C67-9F10-B0B313C1C999",  # 可能需要生成或使用通用值
                "retry": 0,
                "query": search_term,
                "pagenum": 0,  # 第一页
                "isPrefetch": True,
                "pagesize": 30,
                "queryFrom": 0,
                "searchDatakey": "",
                "transInfo": "",
                "isneedQc": True,
                "preQid": "",
                "adClientInfo": "",
                "extraInfo": {
                    "isNewMarkLabel": "1",
                    "multi_terminal_pc": "1",
                    "themeType": "1",
                    "sugRelatedIds": "{}",
                    "appVersion": ""
                }
            }
            
            print(f"腾讯视频 API URL: {api_url}")
            print(f"请求体: {json.dumps(payload, ensure_ascii=False)}")
            
            # 创建一个干净的headers字典，只保留必要的ASCII字段
            clean_headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36",
                "Content-Type": "application/json",
                "Referer": "https://v.qq.com/",
                "Origin": "https://v.qq.com"
            }
            
            # 如果有Cookie，确保它是ASCII兼容的
            if self.tencent_headers.get("Cookie") and self.tencent_headers["Cookie"] != "":
                clean_headers["Cookie"] = self.tencent_headers["Cookie"]
            
            # 发送POST请求
            response = requests.post(api_url, headers=clean_headers, json=payload, timeout=20)
            print(f"腾讯视频API响应状态码: {response.status_code}")
            
            if response.status_code == 200:
                # 解析JSON响应
                data_dict = response.json()
                print("腾讯视频API响应JSON已成功解析")
                
                # 保存调试文件
                try:
                    with open("debug_tencent_response.json", "w", encoding="utf-8") as f:
                        json.dump(data_dict, f, ensure_ascii=False, indent=4)
                    print("完整的腾讯视频API JSON响应已保存到debug_tencent_response.json")
                except Exception as e_save:
                    print(f"警告：保存腾讯视频JSON响应失败: {e_save}")
                
                # 检查API返回是否包含错误
                ret_code = data_dict.get('ret')
                if ret_code != 0 and ret_code is not None:
                    print(f"错误：腾讯API返回错误: ret={ret_code}, msg='{data_dict.get('msg')}'")
                    print(f"--- 发送的 Payload ---:\n{json.dumps(payload, indent=2, ensure_ascii=False)}")
                    print(f"--- 发送的 Headers ---:")
                    for key, value in clean_headers.items():
                        if key.lower() == 'cookie':
                            print(f"  {key}: [Cookie值已隐藏]")
                        else:
                            print(f"  {key}: {value}")
                    
                    # 尝试简化的payload
                    print("\n尝试简化的payload...")
                    simple_payload = {"query": search_term}
                    
                    try:
                        simple_response = requests.post(api_url, headers=clean_headers, json=simple_payload, timeout=20)
                        if simple_response.status_code == 200:
                            simple_data = simple_response.json()
                            simple_ret = simple_data.get('ret')
                            if simple_ret == 0:
                                print("简化请求成功！使用简化请求的结果")
                                data_dict = simple_data
                            else:
                                print("简化请求也失败，可能需要更多调试")
                                return []
                        else:
                            return []
                    except Exception:
                        return []
                
                # 使用正确的JSON路径提取数据
                print("处理腾讯视频搜索结果...")
                full_results_list = []
                try:
                    # 1. 安全地获取 areaBoxList (它是一个列表)
                    area_box_list = data_dict.get('data', {}).get('areaBoxList', [])

                    if area_box_list:
                        # 2. 遍历 areaBoxList 中的每个区域 (通常第一个区域最相关，但也可能需要检查其他区域)
                        #    我们先只处理第一个区域 areaBoxList[0]
                        if len(area_box_list) > 0:
                            item_list_container = area_box_list[0]  # 或者根据 type 等字段选择正确的 areaBox
                            # 3. 安全地获取该区域的 itemList (它也是一个列表)
                            item_list = item_list_container.get('itemList', [])

                            print(f"找到 {len(item_list)} 个 itemList 项目在第一个 areaBox 中。")

                            # 4. 遍历 itemList 中的每个项目
                            for item in item_list:
                                # 5. 安全地获取 videoInfo 对象
                                video_info = item.get('videoInfo')
                                if video_info:
                                    # 6. 从 videoInfo 中提取 title 和 imgUrl
                                    title = video_info.get('title')
                                    img_url = video_info.get('imgUrl')  # 注意键名是 imgUrl
                            
                                    if title and img_url:
                                        # 清理或补全 img_url
                                        if img_url.startswith('//'):
                                            img_url = 'https:' + img_url
                                        full_results_list.append((title, img_url))
                                        # 打印前几个提取结果用于验证
                                        if len(full_results_list) <= 5:
                                            print(f"  提取到: Title='{title}', ImgURL='{img_url}'")
                                    else:
                                        print(f"  警告：在 videoInfo 中未能找到 title 或 imgUrl")
                                else:
                                    print("  警告：itemList 中的项目缺少 videoInfo 对象")
                            else:
                                print("警告：areaBoxList 为空")
                    else:
                        print("警告：在 JSON 响应中未找到 'data' -> 'areaBoxList'")

                    # 根据精确搜索开关决定是否过滤结果
                    if self.precise_search_var.get():
                        print("执行腾讯视频精确匹配过滤...")
                        search_key = search_term
                        # 调用排序和过滤函数
                        results_list = self.filter_tencent_results(full_results_list, search_key)
                        print(f"精确匹配后剩余 {len(results_list)} 个结果。")
                    else:
                        # 宽泛搜索，直接使用全部结果
                        results_list = full_results_list

                except (AttributeError, IndexError, TypeError, KeyError) as e:
                    print(f"错误：解析腾讯视频 JSON 时发生异常: {e}")
                    print(traceback.format_exc())  # 打印详细错误信息
            
            else:
                print(f"腾讯视频API请求失败，状态码: {response.status_code}")
                print(f"响应内容: {response.text[:500]}...")
                raise requests.RequestException(f"腾讯视频API请求失败，状态码: {response.status_code}")
        
        except Exception as e:
            print(f"腾讯视频搜索出错: {e}")
            print(traceback.format_exc())
            raise
        
        print(f"腾讯视频搜索完成，共找到 {len(results_list)} 个结果")
        return results_list
    
    def normalize_text(self, text):
        """标准化文本，用于匹配比较"""
        # 转小写，替换罗马数字，移除空格
        text = text.lower()
        text = text.replace('ii', '2').replace('iii', '3').replace('iv', '4').replace('v', '5')\
               .replace('vi', '6').replace('vii', '7').replace('viii', '8')
        text = text.replace('第一季', '1').replace('第二季', '2').replace('第三季', '3')\
               .replace('第四季', '4').replace('第五季', '5').replace('第六季', '6')
        text = re.sub(r'\s+', '', text)  # 移除空格
        return text

    def calculate_match_score(self, title, search_key):
        """计算匹配分数，分数越低表示匹配度越高"""
        norm_title = self.normalize_text(title)
        norm_key = self.normalize_text(search_key)

        # 完全匹配
        if title == search_key:
            return 0
        
        # 标准化后匹配
        if norm_title == norm_key:
            return 1
        
        # 第一季检查
        is_base_key = not any(s in norm_key for s in ['1', '2', '3', '4', '5', '6', '第一季', '第二季'])
        is_title_s1 = norm_title == norm_key + '1' or norm_title == norm_key + '第一季' or norm_title == norm_key
        if is_base_key and is_title_s1:
            return 2
        
        # 相似度分数（相似度越高越好，转换为越低的分数）
        ratio = difflib.SequenceMatcher(None, norm_key, norm_title).ratio()
        return 3 + (1 - ratio)  # 为较高相似度给较低分数

    def filter_tencent_results(self, results, search_key):
        """过滤并排序腾讯视频结果，返回前3个最匹配的结果"""
        if not results or not search_key:
            return results[:3] if not search_key else results

        scored_results = []
        for title, img_url in results:
            # Use the complex scoring function specific to Tencent
            score = self.calculate_match_score(title, search_key)
            scored_results.append((score, title, img_url))

        # Sort by score ascending (lower score is better for calculate_match_score)
        scored_results.sort(key=lambda x: x[0])

        # Return top 3 results as (title, img_url) tuples
        return [(title, img_url) for score, title, img_url in scored_results[:3]]
    
    def clear_results(self):
        """清除所有结果控件"""
        # 清除框架中的所有控件
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        # 重置结果存储列表
        self.results = []
        self.result_checkboxes = []
        self.result_vars = []
        
        # 重置全选
        self.select_all_state = True
    
    def update_results_list(self):
        """更新GUI显示搜索结果，显示原图、竖图和横图预览"""
        print(f"开始更新GUI显示 {len(self.results)} 个结果...")
        
        if not self.results:
            self.status_label.configure(text="没有找到结果")
            return
        
        # Clear previous results widgets
        self.clear_results_widgets_only()
        
        # 创建表头框架
        header_frame = ctk.CTkFrame(self.results_frame, fg_color=("gray85", "gray25"))
        header_frame.grid(row=0, column=0, columnspan=4, sticky="ew", padx=5, pady=(0, 5))
        # Use the same column configuration as the results_frame for alignment
        header_frame.grid_columnconfigure(0, weight=self.results_frame.grid_columnconfigure(0).get('weight', 4))
        header_frame.grid_columnconfigure(1, weight=self.results_frame.grid_columnconfigure(1).get('weight', 1))
        header_frame.grid_columnconfigure(2, weight=self.results_frame.grid_columnconfigure(2).get('weight', 1))
        header_frame.grid_columnconfigure(3, weight=self.results_frame.grid_columnconfigure(3).get('weight', 1))

        select_all_button = ctk.CTkButton(header_frame, text="全选", width=70, height=25, command=lambda: self.toggle_select_all(select_all_button))
        select_all_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        title_label = ctk.CTkLabel(header_frame, text="标题", font=ctk.CTkFont(weight="bold"))
        title_label.grid(row=0, column=0, padx=(80, 5), pady=5, sticky="w") # Keep padding for select_all

        # --- Center Header Labels (Removed sticky) ---
        orig_preview_header = ctk.CTkLabel(header_frame, text="原图预览", font=ctk.CTkFont(weight="bold"), anchor='center') # Keep anchor
        orig_preview_header.grid(row=0, column=1, padx=5, pady=5) # REMOVED sticky='ew'
        vert_preview_header = ctk.CTkLabel(header_frame, text="竖图预览", font=ctk.CTkFont(weight="bold"), anchor='center') # Keep anchor
        vert_preview_header.grid(row=0, column=2, padx=5, pady=5) # REMOVED sticky='ew'
        horz_preview_header = ctk.CTkLabel(header_frame, text="横图预览", font=ctk.CTkFont(weight="bold"), anchor='center') # Keep anchor
        horz_preview_header.grid(row=0, column=3, padx=5, pady=5) # REMOVED sticky='ew'
        # --- End Header Label Centering ---

        # Reset internal lists before populating
        self.result_checkboxes = []
        self.result_vars = []
        preview_items = []
        
        # 开始逐项添加结果到GUI
        for idx, result_item in enumerate(self.results):
            title = ""
            img_url = "" # For platforms returning only one URL initially
            horz_url = None # For Youku
            vert_url = None # For Youku

            # Unpack based on expected structure for the current platform
            if self.selected_platform == "优酷视频":
                if len(result_item) == 3:
                    title, horz_url, vert_url = result_item
                else:
                     print(f"警告: 优酷结果项格式不符: {result_item}")
                     continue # Skip malformed item
            else: # 爱奇艺, 腾讯视频 (assume (title, img_url))
                if len(result_item) == 2:
                     title, img_url = result_item
                     horz_url = img_url # Treat the single URL as horizontal for preview logic
                     vert_url = img_url # And also vertical
                else:
                     print(f"警告: {self.selected_platform} 结果项格式不符: {result_item}")
                     continue # Skip malformed item


            result_frame = ctk.CTkFrame(self.results_frame)
            result_frame.grid(row=idx + 1, column=0, columnspan=4, sticky="ew", padx=5, pady=2)
            # Use the same column configuration as the results_frame for alignment
            result_frame.grid_columnconfigure(0, weight=self.results_frame.grid_columnconfigure(0).get('weight', 4))
            result_frame.grid_columnconfigure(1, weight=self.results_frame.grid_columnconfigure(1).get('weight', 1))
            result_frame.grid_columnconfigure(2, weight=self.results_frame.grid_columnconfigure(2).get('weight', 1))
            result_frame.grid_columnconfigure(3, weight=self.results_frame.grid_columnconfigure(3).get('weight', 1))

            var = ctk.BooleanVar()
            self.result_vars.append(var)
            
            checkbox = ctk.CTkCheckBox(result_frame, text=title, variable=var)
            checkbox.grid(row=0, column=0, padx=5, pady=5, sticky="w")
            self.result_checkboxes.append(checkbox)
            
            # Determine URL for original preview (use horz_url if available, else img_url)
            orig_preview_display_url = horz_url if horz_url else img_url

            preview_orig = ctk.CTkLabel(result_frame, text="加载中...", height=60, width=100)
            preview_orig.grid(row=0, column=1, padx=5, pady=5, sticky="w")
            # Pass the correct URL for zoom based on platform
            preview_orig.bind("<Button-1>", lambda event, h_url=horz_url, v_url=vert_url, i_url=img_url, t=title, p_type="原图":
                              self.show_zoomed_image(h_url if self.selected_platform == '优酷视频' else i_url,
                                                    v_url if self.selected_platform == '优酷视频' else i_url,
                                                    p_type, t))
            
            preview_vert = ctk.CTkLabel(result_frame, text="加载中...", height=60, width=100)
            preview_vert.grid(row=0, column=2, padx=5, pady=5, sticky="w")
            # Pass the correct URL for zoom based on platform
            preview_vert.bind("<Button-1>", lambda event, h_url=horz_url, v_url=vert_url, i_url=img_url, t=title, p_type="竖图":
                              self.show_zoomed_image(h_url if self.selected_platform == '优酷视频' else i_url,
                                                    v_url if self.selected_platform == '优酷视频' else i_url,
                                                    p_type, t))
            
            preview_horz = ctk.CTkLabel(result_frame, text="加载中...", height=60, width=100)
            preview_horz.grid(row=0, column=3, padx=5, pady=5, sticky="w")
            # --- Conditionally bind click event for horizontal preview ---
            # Only bind if NOT Tencent video
            if self.selected_platform != "腾讯视频":
                preview_horz.bind("<Button-1>", lambda event, h_url=horz_url, v_url=vert_url, i_url=img_url, t=title, p_type="横图":
                                  self.show_zoomed_image(h_url if self.selected_platform == '优酷视频' else i_url,
                                                        v_url if self.selected_platform == '优酷视频' else i_url,
                                                        p_type, t))
            else:
                # For Tencent, explicitly set text to '无' if needed, or let load_preview handle it
                pass # load_preview_image will handle setting text to "无"


            # Store data needed for async loading (pass platform-specific URLs)
            preview_items.append({
                "horz_url": horz_url, # Could be None for non-Youku
                "vert_url": vert_url, # Could be None for non-Youku or if fetch failed
                "img_url": img_url,   # For non-Youku platforms
                "orig_label": preview_orig,
                "vert_label": preview_vert,
                "horz_label": preview_horz
            })
        
        self.status_label.configure(text=f"找到 {len(self.results)} 个结果")
        print("GUI结果显示更新完成。")
        
        # Start background thread to load previews
        threading.Thread(target=self.load_all_previews, args=(preview_items,), daemon=True).start()
    
    def load_all_previews(self, preview_items):
        """异步加载所有预览图。
        - 原图和竖图: 下载原图URL, 本地缩放。
        - 横图: 爱奇艺尝试服务器小图，优酷本地缩放原图，腾讯显示'无'
        """
        print(f"开始加载 {len(preview_items)} 行预览图...")
        # Define the small preview size for iQiyi horizontal attempt
        preview_h_width = 150
        preview_h_height = 90

        for item in preview_items:
            platform = self.selected_platform
            orig_display_url = item["horz_url"] if platform == "优酷视频" else item["img_url"]
            vert_display_url = item["vert_url"] if platform == "优酷视频" else item["img_url"]
            base_horz_url = item["horz_url"] if platform == "优酷视频" else item["img_url"] # Base URL for horz

            # --- Load Original Preview ---
            if orig_display_url:
                 self.load_preview_image(orig_display_url, item["orig_label"], "原图")
            else:
                 self.after(0, lambda lw=item["orig_label"]: lw.configure(text="无图", image=None))

            # --- Load Vertical Preview ---
            if vert_display_url:
                 self.load_preview_image(vert_display_url, item["vert_label"], "竖图")
            else:
                 self.after(0, lambda lw=item["vert_label"]: lw.configure(text="无", image=None))

            # --- Load Horizontal Preview ---
            url_to_load_for_horz = None
            if platform == "爱奇艺" and base_horz_url:
                # --- Build the sized URL specifically for iQiyi horizontal preview ---
                url_to_load_for_horz = self.build_iqiyi_sized_url(base_horz_url, preview_h_width, preview_h_height)
                print(f"  iQiyi Horz Preview: Trying URL: {url_to_load_for_horz[:80]}...")
            elif platform == "优酷视频" and base_horz_url:
                 # Youku uses the original horizontal URL; load_preview_image will scale it
                 url_to_load_for_horz = base_horz_url
            elif platform == "腾讯视频":
                 # Tencent handled directly in load_preview_image to show "无"
                 url_to_load_for_horz = None # Pass None

            # Call load_preview_image with the determined URL
            self.load_preview_image(url_to_load_for_horz, item["horz_label"], "横图")


    def load_preview_image(self, preview_url_to_download, label_widget, img_type):
        """加载单个预览图并更新标签.
        - Scales images proportionally based on TARGET constants.
        - Displays iQiyi horizontal previews using their downloaded small size.
        - Handles Tencent horizontal preview ("无").
        """
        if not label_widget.winfo_exists():
             return

        platform = self.selected_platform

        # Handle Tencent Horizontal Preview explicitly
        if platform == "腾讯视频" and img_type == "横图":
            if label_widget.winfo_exists():
                self.after(0, lambda lw=label_widget: lw.configure(text="无", image=None))
            return

        # Handle cases where no URL is provided
        if not preview_url_to_download:
             if label_widget.winfo_exists():
                 text_to_show = "无URL" # Generic default
                 if platform == "爱奇艺" and img_type == "横图": text_to_show = "无横图URL"
                 elif platform == "优酷视频": text_to_show = f"无{img_type}URL" # Covers both horz/vert
                 self.after(0, lambda lw=label_widget, txt=text_to_show: lw.configure(text=txt, image=None))
             return

        try:
            headers_to_use = self.iqiyi_headers # Default
            if platform == "腾讯视频": headers_to_use = self.tencent_headers
            elif platform == "优酷视频": headers_to_use = self.youku_headers

            response = requests.get(preview_url_to_download, headers=headers_to_use, timeout=10)

            # Handle Download Failure
            if response.status_code != 200:
                error_text = f"失败:{response.status_code}"
                if platform == "爱奇艺" and img_type == "横图": # Specific check for iQiyi sized URL failure
                     error_text = "尺寸无效" if response.status_code == 404 else f"横图:{response.status_code}"

                if label_widget.winfo_exists():
                    self.after(0, lambda lw=label_widget, et=error_text: lw.configure(text=et, image=None))
                # --- DO NOT FALLBACK TO ORIGINAL URL FOR IQIYI HORIZONTAL PREVIEW ---
                return
                
            img_data_bytes = response.content
            if not img_data_bytes:
                error_text = "空数据"
                if label_widget.winfo_exists():
                    self.after(0, lambda lw=label_widget, et=error_text: lw.configure(text=et, image=None))
                return

            # Open Image
            img_data = Image.open(BytesIO(img_data_bytes))
            original_width, original_height = img_data.size

            # --- Determine Display Size and Resize ---
            display_width = original_width
            display_height = original_height
            img_to_display = img_data
            needs_resizing = True # Assume resizing is needed unless it's iQiyi horz

            is_iqiyi_horizontal_preview = (platform == "爱奇艺" and img_type == "横图")

            if is_iqiyi_horizontal_preview:
                # Use downloaded dimensions directly, no further scaling
                display_width = max(1, original_width)
                display_height = max(1, original_height)
                img_to_display = img_data
                needs_resizing = False # Already the size we want
                # print(f"  Preview (iQiyi Horz): Using downloaded size {display_width}x{display_height}")
            else:
                # --- New Scaling Logic for all other previews ---
                if original_height == 0 or original_width == 0: aspect_ratio = 1
                else: aspect_ratio = original_width / original_height

                if img_type == "竖图" or img_type == "原图":
                    target_h = TARGET_PREVIEW_V_HEIGHT
                    display_height = target_h
                    display_width = max(1, int(aspect_ratio * display_height))
                elif img_type == "横图": # This now only applies to Youku horizontal
                    target_w = TARGET_PREVIEW_H_WIDTH
                    display_width = target_w
                    display_height = max(1, int(display_width / aspect_ratio)) if aspect_ratio != 0 else 1

                # Only resize if calculated dimensions differ from original
                if display_width != original_width or display_height != original_height:
                    # print(f"  Preview ('{img_type}', {platform}): Resizing to {display_width}x{display_height}") # Debug log
                    try:
                        img_to_display = img_data.resize((display_width, display_height), Image.Resampling.LANCZOS)
                    except AttributeError:
                        img_to_display = img_data.resize((display_width, display_height), Image.LANCZOS)
                else:
                    needs_resizing = False # No resize needed
                    img_to_display = img_data

            # --- Create CTkImage and Update Label ---
            ctk_img = ctk.CTkImage(light_image=img_to_display, dark_image=img_to_display, size=(display_width, display_height))

            if label_widget.winfo_exists():
                 # Label size configured automatically by CTkImage, ensure text is cleared
                 self.after(0, lambda lw=label_widget, img=ctk_img: lw.configure(image=img, text=""))

        except requests.exceptions.Timeout:
            error_text = "超时"
            if label_widget.winfo_exists():
                 self.after(0, lambda lw=label_widget, et=error_text: lw.configure(text=et, image=None))
        except Exception as e:
            error_text = "加载失败"
            print(f"Error loading preview ({platform}, {img_type}, URL: {preview_url_to_download}): {e}")
            if label_widget.winfo_exists():
                 self.after(0, lambda lw=label_widget, et=error_text: lw.configure(text=et, image=None))
            
    def select_directory(self):
        """选择下载目录"""
        directory = filedialog.askdirectory()
        if directory:
            self.path_entry.delete(0, "end")
            self.path_entry.insert(0, directory)
            
    def sanitize_filename(self, filename):
        """清理文件名，移除不允许的字符"""
        return re.sub(r'[\\/*?:"<>|]', "_", filename)
    
    def smart_resize_and_crop(self, img, target_width, target_height):
        """智能缩放裁剪：先根据基准边等比缩放，再居中裁剪"""
        try:
            original_width, original_height = img.size
            print(f"    原始尺寸: {original_width}x{original_height}")
            
            # 计算两个方向的缩放比例
            width_scale = target_width / original_width
            height_scale = target_height / original_height
            
            # 找到基准边（需要修改幅度最小的那个，即缩放比例更大的那个）
            # 使用较大的缩放比例确保缩放后的图片在两个方向都不小于目标尺寸
            scale_ratio = max(width_scale, height_scale)
            
            print(f"    宽度缩放比: {width_scale:.3f}, 高度缩放比: {height_scale:.3f}")
            print(f"    选择缩放比: {scale_ratio:.3f} ({'基于宽度' if scale_ratio == width_scale else '基于高度'})")
            
            # 等比缩放
            new_width = int(original_width * scale_ratio)
            new_height = int(original_height * scale_ratio)
            print(f"    等比缩放后尺寸: {new_width}x{new_height}")
            
            try:
                scaled_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            except AttributeError:
                scaled_img = img.resize((new_width, new_height), Image.LANCZOS)
            
            # 计算居中裁剪区域
            left = (new_width - target_width) // 2
            top = (new_height - target_height) // 2
            right = left + target_width
            bottom = top + target_height
            
            print(f"    裁剪区域: ({left}, {top}, {right}, {bottom})")
            
            # 确保裁剪区域在图片范围内
            left = max(0, left)
            top = max(0, top)
            right = min(new_width, right)
            bottom = min(new_height, bottom)
            
            # 居中裁剪
            cropped_img = scaled_img.crop((left, top, right, bottom))
            
            final_width, final_height = cropped_img.size
            print(f"    最终尺寸: {final_width}x{final_height}")
            
            return cropped_img
            
        except Exception as e:
            print(f"    智能缩放裁剪失败: {e}")
            print("    回退到简单缩放")
            # 回退到简单缩放
            try:
                return img.resize((target_width, target_height), Image.Resampling.LANCZOS)
            except AttributeError:
                return img.resize((target_width, target_height), Image.LANCZOS)
            
    def download_selected(self):
        """下载选中的图片"""
        selected_indices = [i for i, var in enumerate(self.result_vars) if var.get()]
        
        if not selected_indices:
            messagebox.showerror("错误", "请选择至少一个要下载的项目")
            return
            
        download_path = self.path_entry.get().strip()
        if not download_path:
            messagebox.showerror("错误", "请选择下载路径")
            return
            
        if not os.path.exists(download_path):
            try:
                os.makedirs(download_path)
            except Exception as e:
                messagebox.showerror("错误", f"创建下载目录失败: {str(e)}")
                return
        
        orientation = self.orientation_var.get()
        preset_name = self.preset_combo.get()
        
        # --- Determine download dimensions ---
        target_vertical_size = (0, 0)
        target_horizontal_size = (0, 0)
        size_source = "原始"

        try:
            if preset_name == "自定义尺寸":
                size_source = "自定义"
                v_dim_str = self.v_size_entry.get().strip()
                h_dim_str = self.h_size_entry.get().strip()
                target_vertical_size = self.parse_dimension_string(v_dim_str)
                target_horizontal_size = self.parse_dimension_string(h_dim_str)
                print(f"使用自定义尺寸: 竖={target_vertical_size}, 横={target_horizontal_size}")
            else:
                # Use preset sizes
                size_source = preset_name
                target_vertical_size = self.size_presets[preset_name]["vertical"]
                target_horizontal_size = self.size_presets[preset_name]["horizontal"]
                print(f"使用预设 '{preset_name}': 竖={target_vertical_size}, 横={target_horizontal_size}")

        except ValueError as e:
            messagebox.showerror("输入错误", f"无效的自定义尺寸输入: {e}")
            self.status_label.configure(text="错误：无效的自定义尺寸")
            return # Stop download

        # Start download
        success_count = 0
        failed_items = [] # Keep track of failed titles
        total_to_download = 0
        items_to_process = [] # Store (title, img_url, type, width, height) tuples

        # Prepare list of downloads first
        for selected_idx in selected_indices:
            if selected_idx < len(self.results):
                # --- Unpack result item based on platform ---
                title = ""
                horz_url = None
                vert_url = None
                base_img_url = None # The single URL for iqiyi/tencent

                result_item = self.results[selected_idx]
                platform = self.selected_platform

                if platform == "优酷视频":
                    if len(result_item) == 3:
                        title, horz_url, vert_url = result_item
                    else: continue # Skip malformed
                else: # 爱奇艺, 腾讯视频
                    if len(result_item) == 2:
                        title, base_img_url = result_item
                    else: continue # Skip malformed

                # --- Determine tasks based on orientation ---
                if orientation == "全部":
                    v_w, v_h = target_vertical_size
                    # Vertical task: Use vert_url for Youku, base_img_url otherwise
                    vertical_base_url = vert_url if platform == "优酷视频" else base_img_url
                    if vertical_base_url: # Only add task if URL exists
                        items_to_process.append((title, vertical_base_url, "竖图", v_w, v_h))

                    # --- Added: Conditional Horizontal Task ---
                    # Only add horizontal task if NOT Tencent
                    if self.selected_platform != "腾讯视频":
                        h_w, h_h = target_horizontal_size
                        # Horizontal task: Use horz_url for Youku, base_img_url otherwise
                        horizontal_base_url = horz_url if platform == "优酷视频" else base_img_url
                        if horizontal_base_url: # Only add task if URL exists
                            items_to_process.append((title, horizontal_base_url, "横图", h_w, h_h))
                    # --- End Added Section ---

                elif orientation == "竖图":
                    v_w, v_h = target_vertical_size
                    vertical_base_url = vert_url if platform == "优酷视频" else base_img_url
                    if vertical_base_url:
                        items_to_process.append((title, vertical_base_url, "竖图", v_w, v_h))

                elif orientation == "横图":
                    # --- Added: Skip Tencent Horizontal ---
                    if self.selected_platform != "腾讯视频":
                        h_w, h_h = target_horizontal_size
                        horizontal_base_url = horz_url if platform == "优酷视频" else base_img_url
                        if horizontal_base_url:
                            items_to_process.append((title, horizontal_base_url, "横图", h_w, h_h))
                    # --- End Added Section ---

        total_to_download = len(items_to_process)
        if total_to_download == 0:
            # Check if this was due to missing URLs for Youku vertical or skipped Tencent horizontal
            if self.selected_platform == "优酷视频" and orientation != "横图" and any(len(r)==3 and not r[2] for i, r in enumerate(self.results) if i in selected_indices):
                messagebox.showinfo("提示", "部分选中项缺少竖图URL，无法下载。")
            elif self.selected_platform == "腾讯视频" and orientation == "横图":
                messagebox.showinfo("提示", "腾讯视频不支持仅下载横图。")
            elif self.selected_platform == "腾讯视频" and orientation == "全部" and not any(len(r)==2 and r[1] for i, r in enumerate(self.results) if i in selected_indices):
                messagebox.showinfo("提示", "选中的腾讯视频项缺少基础URL，无法下载竖图。")
            else:
                messagebox.showinfo("提示", "没有需要下载的项目。")
            self.status_label.configure(text="准备就绪")
            return
            
        self.status_label.configure(text=f"准备下载 {total_to_download} 张图片 ({size_source}尺寸)...")
        self.update()
        
        # Process downloads
        for idx, (title, base_url_for_type, img_type_suffix, width, height) in enumerate(items_to_process):
            current_num = idx + 1
            self.status_label.configure(text=f"下载中 ({current_num}/{total_to_download}): {title} ({img_type_suffix})")
            self.update()

            # Note: base_url_for_type is either horz_url or vert_url for Youku,
            # or the single img_url for others, depending on img_type_suffix.

            # Determine size string for filename
            if width == 0 or height == 0:
                size_str = "原始尺寸"
            else:
                size_str = f"{width}x{height}"

            # Determine suffix for filename (only if orientation was '全部')
            # And handle case where only one type exists (e.g. Youku missing vert)
            filename_suffix = None
            if orientation == "全部":
                # Check if both types are actually being downloaded for this item
                has_both_types = False
                if platform == "优酷视频":
                    # Find original item
                    original_item = next((r for i, r in enumerate(self.results) if i in selected_indices and r[0] == title), None)
                    if original_item and original_item[1] and original_item[2]: # Both horz and vert exist
                        has_both_types = True
                elif platform == "爱奇艺": # iQiyi always have base_img_url treated as both
                    has_both_types = True
                # Tencent only downloads Vertical when '全部' is selected, so no suffix needed

                if has_both_types and platform != "腾讯视频": # Add suffix only if both exist AND not tencent
                    filename_suffix = f"_{img_type_suffix}"

            # Call download_image, passing the base URL for the type,
            # target dimensions, platform, and img_type.
            if self.download_image(base_url_for_type, title, download_path,
                                 suffix=filename_suffix, size_str=size_str,
                                 platform=self.selected_platform,
                                 target_width=width, target_height=height,
                                 img_type=img_type_suffix): # Pass img_type
                success_count += 1
            else:
                failed_items.append(f"{title} ({img_type_suffix})")

        # Update status after loop
        failed_count = len(failed_items)
        if failed_count == 0:
            self.status_label.configure(text=f"下载完成: {success_count} 张图片已保存 ({size_source})")
            messagebox.showinfo("完成", f"成功下载 {success_count} 张图片到 {download_path}")
        else:
            self.status_label.configure(text=f"下载完成: {success_count} 张成功, {failed_count} 张失败 ({size_source})")
            # Show only first few failures in message box to avoid overly large dialog
            failures_preview = "\n - ".join(failed_items[:5])
            if failed_count > 5: failures_preview += "\n - ...等等"
            messagebox.showwarning("部分完成", f"成功下载 {success_count} 张图片。\n{failed_count} 张图片下载失败:\n - {failures_preview}")
    
    def download_image(self, base_img_url, title, download_path, suffix=None, size_str=None,
                     platform=None, target_width=0, target_height=0, img_type="", use_cid_filename=False):
        """下载单个图片并保存。Handles server-side sizing (iQiyi attempt) and local scaling (Tencent, Youku forced). Uses configured filename format."""

        if not base_img_url:
            print(f"跳过下载 ({title} - {img_type}): 无基础URL提供")
            return False

        print(f"  准备下载: {title} - {img_type} - {size_str if size_str else '原始尺寸'}")
        print(f"    基础URL: {base_img_url[:80]}...")
        print(f"    目标尺寸: {target_width}x{target_height}")

        download_url_to_try = base_img_url
        attempt_server_side_scaling = False
        needs_forced_local_scaling = False
        headers_to_use = self.iqiyi_headers

        # Determine headers and scaling strategy
        if platform == "腾讯视频":
            headers_to_use = self.tencent_headers
            if target_width > 0 and target_height > 0:
                needs_forced_local_scaling = True
                print("    腾讯平台: 将进行智能缩放裁剪")
        elif platform == "优酷视频":
            headers_to_use = self.youku_headers
            if target_width > 0 and target_height > 0:
                needs_forced_local_scaling = True
                download_url_to_try = base_img_url
                print("    优酷平台: 将下载原图并进行智能缩放裁剪")
        elif platform == "爱奇艺":
            headers_to_use = self.iqiyi_headers
            if target_width > 0 and target_height > 0:
                attempted_sized_url = self.build_iqiyi_sized_url(base_img_url, target_width, target_height)
                if attempted_sized_url != base_img_url:
                    download_url_to_try = attempted_sized_url
                    attempt_server_side_scaling = True
                    print(f"    爱奇艺: 尝试下载尺寸URL: {download_url_to_try[:80]}...")

        # --- Attempt Download ---
        try:
            print(f"  下载URL: {download_url_to_try[:80]}...")
            img_response = requests.get(download_url_to_try, headers=headers_to_use, timeout=30)

            if attempt_server_side_scaling and img_response.status_code != 200:
                print(f"下载爱奇艺尺寸图片失败 ({title} - {img_type}): 服务器返回 {img_response.status_code}")
                print(f"失败的尺寸URL: {download_url_to_try}")
                return False

            if img_response.status_code != 200:
                print(f"下载图片失败 ({title} - {img_type}): 服务器返回 {img_response.status_code}")
                print(f"失败的URL: {download_url_to_try}")
                return False

            img_response.raise_for_status()
            
            # --- Process Content (Potentially Scale Locally) ---
            save_content = img_response.content
            content_type = img_response.headers.get('Content-Type', '').lower()
            save_format = 'JPEG' # Default format if no scaling or type detection occurs

            if needs_forced_local_scaling and target_width > 0 and target_height > 0:
                print(f"  执行 {platform} 智能缩放裁剪至: {target_width}x{target_height}")
                try:
                    img = Image.open(BytesIO(img_response.content))
                    resized_img = self.smart_resize_and_crop(img, target_width, target_height)

                    output_buffer = BytesIO()
                    # Determine save format based on original image if possible
                    img_original_format = img.format
                    if img_original_format:
                        save_format = img_original_format.upper()
                        if save_format not in ['JPEG', 'PNG', 'GIF', 'WEBP']:
                            print(f"    警告: 未知原始格式 '{img_original_format}', 将存为 JPEG.")
                            save_format = 'JPEG' # Fallback
                    elif 'png' in content_type: save_format = 'PNG'
                    elif 'gif' in content_type: save_format = 'GIF'
                    elif 'webp' in content_type: save_format = 'WEBP'
                    else: save_format = 'JPEG' # Default if cannot determine

                    print(f"    原始格式: {img_original_format}, 保存格式: {save_format}")

                    # Handle transparency and color modes correctly
                    if save_format in ['PNG', 'WEBP'] and resized_img.mode in ('RGBA', 'LA'):
                        resized_img.save(output_buffer, format=save_format)
                    elif save_format == 'JPEG' and resized_img.mode != 'RGB':
                        resized_img.convert('RGB').save(output_buffer, format='JPEG', quality=95)
                    else:
                        save_options = {}
                        if save_format == 'JPEG': save_options['quality'] = 95
                        resized_img.save(output_buffer, format=save_format, **save_options)

                    save_content = output_buffer.getvalue()
                    print(f"  智能缩放裁剪完成，格式: {save_format}")

                except Exception as resize_err:
                    print(f"错误：智能缩放裁剪失败 ({title} - {img_type}): {resize_err}")
                    print(f"警告：缩放裁剪失败，将保存原始下载的图片...")
                    save_content = img_response.content # Keep original content
                    # Try to determine format from original content type/url for filename extension
                    if 'png' in content_type: save_format = 'PNG'
                    elif 'gif' in content_type: save_format = 'GIF'
                    elif 'webp' in content_type: save_format = 'WEBP'
                    elif 'jpeg' in content_type or 'jpg' in content_type: save_format = 'JPEG'
                    # Ensure save_format is uppercase for later checks
                    save_format = save_format.upper() if isinstance(save_format, str) else 'JPEG'


            # --- 添加自动压缩功能 ---
            preset_name = self.preset_combo.get() # 获取当前选择的预设
            
            # 根据不同预设设置不同的压缩目标
            if preset_name == "云南尺寸":
                target_filesize_kb = 100  # 云南尺寸压缩到100KB
                size_check_name = "云南尺寸"
            else:
                target_filesize_kb = 300  # 其他尺寸压缩到300KB
                size_check_name = "自动压缩"
            
            # 获取当前 save_content 的大小 (Bytes)
            current_size_bytes = len(save_content)
            current_size_kb = current_size_bytes / 1024

            print(f"  {size_check_name}检查: 当前图片大小 {current_size_kb:.2f} KB (目标 <= {target_filesize_kb} KB)")

            if current_size_kb > target_filesize_kb:
                print(f"  {size_check_name}警告: 图片大小超过 {target_filesize_kb} KB，尝试压缩...")
                try:
                    img_to_compress = Image.open(BytesIO(save_content))
                    # 确定图片格式以便压缩时使用正确的参数
                    # Use the already determined save_format which considers content-type and potential resize format
                    img_compress_format = save_format
                    print(f"    检测到压缩格式: {img_compress_format}")

                    if img_compress_format not in ['JPEG', 'WEBP']: # 只有JPEG和WEBP支持质量压缩
                        print(f"  警告: 图片格式 {img_compress_format} 不支持质量压缩，无法精确控制大小。将保存当前图片。")
                        # 对于不支持压缩的格式（如PNG, GIF），不处理
                    else:
                        # 尝试通过降低质量来压缩图片 (JPEG / WEBP)
                        quality = 85 # 初始压缩质量
                        step = 5    # 每次降低的步长
                        min_quality = 10 # 最低质量

                        compressed_buffer = BytesIO()
                        new_size_kb = current_size_kb # Initialize with current size

                        while quality >= min_quality:
                            compressed_buffer.seek(0) # 重置 buffer
                            compressed_buffer.truncate() # 清空 buffer
                            save_options = {'quality': quality}
                            if img_compress_format == 'WEBP': save_options['lossless'] = False # 确保有损压缩

                            # 转换模式以兼容保存格式
                            save_img = img_to_compress
                            if img_compress_format == 'JPEG' and save_img.mode != 'RGB':
                                print(f"    转换图片模式从 {save_img.mode} 到 RGB 以进行 JPEG 压缩...")
                                save_img = save_img.convert('RGB')
                            elif img_compress_format == 'WEBP' and save_img.mode == 'P':
                                # Handle palette mode for WEBP by converting to RGBA or RGB
                                if 'transparency' in save_img.info:
                                    print(f"    转换图片模式从 {save_img.mode} 到 RGBA 以进行 WEBP 压缩...")
                                    save_img = save_img.convert('RGBA')
                                else:
                                    print(f"    转换图片模式从 {save_img.mode} 到 RGB 以进行 WEBP 压缩...")
                                    save_img = save_img.convert('RGB')


                            save_img.save(compressed_buffer, format=img_compress_format, **save_options)
                            new_size_kb = len(compressed_buffer.getvalue()) / 1024
                            print(f"    尝试质量 {quality}: 新大小 {new_size_kb:.2f} KB")

                            if new_size_kb <= target_filesize_kb:
                                save_content = compressed_buffer.getvalue() # 使用压缩后的数据
                                print(f"  压缩成功！最终大小 {new_size_kb:.2f} KB (质量 {quality})")
                                break # 达到目标，退出循环
                            quality -= step # 未达到目标，降低质量继续尝试
                        else: # 如果循环结束仍未达到目标
                            print(f"  警告: 压缩至最低质量 {min_quality} 后，大小仍为 {new_size_kb:.2f} KB，超过目标。将使用此最低质量图片。")
                            save_content = compressed_buffer.getvalue() # 至少用压缩过的

                except Exception as compress_err:
                    print(f"错误：压缩{size_check_name}图片时出错: {compress_err}")
                    print(traceback.format_exc())
                    # 保留原始（或已缩放）的 save_content
                    print("警告：压缩失败，将保存压缩前的图片。")
            # --- 自动压缩结束 ---


            # --- Determine Filename using Format String and Save ---
            ext = ".jpg" # Default extension
            # Use the potentially updated save_format from compression/resizing logic
            save_format_lower = save_format.lower()
            if save_format_lower == 'jpeg': ext = '.jpg'
            elif save_format_lower == 'png': ext = '.png'
            elif save_format_lower == 'gif': ext = '.gif'
            elif save_format_lower == 'webp': ext = '.webp'
            # Add fallback based on URL if format is still JPEG default but URL suggests otherwise
            elif save_format_lower == 'jpeg':
                 url_path = urllib.parse.urlparse(download_url_to_try).path.lower()
                 if url_path.endswith('.png'): ext = '.png'; save_format_lower = 'png'
                 elif url_path.endswith('.gif'): ext = '.gif'; save_format_lower = 'gif'
                 elif url_path.endswith('.webp'): ext = '.webp'; save_format_lower = 'webp'


            # --- Build filename using format string ---
            if use_cid_filename:
                # 批量模式：直接使用CID作为文件名
                base_filename = self.sanitize_filename(title)
            else:
                # 经典模式：使用格式字符串
                format_data = {
                    "标题": self.sanitize_filename(title),
                    "图片尺寸": f"{target_width}x{target_height}" if target_width > 0 and target_height > 0 else (size_str if size_str else "原始"),
                    "类型": img_type if img_type else "" # Use the passed img_type ('竖图' or '横图')
                }
                # Available placeholders: {标题}, {图片尺寸}, {类型}
                try:
                    # Replace placeholders, providing empty string for missing keys
                    base_filename = self.filename_format.format(
                        **{k: format_data.get(k, '') for k in ['标题', '图片尺寸', '类型']}
                    )
                    # Clean up filename: remove extra/leading/trailing underscores
                    base_filename = re.sub(r'_+', '_', base_filename).strip('_')
                    # Ensure filename is not empty after formatting
                    if not base_filename:
                        base_filename = self.sanitize_filename(title) # Fallback to title
                except KeyError as e:
                     print(f"警告: 文件名格式字符串中存在未知占位符: {e}. 将使用默认格式。")
                     base_filename = f"{format_data['标题']}_{format_data['类型']}_{format_data['图片尺寸']}"
                     base_filename = re.sub(r'_+', '_', base_filename).strip('_')


            file_path = os.path.join(download_path, f"{base_filename}{ext}")

            counter = 1
            original_path_base = os.path.join(download_path, base_filename)
            while os.path.exists(file_path):
                file_path = f"{original_path_base}_{counter}{ext}"
                counter += 1
            
            with open(file_path, 'wb') as f:
                f.write(save_content)
            print(f"  图片已保存: {file_path}")
            return True

        except requests.exceptions.Timeout:
             print(f"下载图片超时 ({title} - {img_type}): {download_url_to_try[:80]}...")
             return False
        except requests.exceptions.RequestException as e:
            print(f"下载图片网络错误 ({title} - {img_type}): {e}")
            print(f"失败URL: {download_url_to_try}")
            return False
        except Exception as e:
            print(f"保存或处理图片失败 ({title} - {img_type}): {e}")
            print(f"处理URL: {download_url_to_try}")
            print(traceback.format_exc())
            return False

    def parse_dimension_string(self, dim_str: str) -> tuple[int, int]:
        """解析 '宽度x高度' 格式的字符串, 返回 (int, int) 或引发 ValueError"""
        if not dim_str:
            raise ValueError("尺寸字符串不能为空")
        try:
            parts = dim_str.lower().split('x')
            if len(parts) != 2:
                raise ValueError("格式应为 '宽度x高度'")
            width = int(parts[0].strip())
            height = int(parts[1].strip())
            if width <= 0 or height <= 0:
                raise ValueError("宽度和高度必须是正数")
            return width, height
        except (ValueError, IndexError) as e:
            # Reraise with a more specific message if it's not already one of ours
            if isinstance(e, ValueError) and str(e) in ["尺寸字符串不能为空", "格式应为 '宽度x高度'", "宽度和高度必须是正数"]:
                 raise e
            else:
                 raise ValueError(f"无效的尺寸格式 '{dim_str}'. 错误: {e}")

    def clear_results_widgets_only(self):
         """Helper to destroy only widgets in results_frame, not reset data lists"""
         print("Clearing results frame widgets...") # Optional: for debugging
         for widget in self.results_frame.winfo_children():
            widget.destroy()
         print("Results frame widgets cleared.") # Optional: for debugging

    def on_platform_change(self):
        """更新当前选择的平台 (No argument needed anymore)"""
        self.selected_platform = self.platform_var.get() # Read from StringVar
        print(f"平台已切换到: {self.selected_platform}")
        # Optionally, clear results when platform changes?
        # self.clear_results()
        # self.status_label.configure(text=f"已切换到 {self.selected_platform}, 请搜索")

    def show_zoomed_image(self, horz_url, vert_url, preview_type, title=""):
        """显示点击预览图的放大视图，展示对应类型和当前选定尺寸的最终图片。"""

        platform = self.selected_platform
        preset_name = self.preset_combo.get()

        print(f"\n--- Zoom Request ---")
        print(f"  Platform: {platform}, Type: {preview_type}, Title: {title}")
        print(f"  Horz URL (Input): {horz_url[:70] if horz_url else 'None'}...")
        print(f"  Vert URL (Input): {vert_url[:70] if vert_url else 'None'}...")
        print(f"  Preset: {preset_name}")

        # 1. Get Original Base URLs (handle None safely)
        original_url_h = horz_url
        original_url_v = vert_url

        # 2. Get Current Target Dimensions from GUI (used for info label and iQiyi sizing)
        target_w = 0
        target_h = 0
        display_w_final = 0
        display_h_final = 0
        size_info_str = "原始尺寸" # Default for info label
        try:
            if preset_name == "自定义尺寸":
                # Get V size for info if Vert/Original, H size if Horz
                if preview_type == "竖图" or preview_type == "原图":
                    v_dim_str = self.v_size_entry.get().strip()
                    if v_dim_str and "x" in v_dim_str:
                        target_w, target_h = self.parse_dimension_string(v_dim_str)
                        display_w_final, display_h_final = target_w, target_h
                        size_info_str = f"自定义 竖图: {target_w}x{target_h}"
                elif preview_type == "横图":
                    h_dim_str = self.h_size_entry.get().strip()
                    if h_dim_str and "x" in h_dim_str:
                        target_w, target_h = self.parse_dimension_string(h_dim_str)
                        display_w_final, display_h_final = target_w, target_h
                        size_info_str = f"自定义 横图: {target_w}x{target_h}"
            else: # Use preset dimensions for info
                if preview_type == "竖图" or preview_type == "原图":
                    target_w, target_h = self.size_presets[preset_name]["vertical"]
                    display_w_final, display_h_final = target_w, target_h
                    size_info_str = f"预设 '{preset_name}' 竖图: {target_w}x{target_h}"
                elif preview_type == "横图":
                    target_w, target_h = self.size_presets[preset_name]["horizontal"]
                    display_w_final, display_h_final = target_w, target_h
                    size_info_str = f"预设 '{preset_name}' 横图: {target_w}x{target_h}"

        except ValueError as e:
            # This error shouldn't prevent zoom, just affects the info label
            print(f"  无法解析尺寸输入用于信息标签: {e}, 将显示原始尺寸信息。")
            size_info_str = "原始尺寸 (解析错误)"

        # --- 3. Determine Base URL to Download ---
        base_url_to_download = None
        final_display_is_original = False # Flag to indicate if we intend to show original

        if preview_type == "原图":
            base_url_to_download = original_url_h # Usually horizontal is the main 'original'
            if not base_url_to_download: base_url_to_download = original_url_v
            final_display_is_original = True
            size_info_str = "原始尺寸" # Override size info for "Original" type
            print(f"  Zoom Strategy: Original -> Download Base: {base_url_to_download[:70] if base_url_to_download else 'None'}...")
        elif preview_type == "竖图":
            base_url_to_download = original_url_v
            final_display_is_original = True # Treat vertical zoom as showing original vertical
            size_info_str = "原始 竖图" # Update info string
            print(f"  Zoom Strategy: Vertical -> Download Base: {base_url_to_download[:70] if base_url_to_download else 'None'}...")
        elif preview_type == "横图":
            # --- Allow Horizontal Zoom for all platforms now ---
            base_url_to_download = original_url_h
            final_display_is_original = True # Treat horizontal zoom as showing original horizontal
            size_info_str = "原始 横图" # Update info string
            print(f"  Zoom Strategy: Horizontal -> Download Base: {base_url_to_download[:70] if base_url_to_download else 'None'}...")

        if not base_url_to_download:
            messagebox.showerror("错误", f"无法显示 {preview_type} 预览，缺少原始 URL。")
            return

        # 4. Determine Actual Download URL and Scaling Needs for Zoom
        url_to_download = base_url_to_download
        local_resize_needed = False # Never force resize to target WxH for zoom
        attempted_server_side = False # Default is not to attempt server resize

        # --- Special Handling for iQiyi Platform Horizontal Preview ---
        if platform == "爱奇艺" and preview_type == "横图" and display_w_final > 0 and display_h_final > 0:
            # Create a sized URL with horizontal dimensions for iQiyi
            url_to_download = self.build_iqiyi_sized_url(base_url_to_download, display_w_final, display_h_final)
            attempted_server_side = True
            local_resize_needed = False
            size_info_str = f"爱奇艺横图: {display_w_final}x{display_h_final}"
            print(f"  爱奇艺横图预览: 使用尺寸URL: {url_to_download[:80]}...")

        print(f"  Zoom Final Prep: Download URL = {url_to_download[:70]}..., Local Resize Needed = {local_resize_needed}")

        # --- Create Popup Window ---
        popup_title = f"{title} - {preview_type} 预览 ({size_info_str})" if title else f"{preview_type} 预览 ({size_info_str})"
        popup = ctk.CTkToplevel(self)
        popup.title(popup_title)
        popup.geometry("600x450") # Larger initial size for better user experience
        popup.attributes('-topmost', True)
        popup.transient(self)

        zoomed_label = ctk.CTkLabel(popup, text="正在加载图片...", height=400, width=550)
        zoomed_label.pack(expand=True, fill="both", padx=10, pady=10)
        note_label = ctk.CTkLabel(popup, text="点击图片关闭", font=("Arial", 10))
        note_label.pack(pady=(0, 10))
        zoomed_label.bind("<Button-1>", lambda event: popup.destroy())

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # --- Async Image Loading Logic ---
        def load_zoomed_image_async():
            pil_image = None
            downloaded_w, downloaded_h = 0, 0 # For actual dimensions of the downloaded image

            try:
                # 1. Download the determined URL
                print(f"  Zoom Download URL: {url_to_download[:80]}...")
                headers_to_use = self.iqiyi_headers
                if platform == "腾讯视频": headers_to_use = self.tencent_headers
                elif platform == "优酷视频": headers_to_use = self.youku_headers

                response = requests.get(url_to_download, headers=headers_to_use, timeout=20)

                if response.status_code != 200:
                    error_msg = f"加载失败: HTTP {response.status_code}"
                    
                    # Special handling for iQiyi attempted server-side sizing failure
                    if attempted_server_side and platform == "爱奇艺":
                        # Try with original URL as fallback for iQiyi
                        print(f"  爱奇艺尺寸URL失败，尝试原始URL: {base_url_to_download[:80]}...")
                        fallback_response = requests.get(base_url_to_download, headers=headers_to_use, timeout=20)
                        
                        if fallback_response.status_code == 200:
                            response = fallback_response
                            print("  成功从原始URL获取图片")
                        else:
                            if popup.winfo_exists() and zoomed_label.winfo_exists():
                                error_msg = f"爱奇艺图片加载失败: HTTP {response.status_code}/{fallback_response.status_code}"
                                self.after(0, lambda label=zoomed_label, msg=error_msg: label.configure(text=msg, image=None))
                            return
                    else:
                        # Non-iQiyi or no fallback needed
                        if popup.winfo_exists() and zoomed_label.winfo_exists():
                            self.after(0, lambda label=zoomed_label, msg=error_msg: label.configure(text=msg, image=None))
                        return

                # 2. Load image data with PIL
                pil_image = Image.open(BytesIO(response.content))
                downloaded_w, downloaded_h = pil_image.size # Get actual dimensions
                print(f"    Downloaded zoom image size: {downloaded_w}x{downloaded_h}")

                # 3. NO forced local scaling to target WxH here for zoom

                # 4. Always use the actual downloaded dimensions as base for screen fitting
                final_w, final_h = downloaded_w, downloaded_h
                
                # 5. Calculate display size fitting to screen (using actual downloaded dimensions)
                popup_display_w = final_w
                popup_display_h = final_h
                max_width = int(screen_width * 0.9)
                max_height = int(screen_height * 0.9)

                if popup_display_w > max_width or popup_display_h > max_height:
                    scale_ratio = min(max_width / popup_display_w, max_height / popup_display_h)
                    popup_display_w = max(1, int(popup_display_w * scale_ratio))
                    popup_display_h = max(1, int(popup_display_h * scale_ratio))
                    print(f"    Adjusting zoom display size for screen fit: {popup_display_w}x{popup_display_h}")

                # 6. Resize image for final popup display IF screen fitting required it
                display_img = pil_image
                if popup_display_w != final_w or popup_display_h != final_h:
                    print("    Resizing final zoom image for popup display...")
                    try:
                        display_img = pil_image.resize((popup_display_w, popup_display_h), Image.Resampling.LANCZOS)
                    except AttributeError:
                        display_img = pil_image.resize((popup_display_w, popup_display_h), Image.LANCZOS)

                # 7. Prepare CTkImage and update UI
                ctk_img = ctk.CTkImage(light_image=display_img, dark_image=display_img,
                                       size=(popup_display_w, popup_display_h))

                # Calculate final window size and position - slightly larger margins for better appearance
                window_width = popup_display_w + 40
                window_height = popup_display_h + 60
                window_x = (screen_width - window_width) // 2
                window_y = (screen_height - window_height) // 2

                # Update UI in main thread
                def update_zoom_ui():
                    if popup.winfo_exists():
                        popup.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
                        if zoomed_label.winfo_exists():
                            zoomed_label.configure(image=ctk_img, text="")
                        # Use actual downloaded dimensions in the info label
                        updated_info = f"实际尺寸: {downloaded_w}x{downloaded_h}"
                        if popup_display_w != downloaded_w or popup_display_h != downloaded_h:
                            updated_info += f" | 显示尺寸: {popup_display_w}x{popup_display_h}"
                        if note_label.winfo_exists():
                            note_label.configure(text=f"点击图片关闭 | {updated_info}")
                        popup.update_idletasks()

                self.after(0, update_zoom_ui)
            
            except Exception as e:
                error_message = f"加载放大图片错误: {str(e)[:100]}"
                print(f"Error loading zoomed image: {str(e)}")
                print(traceback.format_exc())
                if popup.winfo_exists() and zoomed_label.winfo_exists():
                    self.after(0, lambda label=zoomed_label, msg=error_message: label.configure(text=msg, image=None))

        # Start the background thread
        threading.Thread(target=load_zoomed_image_async, daemon=True).start()

    def find_youku_video_nodes(self, node_list, found_count, limit=100):
        """递归遍历节点列表，查找并提取优酷影视结果信息，带有限制"""
        found_results = []
        if not isinstance(node_list, list):
            print("  Debug: find_youku_video_nodes received non-list input.")
            return found_results
        
        # Stop recursion if limit is reached
        if found_count[0] >= limit:
             # print(f"  Recursion limit ({limit}) reached, stopping branch.")
             return found_results

        for node in node_list:
            if not isinstance(node, dict): continue

            # Stop processing this level if limit is reached
            if found_count[0] >= limit:
                 break

            node_data = node.get('data', {})
            node_type = node.get('type')

            is_target_node = isinstance(node_data, dict) and \
                             'titleDTO' in node_data and \
                             'posterDTO' in node_data and \
                             ('thumbUrl' in node_data or node_data.get('posterDTO', {}).get('vThumbUrl'))

            if is_target_node:
                title = node_data.get('titleDTO', {}).get('displayName') or node_data.get('tempTitle')
                vert_url = node_data.get('posterDTO', {}).get('vThumbUrl')
                horz_url = node_data.get('thumbUrl')

                if not horz_url and vert_url: horz_url = vert_url
                if not vert_url and horz_url: vert_url = horz_url

                if title and (horz_url or vert_url):
                    if horz_url and isinstance(horz_url, str) and horz_url.startswith('//'):
                        horz_url = 'https:' + horz_url
                    if vert_url and isinstance(vert_url, str) and vert_url.startswith('//'):
                        vert_url = 'https:' + vert_url

                    found_results.append((title, horz_url, vert_url))
                    found_count[0] += 1 # Increment count
                    print(f"  -> Found Video Node #{found_count[0]} (Type: {node_type}): Title='{title}'")
                    # Early exit check after incrementing
                    if found_count[0] >= limit:
                         # print(f"  Limit ({limit}) reached after finding node.")
                         return found_results # Return immediately after reaching limit

            # --- Recursively Search Child Nodes (only if limit not reached) ---
            if found_count[0] < limit:
                 child_nodes_direct = node.get('nodes', [])
                 child_nodes_in_data = node_data.get('nodes', []) if isinstance(node_data, dict) else []

                 if child_nodes_direct:
                     found_results.extend(self.find_youku_video_nodes(child_nodes_direct, found_count, limit))
                     # Check limit again after direct children recursion
                     if found_count[0] >= limit: return found_results

                 if child_nodes_in_data:
                     found_results.extend(self.find_youku_video_nodes(child_nodes_in_data, found_count, limit))
                     # Check limit again after data children recursion
                     if found_count[0] >= limit: return found_results

        return found_results

    def search_youku(self, search_term):
        """执行优酷视频搜索 (从内嵌 __INITIAL_DATA__ JSON 提取 - 使用递归查找)"""
        initial_results = []
        self.results = []

        try:
            encoded_term = urllib.parse.quote(search_term)
            search_url = f"https://so.youku.com/search/q_{encoded_term}"
            print(f"优酷搜索 URL: {search_url}")

            response = requests.get(search_url, headers=self.youku_headers, timeout=20)

            if response.status_code == 200:
                html_content = response.text
                print(f"优酷搜索页响应状态码: {response.status_code}")

                data_dict = None
                json_string = None
                json_data_match = re.search(r'window\.__INITIAL_DATA__\s*=\s*({.*?})\s*;', html_content, re.DOTALL)

                if json_data_match:
                    json_string = json_data_match.group(1)
                    print("成功匹配到内嵌 JSON 字符串。准备预处理和解析...")
                    print(f"提取到的 JSON 字符串开头: {json_string[:200]}")
                    print(f"提取到的 JSON 字符串结尾: {json_string[-200:]}")
                    print(f"提取到的 JSON 字符串总长度: {len(json_string)}")

                    # --- 添加预处理步骤 ---
                    try:
                        # 1. 替换 new Date("...") 为 "..." (保留日期字符串，加引号)
                        #    使用正则表达式匹配 new Date(...) 结构并提取引号内的内容
                        processed_json_string = re.sub(r'new\s+Date\s*\(\s*"(.*?)"\s*\)', r'"\1"', json_string)

                        # 2. 替换 JavaScript 的 undefined 为 JSON 的 null
                        processed_json_string = processed_json_string.replace('undefined', 'null')

                        # 3. (可选) 检查是否还有其他明显的非 JSON 模式需要替换，例如函数定义等
                        print("JSON 字符串预处理完成。")
                        # --- 预处理结束 ---

                        # 使用处理后的字符串进行解析
                        data_dict = json.loads(processed_json_string)
                        print("成功解析了预处理后的 JSON。")

                    except json.JSONDecodeError as e:
                        print(f"错误：解析预处理后的 JSON 字符串仍然失败: {e}")
                        # 打印详细错误和上下文
                        error_pos = e.pos
                        context_len = 150
                        start = max(0, error_pos - context_len)
                        end = min(len(processed_json_string), error_pos + context_len)
                        print(f"错误发生在字符 {error_pos} 附近 (处理后): ...【{processed_json_string[start:error_pos]}】<<<错误点>>>【{processed_json_string[error_pos:end]}】...")
                        # 保存原始字符串和处理后的字符串以供对比
                        try:
                            with open("debug_youku_raw_json_string.txt", "w", encoding="utf-8") as f: f.write(json_string)
                            with open("debug_youku_processed_fail.txt", "w", encoding="utf-8") as f: f.write(processed_json_string)
                            print("已将原始字符串和处理失败的字符串保存到 debug 文件。")
                        except Exception as e_save:
                            print(f"警告：保存调试文件失败: {e_save}")
                        self.status_label.configure(text="错误：解析优酷页面数据失败 (格式处理后)")
                        messagebox.showerror("解析错误", f"尝试处理后，解析优酷页面数据仍然失败。\n错误: {e}")
                        self.results = []
                        self.update_results_list()
                        return []
                    except Exception as e_other:
                        print(f"错误：预处理 JSON 字符串时发生异常: {e_other}")
                        print(traceback.format_exc())
                        self.status_label.configure(text="错误：处理优酷数据时异常")
                        self.results = []
                        self.update_results_list()
                        return []

                else:
                    print("错误：未能在 HTML 中找到 window.__INITIAL_DATA__ = {...}; 结构")
                    self.status_label.configure(text="错误：无法获取优酷页面核心数据")
                    messagebox.showerror("数据错误", "无法在优酷页面源码中找到 `window.__INITIAL_DATA__ = {...};` 结构。")
                    self.results = []
                    self.update_results_list()
                    return []

                # 此处继续处理成功解析的情况
                if data_dict is None:
                    print("错误: data_dict 未能成功解析，无法继续提取节点。")
                    self.results = []
                    self.update_results_list()
                    return []

                print("开始递归查找视频节点 (限制 50)...")
                try:
                    top_level_nodes = data_dict.get('data', {}).get('nodes', [])
                    found_count = [0]
                    initial_results = self.find_youku_video_nodes(top_level_nodes, found_count, limit=50)
                except Exception as e:
                    print(f"递归查找视频节点时出错: {e}")
                    print(traceback.format_exc())
                    initial_results = []

            else:
                print(f"优酷页面请求失败，状态码: {response.status_code}")
                self.status_label.configure(text=f"错误：无法访问优酷搜索页 ({response.status_code})")
                messagebox.showerror("网络错误", f"访问优酷搜索页失败，状态码: {response.status_code}")
                self.results = []
                self.update_results_list()
                return []

            # --- 精确搜索过滤 (客户端) ---
            if self.precise_search_var.get() and initial_results:
                print("执行优酷精确匹配过滤...")
                self.results = self.filter_results_by_title_similarity(initial_results, search_term, top_n=3)
                print(f"精确匹配后剩余 {len(self.results)} 个结果。")
            else:
                self.results = initial_results

            # --- 更新 GUI ---
            if not self.results:
                self.status_label.configure(text="在优酷中未找到匹配结果")
            else:
                self.status_label.configure(text=f"处理完成 {len(self.results)} 个优酷结果")

            self.update_results_list()

        except requests.exceptions.RequestException as e:
            print(f"优酷搜索网络错误: {e}")
            messagebox.showerror("网络错误", f"无法连接到优酷或请求失败: {e}")
            self.results = []
            self.update_results_list()
            return []
        except Exception as e:
            print(f"优酷搜索处理出错: {e}")
            print(traceback.format_exc())
            messagebox.showerror("处理错误", f"处理优酷搜索结果时出错: {e}")
            self.results = []
            self.update_results_list()
            return []

        return self.results

    def filter_results_by_title_similarity(self, results, search_key, top_n=3):
        """根据标题与搜索词的相似度过滤和排序结果 (通用)"""
        # ... (Implementation remains the same) ...
        if not results or not search_key:
            return results[:top_n] if not search_key else results

        scored_results = []
        norm_key = self.normalize_text(search_key)

        # Adapt to handle different tuple lengths if necessary, but expect (title, url1, url2) for Youku
        for item in results:
            title = item[0] # Assume title is always the first element
            norm_title = self.normalize_text(title)
            score = difflib.SequenceMatcher(None, norm_key, norm_title).ratio()
            scored_results.append((score, item)) # Keep the original item tuple

        scored_results.sort(key=lambda x: x[0], reverse=True)

        # Return top N original items
        return [item for score, item in scored_results[:top_n]]

    def build_youku_vertical_sized_url(self, original_url, width, height):
        """尝试构建优酷带尺寸参数的竖图URL (基于阿里云OSS规则)"""
        if not original_url or width <= 0 or height <= 0:
            return original_url # Return original if params are invalid or no URL

        try:
            parsed_url = urllib.parse.urlparse(original_url)
            query_params = urllib.parse.parse_qs(parsed_url.query)

            # OSS process parameter key
            process_key = 'x-oss-process'
            new_actions = []

            if process_key in query_params:
                # Existing process parameter found
                current_process_val = query_params[process_key][0]
                # Split actions like 'image/resize,w_412,h_600/quality,q_90'
                actions = current_process_val.split('/')
                has_resize = False
                for action in actions:
                    if action.startswith('image/resize'):
                         # Modify existing resize action
                         # Keep existing options like m_fill etc. if present
                         parts = action.split(',')
                         new_parts = [parts[0]] # Keep 'image/resize'
                         # Add new width/height, replacing old ones
                         new_parts.append(f'w_{width}')
                         new_parts.append(f'h_{height}')
                         # Keep other options if they existed
                         for part in parts[1:]:
                              if not part.startswith('w_') and not part.startswith('h_'):
                                   new_parts.append(part)
                         new_actions.append(','.join(new_parts))
                         has_resize = True
                    elif action.strip(): # Keep other non-empty actions
                         new_actions.append(action)

                if not has_resize:
                    # Add new resize action if none existed
                    new_actions.insert(0, f'image/resize,w_{width},h_{height}') # Add at beginning
            else:
                # No existing process parameter, create a new one
                new_actions.append(f'image/resize,w_{width},h_{height}')

            # Reconstruct the process string
            new_process_val = '/'.join(new_actions)

            # Update query parameters
            query_params[process_key] = [new_process_val]

            # Rebuild the query string
            new_query_string = urllib.parse.urlencode(query_params, doseq=True)

            # Reconstruct the full URL
            sized_url = urllib.parse.urlunparse((
                parsed_url.scheme,
                parsed_url.netloc,
                parsed_url.path,
                parsed_url.params,
                new_query_string,
                parsed_url.fragment
            ))
            print(f"    构建的优酷尺寸URL: {sized_url}")
            return sized_url

        except Exception as e:
            print(f"警告: 构建优酷尺寸URL失败 ({original_url}): {e}. 返回原始URL。")
            return original_url # Fallback to original URL on any error

    # --- Settings Methods ---

    def get_default_settings(self):
        """Returns a dictionary with default settings."""
        return {
            "default_platform": "爱奇艺",
            "default_precise": False,
            "default_download_type": "全部",
            "default_path": os.path.join(os.path.expanduser("~"), "Desktop"),
            "default_poster_size": "基础尺寸",
            "default_vertical_size": "412x600",
            "default_horizontal_size": "528x296",
            "filename_format": "{标题}_{类型}_{图片尺寸}",
            "batch_search_priority": [
                "优酷视频-精确搜索", "爱奇艺-精确搜索", "爱奇艺-普通搜索"
            ],
            "batch_horizontal_path": r"D:\海报需求\第十一批\横图",
            "batch_vertical_path": r"D:\海报需求\第十一批\竖图",
            "batch_default_size": "基础尺寸",
            "batch_default_vertical_size": "412x600",
            "batch_default_horizontal_size": "528x296",
            "delete_horizontal_path": r"D:\海报需求\第十一批\横图",
            "delete_vertical_path": r"D:\海报需求\第十一批\竖图",
            "iqiyi_cookie": "", # Start empty
            "tencent_cookie": "",
            "youku_cookie": ""
        }

    def load_settings(self):
        """Loads settings from config.json, returning defaults if file not found or invalid."""
        defaults = self.get_default_settings()
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    loaded_settings = json.load(f)
                # Ensure all keys exist by merging with defaults
                settings = {**defaults, **loaded_settings}
                print("Settings loaded successfully from config.json")
                return settings
            else:
                print("config.json not found, using default settings.")
                return defaults
        except (json.JSONDecodeError, IOError) as e:
            print(f"Error loading {self.config_file}: {e}. Using default settings.")
            return defaults

    def save_settings(self):
        """Saves the current settings from the settings window to config.json."""
        if not hasattr(self, 'settings_window') or not self.settings_window.winfo_exists():
            print("Settings window not open, cannot save.")
            return

        try:
            settings_to_save = {
                "default_platform": self.settings_platform_var.get(),
                "default_precise": self.settings_precise_var.get(),
                "default_download_type": self.settings_download_type_var.get(),
                "default_path": self.settings_path_entry.get().strip(),
                "filename_format": self.settings_filename_format_entry.get().strip(),
                "iqiyi_cookie": self.settings_iqiyi_cookie_box.get("1.0", "end-1c").strip(),
                "tencent_cookie": self.settings_tencent_cookie_box.get("1.0", "end-1c").strip(),
                "youku_cookie": self.settings_youku_cookie_box.get("1.0", "end-1c").strip()
            }

            # Validate filename format roughly
            if not settings_to_save["filename_format"]:
                 settings_to_save["filename_format"] = self.get_default_settings()["filename_format"]
                 self.settings_filename_format_entry.delete(0, 'end')
                 self.settings_filename_format_entry.insert(0, settings_to_save["filename_format"])
                 messagebox.showwarning("格式错误", "文件名格式不能为空，已重置为默认值。", parent=self.settings_window)
                 # Don't proceed with save yet, let user correct if needed or save again
                 return # Stop saving if format is invalid

            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(settings_to_save, f, ensure_ascii=False, indent=4)

            print("Settings saved successfully to config.json")

            # --- Apply changes immediately to the running application ---
            new_default_platform = settings_to_save["default_platform"]
            self.platform_var.set(new_default_platform) # Update the radio button variable
            self.selected_platform = new_default_platform # Update the internal variable used by search etc.

            self.precise_search_var.set(settings_to_save["default_precise"])
            self.orientation_var.set(settings_to_save["default_download_type"])

            self.path_entry.delete(0, 'end')
            self.path_entry.insert(0, settings_to_save["default_path"])

            self.filename_format = settings_to_save["filename_format"]

            self.iqiyi_headers['Cookie'] = settings_to_save["iqiyi_cookie"]
            self.tencent_headers['Cookie'] = settings_to_save["tencent_cookie"]
            self.youku_headers['Cookie'] = settings_to_save["youku_cookie"]

            # Refresh size entries based on current preset (in case defaults changed affecting it)
            self.on_preset_change(self.preset_combo.get())
            # --- End Apply Changes ---

            # Remove success message box
            # messagebox.showinfo("设置已保存", "设置已成功保存并应用。", parent=self.settings_window)

            self.settings_window.destroy() # Close settings window after saving

        except Exception as e:
            messagebox.showerror("保存错误", f"保存设置时出错: {e}", parent=self.settings_window)
            print(f"Error saving settings: {e}")
            print(traceback.format_exc())


    def open_settings_window(self):
        """Opens the settings configuration window."""
        if hasattr(self, 'settings_window') and self.settings_window.winfo_exists():
            self.settings_window.focus() # If already open, bring to front
            return

        # Load settings from file for initial population, but use current headers for cookies
        current_settings_from_file = self.load_settings()

        self.settings_window = ctk.CTkToplevel(self)
        self.settings_window.title("设置")
        self.settings_window.geometry("650x550")
        self.settings_window.attributes('-topmost', True)
        self.settings_window.transient(self)
        self.settings_window.grab_set() # Make modal

        # Create Tabview
        tab_view = ctk.CTkTabview(self.settings_window)
        tab_view.pack(expand=True, fill="both", padx=10, pady=10)

        tab_defaults = tab_view.add("默认设置")
        tab_cookies = tab_view.add("Cookie 设置")

        # --- Defaults Tab ---
        defaults_frame = ctk.CTkFrame(tab_defaults, fg_color="transparent")
        defaults_frame.pack(expand=True, fill="both", padx=5, pady=5)
        defaults_frame.grid_columnconfigure(1, weight=1) # Give entry fields weight

        # Default Platform
        ctk.CTkLabel(defaults_frame, text="默认平台:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        # Use the value currently in the main window's variable for consistency
        self.settings_platform_var = ctk.StringVar(value=self.platform_var.get())
        ctk.CTkComboBox(defaults_frame, values=["爱奇艺", "腾讯视频", "优酷视频"], variable=self.settings_platform_var).grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # Default Precise Search
        ctk.CTkLabel(defaults_frame, text="默认精确搜索:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.settings_precise_var = ctk.BooleanVar(value=self.precise_search_var.get()) # Use current value
        ctk.CTkSwitch(defaults_frame, text="", variable=self.settings_precise_var).grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # Default Download Type
        ctk.CTkLabel(defaults_frame, text="默认下载类型:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.settings_download_type_var = ctk.StringVar(value=self.orientation_var.get()) # Use current value
        radio_frame = ctk.CTkFrame(defaults_frame, fg_color="transparent")
        radio_frame.grid(row=2, column=1, columnspan=2, padx=0, pady=0, sticky="w")
        ctk.CTkRadioButton(radio_frame, text="全部下载", variable=self.settings_download_type_var, value="全部").pack(side="left", padx=5)
        ctk.CTkRadioButton(radio_frame, text="仅下竖图", variable=self.settings_download_type_var, value="竖图").pack(side="left", padx=5)
        ctk.CTkRadioButton(radio_frame, text="仅下横图", variable=self.settings_download_type_var, value="横图").pack(side="left", padx=5)

        # Default Download Path
        ctk.CTkLabel(defaults_frame, text="默认下载路径:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.settings_path_entry = ctk.CTkEntry(defaults_frame)
        self.settings_path_entry.insert(0, self.path_entry.get()) # Use current path entry value
        self.settings_path_entry.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
        def browse_default_path():
            directory = filedialog.askdirectory(parent=self.settings_window) # Ensure dialog is child of settings
            if directory:
                self.settings_path_entry.delete(0, "end")
                self.settings_path_entry.insert(0, directory)
        ctk.CTkButton(defaults_frame, text="浏览", width=60, command=browse_default_path).grid(row=3, column=2, padx=5, pady=5)

        # Filename Format
        ctk.CTkLabel(defaults_frame, text="文件名格式:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.settings_filename_format_entry = ctk.CTkEntry(defaults_frame)
        # Use the current filename_format instance variable
        self.settings_filename_format_entry.insert(0, self.filename_format)
        self.settings_filename_format_entry.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # Filename Format Explanation
        explanation = (
            "可用占位符:\n"
            "  {标题}: 原始标题 (已清理特殊字符)\n"
            "  {图片尺寸}: 图片的目标尺寸 (例如 '412x600' 或 '原始')\n"
            "  {类型}: 图片类型 ('竖图' 或 '横图', 仅当下载类型为'全部'且非腾讯时有效)\n" # Updated explanation
            "示例: {标题}_{类型}_{图片尺寸} -> 示例标题_竖图_412x600.jpg"
        )
        ctk.CTkLabel(defaults_frame, text=explanation, justify="left", anchor="w").grid(row=5, column=0, columnspan=3, padx=5, pady=(10, 5), sticky="w")


        # --- Cookies Tab ---
        cookies_frame = ctk.CTkFrame(tab_cookies, fg_color="transparent")
        cookies_frame.pack(expand=True, fill="both", padx=5, pady=5)
        cookies_frame.grid_columnconfigure(0, weight=1) # Give text boxes weight

        ctk.CTkLabel(cookies_frame, text="爱奇艺 Cookie:").pack(anchor="w", padx=5, pady=(5, 2))
        self.settings_iqiyi_cookie_box = ctk.CTkTextbox(cookies_frame, height=60, wrap="word")
        # --- Load current header value ---
        self.settings_iqiyi_cookie_box.insert("1.0", self.iqiyi_headers.get("Cookie", ""))
        self.settings_iqiyi_cookie_box.pack(expand=True, fill="x", padx=5, pady=(0, 5))

        ctk.CTkLabel(cookies_frame, text="腾讯视频 Cookie:").pack(anchor="w", padx=5, pady=(5, 2))
        self.settings_tencent_cookie_box = ctk.CTkTextbox(cookies_frame, height=60, wrap="word")
        # --- Load current header value ---
        self.settings_tencent_cookie_box.insert("1.0", self.tencent_headers.get("Cookie", ""))
        self.settings_tencent_cookie_box.pack(expand=True, fill="x", padx=5, pady=(0, 5))

        ctk.CTkLabel(cookies_frame, text="优酷视频 Cookie:").pack(anchor="w", padx=5, pady=(5, 2))
        self.settings_youku_cookie_box = ctk.CTkTextbox(cookies_frame, height=60, wrap="word")
        # --- Load current header value ---
        self.settings_youku_cookie_box.insert("1.0", self.youku_headers.get("Cookie", ""))
        self.settings_youku_cookie_box.pack(expand=True, fill="x", padx=5, pady=(0, 5))

        # --- Buttons Frame ---
        button_frame = ctk.CTkFrame(self.settings_window, fg_color="transparent")
        button_frame.pack(fill="x", padx=10, pady=(5, 10))

        # Add buttons using pack for horizontal layout
        close_button = ctk.CTkButton(button_frame, text="关闭", width=100, command=self.settings_window.destroy)
        close_button.pack(side="right", padx=(10, 0))

        save_button = ctk.CTkButton(button_frame, text="保存", width=100, command=self.save_settings)
        save_button.pack(side="right")


    # === 新增的页面功能方法 ===
    
    def setup_table_columns(self, mode):
        """设置表格列配置"""
        if mode == "batch":
            # 批量爬取模式
            columns = ("序号", "影片名称", "CID", "获取图片标题", "处理状态")
            self.batch_tree.configure(columns=columns)
            
            # 定义列标题
            self.batch_tree.heading("序号", text="序号")
            self.batch_tree.heading("影片名称", text="影片名称")
            self.batch_tree.heading("CID", text="CID")
            self.batch_tree.heading("获取图片标题", text="获取图片标题")
            self.batch_tree.heading("处理状态", text="处理状态")
            
            # 设置列宽
            self.batch_tree.column("序号", width=40, minwidth=30)
            self.batch_tree.column("影片名称", width=220, minwidth=150)
            self.batch_tree.column("CID", width=120, minwidth=100)
            self.batch_tree.column("获取图片标题", width=250, minwidth=200)
            self.batch_tree.column("处理状态", width=100, minwidth=80)
            
            # 更新标题
            self.table_title_label.configure(text="📊 批量爬取数据预览")
        
        self.current_table_mode = mode

    def on_batch_preset_change(self, choice):
        """当批量模式尺寸预设变化时更新尺寸显示和输入框状态"""
        if choice == "自定义尺寸":
            # Enable entries
            self.batch_v_size_entry.configure(state="normal")
            self.batch_h_size_entry.configure(state="normal")
            # Set placeholders
            self.batch_v_size_entry.delete(0, 'end')
            self.batch_v_size_entry.insert(0, "例如: 500x700")
            self.batch_h_size_entry.delete(0, 'end')
            self.batch_h_size_entry.insert(0, "例如: 700x500")
        else:
            # Get preset dimensions
            if choice in self.size_presets:
                vertical_size = self.size_presets[choice]["vertical"]
                horizontal_size = self.size_presets[choice]["horizontal"]
                
                # Format text
                v_text = f"{vertical_size[0]}x{vertical_size[1]}"
                h_text = f"{horizontal_size[0]}x{horizontal_size[1]}"

                # Update and disable vertical entry
                self.batch_v_size_entry.configure(state='normal')
                self.batch_v_size_entry.delete(0, 'end')
                self.batch_v_size_entry.insert(0, v_text)
                self.batch_v_size_entry.configure(state='disabled')

                # Update and disable horizontal entry
                self.batch_h_size_entry.configure(state='normal')
                self.batch_h_size_entry.delete(0, 'end')
                self.batch_h_size_entry.insert(0, h_text)
                self.batch_h_size_entry.configure(state='disabled')

    def browse_batch_path(self, entry_widget):
        """为批量模式选择路径"""
        directory = filedialog.askdirectory()
        if directory:
            entry_widget.delete(0, "end")
            entry_widget.insert(0, directory)

    def browse_delete_path(self, entry_widget):
        """为删除模式选择路径"""
        directory = filedialog.askdirectory()
        if directory:
            entry_widget.delete(0, "end")
            entry_widget.insert(0, directory)

    def update_batch_size_entries(self):
        """批量模式中根据选择的下载类型更新尺寸输入框状态"""
        selection = self.batch_orientation_var.get()
        
        if selection == "全部":
            # 全部下载：都启用
            self.batch_v_size_entry.configure(state="normal")
            self.batch_h_size_entry.configure(state="normal")
        elif selection == "竖图":
            # 仅下竖图：只启用竖图尺寸
            self.batch_v_size_entry.configure(state="normal")
            self.batch_h_size_entry.configure(state="disabled")
        elif selection == "横图":
            # 仅下横图：只启用横图尺寸
            self.batch_v_size_entry.configure(state="disabled")
            self.batch_h_size_entry.configure(state="normal")

    def select_excel_file(self):
        """选择Excel文件"""
        filetypes = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=filetypes
        )
        
        if file_path:
            self.excel_file_path = file_path
            self.batch_excel_entry.configure(state="normal")
            self.batch_excel_entry.delete(0, "end")
            self.batch_excel_entry.insert(0, file_path)
            self.batch_excel_entry.configure(state="readonly")
            
            # 加载Sheet名称
            self.load_excel_sheets()

    def load_excel_sheets(self):
        """加载Excel文件的Sheet名称"""
        try:
            # 获取所有sheet名称
            excel_file = pd.ExcelFile(self.excel_file_path)
            sheet_names = excel_file.sheet_names
            
            # 更新ComboBox
            self.batch_sheet_combo.configure(values=sheet_names)
            if sheet_names:
                # 优先选择默认存储sheet，支持多种可能的默认sheet名称
                default_sheets = ["海报爬取结果", "批量爬取结果", "爬取结果", "results", "结果"]
                selected_sheet = None
                
                # 尝试找到匹配的默认sheet
                for default_name in default_sheets:
                    if default_name in sheet_names:
                        selected_sheet = default_name
                        break
                
                # 如果没找到默认sheet，使用第一个
                if selected_sheet is None:
                    selected_sheet = sheet_names[0]
                
                self.batch_sheet_combo.set(selected_sheet)
                self.on_sheet_change(selected_sheet)
            
            self.status_label.configure(text=f"批量爬取Excel文件已加载，找到 {len(sheet_names)} 个Sheet")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载Excel文件失败: {str(e)}")
            self.status_label.configure(text="Excel文件加载失败")

    def on_sheet_change(self, selected_sheet):
        """当Sheet选择变化时，更新表格预览"""
        try:
            print(f"加载Sheet: {selected_sheet}")
            print(f"Excel文件路径: {self.excel_file_path}")
            
            # 读取选中的sheet
            self.batch_df = pd.read_excel(self.excel_file_path, sheet_name=selected_sheet)
            print(f"成功读取数据，行数: {len(self.batch_df)}, 列数: {len(self.batch_df.columns)}")
            print(f"列名: {list(self.batch_df.columns)}")
            
            # 确保有必要的列
            if len(self.batch_df.columns) < 2:
                messagebox.showerror("错误", "Excel表格至少需要2列：影片名称和CID")
                return
            
            # 添加处理状态相关列
            if "获取图片标题" not in self.batch_df.columns:
                self.batch_df["获取图片标题"] = ""
            if "处理状态" not in self.batch_df.columns:
                self.batch_df["处理状态"] = ""
            if "标题不一致" not in self.batch_df.columns:
                self.batch_df["标题不一致"] = False
            
            print(f"添加状态列后的列名: {list(self.batch_df.columns)}")
            
            # 切换到批量爬取模式显示
            if self.current_table_mode != "batch":
                self.setup_table_columns("batch")
            
            # 清空表格
            for item in self.batch_tree.get_children():
                self.batch_tree.delete(item)
            
            # 填充表格数据
            for index, row in self.batch_df.iterrows():
                sequence_num = index + 1  # 序号从1开始
                movie_name = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                cid = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                title = str(row.get("获取图片标题", "")) if pd.notna(row.get("获取图片标题", "")) else ""
                status = str(row.get("处理状态", "")) if pd.notna(row.get("处理状态", "")) else ""
                
                # 插入行数据
                item_id = self.batch_tree.insert("", "end", values=(sequence_num, movie_name, cid, title, status))
                
                # 检查已有数据是否标题不一致
                title_mismatch = row.get("标题不一致", False)
                if title_mismatch and pd.notna(title_mismatch) and title_mismatch:
                    self.batch_tree.item(item_id, tags=("mismatch",))
            
            # 启用开始按钮
            self.batch_start_button.configure(state="normal")
            self.batch_progress_label.configure(text=f"已加载 {len(self.batch_df)} 条数据，可以开始批量爬取")
            
            # 重置进度
            self.current_batch_row = 0
            self.batch_progress_bar.set(0)
            
            print(f"Sheet加载完成，数据行数: {len(self.batch_df)}")
            
        except Exception as e:
            print(f"加载Sheet数据失败: {e}")
            print(f"错误详情: {traceback.format_exc()}")
            messagebox.showerror("错误", f"加载Sheet数据失败: {str(e)}")
            self.status_label.configure(text="Sheet数据加载失败")

    def start_batch_crawling(self):
        """开始批量爬取"""
        try:
            if self.batch_df is None or len(self.batch_df) == 0:
                messagebox.showerror("错误", "没有可处理的数据")
                return
            
            if self.is_batch_processing:
                messagebox.showwarning("提示", "批量爬取正在进行中")
                return
            
            # 获取保存目录（从UI界面读取）
            h_dir = self.batch_h_path_entry.get().strip()
            v_dir = self.batch_v_path_entry.get().strip()
            
            print(f"批量爬取配置:")
            print(f"  横图目录: {h_dir}")
            print(f"  竖图目录: {v_dir}")
            print(f"  数据行数: {len(self.batch_df)}")
            
            try:
                os.makedirs(h_dir, exist_ok=True)
                os.makedirs(v_dir, exist_ok=True)
                print(f"  目录创建成功")
            except Exception as e:
                messagebox.showerror("错误", f"创建保存目录失败: {str(e)}")
                print(f"  目录创建失败: {e}")
                return
            
            # 设置处理状态
            self.is_batch_processing = True
            self.batch_paused = False
            self.current_batch_row = 0  # 重置为从头开始
            
            # 更新按钮状态
            self.batch_start_button.configure(state="disabled")
            self.batch_pause_button.configure(state="normal")
            
            # 启动后台线程
            self.batch_thread = threading.Thread(target=self.batch_crawling_worker, daemon=True)
            self.batch_thread.start()
            
            self.status_label.configure(text="批量爬取已开始...")
            print("批量爬取线程已启动")
            
        except Exception as e:
            print(f"启动批量爬取时出错: {e}")
            print(f"错误详情: {traceback.format_exc()}")
            messagebox.showerror("错误", f"启动批量爬取失败: {str(e)}")
            self.is_batch_processing = False
            self.batch_start_button.configure(state="normal")
            self.batch_pause_button.configure(state="disabled")

    def batch_crawling_worker(self):
        """批量爬取的后台工作线程"""
        try:
            total_rows = len(self.batch_df)
            print(f"开始批量爬取，总行数: {total_rows}")
            
            # 从当前行开始处理
            for i in range(self.current_batch_row, total_rows):
                try:
                    # 检查是否需要暂停
                    if self.batch_paused:
                        self.after(0, lambda: self.update_status("批量爬取已暂停"))
                        # 暂停时需要重置处理状态，以便继续时能重新启动
                        self.is_batch_processing = False
                        # 保存当前行索引，以便继续时从这一行开始
                        self.current_batch_row = i
                        return
                    
                    # 更新当前处理行（在开始处理前更新）
                    self.current_batch_row = i
                    row = self.batch_df.iloc[i]
                    
                    # 获取影片名称和CID
                    movie_name = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                    cid = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                    
                    print(f"处理第 {i+1}/{total_rows} 行: {movie_name} (CID: {cid})")
                    
                    if not movie_name or not cid:
                        # 跳过无效行
                        print(f"跳过无效行: 影片名称='{movie_name}', CID='{cid}'")
                        self.batch_df.at[i, "处理状态"] = "跳过"
                        self.after(0, lambda idx=i: self.update_table_row(idx))
                        continue
                    
                    # 更新状态为处理中
                    self.batch_df.at[i, "处理状态"] = "处理中..."
                    self.after(0, lambda idx=i: self.update_table_row(idx))
                    self.after(0, lambda: self.update_status(f"正在处理: {movie_name}"))
                    
                    # 平台优先级搜索
                    result, successful_platform = self.search_with_priority(movie_name)
                    
                    if result:
                        print(f"搜索成功，开始下载图片: {movie_name}，成功平台: {successful_platform}")
                        # 下载图片
                        success = self.batch_download_images(result, cid, movie_name, successful_platform)
                        if success:
                            obtained_title = result[0]  # 获取到的图片标题
                            self.batch_df.at[i, "获取图片标题"] = obtained_title
                            self.batch_df.at[i, "处理状态"] = "✔成功"
                            
                            # 名字校验：检查影片名称和获取图片标题是否一致
                            if self.normalize_text(movie_name) != self.normalize_text(obtained_title):
                                # 标记为不一致，后续在UI更新时会显示为红色
                                self.batch_df.at[i, "标题不一致"] = True
                                print(f"标题不一致: '{movie_name}' vs '{obtained_title}'")
                            else:
                                self.batch_df.at[i, "标题不一致"] = False
                                print(f"标题一致: '{movie_name}'")
                        else:
                            self.batch_df.at[i, "处理状态"] = "✘失败"
                            print(f"下载失败: {movie_name}")
                    else:
                        self.batch_df.at[i, "处理状态"] = "✘失败"
                        print(f"搜索失败: {movie_name}")
                    
                    # 更新表格
                    self.after(0, lambda idx=i: self.update_table_row(idx))
                    
                    # 更新进度
                    progress = (i + 1) / total_rows
                    self.after(0, lambda p=progress: self.batch_progress_bar.set(p))
                    self.after(0, lambda p=progress: self.batch_progress_label.configure(
                        text=f"已完成 {int(p * 100)}% ({i + 1}/{total_rows})"
                    ))
                    
                    # 每5行保存一次Excel
                    if (i + 1) % 5 == 0 or i == total_rows - 1:
                        self.save_batch_results()
                    
                    # 延迟2-3秒
                    time.sleep(random.uniform(2, 3))
                    
                except Exception as e:
                    print(f"处理第 {i+1} 行时出错: {e}")
                    print(f"错误详情: {traceback.format_exc()}")
                    self.batch_df.at[i, "处理状态"] = "✘错误"
                    self.after(0, lambda idx=i: self.update_table_row(idx))
                    continue
            
            # 处理完成
            print("批量爬取完成")
            self.is_batch_processing = False
            self.after(0, lambda: self.batch_start_button.configure(state="normal"))
            self.after(0, lambda: self.batch_pause_button.configure(state="disabled"))
            self.after(0, lambda: self.update_status("批量爬取完成"))
            
            # 询问是否打开表格文件
            self.after(0, self.ask_open_excel_file)
            
        except Exception as e:
            print(f"批量爬取工作线程出错: {e}")
            print(f"错误详情: {traceback.format_exc()}")
            self.is_batch_processing = False
            self.after(0, lambda: self.batch_start_button.configure(state="normal"))
            self.after(0, lambda: self.batch_pause_button.configure(state="disabled"))
            self.after(0, lambda: self.update_status("批量爬取出错"))

    def search_with_priority(self, movie_name):
        """按设置中的优先级搜索，返回 (result, successful_platform) 元组"""
        # 从设置中获取搜索优先级
        settings = self.load_settings()
        priority_list = settings.get("batch_search_priority", [
            "优酷视频-精确搜索",
            "爱奇艺-精确搜索", 
            "爱奇艺-普通搜索"
        ])
        
        # 转换优先级设置为搜索配置
        search_configs = []
        for priority_item in priority_list:
            if priority_item == "优酷视频-精确搜索":
                search_configs.append(("优酷视频", True))
            elif priority_item == "优酷视频-普通搜索":
                search_configs.append(("优酷视频", False))
            elif priority_item == "爱奇艺-精确搜索":
                search_configs.append(("爱奇艺", True))
            elif priority_item == "爱奇艺-普通搜索":
                search_configs.append(("爱奇艺", False))
            elif priority_item == "腾讯视频-精确搜索":
                search_configs.append(("腾讯视频", True))
            elif priority_item == "腾讯视频-普通搜索":
                search_configs.append(("腾讯视频", False))
        
        # 保存原始设置
        old_platform = self.selected_platform
        old_precise = self.precise_search_var.get()
        
        for platform, precise in search_configs:
            # 检查是否暂停
            if self.batch_paused:
                # 恢复原设置
                self.selected_platform = old_platform
                self.precise_search_var.set(old_precise)
                return None, None
            
            # 设置平台和精确搜索模式
            self.selected_platform = platform
            self.precise_search_var.set(precise)
            
            try:
                print(f"批量搜索: {movie_name} - 平台: {platform} - 精确搜索: {precise}")
                
                # 执行搜索
                if platform == "优酷视频":
                    results = self.search_youku(movie_name)
                elif platform == "爱奇艺":
                    results = self.search_iqiyi(movie_name)
                elif platform == "腾讯视频":
                    results = self.search_tencent(movie_name)
                else:
                    results = []
                
                # 如果找到结果，返回第一个结果和成功的平台
                if results and len(results) > 0:
                    print(f"找到结果: {results[0][0] if results[0] else '无标题'}")
                    # 不恢复原设置，保持成功搜索的平台
                    return results[0], platform
                else:
                    print(f"平台 {platform} 未找到结果")
                
                # 平台切换延迟1-2秒
                time.sleep(random.uniform(1, 2))
                
            except Exception as e:
                print(f"搜索平台 {platform} 出错: {e}")
                print(f"错误详情: {traceback.format_exc()}")
                continue
        
        # 所有平台都未找到结果，恢复原设置
        self.selected_platform = old_platform
        self.precise_search_var.set(old_precise)
        print(f"所有平台都未找到结果: {movie_name}")
        return None, None

    def batch_download_images(self, result, cid, movie_name, platform=None):
        """批量模式的图片下载 - 使用经典模式的下载逻辑"""
        try:
            print(f"开始批量下载图片: {movie_name}")
            print(f"结果格式: {type(result)}, 长度: {len(result) if hasattr(result, '__len__') else 'N/A'}")
            print(f"结果内容: {result}")
            
            # 获取当前选择的尺寸
            selected_preset = self.batch_preset_combo.get()
            target_vertical_size = self.size_presets[selected_preset]["vertical"]
            target_horizontal_size = self.size_presets[selected_preset]["horizontal"]
            
            print(f"选择的预设: {selected_preset}")
            print(f"竖图尺寸: {target_vertical_size}")
            print(f"横图尺寸: {target_horizontal_size}")
            
            # 从UI界面获取下载路径
            h_dir = self.batch_h_path_entry.get().strip()
            v_dir = self.batch_v_path_entry.get().strip()
            
            
            print(f"横图目录: {h_dir}")
            print(f"竖图目录: {v_dir}")
            
            # 获取下载类型选择
            download_type = self.batch_orientation_var.get()
            print(f"下载类型: {download_type}")
            
            # 解析结果 - 使用传入的platform参数，如果没有则使用当前选择的平台
            if platform is None:
                platform = self.selected_platform
            print(f"使用平台: {platform}")
            
            # 使用经典模式的逻辑处理单个结果
            title = ""
            horz_url = None
            vert_url = None
            base_img_url = None
            
            if platform == "优酷视频" and len(result) == 3:
                title, horz_url, vert_url = result
                print(f"优酷结果解析: 标题={title}, 横图={bool(horz_url)}, 竖图={bool(vert_url)}")
            elif len(result) == 2:  # 爱奇艺/腾讯
                title, base_img_url = result
                print(f"爱奇艺/腾讯结果解析: 标题={title}, 图片URL={bool(base_img_url)}")
            else:
                print(f"不支持的结果格式: 平台={platform}, 结果长度={len(result) if hasattr(result, '__len__') else 'N/A'}")
                return False
            
            # 准备下载任务列表（使用经典模式的逻辑）
            items_to_process = []
            
            if download_type == "全部":
                v_w, v_h = target_vertical_size
                # 竖图任务
                vertical_base_url = vert_url if platform == "优酷视频" else base_img_url
                if vertical_base_url:
                    items_to_process.append((title, vertical_base_url, "竖图", v_w, v_h, v_dir))
                
                # 横图任务（腾讯除外）
                if platform != "腾讯视频":
                    h_w, h_h = target_horizontal_size
                    horizontal_base_url = horz_url if platform == "优酷视频" else base_img_url
                    if horizontal_base_url:
                        items_to_process.append((title, horizontal_base_url, "横图", h_w, h_h, h_dir))
                        
            elif download_type == "竖图":
                v_w, v_h = target_vertical_size
                vertical_base_url = vert_url if platform == "优酷视频" else base_img_url
                if vertical_base_url:
                    items_to_process.append((title, vertical_base_url, "竖图", v_w, v_h, v_dir))
                    
            elif download_type == "横图":
                if platform != "腾讯视频":
                    h_w, h_h = target_horizontal_size
                    horizontal_base_url = horz_url if platform == "优酷视频" else base_img_url
                    if horizontal_base_url:
                        items_to_process.append((title, horizontal_base_url, "横图", h_w, h_h, h_dir))
            
            print(f"准备下载 {len(items_to_process)} 个任务")
            
            # 执行下载（使用经典模式的download_image函数）
            success_count = 0
            for title, base_url, img_type, width, height, download_path in items_to_process:
                print(f"下载任务: {title} - {img_type} - {width}x{height}")
                
                # 确定尺寸字符串
                if width == 0 or height == 0:
                    size_str = "原始尺寸"
                else:
                    size_str = f"{width}x{height}"
                
                # 确定文件名后缀
                filename_suffix = None
                if download_type == "全部":
                    # 检查是否两种类型都存在
                    has_both_types = False
                    if platform == "优酷视频":
                        has_both_types = bool(horz_url and vert_url)
                    elif platform == "爱奇艺":
                        has_both_types = True
                    
                    if has_both_types and platform != "腾讯视频":
                        filename_suffix = f"_{img_type}"
                
                # 调用经典模式的download_image函数
                if self.download_image(
                    base_url, 
                    cid,  # 使用CID作为标题
                    download_path,
                    suffix=filename_suffix, 
                    size_str=size_str,
                    platform=platform,
                    target_width=width, 
                    target_height=height,
                    img_type=img_type,
                    use_cid_filename=True  # 使用CID作为文件名
                ):
                    success_count += 1
                    print(f"✓ {img_type} 下载成功")
                else:
                    print(f"✗ {img_type} 下载失败")
            
            print(f"批量下载完成: {success_count}/{len(items_to_process)} 成功")
            return success_count > 0
            
        except Exception as e:
            print(f"批量下载图片失败 ({movie_name}): {e}")
            print(f"错误详情: {traceback.format_exc()}")
            return False



    def update_table_row(self, row_index):
        """更新表格中指定行的数据"""
        try:
            # 只有在批量爬取模式下才更新行数据
            if self.current_table_mode != "batch":
                return
                
            # 获取表格中的所有项目
            items = self.batch_tree.get_children()
            if row_index < len(items):
                item = items[row_index]
                row_data = self.batch_df.iloc[row_index]
                
                # 更新表格行
                sequence_num = row_index + 1  # 序号从1开始
                movie_name = str(row_data.iloc[0]) if pd.notna(row_data.iloc[0]) else ""
                cid = str(row_data.iloc[1]) if pd.notna(row_data.iloc[1]) else ""
                title = str(row_data.get("获取图片标题", "")) if pd.notna(row_data.get("获取图片标题", "")) else ""
                status = str(row_data.get("处理状态", "")) if pd.notna(row_data.get("处理状态", "")) else ""
                
                self.batch_tree.item(item, values=(sequence_num, movie_name, cid, title, status))
                
                # 检查是否标题不一致，如果是则设置红色标签
                title_mismatch = row_data.get("标题不一致", False)
                if title_mismatch and pd.notna(title_mismatch) and title_mismatch:
                    # 设置红色加粗标签用于获取图片标题列
                    self.batch_tree.set(item, "获取图片标题", title)
                    # 注意：ttk.Treeview不能对单个单元格设置样式，但我们可以在文本中添加标记
                    # 或者使用tags来标记整行
                    self.batch_tree.item(item, tags=("mismatch",))
                else:
                    self.batch_tree.item(item, tags=())
                
                # 滚动到当前行
                self.batch_tree.see(item)
                
        except Exception as e:
            print(f"更新表格行失败: {e}")

    def update_status(self, message):
        """更新状态标签"""
        self.status_label.configure(text=message)

    def toggle_batch_pause(self):
        """切换暂停/继续状态"""
        if self.batch_paused:
            # 继续
            self.batch_paused = False
            self.batch_pause_button.configure(text="暂停爬取")
            
            # 重新启动线程
            if not self.is_batch_processing:
                self.is_batch_processing = True
                self.batch_thread = threading.Thread(target=self.batch_crawling_worker, daemon=True)
                self.batch_thread.start()
            
            self.status_label.configure(text="批量爬取已继续...")
        else:
            # 暂停
            self.batch_paused = True
            self.batch_pause_button.configure(text="继续爬取")
            self.status_label.configure(text="正在暂停批量爬取...")
            
            # 保存当前进度
            self.save_batch_results()

    def save_batch_results(self):
        """保存批量处理结果到Excel（覆盖原文件）"""
        try:
            if self.batch_df is None:
                return
            
            # 获取当前选中的sheet名称
            current_sheet = self.batch_sheet_combo.get()
            
            # 使用openpyxl引擎保存Excel并设置格式
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            # 保存数据到Excel
            with pd.ExcelWriter(self.excel_file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                self.batch_df.to_excel(writer, sheet_name=current_sheet, index=False)
            
            # 重新加载工作簿以设置格式
            wb = load_workbook(self.excel_file_path)
            ws = wb[current_sheet]
            
            # 设置标题不一致的整行为红色加粗
            for index, row in self.batch_df.iterrows():
                title_mismatch = row.get("标题不一致", False)
                if title_mismatch and pd.notna(title_mismatch) and title_mismatch:
                    # Excel行号从2开始（第1行是表头）
                    excel_row = index + 2
                    # 对整行的所有列都设置红色加粗
                    for col_num in range(1, len(self.batch_df.columns) + 1):
                        cell = ws.cell(row=excel_row, column=col_num)
                        cell.font = Font(color="FF0000", bold=True)  # 红色加粗
            
            # 保存工作簿
            wb.save(self.excel_file_path)
            wb.close()
            
            print(f"批量结果已保存到原文件: {self.excel_file_path} (Sheet: {current_sheet})")
            
        except Exception as e:
            print(f"保存批量结果失败: {e}")

    def ask_open_excel_file(self):
        """询问是否打开Excel文件"""
        try:
            result = messagebox.askyesno("批量爬取完成", "批量爬取已完成！是否立即打开Excel表格查看结果？")
            if result and hasattr(self, 'excel_file_path') and self.excel_file_path:
                # 使用系统默认程序打开Excel文件
                import subprocess
                import platform
                
                if platform.system() == "Windows":
                    subprocess.Popen(['start', self.excel_file_path], shell=True)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.Popen(['open', self.excel_file_path])
                else:  # Linux
                    subprocess.Popen(['xdg-open', self.excel_file_path])
                    
        except Exception as e:
            print(f"打开Excel文件失败: {e}")

    # === 删除页面方法 ===
    def browse_delete_excel_file(self):
        """浏览选择要删除的Excel文件"""
        filetypes = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=filetypes
        )
        
        if file_path:
            self.delete_excel_file_path = file_path
            self.delete_excel_path_entry.configure(state="normal")
            self.delete_excel_path_entry.delete(0, "end")
            self.delete_excel_path_entry.insert(0, file_path)
            self.delete_excel_path_entry.configure(state="readonly")
            
            # 加载Sheet名称
            self.load_delete_excel_sheets()

    def load_delete_excel_sheets(self):
        """加载删除页面的Excel工作表"""
        try:
            excel_file = pd.ExcelFile(self.delete_excel_file_path)
            sheet_names = excel_file.sheet_names
            self.delete_sheet_combo.configure(values=sheet_names)
            if sheet_names:
                # 优先选择默认存储sheet，支持多种可能的默认sheet名称
                default_sheets = ["海报爬取结果", "批量爬取结果", "爬取结果", "results", "结果"]
                selected_sheet = None
                
                # 尝试找到匹配的默认sheet
                for default_name in default_sheets:
                    if default_name in sheet_names:
                        selected_sheet = default_name
                        break
                
                # 如果没找到默认sheet，使用第一个
                if selected_sheet is None:
                    selected_sheet = sheet_names[0]
                
                self.delete_sheet_combo.set(selected_sheet)
                self.on_delete_sheet_change(selected_sheet)
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def on_delete_sheet_change(self, sheet_name):
        """当删除页面工作表选择改变时"""
        if hasattr(self, 'delete_excel_file_path') and self.delete_excel_file_path:
            self.load_delete_excel_preview(sheet_name)

    def load_delete_excel_preview(self, sheet_name):
        """加载删除页面Excel预览"""
        try:
            self.delete_df = pd.read_excel(self.delete_excel_file_path, sheet_name=sheet_name)
            self.update_delete_table()
            # 移除状态标签更新，因为在删除页面UI中已被注释
            print(f"已加载删除页面Excel数据: {len(self.delete_df)} 行")
        except Exception as e:
            # 需求1: 取消弹窗，改为控制台输出
            print(f"读取工作表失败: {str(e)}")
            # 清空表格
            self.delete_df = None
            for item in self.delete_table.get_children():
                self.delete_table.delete(item)

    def update_delete_table(self):
        """更新删除表格显示"""
        # 清除现有数据
        for item in self.delete_table.get_children():
            self.delete_table.delete(item)
        
        # 清除列头选择框
        for widget in self.delete_column_frame.winfo_children():
            if isinstance(widget, ctk.CTkCheckBox):
                widget.destroy()
        self.delete_column_checkboxes.clear()
        
        if hasattr(self, 'delete_df') and self.delete_df is not None:
            # 设置列
            columns = list(self.delete_df.columns)
            self.delete_table["columns"] = columns
            
            # 创建列头选择复选框
            for col in columns:
                self.delete_table.heading(col, text=col)
                self.delete_table.column(col, width=100)
                
                # 为每个列创建复选框
                var = ctk.BooleanVar()
                checkbox = ctk.CTkCheckBox(
                    self.delete_column_frame, 
                    text=col,
                    variable=var,
                    command=lambda c=col, v=var: self.on_delete_column_select(c, v)
                )
                checkbox.pack(side="left", padx=5, pady=5)
                self.delete_column_checkboxes[col] = var
            
            # 添加数据
            for index, row in self.delete_df.iterrows():
                values = [str(row[col]) for col in columns]
                item_id = self.delete_table.insert("", "end", values=values)
            
            # 需求4: 自动选择"标题不一致"列为True的行 - 延迟执行确保UI组件已初始化
            self.after(100, self.auto_select_mismatch_rows)

    def on_delete_column_select(self, column_name, var):
        """当选择删除标准列时"""
        if var.get():
            # 取消其他列的选择
            for col, col_var in self.delete_column_checkboxes.items():
                if col != column_name:
                    col_var.set(False)
            
            self.selected_delete_key_column = column_name
            print(f"选择删除标准列: {column_name}")
        else:
            if self.selected_delete_key_column == column_name:
                self.selected_delete_key_column = None
                print("取消删除标准列选择")

    def on_delete_table_click(self, event):
        """处理表格点击事件，实现多选功能"""
        # 获取点击的项目
        item = self.delete_table.identify_row(event.y)
        if item:
            # 检查是否已选中
            current_selection = list(self.delete_table.selection())
            if item in current_selection:
                # 如果已选中，则取消选择
                self.delete_table.selection_remove(item)
            else:
                # 如果未选中，则添加到选择
                self.delete_table.selection_add(item)
        
        # 阻止默认行为
        return "break"

    def auto_select_mismatch_rows(self):
        """自动选择标题不一致的行"""
        if self.delete_df is None:
            return
        
        # 检查是否有"标题不一致"列
        mismatch_columns = [col for col in self.delete_df.columns if "标题不一致" in col]
        
        if mismatch_columns:
            mismatch_col = mismatch_columns[0]
            print(f"发现标题不一致列: {mismatch_col}")
            
            # 自动选择该列作为删除标准
            if mismatch_col in self.delete_column_checkboxes:
                self.delete_column_checkboxes[mismatch_col].set(True)
                # 调用列选择回调来确保正确设置选择状态
                self.on_delete_column_select(mismatch_col, self.delete_column_checkboxes[mismatch_col])
            
            # 清除当前选择
            self.delete_table.selection_remove(self.delete_table.selection())
            
            # 选择所有值为True的行
            items = self.delete_table.get_children()
            for i, item in enumerate(items):
                if i < len(self.delete_df):
                    row_data = self.delete_df.iloc[i]
                    value = row_data.get(mismatch_col, False)
                    # 检查是否为True（可能是字符串或布尔值）
                    if (isinstance(value, bool) and value) or (isinstance(value, str) and value.lower() == "true"):
                        self.delete_table.selection_add(item)
            
            selected_count = len(self.delete_table.selection())
            print(f"自动选择了 {selected_count} 行标题不一致的记录")

    def delete_selected_files(self):
        """删除选中的文件"""
        selected_items = self.delete_table.selection()
        if not selected_items:
            messagebox.showwarning("警告", "请先选择要删除的项目")
            return
        
        if not self.selected_delete_key_column:
            messagebox.showwarning("警告", "请先选择删除标准列")
            return
        
        # 获取文件路径
        h_dir = self.delete_h_path_entry.get().strip()
        v_dir = self.delete_v_path_entry.get().strip()
        
        if not h_dir or not v_dir:
            messagebox.showwarning("警告", "请先设置横图和竖图路径")
            return
        
        # 获取要删除的文件名
        delete_files = []
        for item in selected_items:
            # 获取行在DataFrame中的索引
            item_index = self.delete_table.index(item)
            if item_index < len(self.delete_df):
                row_data = self.delete_df.iloc[item_index]
                key_value = str(row_data.get(self.selected_delete_key_column, "")).strip()
                if key_value and key_value != "nan":
                    delete_files.append(key_value)
        
        if not delete_files:
            messagebox.showwarning("警告", "没有找到有效的删除文件名")
            return
        
        result = messagebox.askyesno("确认删除", 
                                   f"确定要删除选中的 {len(selected_items)} 个项目对应的文件吗？\n"
                                   f"删除标准列: {self.selected_delete_key_column}\n"
                                   f"将在横图和竖图文件夹中查找对应文件进行删除")
        
        if result:
            self.perform_file_deletion(delete_files, h_dir, v_dir)

    def perform_file_deletion(self, delete_files, h_dir, v_dir):
        """执行实际的文件删除操作"""
        import os
        import threading
        
        def delete_worker():
            deleted_count = 0
            failed_files = []
            
            for filename in delete_files:
                # 清理文件名
                clean_filename = self.sanitize_filename(filename)
                
                # 常见的图片扩展名
                extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp']
                
                # 在横图文件夹中查找并删除
                for ext in extensions:
                    file_path = os.path.join(h_dir, clean_filename + ext)
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            deleted_count += 1
                            print(f"已删除横图文件: {file_path}")
                            break
                        except Exception as e:
                            failed_files.append(f"横图-{clean_filename + ext}: {str(e)}")
                
                # 在竖图文件夹中查找并删除
                for ext in extensions:
                    file_path = os.path.join(v_dir, clean_filename + ext)
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            deleted_count += 1
                            print(f"已删除竖图文件: {file_path}")
                            break
                        except Exception as e:
                            failed_files.append(f"竖图-{clean_filename + ext}: {str(e)}")
            
            # 在主线程中显示结果
            self.after(0, lambda: self.show_delete_result(deleted_count, failed_files))
        
        # 在后台线程中执行删除操作
        delete_thread = threading.Thread(target=delete_worker, daemon=True)
        delete_thread.start()

    def show_delete_result(self, deleted_count, failed_files):
        """显示删除结果"""
        if failed_files:
            failed_info = "\n".join(failed_files[:10])  # 最多显示10个失败的文件
            if len(failed_files) > 10:
                failed_info += f"\n... 还有 {len(failed_files) - 10} 个文件删除失败"
            
            messagebox.showinfo("删除完成", 
                              f"删除完成!\n"
                              f"成功删除: {deleted_count} 个文件\n"
                              f"删除失败: {len(failed_files)} 个文件\n\n"
                              f"失败详情:\n{failed_info}")
        else:
            messagebox.showinfo("删除完成", f"删除完成!\n成功删除: {deleted_count} 个文件")

    def toggle_select_all_delete(self):
        """切换删除页面的全选状态"""
        if self.select_all_delete_button.cget("text") == "全选":
            # 全选
            for item in self.delete_table.get_children():
                self.delete_table.selection_add(item)
            self.select_all_delete_button.configure(text="取消全选")
        else:
            # 取消全选
            self.delete_table.selection_remove(self.delete_table.selection())
            self.select_all_delete_button.configure(text="全选")

    # === 设置页面方法 ===
    def create_basic_settings_tab(self, parent):
        """创建基础设置选项卡"""
        # 加载当前设置用于初始值
        current_settings_from_file = self.load_settings()
        
        # --- Defaults Tab ---
        defaults_frame = ctk.CTkFrame(parent, fg_color="transparent")
        defaults_frame.pack(expand=True, fill="both", padx=5, pady=5)
        defaults_frame.grid_columnconfigure(1, weight=1) # Give entry fields weight
        defaults_frame.grid_columnconfigure(2, weight=1) # Give entry fields weight for middle column
        defaults_frame.grid_columnconfigure(3, weight=1) # Give entry fields weight for right column

        # Default Platform
        ctk.CTkLabel(defaults_frame, text="默认平台:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        # Use the value currently in the main window's variable for consistency
        self.settings_platform_var = ctk.StringVar(value=self.platform_var.get())
        ctk.CTkComboBox(defaults_frame, values=["爱奇艺", "腾讯视频", "优酷视频"], variable=self.settings_platform_var).grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky="ew")

        # Default Precise Search
        ctk.CTkLabel(defaults_frame, text="默认精确搜索:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.settings_precise_var = ctk.BooleanVar(value=self.precise_search_var.get()) # Use current value
        ctk.CTkSwitch(defaults_frame, text="", variable=self.settings_precise_var).grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")

        # Default Download Type
        ctk.CTkLabel(defaults_frame, text="默认下载类型:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.settings_download_type_var = ctk.StringVar(value=self.orientation_var.get()) # Use current value
        radio_frame = ctk.CTkFrame(defaults_frame, fg_color="transparent")
        radio_frame.grid(row=2, column=1, columnspan=3, padx=0, pady=0, sticky="w")
        ctk.CTkRadioButton(radio_frame, text="全部下载", variable=self.settings_download_type_var, value="全部").pack(side="left", padx=5)
        ctk.CTkRadioButton(radio_frame, text="仅下竖图", variable=self.settings_download_type_var, value="竖图").pack(side="left", padx=5)
        ctk.CTkRadioButton(radio_frame, text="仅下横图", variable=self.settings_download_type_var, value="横图").pack(side="left", padx=5)

        # Default Download Path
        ctk.CTkLabel(defaults_frame, text="默认下载路径:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.settings_path_entry = ctk.CTkEntry(defaults_frame)
        self.settings_path_entry.insert(0, self.path_entry.get()) # Use current path entry value
        self.settings_path_entry.grid(row=3, column=1, padx=(5, 0), pady=5, sticky="ew")
        def browse_default_path():
            directory = filedialog.askdirectory()
            if directory:
                self.settings_path_entry.delete(0, "end")
                self.settings_path_entry.insert(0, directory)
        ctk.CTkButton(defaults_frame, text="浏览", width=100, command=browse_default_path).grid(row=3, column=2, padx=(0, 5), pady=5)

        # Default Poster Size
        ctk.CTkLabel(defaults_frame, text="默认海报尺寸:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.settings_poster_size_var = ctk.StringVar(value=current_settings_from_file.get("default_poster_size", "基础尺寸"))
        preset_options = list(self.size_presets.keys())
        ctk.CTkComboBox(defaults_frame, values=preset_options, variable=self.settings_poster_size_var, command=self.on_settings_poster_size_change).grid(row=4, column=1, columnspan=3, padx=5, pady=5, sticky="ew")

        # Default Vertical Size
        ctk.CTkLabel(defaults_frame, text="默认竖图尺寸:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.settings_vertical_size_entry = ctk.CTkEntry(defaults_frame, placeholder_text="宽度x高度")
        self.settings_vertical_size_entry.insert(0, current_settings_from_file.get("default_vertical_size", "412x600"))
        self.settings_vertical_size_entry.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

        # Default Horizontal Size
        ctk.CTkLabel(defaults_frame, text="默认横图尺寸:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        self.settings_horizontal_size_entry = ctk.CTkEntry(defaults_frame, placeholder_text="宽度x高度")
        self.settings_horizontal_size_entry.insert(0, current_settings_from_file.get("default_horizontal_size", "528x296"))
        self.settings_horizontal_size_entry.grid(row=6, column=1, padx=5, pady=5, sticky="ew")

        # Filename Format
        ctk.CTkLabel(defaults_frame, text="文件名格式:").grid(row=7, column=0, padx=5, pady=5, sticky="w")
        self.settings_filename_format_entry = ctk.CTkEntry(defaults_frame)
        # Use the current filename_format instance variable
        self.settings_filename_format_entry.insert(0, self.filename_format)
        self.settings_filename_format_entry.grid(row=7, column=1, columnspan=3, padx=5, pady=5, sticky="ew")

        # Filename Format Explanation
        explanation = (
            "可用占位符:\n"
            "  {标题}: 原始标题 (已清理特殊字符)\n"
            "  {图片尺寸}: 图片的目标尺寸 (例如 '412x600' 或 '原始')\n"
            "  {类型}: 图片类型 ('竖图' 或 '横图', 仅当下载类型为'全部'且非腾讯时有效)\n"
            "示例: {标题}_{类型}_{图片尺寸} -> 示例标题_竖图_412x600.jpg"
        )
        ctk.CTkLabel(defaults_frame, text=explanation, justify="left", anchor="w").grid(row=8, column=0, columnspan=4, padx=5, pady=(10, 5), sticky="w")

    def create_batch_settings_tab(self, parent):
        """创建批量设置选项卡"""
        # 加载当前设置用于初始值
        current_settings_from_file = self.load_settings()
        
        # --- Batch Settings Tab ---
        batch_frame = ctk.CTkFrame(parent, fg_color="transparent")
        batch_frame.pack(expand=True, fill="both", padx=5, pady=5)
        batch_frame.grid_columnconfigure(1, weight=1)

        # Batch Search Priority
        ctk.CTkLabel(batch_frame, text="搜索优先级:").grid(row=0, column=0, padx=5, pady=5, sticky="nw")
        
        # Priority listbox and controls
        priority_control_frame = ctk.CTkFrame(batch_frame, fg_color="transparent")
        priority_control_frame.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        
        # Available options
        self.available_priorities = [
            "优酷视频-精确搜索", "优酷视频-普通搜索",
            "爱奇艺-精确搜索", "爱奇艺-普通搜索", 
            "腾讯视频-精确搜索", "腾讯视频-普通搜索"
        ]
        
        # Current priority list
        current_priorities = current_settings_from_file.get("batch_search_priority", [
            "优酷视频-精确搜索", "爱奇艺-精确搜索", "爱奇艺-普通搜索"
        ])
        
        # Priority display and controls
        priority_display_frame = ctk.CTkFrame(priority_control_frame)
        priority_display_frame.pack(fill="x", pady=5)
        
        ctk.CTkLabel(priority_display_frame, text="当前优先级顺序:").pack(anchor="w")
        
        # 创建滚动框架用于显示优先级项目
        self.priority_scroll_frame = ctk.CTkScrollableFrame(priority_display_frame, height=120)
        self.priority_scroll_frame.pack(fill="x", pady=5)
        
        # Priority control buttons
        priority_buttons_frame = ctk.CTkFrame(priority_control_frame, fg_color="transparent")
        priority_buttons_frame.pack(fill="x", pady=5)
        
        self.priority_add_combo = ctk.CTkComboBox(priority_buttons_frame, values=self.available_priorities, width=200)
        self.priority_add_combo.pack(side="left", padx=5)
        
        ctk.CTkButton(priority_buttons_frame, text="添加", width=60, command=self.add_priority).pack(side="left", padx=5)
        self.remove_priority_button = ctk.CTkButton(priority_buttons_frame, text="删除", width=60, command=self.remove_priority, state="disabled")
        self.remove_priority_button.pack(side="left", padx=5)
        self.move_up_button = ctk.CTkButton(priority_buttons_frame, text="上移", width=60, command=self.move_priority_up, state="disabled")
        self.move_up_button.pack(side="left", padx=5)
        self.move_down_button = ctk.CTkButton(priority_buttons_frame, text="下移", width=60, command=self.move_priority_down, state="disabled")
        self.move_down_button.pack(side="left", padx=5)
        
        # Update priority display (在按钮创建后)
        self.current_priorities = current_priorities.copy()
        self.update_priority_display()

        # Batch Default Paths
        ctk.CTkLabel(batch_frame, text="默认横图路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.settings_batch_h_path_entry = ctk.CTkEntry(batch_frame)
        self.settings_batch_h_path_entry.insert(0, current_settings_from_file.get("batch_horizontal_path", r"D:\海报需求\第十一批\横图"))
        self.settings_batch_h_path_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        def browse_batch_h_path():
            directory = filedialog.askdirectory()
            if directory:
                self.settings_batch_h_path_entry.delete(0, "end")
                self.settings_batch_h_path_entry.insert(0, directory)
        ctk.CTkButton(batch_frame, text="浏览", width=60, command=browse_batch_h_path).grid(row=1, column=2, padx=5, pady=5)

        ctk.CTkLabel(batch_frame, text="默认竖图路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.settings_batch_v_path_entry = ctk.CTkEntry(batch_frame)
        self.settings_batch_v_path_entry.insert(0, current_settings_from_file.get("batch_vertical_path", r"D:\海报需求\第十一批\竖图"))
        self.settings_batch_v_path_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        def browse_batch_v_path():
            directory = filedialog.askdirectory()
            if directory:
                self.settings_batch_v_path_entry.delete(0, "end")
                self.settings_batch_v_path_entry.insert(0, directory)
        ctk.CTkButton(batch_frame, text="浏览", width=60, command=browse_batch_v_path).grid(row=2, column=2, padx=5, pady=5)

        # Batch Default Size
        ctk.CTkLabel(batch_frame, text="批量默认尺寸:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.settings_batch_size_var = ctk.StringVar(value=current_settings_from_file.get("batch_default_size", "基础尺寸"))
        preset_options = list(self.size_presets.keys())
        ctk.CTkComboBox(batch_frame, values=preset_options, variable=self.settings_batch_size_var, command=self.on_settings_batch_size_change).grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # Batch Default Vertical Size
        ctk.CTkLabel(batch_frame, text="批量默认竖图尺寸:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.settings_batch_vertical_size_entry = ctk.CTkEntry(batch_frame, placeholder_text="宽度x高度")
        self.settings_batch_vertical_size_entry.insert(0, current_settings_from_file.get("batch_default_vertical_size", "412x600"))
        self.settings_batch_vertical_size_entry.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # Batch Default Horizontal Size
        ctk.CTkLabel(batch_frame, text="批量默认横图尺寸:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.settings_batch_horizontal_size_entry = ctk.CTkEntry(batch_frame, placeholder_text="宽度x高度")
        self.settings_batch_horizontal_size_entry.insert(0, current_settings_from_file.get("batch_default_horizontal_size", "528x296"))
        self.settings_batch_horizontal_size_entry.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

    def create_cookies_settings_tab(self, parent):
        """创建Cookie设置选项卡"""
        # --- Cookies Tab ---
        cookies_frame = ctk.CTkFrame(parent, fg_color="transparent")
        cookies_frame.pack(expand=True, fill="both", padx=5, pady=5)
        cookies_frame.grid_columnconfigure(0, weight=1) # Give text boxes weight

        ctk.CTkLabel(cookies_frame, text="爱奇艺 Cookie:").pack(anchor="w", padx=5, pady=(5, 2))
        self.settings_iqiyi_cookie_box = ctk.CTkTextbox(cookies_frame, height=60, wrap="word")
        # --- Load current header value ---
        self.settings_iqiyi_cookie_box.insert("1.0", self.iqiyi_headers.get("Cookie", ""))
        self.settings_iqiyi_cookie_box.pack(expand=True, fill="x", padx=5, pady=(0, 5))

        ctk.CTkLabel(cookies_frame, text="腾讯视频 Cookie:").pack(anchor="w", padx=5, pady=(5, 2))
        self.settings_tencent_cookie_box = ctk.CTkTextbox(cookies_frame, height=60, wrap="word")
        # --- Load current header value ---
        self.settings_tencent_cookie_box.insert("1.0", self.tencent_headers.get("Cookie", ""))
        self.settings_tencent_cookie_box.pack(expand=True, fill="x", padx=5, pady=(0, 5))

        ctk.CTkLabel(cookies_frame, text="优酷视频 Cookie:").pack(anchor="w", padx=5, pady=(5, 2))
        self.settings_youku_cookie_box = ctk.CTkTextbox(cookies_frame, height=60, wrap="word")
        # --- Load current header value ---
        self.settings_youku_cookie_box.insert("1.0", self.youku_headers.get("Cookie", ""))
        self.settings_youku_cookie_box.pack(expand=True, fill="x", padx=5, pady=(0, 5))

    def save_page_settings(self):
        """保存页面设置"""
        try:
            # 获取当前设置的默认路径
            current_default_path = self.settings_path_entry.get().strip()
            
            # 检查默认路径是否发生变化，如果变化则同步更新批量和删除路径
            current_settings = self.load_settings()
            old_default_path = current_settings.get("default_path", "")
            
            # 如果默认路径发生变化，自动更新相关路径
            if current_default_path != old_default_path and current_default_path:
                # 生成新的横图和竖图路径
                new_h_path = os.path.join(current_default_path, "横图")
                new_v_path = os.path.join(current_default_path, "竖图")
                
                # 更新设置窗口中的批量路径
                self.settings_batch_h_path_entry.delete(0, "end")
                self.settings_batch_h_path_entry.insert(0, new_h_path)
                self.settings_batch_v_path_entry.delete(0, "end")
                self.settings_batch_v_path_entry.insert(0, new_v_path)
                
                print(f"默认路径已更改，自动同步批量路径:")
                print(f"  横图路径: {new_h_path}")
                print(f"  竖图路径: {new_v_path}")
            
            settings_to_save = {
                "default_platform": self.settings_platform_var.get(),
                "default_precise": self.settings_precise_var.get(),
                "default_download_type": self.settings_download_type_var.get(),
                "default_path": current_default_path,
                "default_poster_size": self.settings_poster_size_var.get(),
                "default_vertical_size": self.settings_vertical_size_entry.get().strip(),
                "default_horizontal_size": self.settings_horizontal_size_entry.get().strip(),
                "filename_format": self.settings_filename_format_entry.get().strip(),
                "batch_search_priority": self.current_priorities,
                "batch_horizontal_path": self.settings_batch_h_path_entry.get().strip(),
                "batch_vertical_path": self.settings_batch_v_path_entry.get().strip(),
                "batch_default_size": self.settings_batch_size_var.get(),
                "batch_default_vertical_size": self.settings_batch_vertical_size_entry.get().strip(),
                "batch_default_horizontal_size": self.settings_batch_horizontal_size_entry.get().strip(),
                "delete_horizontal_path": self.settings_batch_h_path_entry.get().strip(),  # 使用相同的路径
                "delete_vertical_path": self.settings_batch_v_path_entry.get().strip(),    # 使用相同的路径
                "iqiyi_cookie": self.settings_iqiyi_cookie_box.get("1.0", "end-1c").strip(),
                "tencent_cookie": self.settings_tencent_cookie_box.get("1.0", "end-1c").strip(),
                "youku_cookie": self.settings_youku_cookie_box.get("1.0", "end-1c").strip()
            }

            # Validate filename format roughly
            if not settings_to_save["filename_format"]:
                 settings_to_save["filename_format"] = self.get_default_settings()["filename_format"]
                 self.settings_filename_format_entry.delete(0, 'end')
                 self.settings_filename_format_entry.insert(0, settings_to_save["filename_format"])
                 messagebox.showwarning("格式错误", "文件名格式不能为空，已重置为默认值。")
                 return

            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(settings_to_save, f, ensure_ascii=False, indent=4)

            print("Settings saved successfully to config.json")

            # --- Apply changes immediately to the running application ---
            new_default_platform = settings_to_save["default_platform"]
            self.platform_var.set(new_default_platform)
            self.selected_platform = new_default_platform

            self.precise_search_var.set(settings_to_save["default_precise"])
            self.orientation_var.set(settings_to_save["default_download_type"])

            self.path_entry.delete(0, 'end')
            self.path_entry.insert(0, settings_to_save["default_path"])

            self.filename_format = settings_to_save["filename_format"]

            # 更新批量模式的默认路径
            self.batch_h_path_entry.delete(0, 'end')
            self.batch_h_path_entry.insert(0, settings_to_save["batch_horizontal_path"])
            self.batch_v_path_entry.delete(0, 'end')
            self.batch_v_path_entry.insert(0, settings_to_save["batch_vertical_path"])
            
            # 更新批量删除的默认路径
            self.delete_h_path_entry.delete(0, 'end')
            self.delete_h_path_entry.insert(0, settings_to_save["delete_horizontal_path"])
            self.delete_v_path_entry.delete(0, 'end')
            self.delete_v_path_entry.insert(0, settings_to_save["delete_vertical_path"])

            self.iqiyi_headers['Cookie'] = settings_to_save["iqiyi_cookie"]
            self.tencent_headers['Cookie'] = settings_to_save["tencent_cookie"]
            self.youku_headers['Cookie'] = settings_to_save["youku_cookie"]

            # Refresh size entries based on current preset
            self.on_preset_change(self.preset_combo.get())

            # 准备保存成功消息
            success_message = "设置已成功保存并应用。"
            if current_default_path != old_default_path and current_default_path:
                success_message += "\n\n已自动同步路径设置：\n• 批量爬取路径已更新\n• 批量删除路径已更新"

            messagebox.showinfo("设置已保存", success_message)

        except Exception as e:
            messagebox.showerror("保存错误", f"保存设置时出错: {e}")
            print(f"Error saving settings: {e}")
            print(traceback.format_exc())

    def update_priority_display(self):
        """更新优先级显示"""
        # 清除现有的优先级框架
        for frame in self.priority_frames:
            frame.destroy()
        self.priority_frames.clear()
        
        # 重新创建优先级项目
        for i, priority in enumerate(self.current_priorities):
            self.create_priority_item(i, priority)
        
        # 更新按钮状态
        self.update_priority_buttons()
    
    def create_priority_item(self, index, priority_text):
        """创建单个优先级项目"""
        # 创建项目框架
        item_frame = ctk.CTkFrame(self.priority_scroll_frame, height=35)
        item_frame.pack(fill="x", padx=2, pady=2)
        item_frame.grid_propagate(False)
        item_frame.grid_columnconfigure(1, weight=1)
        
        # 序号标签
        number_label = ctk.CTkLabel(item_frame, text=f"{index + 1}.", width=30)
        number_label.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w")
        
        # 优先级文本标签
        text_label = ctk.CTkLabel(item_frame, text=priority_text, anchor="w")
        text_label.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        # 绑定点击事件
        def on_click(event, idx=index):
            self.select_priority_item(idx)
        
        item_frame.bind("<Button-1>", on_click)
        number_label.bind("<Button-1>", on_click)
        text_label.bind("<Button-1>", on_click)
        
        # 存储框架引用
        self.priority_frames.append(item_frame)
        
        return item_frame
    
    def select_priority_item(self, index):
        """选择优先级项目"""
        # 重置所有项目的颜色
        for i, frame in enumerate(self.priority_frames):
            if i == index:
                # 选中状态：深色背景
                frame.configure(fg_color=("gray75", "gray25"))
                self.selected_priority_index = index
            else:
                # 未选中状态：默认背景
                frame.configure(fg_color=("gray92", "gray14"))
        
        # 更新按钮状态
        self.update_priority_buttons()
    
    def update_priority_buttons(self):
        """更新按钮状态"""
        # 检查按钮是否已创建
        if not hasattr(self, 'remove_priority_button'):
            return
            
        has_selection = self.selected_priority_index >= 0
        has_items = len(self.current_priorities) > 0
        
        # 删除按钮：有选择时启用
        if has_selection:
            self.remove_priority_button.configure(state="normal")
        else:
            self.remove_priority_button.configure(state="disabled")
        
        # 上移下移按钮：有选择且有多个项目时启用
        if has_selection and len(self.current_priorities) > 1:
            self.move_up_button.configure(state="normal")
            self.move_down_button.configure(state="normal")
        else:
            self.move_up_button.configure(state="disabled")
            self.move_down_button.configure(state="disabled")
    
    def add_priority(self):
        """添加搜索优先级"""
        selected = self.priority_add_combo.get()
        if selected and selected not in self.current_priorities:
            self.current_priorities.append(selected)
            self.selected_priority_index = -1  # 清除选择
            self.update_priority_display()
    
    def move_priority_up(self):
        """上移选中的优先级"""
        if self.selected_priority_index <= 0:
            return
        
        # 交换当前项和上一项
        current_index = self.selected_priority_index
        self.current_priorities[current_index], self.current_priorities[current_index - 1] = \
            self.current_priorities[current_index - 1], self.current_priorities[current_index]
        
        # 更新选中索引
        self.selected_priority_index = current_index - 1
        self.update_priority_display()
        
        # 重新选中移动后的项目
        self.select_priority_item(self.selected_priority_index)
    
    def move_priority_down(self):
        """下移选中的优先级"""
        if self.selected_priority_index < 0 or self.selected_priority_index >= len(self.current_priorities) - 1:
            return
        
        # 交换当前项和下一项
        current_index = self.selected_priority_index
        self.current_priorities[current_index], self.current_priorities[current_index + 1] = \
            self.current_priorities[current_index + 1], self.current_priorities[current_index]
        
        # 更新选中索引
        self.selected_priority_index = current_index + 1
        self.update_priority_display()
        
        # 重新选中移动后的项目
        self.select_priority_item(self.selected_priority_index)
    
    def remove_priority(self):
        """删除选中的优先级"""
        if self.selected_priority_index < 0 or self.selected_priority_index >= len(self.current_priorities):
            return
        
        # 删除选中的项目
        self.current_priorities.pop(self.selected_priority_index)
        
        # 调整选中索引
        if self.selected_priority_index >= len(self.current_priorities):
            self.selected_priority_index = len(self.current_priorities) - 1
        if len(self.current_priorities) == 0:
            self.selected_priority_index = -1
        
        self.update_priority_display()
        
        # 如果还有项目，重新选中调整后的项目
        if self.selected_priority_index >= 0:
            self.select_priority_item(self.selected_priority_index)

    def on_settings_poster_size_change(self, choice):
        """当基础设置中的默认海报尺寸改变时，更新竖图和横图尺寸输入框"""
        if choice in self.size_presets:
            vertical_size = self.size_presets[choice]["vertical"]
            horizontal_size = self.size_presets[choice]["horizontal"]
            
            # 更新输入框
            if hasattr(self, 'settings_vertical_size_entry'):
                self.settings_vertical_size_entry.delete(0, "end")
                if vertical_size != (0, 0):  # 不是自定义尺寸
                    self.settings_vertical_size_entry.insert(0, f"{vertical_size[0]}x{vertical_size[1]}")
            
            if hasattr(self, 'settings_horizontal_size_entry'):
                self.settings_horizontal_size_entry.delete(0, "end")
                if horizontal_size != (0, 0):  # 不是自定义尺寸
                    self.settings_horizontal_size_entry.insert(0, f"{horizontal_size[0]}x{horizontal_size[1]}")

    def on_settings_batch_size_change(self, choice):
        """当批量设置中的默认尺寸改变时，更新竖图和横图尺寸输入框"""
        if choice in self.size_presets:
            vertical_size = self.size_presets[choice]["vertical"]
            horizontal_size = self.size_presets[choice]["horizontal"]
            
            # 更新输入框
            if hasattr(self, 'settings_batch_vertical_size_entry'):
                self.settings_batch_vertical_size_entry.delete(0, "end")
                if vertical_size != (0, 0):  # 不是自定义尺寸
                    self.settings_batch_vertical_size_entry.insert(0, f"{vertical_size[0]}x{vertical_size[1]}")
            
            if hasattr(self, 'settings_batch_horizontal_size_entry'):
                self.settings_batch_horizontal_size_entry.delete(0, "end")
                if horizontal_size != (0, 0):  # 不是自定义尺寸
                    self.settings_batch_horizontal_size_entry.insert(0, f"{horizontal_size[0]}x{horizontal_size[1]}")

if __name__ == "__main__":
    # Ensure PIL uses modern resampling if available
    if not hasattr(Image, 'Resampling'):
        Image.Resampling = Image # Patch for older PIL versions

    app = MultiPlatformImageDownloader()
    app.mainloop()