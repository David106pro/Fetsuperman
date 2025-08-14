import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import datetime
import re
from tqdm import tqdm
from openpyxl import load_workbook
import subprocess
import platform

class TimeFormatConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("时间格式转换工具")
        self.root.geometry("900x650")
        
        # 设置主题样式
        self.style = ttk.Style()
        self.style.configure("TButton", font=("等线", 10))
        self.style.configure("TLabel", font=("等线", 10))
        self.style.configure("TLabelframe", font=("等线", 10))
        self.style.configure("TLabelframe.Label", font=("等线", 10, "bold"))
        
        # 自定义进度条样式
        self.style.configure("Custom.Horizontal.TProgressbar", 
                        thickness=25,
                        background='#4CAF50')
        
        self.file_path = tk.StringVar()
        self.selected_columns = []
        self.format_option = tk.StringVar(value="YYYY-MM-DD HH:MM:SS")
        self.sheet_option = tk.StringVar()
        self.sheet_names = []
        self.df = None
        
        self.create_widgets()
    
    def create_widgets(self):
        # 主容器框架
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding=10)
        file_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(file_frame, text="Excel文件路径:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.file_path, width=60).grid(row=0, column=1, padx=5, pady=5)
        browse_btn = ttk.Button(file_frame, text="浏览", command=self.browse_file)
        browse_btn.grid(row=0, column=2, padx=5, pady=5)
        
        # 格式选择区域
        options_frame = ttk.LabelFrame(main_frame, text="选项设置", padding=10)
        options_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 时间格式选择
        ttk.Label(options_frame, text="时间格式:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        format_options = ["YYYY-MM-DD HH:MM:SS", "YYYY-MM-DD"]
        format_combo = ttk.Combobox(options_frame, textvariable=self.format_option, values=format_options, state="readonly", width=20)
        format_combo.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        format_combo.current(0)
        
        # Sheet选择
        ttk.Label(options_frame, text="工作表:").grid(row=0, column=2, sticky=tk.W, padx=20, pady=5)
        self.sheet_combo = ttk.Combobox(options_frame, textvariable=self.sheet_option, state="readonly", width=20)
        self.sheet_combo.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        
        # 表格预览区域
        preview_frame = ttk.LabelFrame(main_frame, text="Excel预览", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 创建Treeview来显示Excel数据
        self.tree_frame = ttk.Frame(preview_frame)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # 提示信息标签
        self.info_label = ttk.Label(main_frame, text="点击列标题可选择/取消选择该列进行转换", font=("微软雅黑", 9, "italic"))
        self.info_label.pack(fill=tk.X, padx=5, pady=(0, 5))
        
        # 底部控制区域
        control_frame = ttk.Frame(main_frame, padding=10)
        control_frame.pack(fill=tk.X, padx=5, pady=(5, 10))
        
        # 更美观的进度条
        progress_frame = ttk.Frame(control_frame)
        progress_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.progress_text = ttk.Label(progress_frame, text="准备就绪", anchor="w")
        self.progress_text.pack(fill=tk.X, padx=5, pady=(0, 3))
        
        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, 
                                       length=200, mode='determinate', 
                                       style="Custom.Horizontal.TProgressbar")
        self.progress.pack(fill=tk.X, expand=True, padx=5)
        
        # 更美观的开始按钮
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(side=tk.RIGHT, padx=5)
        
        # 自定义按钮样式
        self.style.configure("Start.TButton", font=("微软雅黑", 11, "bold"), background="#4CAF50")
        self.start_button = ttk.Button(btn_frame, text="开始转换", command=self.start_conversion,
                                      style="Start.TButton", width=15)
        self.start_button.pack(pady=5)
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.file_path.set(file_path)
            self.load_sheet_names()
    
    def load_sheet_names(self):
        try:
            # 读取Excel文件的所有sheet名称
            xls = pd.ExcelFile(self.file_path.get())
            self.sheet_names = xls.sheet_names
            self.sheet_combo['values'] = self.sheet_names
            
            # 选择第一个sheet
            if self.sheet_names:
                self.sheet_option.set(self.sheet_names[0])
                self.load_excel_preview()
                
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel工作表时出错: {str(e)}")
    
    def on_sheet_selected(self, event=None):
        if self.sheet_option.get():
            self.load_excel_preview()
    
    def load_excel_preview(self):
        try:
            # 清除旧的treeview
            for widget in self.tree_frame.winfo_children():
                widget.destroy()
            
            # 读取Excel文件的指定sheet
            sheet_name = self.sheet_option.get()
            self.df = pd.read_excel(self.file_path.get(), sheet_name=sheet_name)
            
            # 创建带滚动条的Treeview
            self.tree_scroll_y = ttk.Scrollbar(self.tree_frame, orient=tk.VERTICAL)
            self.tree_scroll_x = ttk.Scrollbar(self.tree_frame, orient=tk.HORIZONTAL)
            self.tree = ttk.Treeview(self.tree_frame, yscrollcommand=self.tree_scroll_y.set, 
                                     xscrollcommand=self.tree_scroll_x.set, selectmode='none')
            
            self.tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
            self.tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
            self.tree.pack(expand=True, fill=tk.BOTH)
            
            self.tree_scroll_y.config(command=self.tree.yview)
            self.tree_scroll_x.config(command=self.tree.xview)
            
            # 设置列
            self.tree['columns'] = list(self.df.columns)
            self.tree['show'] = 'headings'
            
            # 重置选中的列
            self.selected_columns = []
            
            # 添加列标题和设置点击事件
            for column in self.df.columns:
                col_id = str(column)
                self.tree.heading(col_id, text=str(column))
                self.tree.column(col_id, width=100, anchor='center')
                
                # 为标题添加点击事件
                self.tree.heading(col_id, command=lambda _col=col_id: self.toggle_column_selection(_col))
            
            # 添加数据行 - 使用交替行颜色
            for i, row in self.df.head(100).iterrows():  # 只预览前100行
                values = [str(row[col]) for col in self.df.columns]
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                self.tree.insert('', 'end', values=values, tags=(tag,))
            
            # 设置交替行颜色
            self.tree.tag_configure('oddrow', background='#f0f0f0')
            self.tree.tag_configure('evenrow', background='white')
            
            self.progress_text.config(text=f"已加载工作表 '{sheet_name}'，共{len(self.df)}行数据，预览显示前100行")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载Excel文件时出错: {str(e)}")
    
    def toggle_column_selection(self, col_id):
        if col_id in self.selected_columns:
            self.selected_columns.remove(col_id)
            # 恢复标题颜色
            self.tree.heading(col_id, text=col_id)
        else:
            self.selected_columns.append(col_id)
            # 改变标题颜色或添加标记
            self.tree.heading(col_id, text=f"✓ {col_id}")
        
        # 更新选择状态提示
        if self.selected_columns:
            self.progress_text.config(text=f"已选择{len(self.selected_columns)}列进行转换")
        else:
            self.progress_text.config(text="请选择要转换的列")
    
    def convert_time_format(self, value):
        if pd.isna(value):
            return value
        
        value = str(value).strip()
        
        # 处理特殊情况 "超过2年"
        if "超过2年" in value:
            return "3000-01-01"
        
        # 尝试匹配常见日期格式
        date_formats = [
            # 标准日期时间格式
            r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?',
            # 中文日期格式
            r'(\d{4})年(\d{1,2})月(\d{1,2})日(?:\s*(\d{1,2})时(\d{1,2})分(?:(\d{1,2})秒)?)?',
            # 短年份格式
            r'(\d{2})[/-](\d{1,2})[/-](\d{1,2})(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?',
        ]
        
        for pattern in date_formats:
            match = re.search(pattern, value)
            if match:
                groups = match.groups()
                year = groups[0]
                if len(year) == 2:  # 处理短年份
                    year = f"20{year}" if int(year) < 50 else f"19{year}"
                month = groups[1].zfill(2)
                day = groups[2].zfill(2)
                
                if self.format_option.get() == "YYYY-MM-DD":
                    return f"{year}-{month}-{day}"
                else:  # YYYY-MM-DD HH:MM:SS
                    hour = "00"
                    minute = "00"
                    second = "00"
                    
                    if groups[3] is not None:
                        hour = groups[3].zfill(2)
                    if groups[4] is not None:
                        minute = groups[4].zfill(2)
                    if groups[5] is not None:
                        second = groups[5].zfill(2)
                    
                    return f"{year}-{month}-{day} {hour}:{minute}:{second}"
        
        # 尝试用pandas解析日期
        try:
            dt = pd.to_datetime(value)
            if self.format_option.get() == "YYYY-MM-DD":
                return dt.strftime("%Y-%m-%d")
            else:  # YYYY-MM-DD HH:MM:SS
                return dt.strftime("%Y-%m-%d %H:%M:%S")
        except:
            # 无法解析时返回原值
            return value
    
    def start_conversion(self):
        if self.df is None:
            messagebox.showerror("错误", "请先加载Excel文件!")
            return
        
        if not self.selected_columns:
            messagebox.showerror("错误", "请选择至少一列进行转换!")
            return
        
        try:
            # 更新开始按钮状态
            self.start_button.config(state="disabled")
            self.progress_text.config(text="正在转换中...")
            self.root.update()
            
            # 获取文件路径
            file_path = self.file_path.get()
            file_name, file_ext = os.path.splitext(file_path)
            output_path = f"{file_name}_转换后{file_ext}"
            
            # 如果是新文件，先复制原始文件
            if not os.path.exists(output_path):
                import shutil
                shutil.copy2(file_path, output_path)
            
            # 使用openpyxl加载工作簿，以保留格式
            wb = load_workbook(output_path)
            sheet_name = self.sheet_option.get()
            ws = wb[sheet_name]
            
            # 初始化进度条
            total_cells = len(self.df) * len(self.selected_columns)
            self.progress['maximum'] = total_cells
            self.progress['value'] = 0
            self.root.update()
            
            # 获取列名到列索引的映射
            col_indices = {}
            for i, col in enumerate(ws[1]):
                if col.value:
                    col_indices[str(col.value)] = i
            
            # 转换每个选中的列
            for col_idx, col_name in enumerate(self.selected_columns):
                if col_name not in col_indices:
                    continue
                
                # 获取列索引（0基）
                sheet_col_idx = col_indices[col_name]
                new_col_name = f"{col_name}_转换后"
                
                # 更新进度文本
                self.progress_text.config(text=f"正在处理列: {col_name} ({col_idx+1}/{len(self.selected_columns)})")
                self.root.update()
                
                # 插入新列
                ws.insert_cols(sheet_col_idx + 2)  # +2因为openpyxl列是从1开始的，且我们要在后面插入
                
                # 设置列标题
                title_cell = ws.cell(row=1, column=sheet_col_idx + 2)
                title_cell.value = new_col_name
                
                # 应用转换
                for idx, row in enumerate(self.df[col_name].items(), 2):  # 从第2行开始（跳过标题行）
                    value = row[1]
                    converted = self.convert_time_format(value)
                    ws.cell(row=idx, column=sheet_col_idx + 2).value = converted
                    
                    # 更新进度条
                    self.progress['value'] += 1
                    if idx % 100 == 0:  # 避免过于频繁的UI更新
                        self.progress_text.config(text=f"处理列: {col_name} - {idx}/{len(self.df)}行")
                        self.root.update()
                
                # 更新列索引，因为我们插入了新列
                for k in list(col_indices.keys()):
                    if col_indices[k] > sheet_col_idx:
                        col_indices[k] += 1
            
            # 保存工作簿
            self.progress_text.config(text="正在保存文件...")
            self.root.update()
            wb.save(output_path)
            
            self.progress['value'] = total_cells  # 完成进度条
            self.progress_text.config(text=f"转换完成! 已保存到: {os.path.basename(output_path)}")
            
            # 添加自动打开Excel文件的功能
            self.open_excel_file(output_path)
            
            messagebox.showinfo("成功", f"时间格式转换完成！\n结果已保存到: {output_path}")
            
        except Exception as e:
            messagebox.showerror("错误", f"转换过程中出错: {str(e)}")
            self.progress_text.config(text="转换失败")
            
        finally:
            # 恢复开始按钮状态
            self.start_button.config(state="normal")
    
    def open_excel_file(self, file_path):
        """根据操作系统打开Excel文件"""
        try:
            # 根据不同操作系统使用不同的打开方式
            system = platform.system()
            if system == 'Windows':
                os.startfile(file_path)
            elif system == 'Darwin':  # macOS
                subprocess.call(['open', file_path])
            else:  # Linux and other systems
                subprocess.call(['xdg-open', file_path])
            
            self.progress_text.config(text=f"已自动打开文件: {os.path.basename(file_path)}")
        except Exception as e:
            self.progress_text.config(text=f"文件已保存，但无法自动打开: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = TimeFormatConverterApp(root)
    root.mainloop()