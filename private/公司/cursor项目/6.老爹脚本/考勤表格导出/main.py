# -*- coding: utf-8 -*-
# 使用中文注释说明代码功能
# main.py：公司考勤表格自动生成脚本，支持UI操作，自动读取日明细报表和休假出差表，输出员工考勤表

import os
import sys
import json
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import pandas as pd
from datetime import datetime, timedelta
import calendar
import shutil
import threading

# 配置文件路径
CONFIG_DIR = os.path.join(os.path.dirname(__file__), 'setting')
CONFIG_PATH = os.path.join(CONFIG_DIR, 'config.json')

# 默认路径
DEFAULT_DETAIL_PATH = r'C:\Users\fucha\PycharmProjects\python\公司\cursor项目\6.老爹脚本\考勤表格导出\日明细报表_20250603101227.xlsx'
DEFAULT_LEAVE_PATH = r'C:\Users\fucha\PycharmProjects\python\公司\cursor项目\6.老爹脚本\考勤表格导出\休假出差表_20250603.xlsx'
DEFAULT_OUTPUT_DIR = r'C:\Users\fucha\PycharmProjects\python\公司\cursor项目\6.老爹脚本\考勤表格导出\生成表格'

# 创建配置文件夹和默认配置
if not os.path.exists(CONFIG_DIR):
    os.makedirs(CONFIG_DIR)
if not os.path.exists(CONFIG_PATH):
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump({
            'detail_path': DEFAULT_DETAIL_PATH,
            'leave_path': DEFAULT_LEAVE_PATH,
            'output_dir': DEFAULT_OUTPUT_DIR
        }, f, ensure_ascii=False, indent=2)

def load_config():
    # 加载配置参数
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print('配置文件读取失败:', e)
        return {
            'detail_path': DEFAULT_DETAIL_PATH,
            'leave_path': DEFAULT_LEAVE_PATH,
            'output_dir': DEFAULT_OUTPUT_DIR
        }

def save_config(cfg):
    # 保存配置参数
    try:
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print('配置文件保存失败:', e)

# 工具函数：星期数字转中文
WEEKDAY_MAP = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
def weekday_to_cn(date_str):
    # 输入日期字符串，返回中文星期
    dt = pd.to_datetime(date_str)
    return WEEKDAY_MAP[dt.weekday()]

def parse_time(tstr):
    # 解析时间字符串，返回datetime.time对象
    try:
        if tstr == '?' or not tstr:
            return None
        return datetime.strptime(tstr, '%H:%M').time()
    except:
        return None

def time_diff_minutes(t1, t2):
    # 计算两个时间的分钟差
    if not t1 or not t2:
        return None
    dt1 = datetime.combine(datetime.today(), t1)
    dt2 = datetime.combine(datetime.today(), t2)
    return int((dt1 - dt2).total_seconds() // 60)

def get_days_in_month(year, month):
    # 获取某年某月的所有日期字符串
    days = calendar.monthrange(year, month)[1]
    return [datetime(year, month, d).strftime('%Y/%m/%d') for d in range(1, days+1)]

def read_detail_table(path):
    # 读取日明细报表，返回DataFrame
    try:
        df = pd.read_excel(path, header=2, dtype=str)
        # 只保留有人员编号的行
        df = df[df['人员编号'].notna() & df['人员编号'].astype(str).str.match(r'\d{5}', na=False)]
        return df
    except Exception as e:
        raise Exception(f'日明细报表读取失败: {e}')

def read_leave_table(path):
    # 读取休假出差表，返回DataFrame
    try:
        df = pd.read_excel(path, dtype=str)
        # 只保留有人员编号的行
        df = df[df['人员编号'].notna() & df['人员编号'].astype(str).str.match(r'\d{5}', na=False)]
        return df
    except Exception as e:
        raise Exception(f'休假出差表读取失败: {e}')

def build_attendance_dict(detail_df):
    # 构建员工考勤字典
    att_dict = {}
    for _, row in detail_df.iterrows():
        pid = str(row['人员编号'])
        name = row['姓名']
        dept = row['名称']
        date = row.get('Unnamed: 4', '')  # 可能为考勤日期
        if not date:
            date = row.get('考勤日期', '')
        record = row.get('打卡数据', '')
        if pid not in att_dict:
            att_dict[pid] = {'name': name, 'dept': dept, 'records': {}}
        att_dict[pid]['records'][date] = record
    return att_dict

def build_leave_dict(leave_df):
    # 构建员工请假字典
    leave_dict = {}
    for _, row in leave_df.iterrows():
        pid = str(row['人员编号'])
        reason = row['原因']
        start = row['起始日期']
        end = row['结束日期']
        half = row.get('半天', '')
        if pid not in leave_dict:
            leave_dict[pid] = []
        leave_dict[pid].append({'reason': reason, 'start': start, 'end': end, 'half': half})
    return leave_dict

def generate_attendance_table(pid, att_info, leave_info, year, month, output_dir):
    # 生成单个员工考勤表
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    # 表头
    title = f"{year}年{month}月考勤表"
    name = att_info['name']
    dept = att_info['dept']
    days = get_days_in_month(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = '考勤表'
    ws.append([title])
    ws.append(['姓名', name, '部门', dept])
    ws.append(['星期', '日期', '上班', '下班', '迟到', '早退', '加班', '备注'])
    for d in days:
        weekday = weekday_to_cn(d)
        record = att_info['records'].get(d.replace('/', '-'), '')
        on_time, off_time = '', ''
        if record:
            if '-' in record:
                on_time, off_time = record.split('-')
                on_time = on_time.strip(';')
                off_time = off_time.strip(';')
            else:
                on_time = record.strip(';')
        # 迟到、早退、加班、备注逻辑
        late, early, overtime, note = '', '', '', ''
        on_dt = parse_time(on_time)
        off_dt = parse_time(off_time)
        # 迟到
        if on_dt:
            late_min = time_diff_minutes(on_dt, datetime.strptime('09:15', '%H:%M').time())
            if late_min > 0:
                late = str(late_min)
        # 早退/加班
        if on_dt and off_dt:
            work_min = time_diff_minutes(off_dt, on_dt)
            if not late:
                if work_min < 480:
                    early = str(480 - work_min)
                elif work_min > 480:
                    overtime = str(work_min - 480)
            else:
                cmp = time_diff_minutes(off_dt, datetime.strptime('17:45', '%H:%M').time())
                if cmp > 0:
                    overtime = str(cmp)
                elif cmp < 0:
                    early = str(-cmp)
        # 备注
        if not on_time or on_time == '?':
            note = '上班缺卡'
        if not off_time or off_time == '?':
            note = '下班缺卡' if note == '' else note + '，下班缺卡'
        if not on_time and not off_time:
            note = '旷工'
        if late and early:
            note = '迟到+早退'
        elif late:
            note = '迟到'
        elif early:
            note = '早退'
        ws.append([weekday, d, on_time, off_time, late, early, overtime, note])
    # 休假/出差备注覆盖
    if leave_info:
        for leave in leave_info:
            s = pd.to_datetime(leave['start']).strftime('%Y/%m/%d')
            e = pd.to_datetime(leave['end']).strftime('%Y/%m/%d')
            for i, row in enumerate(ws.iter_rows(min_row=4, max_row=3+len(days)), start=4):
                date_cell = row[1].value
                if s <= date_cell <= e:
                    ws.cell(row=i, column=8, value=leave['reason'])
    # 保存表格
    fname = f"{name}_{year}{str(month).zfill(2)}_考勤表.xlsx"
    out_path = os.path.join(output_dir, fname)
    wb.save(out_path)
    return out_path

def open_file_or_dir(path):
    # 打开文件或文件夹
    try:
        if os.path.isfile(path):
            os.startfile(path)
        elif os.path.isdir(path):
            os.startfile(path)
    except Exception as e:
        print('打开文件/文件夹失败:', e)

# 主UI类
class AttendanceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('考勤表格自动生成工具')
        self.geometry('900x600')
        self.resizable(False, False)
        self.config = load_config()
        # 输入路径
        self.detail_path = ctk.StringVar(value=self.config['detail_path'])
        self.leave_path = ctk.StringVar(value=self.config['leave_path'])
        self.output_dir = ctk.StringVar(value=self.config['output_dir'])
        # 输入框
        self.pid_var = ctk.StringVar()
        self.name_var = ctk.StringVar()
        # 进度与反馈
        self.status_var = ctk.StringVar(value='等待操作...')
        self.preview_df = None
        self.preview_table = None
        self.table_files = []
        self.current_table_idx = 0
        self.create_widgets()

    def create_widgets(self):
        # 路径选择
        frame1 = ctk.CTkFrame(self)
        frame1.pack(fill='x', padx=10, pady=5)
        ctk.CTkLabel(frame1, text='日明细报表:').pack(side='left')
        ctk.CTkEntry(frame1, textvariable=self.detail_path, width=400).pack(side='left', padx=5)
        ctk.CTkButton(frame1, text='选择文件', command=self.select_detail).pack(side='left', padx=5)
        ctk.CTkLabel(frame1, text='休假出差表:').pack(side='left', padx=10)
        ctk.CTkEntry(frame1, textvariable=self.leave_path, width=400).pack(side='left', padx=5)
        ctk.CTkButton(frame1, text='选择文件', command=self.select_leave).pack(side='left', padx=5)
        # 输入框
        frame2 = ctk.CTkFrame(self)
        frame2.pack(fill='x', padx=10, pady=5)
        ctk.CTkLabel(frame2, text='员工编号:').pack(side='left')
        ctk.CTkEntry(frame2, textvariable=self.pid_var, width=120).pack(side='left', padx=5)
        ctk.CTkLabel(frame2, text='员工姓名:').pack(side='left', padx=10)
        ctk.CTkEntry(frame2, textvariable=self.name_var, width=120).pack(side='left', padx=5)
        # 进度与反馈
        frame3 = ctk.CTkFrame(self)
        frame3.pack(fill='x', padx=10, pady=5)
        ctk.CTkLabel(frame3, textvariable=self.status_var, text_color='blue').pack(side='left')
        # 表格预览
        self.preview_table = tk.Text(self, height=20, width=120, font=('Consolas', 10))
        self.preview_table.pack(padx=10, pady=5)
        # 表格切换
        frame4 = ctk.CTkFrame(self)
        frame4.pack(fill='x', padx=10, pady=5)
        ctk.CTkButton(frame4, text='上一表格', command=self.prev_table).pack(side='left', padx=5)
        ctk.CTkButton(frame4, text='下一表格', command=self.next_table).pack(side='left', padx=5)
        # 输出路径
        frame5 = ctk.CTkFrame(self)
        frame5.pack(fill='x', padx=10, pady=5)
        ctk.CTkLabel(frame5, text='输出目录:').pack(side='left')
        ctk.CTkEntry(frame5, textvariable=self.output_dir, width=400).pack(side='left', padx=5)
        ctk.CTkButton(frame5, text='选择目录', command=self.select_output).pack(side='left', padx=5)
        # 生成按钮
        ctk.CTkButton(self, text='开始生成', command=self.start_generate).pack(pady=10)

    def select_detail(self):
        # 选择日明细报表
        path = filedialog.askopenfilename(title='选择日明细报表', filetypes=[('Excel', '*.xlsx')])
        if path:
            self.detail_path.set(path)

    def select_leave(self):
        # 选择休假出差表
        path = filedialog.askopenfilename(title='选择休假出差表', filetypes=[('Excel', '*.xlsx')])
        if path:
            self.leave_path.set(path)

    def select_output(self):
        # 选择输出目录
        path = filedialog.askdirectory(title='选择输出目录')
        if path:
            self.output_dir.set(path)

    def prev_table(self):
        # 上一表格
        if self.table_files and self.current_table_idx > 0:
            self.current_table_idx -= 1
            self.show_table_preview(self.table_files[self.current_table_idx])

    def next_table(self):
        # 下一表格
        if self.table_files and self.current_table_idx < len(self.table_files) - 1:
            self.current_table_idx += 1
            self.show_table_preview(self.table_files[self.current_table_idx])

    def show_table_preview(self, file_path):
        # 预览表格内容
        try:
            df = pd.read_excel(file_path, header=None)
            self.preview_table.delete('1.0', tk.END)
            for row in df.values.tolist():
                self.preview_table.insert(tk.END, '\t'.join([str(x) for x in row]) + '\n')
        except Exception as e:
            self.preview_table.delete('1.0', tk.END)
            self.preview_table.insert(tk.END, f'表格预览失败: {e}')

    def start_generate(self):
        # 启动生成线程，避免UI卡死
        threading.Thread(target=self.generate_tables, daemon=True).start()

    def generate_tables(self):
        # 生成考勤表主逻辑
        try:
            self.status_var.set('正在读取表格...')
            pid = self.pid_var.get().strip()
            name = self.name_var.get().strip()
            if pid and name:
                messagebox.showwarning('输入错误', '员工编号和员工姓名仅允许输入一个')
                self.status_var.set('等待操作...')
                return
            # 读取表格
            detail_df = read_detail_table(self.detail_path.get())
            leave_df = read_leave_table(self.leave_path.get())
            att_dict = build_attendance_dict(detail_df)
            leave_dict = build_leave_dict(leave_df)
            # 获取年月
            if not detail_df.empty:
                first_date = detail_df['Unnamed: 4'].iloc[0]
                year, month = pd.to_datetime(first_date).year, pd.to_datetime(first_date).month
            else:
                self.status_var.set('日明细报表无数据')
                return
            # 输出目录
            out_dir = self.output_dir.get()
            if not os.path.exists(out_dir):
                os.makedirs(out_dir)
            self.table_files = []
            # 生成表格
            for k, v in att_dict.items():
                if (pid and k != pid) or (name and v['name'] != name):
                    continue
                leave_info = leave_dict.get(k, [])
                fpath = generate_attendance_table(k, v, leave_info, year, month, out_dir)
                self.table_files.append(fpath)
            if not self.table_files:
                self.status_var.set('未找到符合条件的员工')
                return
            self.current_table_idx = 0
            self.show_table_preview(self.table_files[0])
            self.status_var.set(f'生成完成，共{len(self.table_files)}个表格')
            # 打开文件夹或单个表格
            if len(self.table_files) == 1:
                open_file_or_dir(self.table_files[0])
            else:
                open_file_or_dir(out_dir)
            # 保存配置
            save_config({
                'detail_path': self.detail_path.get(),
                'leave_path': self.leave_path.get(),
                'output_dir': self.output_dir.get()
            })
        except Exception as e:
            self.status_var.set('生成失败，请检查日志')
            with open('error.log', 'a', encoding='utf-8') as f:
                f.write(traceback.format_exc())
            messagebox.showerror('错误', f'生成失败: {e}')

if __name__ == '__main__':
    # 启动UI
    app = AttendanceApp()
    app.mainloop() 