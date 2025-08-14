from openpyxl import load_workbook

xlsx = "C:\\Users\\fucha\\Desktop\爬取误差比对.xlsx"
wb = load_workbook(xlsx)  # 加载工作簿
ws = wb.active  # 获取活动工作表

# 数组拆分比对
def compare_actor(entry1, entry2):
    if not entry1 or not entry2:
        return 0
    array1 = [item.strip() for item in entry1.split(',')]
    array2 = [item.strip() for item in entry2.split('/')]
    count = 0
    for item1 in array1:
        for item2 in array2:
            if item1 == item2:
                count += 1
    return count
# 演员比对
def diff_actor(ws):
    actor1 = ws.cell(row=row_index, column=3).value
    actor2 = ws.cell(row=row_index, column=13).value
    count = compare_actor(actor1, actor2)
    print(f"第 {row_index + 1} 行的相同元素个数: {count}")
    ws.cell(row=row_index, column=21, value=count)  # 将计数器的值写入第 14 列

for row_index in range(2, ws.max_row + 1):  # 从第二行开始，跳过标题行
    n = ws.cell(row=row_index, column=1).value
    if n:
        diff_actor(ws)


wb.save(xlsx)  # 保存工作簿